"""
Inspect the transition section (rows 276-300) and the flat FDH section.
Also check the row pattern: are rows always in pairs (FDH + next row)?
"""
import openpyxl
from openpyxl.utils import get_column_letter

FILE_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build\CompleteEncinitasView_PBI.xlsx"

wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False)

ws_sv = wb_values["Summary"]
ws_sf = wb_formulas["Summary"]

print("=" * 80)
print("TRANSITION SECTION (rows 276-310)")
print("=" * 80)

for row in range(276, 311):
    row_content = []
    for col in range(1, 11):
        val_cell = ws_sv.cell(row=row, column=col)
        form_cell = ws_sf.cell(row=row, column=col)
        letter = get_column_letter(col)
        val = val_cell.value
        form = form_cell.value
        is_formula = isinstance(form, str) and form.startswith("=")
        if val is not None or is_formula:
            indent = val_cell.alignment.indent or 0
            entry = f"{letter}={repr(val)}"
            if is_formula:
                entry += f" [F]"
            if indent:
                entry += f" [i={indent}]"
            row_content.append(entry)
    if row_content:
        print(f"  Row {row}: {' | '.join(row_content)}")
    else:
        print(f"  Row {row}: (empty)")

# Analyze the flat section (280+) pattern
print("\n\n" + "=" * 80)
print("FLAT FDH SECTION PATTERN (rows 280-381)")
print("=" * 80)
print("This section has NO Phase numbers. Each FDH appears as 2 rows:")
print("  Row N: FDH Name (indent=1) with counts and G formula")
print("  Row N+1: counts repeated (no B value, no G formula result)")
print()

# Verify this pattern
fdh_rows = []
for row in range(280, 382):
    b_val = ws_sv.cell(row=row, column=2).value
    indent = ws_sv.cell(row=row, column=2).alignment.indent or 0
    if b_val is not None and indent == 1:
        fdh_rows.append(row)

print(f"FDH rows in flat section: {len(fdh_rows)}")
print("Verifying each FDH row is followed by a duplicate-counts row:")
for fr in fdh_rows[:10]:
    next_row = fr + 1
    b_next = ws_sv.cell(row=next_row, column=2).value
    e_fdh = ws_sv.cell(row=fr, column=5).value
    e_next = ws_sv.cell(row=next_row, column=5).value
    g_fdh_form = ws_sf.cell(row=fr, column=7).value
    g_next_form = ws_sf.cell(row=next_row, column=7).value
    print(f"  Row {fr}: B={ws_sv.cell(row=fr, column=2).value}, E={e_fdh}")
    print(f"  Row {next_row}: B={b_next}, E={e_next}, G_formula={isinstance(g_next_form, str) and g_next_form.startswith('=')}")
    print()

# Check row 280 specifically - what is this empty-string Phase?
print("\n\n" + "=" * 80)
print("ROW 280 - THE BLANK PHASE ROW")
print("=" * 80)
for col in range(1, 11):
    val = ws_sv.cell(row=280, column=col).value
    letter = get_column_letter(col)
    indent = ws_sv.cell(row=280, column=col).alignment.indent or 0
    if val is not None:
        print(f"  {letter}: value={repr(val)}, indent={indent}")

# The flat section seems to be addresses where Phase is blank/empty
# Let's verify by checking indent=0 rows in the flat section
print("\n\nIndent=0 rows in 280-381:")
for row in range(280, 382):
    cell = ws_sv.cell(row=row, column=2)
    val = cell.value
    indent = cell.alignment.indent or 0
    if val is not None and indent == 0:
        print(f"  Row {row}: value={repr(val)}")


# Finally, check the structure of how the "2-row per FDH" pattern works in
# the HIERARCHICAL section (rows 6-279) -- do date rows also repeat counts?
print("\n\n" + "=" * 80)
print("HIERARCHICAL SECTION: DO DATE ROWS ALSO HAVE FOLLOWING ROWS?")
print("=" * 80)
# Look at Phase 4 (row 9) for example
for row in range(9, 17):
    b_val = ws_sv.cell(row=row, column=2).value
    indent = ws_sv.cell(row=row, column=2).alignment.indent or 0
    c_val = ws_sv.cell(row=row, column=3).value
    d_val = ws_sv.cell(row=row, column=4).value
    e_val = ws_sv.cell(row=row, column=5).value
    g_val = ws_sv.cell(row=row, column=7).value
    g_form = ws_sf.cell(row=row, column=7).value
    has_g_formula = isinstance(g_form, str) and g_form.startswith("=")
    print(f"  Row {row}: B={repr(b_val)} [i={indent}] | C={c_val} | D={d_val} | E={e_val} | "
          f"G={g_val} [formula={has_g_formula}]")

# And Phase 5 (row 17)
print("\nPhase 5:")
for row in range(17, 24):
    b_val = ws_sv.cell(row=row, column=2).value
    indent = ws_sv.cell(row=row, column=2).alignment.indent or 0
    c_val = ws_sv.cell(row=row, column=3).value
    d_val = ws_sv.cell(row=row, column=4).value
    e_val = ws_sv.cell(row=row, column=5).value
    g_val = ws_sv.cell(row=row, column=7).value
    g_form = ws_sf.cell(row=row, column=7).value
    has_g_formula = isinstance(g_form, str) and g_form.startswith("=")
    print(f"  Row {row}: B={repr(b_val)} [i={indent}] | C={c_val} | D={d_val} | E={e_val} | "
          f"G={g_val} [formula={has_g_formula}]")

wb_values.close()
wb_formulas.close()
print("\n\nDone.")
