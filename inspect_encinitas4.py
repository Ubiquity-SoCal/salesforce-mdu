"""
Final details: fill/theme colors, conditional formatting details,
and the tail section of Summary where structure changes.
"""
import openpyxl
from openpyxl.utils import get_column_letter

FILE_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build\CompleteEncinitasView_PBI.xlsx"

wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False)

ws_sv = wb_values["Summary"]
ws_sf = wb_formulas["Summary"]

# 1. Detailed fill/font color info for key cells
print("=" * 80)
print("DETAILED COLOR/FILL INFO FOR KEY CELLS")
print("=" * 80)

for row in [2, 4, 5, 6, 7, 8]:
    for col in range(1, 11):
        cell = ws_sv.cell(row=row, column=col)
        letter = get_column_letter(col)
        if cell.value is not None:
            font = cell.font
            fill = cell.fill
            print(f"\nRow {row} {letter}: value={repr(cell.value)}")
            print(f"  Font: name={font.name}, size={font.size}, bold={font.bold}, italic={font.italic}")
            print(f"  Font color: type={font.color.type if font.color else None}, "
                  f"theme={font.color.theme if font.color else None}, "
                  f"rgb={font.color.rgb if font.color and hasattr(font.color, 'rgb') else None}, "
                  f"tint={font.color.tint if font.color else None}, "
                  f"indexed={font.color.indexed if font.color else None}")
            print(f"  Fill: patternType={fill.patternType}")
            if fill.fgColor:
                print(f"  Fill fgColor: type={fill.fgColor.type}, theme={fill.fgColor.theme}, "
                      f"rgb={fill.fgColor.rgb if hasattr(fill.fgColor, 'rgb') else None}, "
                      f"tint={fill.fgColor.tint}, indexed={fill.fgColor.indexed}")
            if fill.bgColor:
                print(f"  Fill bgColor: type={fill.bgColor.type}, theme={fill.bgColor.theme}, "
                      f"rgb={fill.bgColor.rgb if hasattr(fill.bgColor, 'rgb') else None}, "
                      f"tint={fill.bgColor.tint}, indexed={fill.bgColor.indexed}")

# 2. The tail section where hierarchy changes (rows 355-381)
print("\n\n" + "=" * 80)
print("SUMMARY ROWS 340-381: TAIL SECTION DETAIL")
print("=" * 80)

for row in range(340, 382):
    row_content = []
    for col in range(1, 11):
        val_cell = ws_sv.cell(row=row, column=col)
        form_cell = ws_sf.cell(row=row, column=col)
        letter = get_column_letter(col)
        val = val_cell.value
        form = form_cell.value
        is_formula = isinstance(form, str) and form.startswith("=")
        if val is not None or is_formula:
            indent = val_cell.alignment.indent or 0
            bold = val_cell.font.bold
            entry = f"{letter}={repr(val)}"
            if is_formula:
                entry += f" [F:{form}]"
            if indent:
                entry += f" [i={indent}]"
            if bold:
                entry += " [BOLD]"
            row_content.append(entry)
    if row_content:
        print(f"  Row {row}: {' | '.join(row_content)}")

# 3. Check: around row 350+ where it seems Phase numbers disappear
# and only FDH rows exist
print("\n\n" + "=" * 80)
print("PHASE STRUCTURE - WHERE PHASES END AND FDH-ONLY BEGINS")
print("=" * 80)

last_phase_row = None
for row in range(6, ws_sv.max_row + 1):
    cell = ws_sv.cell(row=row, column=2)
    val = cell.value
    indent = cell.alignment.indent if cell.alignment.indent else 0
    if indent == 0 and isinstance(val, int):
        last_phase_row = row
        last_phase_val = val

print(f"Last Phase row: {last_phase_row}, Phase value: {last_phase_val}")

# Show 5 rows before and 10 after the last phase
print(f"\nRows around last phase (row {last_phase_row}):")
for row in range(last_phase_row - 3, min(last_phase_row + 15, ws_sv.max_row + 1)):
    cell = ws_sv.cell(row=row, column=2)
    val = cell.value
    indent = cell.alignment.indent if cell.alignment.indent else 0
    if val is not None:
        print(f"  Row {row}: indent={indent}, value={repr(val)}")

# 4. What's in the "Grand Total" row?
print("\n\n" + "=" * 80)
print("GRAND TOTAL ROW (381)")
print("=" * 80)
for col in range(1, 11):
    cell = ws_sv.cell(row=381, column=col)
    letter = get_column_letter(col)
    val = cell.value
    if val is not None:
        font = cell.font
        print(f"  {letter}: value={repr(val)}, bold={font.bold}, size={font.size}")

# 5. Conditional formatting detail
print("\n\n" + "=" * 80)
print("CONDITIONAL FORMATTING DETAIL")
print("=" * 80)
for cf in ws_sv.conditional_formatting:
    print(f"Cell ranges: {cf.cells}")
    for rule in cf.rules:
        print(f"  Rule type: {rule.type}")
        print(f"  Operator: {rule.operator}")
        print(f"  Formula: {rule.formula}")
        print(f"  Priority: {rule.priority}")
        if rule.dxf:
            dxf = rule.dxf
            if dxf.font:
                f = dxf.font
                print(f"  DXF Font: bold={f.bold}, color type={f.color.type if f.color else None}, "
                      f"theme={f.color.theme if f.color else None}, rgb={f.color.rgb if f.color and hasattr(f.color, 'rgb') else None}")
            if dxf.fill:
                fl = dxf.fill
                print(f"  DXF Fill: pattern={fl.patternType}")
                if fl.fgColor:
                    print(f"    fgColor: type={fl.fgColor.type}, theme={fl.fgColor.theme}, "
                          f"rgb={fl.fgColor.rgb if hasattr(fl.fgColor, 'rgb') else None}")

# 6. G4:G5 merged cell content
print("\n\n" + "=" * 80)
print("MERGED CELL G4:G5")
print("=" * 80)
g4 = ws_sv.cell(row=4, column=7)
g5 = ws_sv.cell(row=5, column=7)
g4f = ws_sf.cell(row=4, column=7)
g5f = ws_sf.cell(row=5, column=7)
print(f"G4 value: {repr(g4.value)}, formula: {repr(g4f.value)}")
print(f"G5 value: {repr(g5.value)}, formula: {repr(g5f.value)}")
print(f"G4 alignment: h={g4.alignment.horizontal}, v={g4.alignment.vertical}, wrap={g4.alignment.wrap_text}")

# 7. Row 380 has no G formula? Check where G formulas end
print("\n\n" + "=" * 80)
print("G COLUMN FORMULA EXTENT")
print("=" * 80)
last_g_formula_row = None
for row in range(6, ws_sv.max_row + 1):
    form = ws_sf.cell(row=row, column=7).value
    if isinstance(form, str) and form.startswith("="):
        last_g_formula_row = row
print(f"Last row with G formula: {last_g_formula_row}")

# Show around the transition
print(f"\nRows {last_g_formula_row - 2} to {last_g_formula_row + 3}:")
for row in range(last_g_formula_row - 2, last_g_formula_row + 4):
    g_form = ws_sf.cell(row=row, column=7).value
    b_val = ws_sv.cell(row=row, column=2).value
    is_f = isinstance(g_form, str) and g_form.startswith("=")
    print(f"  Row {row}: B={repr(b_val)}, G formula={is_f}")


wb_values.close()
wb_formulas.close()
print("\n\nDone.")
