"""
Enrich Carlsbad and Solana Beach PBI data with Invoice Support data,
matching on Circuit ID -- same process as the completed Encinitas file.

Pulls in: Address type, Phase, Serviceable from date
Then builds a Summary sheet with real Excel PivotTables.
"""

import os

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build"

CITIES = [
    {
        "name": "Carlsbad",
        "pbi_file": "ComprehensiveCarlsbadView_PBI.xlsx",
        "invoice_file": "Carlsbad Invoice Support_February - March ViewCopy.xlsm",
        "invoice_sheet": "Serviceable Doors in Jan",
        "invoice_header_row": 2,
        "output_file": "CompleteCarlsbadView_PBI.xlsx",
    },
    {
        "name": "Solana Beach",
        "pbi_file": "ComprehensiveSolanaBeachView_PBI.xlsx",
        "invoice_file": "Solana Beach Invoice Support_February_MarchView.xlsm",
        "invoice_sheet": "Serviceable Doors - in Jan ",
        "invoice_header_row": 2,
        "output_file": "CompleteSolanaBeachView_PBI.xlsx",
    },
]


def read_invoice_data(filepath, sheet_name, header_row):
    """Read invoice support data and build a Circuit ID -> row dict."""
    print(f"  Reading invoice: {os.path.basename(filepath)} [{sheet_name}]")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    # Read headers
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        headers.append(val)

    # Find key columns
    def find_col(name):
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i
        return None

    cid_idx = find_col("Circuit ID")
    atype_idx = find_col("Address type")
    phase_idx = find_col("phase")
    svc_date_idx = find_col("Serviceable from date")

    print(f"    Columns found: Circuit ID={cid_idx}, Address type={atype_idx}, Phase={phase_idx}, Serviceable from date={svc_date_idx}")

    if cid_idx is None:
        print("    ERROR: Circuit ID column not found!")
        wb.close()
        return {}

    # Build lookup: Circuit ID -> {address_type, phase, serviceable_from_date}
    lookup = {}
    for row in range(header_row + 1, ws.max_row + 1):
        cid = ws.cell(row=row, column=cid_idx + 1).value
        if cid:
            cid = str(cid).strip()
            lookup[cid] = {
                "Address type": ws.cell(row=row, column=atype_idx + 1).value if atype_idx is not None else None,
                "Phase": ws.cell(row=row, column=phase_idx + 1).value if phase_idx is not None else None,
                "Serviceable from date": ws.cell(row=row, column=svc_date_idx + 1).value if svc_date_idx is not None else None,
            }

    print(f"    Loaded {len(lookup)} invoice records")
    wb.close()
    return lookup


def col_num_to_letter(n):
    """Convert 1-based column number to Excel column letter (1='A', 27='AA')."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def create_summary_pivots(filepath, city_name, fdh_col_letter, act_col_letter):
    """Create real Excel PivotTables on a Summary sheet using COM automation."""
    import win32com.client

    print("  Creating PivotTables via Excel...")

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        abs_path = os.path.abspath(filepath)
        wb = excel.Workbooks.Open(abs_path)
        ws_data = wb.Sheets("PowerBI_Data")

        # Delete existing Summary if present
        for i in range(wb.Sheets.Count, 0, -1):
            if wb.Sheets(i).Name == "Summary":
                wb.Sheets(i).Delete()

        # Create Summary sheet at end
        ws = wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count))
        ws.Name = "Summary"

        # --- Title ---
        ws.Range("B2").Value = city_name.upper()
        ws.Range("B2").Font.Bold = True
        ws.Range("B2").Font.Color = 0xC47244   # RGB(68,114,196) blue in BGR
        ws.Range("B2").Interior.Color = 0xCCF2FF  # RGB(255,242,204) light gold in BGR

        # --- Data range (R1C1 string reference) ---
        last_row = ws_data.Cells(ws_data.Rows.Count, 1).End(-4162).Row   # xlUp
        last_col = ws_data.Cells(1, ws_data.Columns.Count).End(-4159).Column  # xlToLeft
        source = f"PowerBI_Data!R1C1:R{last_row}C{last_col}"

        # --- PivotCache ---
        cache = wb.PivotCaches().Create(SourceType=1, SourceData=source)  # xlDatabase=1

        # ==========================================
        # LEFT PIVOT: Phase > FDH Name > Svc Date
        #   Column field: Address Status
        #   Value: Count of Address
        # ==========================================
        pt1 = cache.CreatePivotTable(
            TableDestination=ws.Range("B4"),
            TableName="AddressPivot"
        )

        # Row fields
        pf = pt1.PivotFields("Phase")
        pf.Orientation = 1   # xlRowField
        pf.Position = 1

        pf = pt1.PivotFields("FDH Name")
        pf.Orientation = 1
        pf.Position = 2

        pf = pt1.PivotFields("Serviceable from date")
        pf.Orientation = 1
        pf.Position = 3

        # Column field
        pf = pt1.PivotFields("Address Status")
        pf.Orientation = 2   # xlColumnField

        # Value: Count of Address
        pt1.AddDataField(pt1.PivotFields("Address"), "Count of Address", -4112)  # xlCount

        # Style (blue banding matching Encinitas screenshot)
        pt1.TableStyle2 = "PivotStyleLight16"

        # --- Determine left pivot extent ---
        r1 = pt1.TableRange1
        pivot_first_row = r1.Row
        pivot_last_row = r1.Row + r1.Rows.Count - 1
        pivot_last_col = r1.Column + r1.Columns.Count - 1

        # ==========================================
        # XLOOKUP: PowerBI FDH Activation Date
        #   Placed 2 columns after left pivot
        # ==========================================
        xlu_col = pivot_last_col + 2
        xlu = col_num_to_letter(xlu_col)

        # Header (merged across 2 rows to match Encinitas)
        ws.Range(f"{xlu}{pivot_first_row}:{xlu}{pivot_first_row + 1}").Merge()
        ws.Range(f"{xlu}{pivot_first_row}").Value = "PowerBI FDH Activation Date"
        ws.Range(f"{xlu}{pivot_first_row}").HorizontalAlignment = -4108  # xlCenter
        ws.Range(f"{xlu}{pivot_first_row}").VerticalAlignment = -4108
        ws.Range(f"{xlu}{pivot_first_row}").WrapText = True
        ws.Columns(xlu).ColumnWidth = 18

        # XLOOKUP formulas for each data row
        data_start = pivot_first_row + 2  # skip header rows
        for r in range(data_start, pivot_last_row + 1):
            ws.Range(f"{xlu}{r}").Formula = (
                f'=IFERROR(_xlfn.XLOOKUP(B{r},'
                f'PowerBI_Data!{fdh_col_letter}:{fdh_col_letter},'
                f'PowerBI_Data!{act_col_letter}:{act_col_letter},'
                f'"",0,1),"")'
            )
            ws.Range(f"{xlu}{r}").HorizontalAlignment = -4108

        # --- Note ---
        note_col = col_num_to_letter(xlu_col + 2)
        ws.Range(f"{note_col}2").Value = "*data pulled is only from serviceable (Build)"
        ws.Range(f"{note_col}2").Font.Italic = True
        ws.Range(f"{note_col}2").Font.Color = 0x0000FF  # Red in BGR

        # ==========================================
        # RIGHT PIVOT: FDH Name > FDH Activation Date
        #   Value: Count of Address
        # ==========================================
        right_col = col_num_to_letter(xlu_col + 3)

        pt2 = cache.CreatePivotTable(
            TableDestination=ws.Range(f"{right_col}4"),
            TableName="FDHPivot"
        )

        pf = pt2.PivotFields("FDH Name")
        pf.Orientation = 1   # xlRowField
        pf.Position = 1

        pf = pt2.PivotFields("FDH Activation Date")
        pf.Orientation = 1
        pf.Position = 2

        pt2.AddDataField(pt2.PivotFields("Address"), "Count of Address", -4112)

        pt2.TableStyle2 = "PivotStyleLight16"

        # --- Column widths ---
        ws.Columns("A").ColumnWidth = 4
        ws.Columns("B").ColumnWidth = 55

        # Widen right pivot label column
        ws.Columns(right_col).ColumnWidth = 45

        wb.Save()
        print("  PivotTables created successfully!")

    except Exception as e:
        print(f"  ERROR creating PivotTables: {e}")
        import traceback
        traceback.print_exc()
    finally:
        try:
            wb.Close(False)
        except Exception:
            pass
        excel.Quit()


def enrich_pbi(city_config):
    """Enrich a PBI file with invoice data and build Summary with PivotTables."""
    name = city_config["name"]
    pbi_path = os.path.join(BASE_DIR, city_config["pbi_file"])
    invoice_path = os.path.join(BASE_DIR, city_config["invoice_file"])
    output_path = os.path.join(BASE_DIR, city_config["output_file"])

    print(f"\n{'='*60}")
    print(f"Processing {name}")
    print(f"{'='*60}")

    # Read invoice lookup
    invoice_lookup = read_invoice_data(
        invoice_path,
        city_config["invoice_sheet"],
        city_config["invoice_header_row"],
    )

    # Open PBI workbook
    print(f"  Reading PBI: {city_config['pbi_file']}")
    wb = openpyxl.load_workbook(pbi_path)
    ws = wb.active

    # Rename sheet to PowerBI_Data
    ws.title = "PowerBI_Data"

    # Read headers from row 1
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col).value)

    # Find Circuit ID column in PBI
    cid_col = None
    for i, h in enumerate(headers):
        if h and "circuit id" in str(h).lower():
            cid_col = i
            break

    if cid_col is None:
        print("  ERROR: Circuit ID not found in PBI data!")
        return

    print(f"  Circuit ID at column {cid_col} ({headers[cid_col]})")

    # Add 3 new columns: Address type, Phase, Serviceable from date
    new_cols = ["Address type", "Phase", "Serviceable from date"]
    start_col = ws.max_column + 1

    # Write new headers
    header_font = Font(bold=True)
    for i, col_name in enumerate(new_cols):
        cell = ws.cell(row=1, column=start_col + i, value=col_name)
        cell.font = header_font

    # Enrich each row
    matched = 0
    unmatched = 0
    total = 0
    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row=row, column=cid_col + 1).value
        total += 1
        if cid:
            cid = str(cid).strip()
            invoice_row = invoice_lookup.get(cid)
            if invoice_row:
                matched += 1
                ws.cell(row=row, column=start_col, value=invoice_row["Address type"])
                ws.cell(row=row, column=start_col + 1, value=invoice_row["Phase"])
                ws.cell(row=row, column=start_col + 2, value=invoice_row["Serviceable from date"])
            else:
                unmatched += 1
        else:
            unmatched += 1

    print(f"  Enriched: {matched}/{total} matched, {unmatched} unmatched")

    # Find key columns for Match Report and XLOOKUP
    fdh_col = None
    status_col = None
    activation_col = None
    for i, h in enumerate(headers):
        if h and "fdh name" in str(h).lower():
            fdh_col = i
        if h and "address status" in str(h).lower():
            status_col = i
        if h and "fdh activation date" in str(h).lower():
            activation_col = i

    fdh_col_letter = get_column_letter(fdh_col + 1) if fdh_col is not None else "D"
    act_col_letter = get_column_letter(activation_col + 1) if activation_col is not None else "C"

    # === MATCH REPORT SHEET ===
    bold = Font(bold=True)
    if "Match Report" in wb.sheetnames:
        del wb["Match Report"]
    match_ws = wb.create_sheet("Match Report")

    match_ws.cell(row=1, column=1, value="Circuit ID Match Report").font = Font(bold=True, size=14)
    match_ws.cell(row=3, column=1, value="Metric").font = bold
    match_ws.cell(row=3, column=2, value="Count").font = bold
    match_ws.cell(row=4, column=1, value="PBI records")
    match_ws.cell(row=4, column=2, value=total)
    match_ws.cell(row=5, column=1, value="Matched to Invoice")
    match_ws.cell(row=5, column=2, value=matched)
    match_ws.cell(row=6, column=1, value="Not in Invoice")
    match_ws.cell(row=6, column=2, value=unmatched)
    match_ws.cell(row=7, column=1, value="Match Rate")
    match_ws.cell(row=7, column=2, value=f"{matched/total*100:.1f}%" if total else "N/A")
    match_ws.cell(row=8, column=1, value="Invoice records")
    match_ws.cell(row=8, column=2, value=len(invoice_lookup))

    # PBI circuits not in invoice
    pbi_not_in_invoice = []
    invoice_cids = set(invoice_lookup.keys())
    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row=row, column=cid_col + 1).value
        if cid and str(cid).strip() not in invoice_cids:
            addr = ws.cell(row=row, column=1).value
            fdh = ws.cell(row=row, column=fdh_col + 1).value if fdh_col is not None else None
            status = ws.cell(row=row, column=status_col + 1).value if status_col is not None else None
            pbi_not_in_invoice.append((str(cid).strip(), addr, fdh, status))

    match_ws.cell(row=10, column=1, value="PBI Circuits NOT in Invoice").font = Font(bold=True, size=12)
    match_ws.cell(row=11, column=1, value="Circuit ID").font = bold
    match_ws.cell(row=11, column=2, value="Address").font = bold
    match_ws.cell(row=11, column=3, value="FDH Name").font = bold
    match_ws.cell(row=11, column=4, value="Address Status").font = bold

    for i, (cid, addr, fdh, status) in enumerate(pbi_not_in_invoice):
        match_ws.cell(row=12 + i, column=1, value=cid)
        match_ws.cell(row=12 + i, column=2, value=addr)
        match_ws.cell(row=12 + i, column=3, value=fdh)
        match_ws.cell(row=12 + i, column=4, value=status)

    match_ws.column_dimensions["A"].width = 25
    match_ws.column_dimensions["B"].width = 45
    match_ws.column_dimensions["C"].width = 35
    match_ws.column_dimensions["D"].width = 20

    # Save with openpyxl (PowerBI_Data + Match Report, no Summary yet)
    print(f"  Saving enriched data to: {city_config['output_file']}")
    wb.save(output_path)
    wb.close()

    # Create real PivotTables via Excel COM automation
    create_summary_pivots(output_path, name, fdh_col_letter, act_col_letter)

    print(f"  Done! {matched} matched, {unmatched} unmatched, {len(pbi_not_in_invoice)} PBI-only circuits")


def main():
    print("=" * 60)
    print("PBI + Invoice Enrichment")
    print("=" * 60)

    for city in CITIES:
        enrich_pbi(city)

    print("\n" + "=" * 60)
    print("COMPLETE")
    print("=" * 60)
    print(f"\nOutput files in: {BASE_DIR}")
    for city in CITIES:
        print(f"  {city['output_file']}")


if __name__ == "__main__":
    main()
