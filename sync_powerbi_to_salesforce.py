"""
Sync PowerBI export data to Salesforce.

Reads the latest PowerBI Excel export from PowerBI_Report/, deduplicates
locations, and upserts Property_Location__c and Property_Unit__c records
to Salesforce via the Bulk API.
"""

import csv
import glob
import os
import re
import shutil
import sys
from datetime import datetime, timezone

import openpyxl
from simple_salesforce import Salesforce

# ── Configuration ──────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
POWERBI_DIR = os.path.join(BASE_DIR, "PowerBI_Report")
ARCHIVE_DIR = os.path.join(POWERBI_DIR, "Previously_Imported")

SF_USERNAME = "cass1@ubiquitygp.com"
SF_PASSWORD = "Karate88!"
SF_SECURITY_TOKEN = "Ktc1n9mLmD9vwEcVcl45q0iAD"

BULK_BATCH_SIZE = 10_000
HEADER_ROW = 3
DATA_START_ROW = 4


# ── Helpers ────────────────────────────────────────────────────────────────

def find_latest_export():
    """Find the most recently modified .xlsx file in the PowerBI_Report folder."""
    pattern = os.path.join(POWERBI_DIR, "*.xlsx")
    files = glob.glob(pattern)
    if not files:
        print(f"ERROR: No .xlsx files found in {POWERBI_DIR}")
        sys.exit(1)
    latest = max(files, key=os.path.getmtime)
    print(f"Using export file: {os.path.basename(latest)}")
    return latest


def to_date_str(value):
    """Convert a datetime value to YYYY-MM-DD string, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    # If it's already a date object
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    # If it's a string, try to parse common formats
    s = str(value).strip()
    if not s:
        return None
    return s


def to_str(value):
    """Convert value to stripped string, or None if blank."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def normalize_address(value):
    """Normalize an address: strip and collapse multiple spaces into one."""
    s = to_str(value)
    if not s:
        return None
    return re.sub(r"\s+", " ", s)


def read_excel(filepath):
    """Read the PowerBI export. Returns (headers, rows) where each row is a dict."""
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=HEADER_ROW, column=col).value)

    rows = []
    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers):
            if header is not None:
                row_data[header] = ws.cell(row=row_num, column=col_idx + 1).value
        rows.append(row_data)

    wb.close()
    print(f"  Read {len(rows)} data rows, {len(headers)} columns")
    return rows


def build_location_records(rows):
    """Deduplicate by Business Base Address and build Property_Location__c records."""
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    seen = set()
    records = []
    bad_rows = []

    for row_num, row in enumerate(rows, start=DATA_START_ROW):
        bba = normalize_address(row.get("Business Base Address"))
        cid = to_str(row.get("Circuit ID"))

        if not bba or not cid:
            bad_rows.append((row_num, bba, cid))
            continue
        if bba in seen:
            continue
        seen.add(bba)

        state = to_str(row.get("State"))
        market = to_str(row.get("Market"))
        hold = state == "CA" and market != "OCEANSIDE"

        records.append({
            "Business_Base_Address__c": bba,
            "Name": bba[:80],  # SF Name field max 80 chars
            "Market__c": market,
            "State__c": state,
            "FDH_Activated_Date__c": to_date_str(row.get("FDH Activated Date")),
            "FDH_Name__c": to_str(row.get("FDH Name")),
            "Serving_Area__c": to_str(row.get("Serving Area")),
            "City__c": to_str(row.get("City")),
            "Business_Building_Id__c": to_str(row.get("Business Building Id")),
            "Circuit_ID__c": to_str(row.get("Circuit ID")),
            "Hold__c": hold,
            "Import_DateTime__c": now_str,
        })

    if bad_rows:
        print(f"\n  ERROR: {len(bad_rows)} rows have blank Business Base Address or Circuit ID:")
        for row_num, bba, cid in bad_rows[:20]:
            missing = []
            if not bba:
                missing.append("Business Base Address")
            if not cid:
                missing.append("Circuit ID")
            print(f"    Row {row_num}: missing {', '.join(missing)}")
        if len(bad_rows) > 20:
            print(f"    ... and {len(bad_rows) - 20} more")
        print("\n  Fix the source data and re-export from PowerBI.")
        sys.exit(1)

    print(f"  {len(records)} unique locations")
    return records


def build_unit_records(rows):
    """Build Property_Unit__c records from all rows."""
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    records = []
    bad_rows = []

    for row_num, row in enumerate(rows, start=DATA_START_ROW):
        cid = to_str(row.get("Circuit ID"))
        bba = normalize_address(row.get("Business Base Address"))

        if not cid or not bba:
            bad_rows.append((row_num, bba, cid))
            continue

        records.append({
            "Circuit_ID__c": cid,
            "Name": (to_str(row.get("Address")) or "")[:80],
            "Unit__c": to_str(row.get("Unit #")),
            "Activated__c": to_str(row.get("Activated")),
            "Coho__c": to_str(row.get("Cohort Map Legend")),
            "Address_Activation_Date__c": to_date_str(row.get("Address Activation Date")),
            "ValidForFF__c": to_str(row.get("ValidForFF")),
            "Address_De_activation_Date__c": to_date_str(row.get("Address De-activation Date")),
            "Address_Deactivated__c": to_str(row.get("Address Deactivated")),
            "Ordered_Product__c": to_str(row.get("Ordered Product")),
            "AreaId__c": to_str(row.get("AreaId")),
            "Property_Location__r": {"Business_Base_Address__c": bba},
            "Import_DateTime__c": now_str,
        })

    if bad_rows:
        print(f"\n  ERROR: {len(bad_rows)} rows have blank Circuit ID or Business Base Address:")
        for row_num, bba, cid in bad_rows[:20]:
            missing = []
            if not cid:
                missing.append("Circuit ID")
            if not bba:
                missing.append("Business Base Address")
            print(f"    Row {row_num}: missing {', '.join(missing)}")
        if len(bad_rows) > 20:
            print(f"    ... and {len(bad_rows) - 20} more")
        print("\n  Fix the source data and re-export from PowerBI.")
        sys.exit(1)

    print(f"  {len(records)} unit records")
    return records


def fix_dirty_addresses(sf):
    """Fix existing SF locations that have double spaces in Business_Base_Address__c."""
    print("Checking for dirty addresses in Salesforce...")
    result = sf.query_all("SELECT Id, Business_Base_Address__c FROM Property_Location__c")

    updates = []
    for r in result["records"]:
        bba = r["Business_Base_Address__c"]
        if bba:
            clean = re.sub(r"\s+", " ", bba.strip())
            if clean != bba:
                updates.append({"Id": r["Id"], "Business_Base_Address__c": clean, "Name": clean[:80]})

    if not updates:
        print("  No dirty addresses found")
        return 0

    print(f"  Found {len(updates)} locations with whitespace issues, fixing...")
    results = sf.bulk.Property_Location__c.update(updates)
    errors = [r for r in results if not r.get("success")]
    fixed = len(updates) - len(errors)
    print(f"  Fixed: {fixed}")
    if errors:
        print(f"  Errors: {len(errors)}")
        for err in errors[:5]:
            print(f"    {err}")
    return fixed


def reparent_units(sf, unit_records):
    """Fix units whose parent location changed in PowerBI."""
    print("Checking for units that need reparenting...")

    # Build map of what the export says: circuit_id -> expected parent BBA
    export_parents = {}
    for r in unit_records:
        cid = r["Circuit_ID__c"]
        bba = r["Property_Location__r"]["Business_Base_Address__c"]
        export_parents[cid] = bba

    # Get current parent BBAs from Salesforce
    result = sf.query_all(
        "SELECT Id, Circuit_ID__c, Property_Location__r.Business_Base_Address__c "
        "FROM Property_Unit__c"
    )

    mismatched = []
    for r in result["records"]:
        cid = r["Circuit_ID__c"]
        parent_ref = r.get("Property_Location__r")
        sf_bba = parent_ref.get("Business_Base_Address__c") if parent_ref else None
        expected_bba = export_parents.get(cid)
        if expected_bba and sf_bba and expected_bba != sf_bba:
            mismatched.append({"unit_id": r["Id"], "cid": cid, "expected_bba": expected_bba})

    if not mismatched:
        print("  No reparenting needed")
        return 0

    # Look up correct parent location IDs
    needed_bbas = {m["expected_bba"] for m in mismatched}
    bba_to_loc_id = {}
    for bba in needed_bbas:
        escaped = bba.replace("'", "\\'")
        r = sf.query(f"SELECT Id FROM Property_Location__c WHERE Business_Base_Address__c = '{escaped}'")
        if r["records"]:
            bba_to_loc_id[bba] = r["records"][0]["Id"]

    updates = []
    skipped = 0
    for m in mismatched:
        new_parent_id = bba_to_loc_id.get(m["expected_bba"])
        if new_parent_id:
            updates.append({"Id": m["unit_id"], "Property_Location__c": new_parent_id})
        else:
            skipped += 1

    print(f"  Reparenting {len(updates)} units...")
    results = sf.bulk.Property_Unit__c.update(updates)
    success = sum(1 for r in results if r.get("success"))
    errors = [r for r in results if not r.get("success")]
    print(f"  Success: {success}")
    if errors:
        print(f"  Errors: {len(errors)}")
        for e in errors[:5]:
            print(f"    {e}")
    if skipped:
        print(f"  Skipped {skipped} (parent location not found)")
    return success


def generate_review_csvs(sf, location_records, unit_records):
    """Export CSVs of SF records that were NOT in the PowerBI export for review."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Build sets of what was synced
    synced_bbas = {r["Business_Base_Address__c"] for r in location_records}
    synced_cids = {r["Circuit_ID__c"] for r in unit_records}

    # --- Unsynced Locations ---
    print("Checking for SF locations not in this export...")
    loc_result = sf.query_all(
        "SELECT Id, Name, Business_Base_Address__c, Market__c, State__c, "
        "City__c, Circuit_ID__c, FDH_Name__c, Serving_Area__c, Import_DateTime__c "
        "FROM Property_Location__c"
    )

    unsynced_locs = []
    for r in loc_result["records"]:
        if r["Business_Base_Address__c"] not in synced_bbas:
            r.pop("attributes", None)
            unsynced_locs.append(r)

    if unsynced_locs:
        loc_csv = os.path.join(BASE_DIR, f"unsynced_locations_{timestamp}.csv")
        with open(loc_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=unsynced_locs[0].keys())
            writer.writeheader()
            writer.writerows(unsynced_locs)
        print(f"  {len(unsynced_locs)} locations not in export -> {os.path.basename(loc_csv)}")
    else:
        print("  All SF locations were in the export")

    # --- Unsynced Units ---
    print("Checking for SF units not in this export...")
    unit_result = sf.query_all(
        "SELECT Id, Name, Circuit_ID__c, Unit__c, Activated__c, "
        "Address_Activation_Date__c, Address_Deactivated__c, "
        "Property_Location__r.Business_Base_Address__c, Import_DateTime__c "
        "FROM Property_Unit__c"
    )

    unsynced_units = []
    for r in unit_result["records"]:
        if r["Circuit_ID__c"] not in synced_cids:
            parent_ref = r.pop("Property_Location__r", None)
            r["Parent_Business_Base_Address"] = (
                parent_ref.get("Business_Base_Address__c") if parent_ref else None
            )
            r.pop("attributes", None)
            unsynced_units.append(r)

    if unsynced_units:
        unit_csv = os.path.join(BASE_DIR, f"unsynced_units_{timestamp}.csv")
        with open(unit_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=unsynced_units[0].keys())
            writer.writeheader()
            writer.writerows(unsynced_units)
        print(f"  {len(unsynced_units)} units not in export -> {os.path.basename(unit_csv)}")
    else:
        print("  All SF units were in the export")


def bulk_upsert(sf, object_name, external_id_field, records):
    """Upsert records via Bulk API. Returns (created, updated, errors)."""
    if not records:
        print(f"  No records to upsert for {object_name}")
        return 0, 0, []

    sf_bulk = getattr(sf.bulk, object_name)

    # Process in batches
    all_results = []
    for i in range(0, len(records), BULK_BATCH_SIZE):
        batch = records[i:i + BULK_BATCH_SIZE]
        batch_num = i // BULK_BATCH_SIZE + 1
        total_batches = (len(records) + BULK_BATCH_SIZE - 1) // BULK_BATCH_SIZE
        print(f"  Upserting batch {batch_num}/{total_batches} ({len(batch)} records)...")
        results = sf_bulk.upsert(batch, external_id_field)
        all_results.extend(results)

    created = sum(1 for r in all_results if r.get("created"))
    updated = sum(1 for r in all_results if r.get("success") and not r.get("created"))
    errors = [r for r in all_results if not r.get("success")]

    return created, updated, errors


def archive_file(filepath):
    """Move processed file to the archive folder with a timestamp."""
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    basename = os.path.basename(filepath)
    name, ext = os.path.splitext(basename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_name = f"{name}_imported_{timestamp}{ext}"
    dest = os.path.join(ARCHIVE_DIR, new_name)
    shutil.move(filepath, dest)
    print(f"Archived to: {os.path.basename(dest)}")


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("PowerBI -> Salesforce Sync")
    print("=" * 60)
    print()

    # Step 1: Find and read the export file
    filepath = find_latest_export()
    rows = read_excel(filepath)
    print()

    # Step 2: Build records
    print("Building Property Location records...")
    location_records = build_location_records(rows)
    print()

    print("Building Property Unit records...")
    unit_records = build_unit_records(rows)
    print()

    # Step 3: Connect to Salesforce
    print("Connecting to Salesforce...")
    sf = Salesforce(
        username=SF_USERNAME,
        password=SF_PASSWORD,
        security_token=SF_SECURITY_TOKEN,
    )
    print(f"  Connected as {SF_USERNAME}")
    print()

    # Step 4: Fix dirty addresses from previous imports
    fixed_count = fix_dirty_addresses(sf)
    print()

    # Step 5: Upsert Property Locations
    print("Upserting Property Locations...")
    loc_created, loc_updated, loc_errors = bulk_upsert(
        sf, "Property_Location__c", "Business_Base_Address__c", location_records
    )
    print(f"  Created: {loc_created}, Updated: {loc_updated}, Errors: {len(loc_errors)}")
    print()

    # Step 6: Reparent units whose location changed in PowerBI
    reparented_count = reparent_units(sf, unit_records)
    print()

    # Step 7: Upsert Property Units
    print("Upserting Property Units...")
    unit_created, unit_updated, unit_errors = bulk_upsert(
        sf, "Property_Unit__c", "Circuit_ID__c", unit_records
    )
    print(f"  Created: {unit_created}, Updated: {unit_updated}, Errors: {len(unit_errors)}")
    print()

    # Step 8: Summary
    print("=" * 60)
    print("SUMMARY")
    print("=" * 60)
    if fixed_count:
        print(f"Address cleanup:    {fixed_count} dirty addresses fixed")
    if reparented_count:
        print(f"Units reparented:   {reparented_count} units moved to correct location")
    print(f"Property Locations: {loc_created} created, {loc_updated} updated")
    print(f"Property Units:     {unit_created} created, {unit_updated} updated")
    print()

    if loc_errors or unit_errors:
        print(f"ERRORS: {len(loc_errors)} location errors, {len(unit_errors)} unit errors")
        if loc_errors:
            print("\n  Location errors (first 10):")
            for err in loc_errors[:10]:
                print(f"    {err}")
        if unit_errors:
            print("\n  Unit errors (first 10):")
            for err in unit_errors[:10]:
                print(f"    {err}")
        print()

    # Step 9: Generate review CSVs for records in SF but not in this export
    generate_review_csvs(sf, location_records, unit_records)
    print()

    # Step 10: Archive the processed file
    archive_file(filepath)

    print()
    print("Done!")


if __name__ == "__main__":
    main()
