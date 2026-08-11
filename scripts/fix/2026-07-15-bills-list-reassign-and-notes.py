"""
From Bill's active list (Bills List of 56 Active.xlsx), two actions against Salesforce:

  1. REASSIGN the opps currently owned by a DEACTIVATED user (Chuck McNeely / Jeff Chao)
     to Bill Holick, so they show under his filter and he can work them.
  2. ADD NOTES: push the useful tracker content (owner/PM contact, management company,
     existing fiber, dated outreach history) onto the matched ACTIVE opps as a Salesforce
     Note, so the contact info lives in SF (what the 7/15 meeting asked for). Skips Closed
     Lost and rows with no content; skips an opp that already has this note (re-run safe).

Dry-run by default. --write snapshots owners (rollback), reassigns, adds notes, and audits.
"""
import argparse
import base64
import csv
import re
import sys
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "analysis"))
from enrich_omaha_onnet_mdus import creds  # noqa: E402
from lookup_agree_names_for_unlinked import house, st_tokens  # noqa: E402
from simple_salesforce import Salesforce  # noqa: E402

SRC = Path(r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Bills List of 56 Active.xlsx")
OUT = Path(r"C:\Users\cass\Work_Projects\SalesForce\data\output\audit_logs")
SNAP = OUT / "2026-07-15-bills-list-owner-rollback.csv"
AUDIT_R = OUT / "2026-07-15-bills-list-reassign-audit.csv"
AUDIT_N = OUT / "2026-07-15-bills-list-notes-audit.csv"
BILL_ID = "005WR00000DEU6oYAH"
DEACT = {"Chuck McNeely", "Jeff Chao"}
NOTE_TITLE = "MDU tracker contact/outreach notes (Bill's list) - 2026-07-15"
F = "Id, Name, Owner.Name, StageName, Property_City__c, Property_Address__c"
ncity = lambda s: re.sub(r"[^a-z]", "", (s or "").lower())


def clean(v):
    if v in (None, ""):
        return ""
    return re.sub(r"\s*\n\s*", " / ", str(v).replace("_x000D_", "\n").strip()).strip(" /")


def load_rows():
    ws = openpyxl.load_workbook(SRC, data_only=True)["Sheet1"]
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    g = lambda r, name: ws.cell(r, hdr[name]).value if name in hdr else None
    out = []
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        out.append({
            "name": str(ws.cell(r, 1).value).strip(),
            "units": g(r, "Units"), "city": str(g(r, "City") or "").strip(),
            "addr": str(g(r, "Address") or "").strip(), "state": g(r, "State"),
            "contact": clean(g(r, "RE Contact Found")), "mgmt": clean(g(r, "Management Company")),
            "fiber": clean(g(r, "Existing Fiber Provider")), "notes": clean(g(r, "Notes")),
        })
    return out


def stset(a):
    return {t for t in st_tokens(a) if not t.isdigit()}


def note_html(brow):
    bits = []
    if brow["contact"]:
        bits.append(f"<b>Owner/PM contact (from RE research):</b> {esc(brow['contact'])}")
    if brow["mgmt"]:
        bits.append(f"<b>Management company:</b> {esc(brow['mgmt'])}")
    if brow["fiber"]:
        bits.append(f"<b>Existing fiber provider:</b> {esc(brow['fiber'])}")
    if brow["notes"]:
        bits.append(f"<b>Outreach history:</b> {esc(brow['notes'])}")
    if not bits:
        return None
    return "<p>" + "</p><p>".join(bits) + "</p>"


def esc(s):
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()

    rows = load_rows()
    print(f"Bill's list rows: {len(rows)}")
    sf = Salesforce(*creds())

    names = sorted({r["name"] for r in rows})
    byname = defaultdict(list)
    for i in range(0, len(names), 150):
        vals = "','".join(n.replace("\\", "\\\\").replace("'", "\\'") for n in names[i:i + 150])
        for o in sf.query_all(f"SELECT {F} FROM Opportunity WHERE Name IN ('{vals}')")["records"]:
            byname[(o["Name"] or "").strip().lower()].append(o)

    def addr_lookup(r):
        h = house(r["addr"])
        if not h or not r["city"]:
            return None
        try:
            cands = sf.query(f"SELECT {F} FROM Opportunity WHERE Property_City__c LIKE "
                             f"'{r['city'][:18]}%' AND Property_Address__c LIKE '{h} %' LIMIT 6")["records"]
        except Exception:
            return None
        gt = stset(r["addr"])
        return next((o for o in cands if gt & stset(o["Property_Address__c"] or "")), None)

    # opp_id -> (opp, richest bill row)
    by_opp = {}
    for r in rows:
        hits = byname.get(r["name"].lower())
        o = hits[0] if hits else addr_lookup(r)
        if not o:
            continue
        prev = by_opp.get(o["Id"])
        if not prev or len(r["contact"]) + len(r["notes"]) > len(prev[1]["contact"]) + len(prev[1]["notes"]):
            by_opp[o["Id"]] = (o, r)

    own = lambda o: (o.get("Owner") or {}).get("Name") or "(none)"
    is_closed = lambda o: "Closed" in (o["StageName"] or "")

    # ---- phase 1: reassign deactivated-owner opps to Bill ----
    reassign = [(oid, o) for oid, (o, r) in by_opp.items() if own(o) in DEACT]
    print(f"\n[1] REASSIGN to Bill Holick (owned by {'/'.join(DEACT)}): {len(reassign)} opps")
    for oid, o in sorted(reassign, key=lambda x: own(x[1])):
        print(f"      {o['Name'][:34]:34} | {own(o):16} -> Bill | {o['StageName']}")

    # ---- phase 2: notes on active matched opps with content ----
    note_targets = [(oid, o, r) for oid, (o, r) in by_opp.items()
                    if not is_closed(o) and note_html(r)]
    print(f"\n[2] ADD NOTES to active opps with tracker content: {len(note_targets)}")
    print(f"      (skipped {sum(1 for _, (o, r) in by_opp.items() if is_closed(o))} closed, "
          f"{sum(1 for _, (o, r) in by_opp.items() if not note_html(r))} with no content)")

    if not args.write:
        print("\nDRY-RUN. Re-run with --write to reassign + add notes.")
        return

    OUT.mkdir(parents=True, exist_ok=True)
    # snapshot owners for rollback
    with SNAP.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "Name", "old_owner_name"])
        for oid, o in reassign:
            w.writerow([oid, o["Name"], own(o)])

    ra, rerr = 0, 0
    with AUDIT_R.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "Name", "old_owner", "new_owner", "result"])
        for oid, o in reassign:
            try:
                sf.Opportunity.update(oid, {"OwnerId": BILL_ID})
                ra += 1
                w.writerow([oid, o["Name"], own(o), "Bill Holick", "OK"])
            except Exception as e:
                rerr += 1
                w.writerow([oid, o["Name"], own(o), "Bill Holick", f"ERROR: {e}"])
                print(f"  reassign ERROR {oid}: {e}")
    print(f"\nreassigned: {ra}  |  errors: {rerr}")

    # add notes (skip if this note already on the opp)
    na, nskip, nerr = 0, 0, 0
    with AUDIT_N.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "Name", "result"])
        title_soql = NOTE_TITLE.replace("\\", "\\\\").replace("'", "\\'")
        for oid, o, r in note_targets:
            try:
                existing = sf.query(f"SELECT Id FROM ContentDocumentLink WHERE LinkedEntityId='{oid}' "
                                    f"AND ContentDocument.Title='{title_soql}'")["records"]
                if existing:
                    nskip += 1
                    w.writerow([oid, o["Name"], "skip-exists"])
                    continue
                res = sf.ContentNote.create({
                    "Title": NOTE_TITLE,
                    "Content": base64.b64encode(note_html(r).encode("utf-8")).decode()})
                sf.ContentDocumentLink.create({"ContentDocumentId": res["id"], "LinkedEntityId": oid,
                                               "ShareType": "V", "Visibility": "AllUsers"})
                na += 1
                w.writerow([oid, o["Name"], "note-ok"])
            except Exception as e:
                nerr += 1
                w.writerow([oid, o["Name"], f"ERROR: {e}"])
                print(f"  note ERROR {oid} {o['Name']}: {e}")
    print(f"notes added: {na}  |  skipped (already had it): {nskip}  |  errors: {nerr}")
    print(f"\naudits: {AUDIT_R}\n        {AUDIT_N}\nrollback owners from {SNAP}")


if __name__ == "__main__":
    main()
