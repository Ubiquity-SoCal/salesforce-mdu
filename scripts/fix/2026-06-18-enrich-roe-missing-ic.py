"""One-off: enrich ROEs_Missing IC#_enriched.xlsx with IronClad IDs (Omaha/NE batch).

The sheet "ROEs Missing IC#" lists ROE agreements whose IronClad ID (col F) is blank.
Koa pasted the enriched values (Rosemarie Shortino's NE book). Match by Agreement
Number (col B) and fill IronClad ID; also fill Property Address (col J) where blank.
Never overwrite a non-empty cell — mismatches are flagged for human review.

Run dry (default) to preview; pass --apply to write (backs up the file first)."""
import sys, re, shutil, datetime
import openpyxl

XLSX = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\ROEs_Missing IC#_enriched.xlsx"
APPLY = "--apply" in sys.argv

# Raw paste — Name | AGR | ROE | Completed | Date | IC# | RE | State | Owner | [Address]
RAW = """\
Omaha_MDU_5006 Davenport St\tAGR-1357\tROE\tCompleted\t3/27/2025\tIC-2727\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_St Frances Apartments\tAGR-1366\tROE\tCompleted\t8/22/2024\tIC-2635\tRosemarie Shortino\tNE\tRosemarie Shortino\t8929 Cuming Street Omaha, Nebraska 68114
Omaha_MDU_Taylor Apartments\tAGR-1355\tROE\tCompleted\t12/18/2024\tIC-2637\tRosemarie Shortino\tNE\tRosemarie Shortino\t4854 Taylor Street Omaha, Nebraska 68104
Omaha_MDU_Caroline\tAGR-1327\tROE\tCompleted\t2/23/2026\tIC-2272\tRosemarie Shortino\tNE\tRosemarie Shortino\t1616 & 1618 Military Avenue, Omaha, NE 68111
Omaha_MDU_2313 N 72nd St apts\tAGR-1328\tROE\tCompleted\t1/27/2026\tIC-2278\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_Benson Gardens\tAGR-1329\tROE\tCompleted\t5/23/2025\tIC-2638\tRosemarie Shortino\tNE\tRosemarie Shortino\t2017 Benson Gardens Boulevard Omaha, Nebraska 68134
Omaha_MDU_360 Skyview\tAGR-1330\tROE\tCompleted\t2/11/2026\tIC-2679\tRosemarie Shortino\tNE\tRosemarie Shortino\t3025 Meredith Avenue Omaha, NE  68111
Omaha_MDU_Arbor Creek Apartments\tAGR-1331\tROE\tCompleted\t8/22/2024\tIC-2662\tRosemarie Shortino\tNE\tRosemarie Shortino\tArbor Creek Apartments - 345 S 78th St Omaha, NE 68114  AKA Farnam Flats
Omaha_MDU_4015 HAMILTON ST\tAGR-1333\tROE\tCompleted\t11/26/2025\tIC-1858\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4023 HAMILTON ST\tAGR-1334\tROE\tCompleted\t11/26/2025\tIC-1859\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4803 Boyd St Apartments\tAGR-1335\tROE\tCompleted\t12/18/2024\tIC-2648\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_Birchwood Apts\tAGR-1336\tROE\tCompleted\t8/26/2024\tIC-679\tRosemarie Shortino\tNE\tRosemarie Shortino\t4306 North 88th Plaza, Omaha, Nebraska 68134
Omaha_MDU_4313 N 65th St\tAGR-1337\tROE\tCompleted\t10/18/2024\tIC-2649\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4455 Franklin St\tAGR-1340\tROE\tCompleted\t3/31/2026\tIC-2656\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4484 Redman Ave\tAGR-1342\tROE\tCompleted\t12/18/2024\tIC-2651\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4527 Decatur St\tAGR-1344\tROE\tCompleted\t3/31/2026\tIC-2659\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4616 Redman Ave\tAGR-1345\tROE\tCompleted\t12/18/2024\tIC-2645\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4606 Redman Ave\tAGR-1346\tROE\tCompleted\t12/18/2024\tIC-2652\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_4618 Redman Ave\tAGR-1347\tROE\tCompleted\t12/18/2024\tIC-2646\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_The Wakeley Pointe\tAGR-1348\tROE\tCompleted\t2/7/2026\tIC-2108\tRosemarie Shortino\tNE\tRosemarie Shortino\t4715 Wakeley Street,Omaha, Nebraska 68132 PIN 1627250000
Omaha_MDU_The Rose Apartments\tAGR-1349\tROE\tCompleted\t11/28/2025\tIC-1817\tRosemarie Shortino\tNE\tRosemarie Shortino\t4729 California St,Omaha, Nebraska 68132 PIN 1626610000
Omaha_MDU_4810-4812 Capital Ave\tAGR-1350\tROE\tCompleted\t2/7/2026\tIC-2666\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_The Frederick\tAGR-1351\tROE\tCompleted\t12/9/2025\tIC-1808\tRosemarie Shortino\tNE\tRosemarie Shortino\t4813 & 4815 Underwood Avenue, Omaha, Nebraska 68132 PIN 953560002
Omaha_MDU_4822-4824 Cass St\tAGR-1354\tROE\tCompleted\t1/30/2026\tIC-2293\tRosemarie Shortino\tNE\tRosemarie Shortino\t4824 Cass Street, Omaha, Nebraska 68132
Omaha_MDU_4907 Ames Ave Apartments\tAGR-1356\tROE\tCompleted\t12/12/2024\tIC-2644\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_Ellington Apartments\tAGR-1358\tROE\tCompleted\t2/3/2026\tIC-2115\tRosemarie Shortino\tNE\tRosemarie Shortino\t4960 Northwest Radial Highway,Omaha, Nebraska 68104 PIN 1821970000
Omaha_MDU_California Place\tAGR-1359\tROE\tCompleted\t2/23/2026\tIC-2727\tRosemarie Shortino\tNE\tRosemarie Shortino\t5006 California Street, Omaha, Nebraska 68132
Omaha_MDU_6302 Boyd St\tAGR-1361\tROE\tCompleted\t1/7/2025\tIC-2643\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_6303 Taylor Cir\tAGR-1362\tROE\tCompleted\t12/18/2024\tIC-2647\tRosemarie Shortino\tNE\tRosemarie Shortino\t
Omaha_MDU_The Richards Apartments\tAGR-1365\tROE\tCompleted\t9/4/2024\tIC-2639\tRosemarie Shortino\tNE\tRosemarie Shortino\t7402 Blondo St. Omaha NE
New Added Site\tAGR-1367\tROE\tCompleted\t3/31/2026\t?\tRosemarie Shortino\tNE\tRosemarie Shortino\t
78th Place Apartments\tAGR-1463\tROE\tCompleted\t1/26/2026\tIC-2271\tRosemarie Shortino\tNE\tRosemarie Shortino\t343 S 78th St, Omaha, NE 68114"""

rows = []
for line in RAW.splitlines():
    p = line.split("\t")
    rows.append({"name": p[0].strip(), "agr": p[1].strip(), "ic": p[5].strip(),
                 "addr": (p[9].strip() if len(p) > 9 else "")})

# Flag duplicate IC#s within the paste (shared IronClad across agreements).
from collections import Counter
ic_counts = Counter(r["ic"] for r in rows if r["ic"] and r["ic"] != "?")
dups = {ic: n for ic, n in ic_counts.items() if n > 1}

wb = openpyxl.load_workbook(XLSX)
ws = wb["ROEs Missing IC#"]
IC_COL, AGR_COL, ADDR_COL = 6, 2, 10  # F, B, J

agr_to_row = {}
for r in range(2, ws.max_row + 1):
    a = ws.cell(r, AGR_COL).value
    if a:
        agr_to_row[str(a).strip()] = r

filled_ic, filled_addr, skipped, conflicts, unmatched, no_ic = [], [], [], [], [], []
for rec in rows:
    if rec["ic"] == "?" or not rec["ic"]:
        no_ic.append(rec)
    r = agr_to_row.get(rec["agr"])
    if not r:
        unmatched.append(rec)
        continue
    cur = ws.cell(r, IC_COL).value
    if rec["ic"] and rec["ic"] != "?":
        if cur in (None, ""):
            if APPLY:
                ws.cell(r, IC_COL).value = rec["ic"]
            filled_ic.append((rec["agr"], rec["ic"]))
        elif str(cur).strip() != rec["ic"]:
            conflicts.append((rec["agr"], cur, rec["ic"]))
        else:
            skipped.append((rec["agr"], rec["ic"]))
    if rec["addr"]:
        curaddr = ws.cell(r, ADDR_COL).value
        if curaddr in (None, ""):
            if APPLY:
                ws.cell(r, ADDR_COL).value = rec["addr"]
            filled_addr.append((rec["agr"], rec["addr"][:40]))

print(f"MODE: {'APPLY' if APPLY else 'DRY RUN'}  | paste rows: {len(rows)} | sheet rows: {ws.max_row-1}\n")
print(f"IC# to fill (blank in sheet): {len(filled_ic)}")
for a, ic in filled_ic:
    print(f"   {a} -> {ic}")
print(f"\nAddress to fill (blank in sheet): {len(filled_addr)}")
for a, ad in filled_addr:
    print(f"   {a} -> {ad}")
if skipped:
    print(f"\nAlready had matching IC# (no change): {len(skipped)}  {[a for a,_ in skipped]}")
if conflicts:
    print(f"\n!! CONFLICT — sheet already has a DIFFERENT IC# (NOT overwritten): {len(conflicts)}")
    for a, cur, new in conflicts:
        print(f"   {a}: sheet={cur}  paste={new}")
if unmatched:
    print(f"\n!! AGR not found in sheet (NOT added): {len(unmatched)}")
    for rec in unmatched:
        print(f"   {rec['agr']}  {rec['name']}")
if no_ic:
    print(f"\n!! No IC# in paste (can't fill): {len(no_ic)}")
    for rec in no_ic:
        print(f"   {rec['agr']}  {rec['name']}  (IC = '{rec['ic']}')")
if dups:
    print(f"\n!! Duplicate IC# within paste (same IronClad on multiple AGRs — verify): {dups}")
    for ic in dups:
        print(f"   {ic}: " + ", ".join(r['agr'] for r in rows if r['ic'] == ic))

# Remaining blanks after this batch
remaining = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r, IC_COL).value in (None, ""))
print(f"\nRows still missing IC# in sheet (current state): {remaining}")

if APPLY:
    ts = datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S")
    bak = XLSX.replace(".xlsx", f".backup-{ts}.xlsx")
    shutil.copy2(XLSX, bak)
    wb.save(XLSX)
    print(f"\nSAVED. Backup: {bak}")
else:
    print("\n(dry run — no file written; re-run with --apply)")
