"""
Backfill IronClad_ID__c on the Omaha ROE Agreements from "ROEs_Missing IC#_enriched.xlsx".
These rows carry an IC-#### in the spreadsheet but were never pushed to Salesforce.

Mapping is read from the spreadsheet (source of truth). Targets only agreements whose
SF IronClad_ID__c is currently blank. Sets ONLY IronClad_ID__c (text); parent linking is
a separate step (sync/link_orphan_ironclad_agreements.py --apply).

HOLD: AGR-1357 (5006 Davenport St) excluded -- its spreadsheet value IC-2727 actually
belongs to AGR-1359 (California Place, confirmed in IronClad). AGR-1357's correct IC is
unresolved (only Davenport agreement in IronClad is IC-712, a PAL). Pending Koa's call.

PREVIEW only by default. Run with --apply to write.
Audit: SalesForce/data/output/audit_logs/set_omaha_ironclad_ids_<TS>.csv
"""
import sys, csv
from collections import Counter
from datetime import datetime
from pathlib import Path
import openpyxl
from simple_salesforce import Salesforce

USER = _SF["username"]; PW = _SF["password"]; TOK = _SF["token"]
XLSX = r"C:/Users/cass/OneDrive - Ubiquity Management/Desktop/ROEs_Missing IC#_enriched.xlsx"
LOG_DIR = Path(r"C:\Users\cass\Work_Projects\SalesForce\data\output\audit_logs")
LOG_DIR.mkdir(parents=True, exist_ok=True)
APPLY = "--apply" in sys.argv
SOURCE = "2026-06-22-set-omaha-ironclad-ids.py"
HOLD = {"AGR-1357"}

# --- read mapping from spreadsheet ---
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active
file_ic = {}
for r in range(2, ws.max_row + 1):
    agr = ws.cell(row=r, column=2).value
    ic = ws.cell(row=r, column=6).value
    if agr and ic and str(ic).strip().upper().startswith("IC-"):
        file_ic[str(agr).strip()] = str(ic).strip()
print(f"Rows in spreadsheet with an IC-#### value: {len(file_ic)}")

sf = Salesforce(username=USER, password=PW, security_token=TOK)
names = list(file_ic.keys())
recs = {x["Name"]: x for x in sf.query_all(
    "SELECT Id, Name, IronClad_ID__c, IronClad_Record__c, Opportunity__r.Name "
    "FROM Agreement__c WHERE Name IN ('" + "','".join(names) + "')")["records"]}

to_set, held, already, conflict, missing = [], [], [], [], []
for agr, ic in file_ic.items():
    if agr in HOLD:
        held.append((agr, ic)); continue
    r = recs.get(agr)
    if not r:
        missing.append((agr, ic)); continue
    cur = r.get("IronClad_ID__c")
    if cur == ic:
        already.append((agr, ic))
    elif cur and cur != ic:
        conflict.append((agr, cur, ic, r))
    else:
        to_set.append((agr, ic, r))

# parent existence
target_ics = [ic for _, ic, _ in to_set]
parents = set()
if target_ics:
    parents = {p["IronClad_Id__c"] for p in sf.query_all(
        "SELECT IronClad_Id__c FROM IronClad__c WHERE IronClad_Id__c IN ('" + "','".join(set(target_ics)) + "')")["records"]}

# duplicate IC across targets (should be none)
dups = {ic: n for ic, n in Counter(target_ics).items() if n > 1}

print(f"\nWILL SET (blank in SF):  {len(to_set)}")
for agr, ic, r in to_set:
    flag = "" if ic in parents else "  [NO PARENT]"
    print(f"   {agr}  ->  {ic}{flag}   ({(r.get('Opportunity__r') or {}).get('Name')})")
print(f"\nHELD (excluded):         {len(held)}  {held}")
print(f"Already set (skip):      {len(already)}  {[a for a,_ in already]}")
print(f"Conflict (diff IC set):  {len(conflict)}  {[(a,c,n) for a,c,n,_ in conflict]}")
print(f"Not found in SF:         {len(missing)}  {missing}")
print(f"Parents present:         {len([1 for _,ic,_ in to_set if ic in parents])}/{len(to_set)}")
print(f"DUPLICATE IC in targets: {dups or 'none'}")

if dups:
    print("\n! Duplicate IC across targets -- aborting before any write. Resolve first.")
    sys.exit(1)

ts = datetime.now().strftime("%Y%m%d-%H%M%S")
audit = LOG_DIR / f"set_omaha_ironclad_ids_{ts}.csv"
with open(audit, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["Action", "SF_Id", "Name", "Field", "Before", "After", "Parent_Exists", "Source", "Timestamp"])
    act = "UPDATE" if APPLY else "PREVIEW"
    for agr, ic, r in to_set:
        w.writerow([act, r["Id"], agr, "IronClad_ID__c", r.get("IronClad_ID__c") or "", ic,
                    ic in parents, SOURCE, datetime.now().isoformat()])
    for agr, ic in held:
        w.writerow(["HOLD", recs.get(agr, {}).get("Id", ""), agr, "IronClad_ID__c", "", ic, "", SOURCE, datetime.now().isoformat()])
print(f"\nAudit: {audit}")

if not APPLY:
    print("\nPREVIEW only. Re-run with --apply to write.")
    sys.exit(0)

import requests

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()

headers = {"Authorization": f"Bearer {sf.session_id}", "Content-Type": "application/json"}
url = f"{sf.base_url}composite/sobjects"
records = [{"attributes": {"type": "Agreement__c"}, "Id": r["Id"], "IronClad_ID__c": ic} for agr, ic, r in to_set]
ok = fail = 0
for i in range(0, len(records), 200):
    chunk = records[i:i+200]
    resp = requests.patch(url, headers=headers, json={"allOrNone": False, "records": chunk}, timeout=120)
    if resp.status_code == 200:
        for res in resp.json():
            if res.get("success"):
                ok += 1
            else:
                fail += 1; print("  !", res.get("errors"))
    else:
        fail += len(chunk); print(f"  ! HTTP {resp.status_code}: {resp.text[:300]}")
print(f"\nSet IronClad_ID__c: ok={ok} fail={fail}")
print("Next: run sync/link_orphan_ironclad_agreements.py --apply to link parents.")
