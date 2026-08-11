"""Stage the TX-NE On-Net MDU status changes marked up on the 2026-07-15 call.

DRY RUN BY DEFAULT. Writes a rollback snapshot and a staged diff; touches nothing in Salesforce
unless --execute is passed, and --execute refuses to run without a snapshot on disk.

Source of truth is the 'Updates' column of the marked-up workbook, 134 rows across 3 actions:

  73  'Move to Propspects'                    Closed Lost / No Contact Info -> Prospects
  47  'Mark as Prospects - Assign to ?'       PARKED - owner is literally '?' on the sheet
  14  'Change to Closed Lost - Existing BULK' -> Closed Lost / Existing Contract, keep owner

Decisions taken from the call and the 2026-04-29 stage restructure, in case anyone re-reads this:
  * 'Prospects' not 'Prospecting'. Canonical split is Prospects = cold/no activity, Prospecting =
    active with 2026 activity. Speaker 00 wanted "the lowest bucket you can put it in".
  * Loss_Reason__c has no 'Existing BULK' value. All 14 rows' notes describe a DP Bulk Agreement
    with Spectrum, so 'Existing Contract' is the fit. Not 'Lost to Competitor' - we never
    competed, the bulk pre-dated us.
  * Owners on the Chuck McNeely / Jeff Chao rows are KEPT. Speaker 00 floated stripping them,
    Koa pushed back that the name is the only breadcrumb until reassignment, and he agreed.
  * The 73 are a reversal of the SMB ROE mapping that put them here
    ('Closed - Contact Info' -> Closed Lost + Loss_Reason 'No Contact Info').
"""

import argparse
import csv
import datetime
import json
import re
import subprocess
import sys
from pathlib import Path
from collections import Counter

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT.parent / "MDU" / "data" / "input" / "tx-ne-on-net-mdus-with-comments-2026-07-15.xlsx"
AUDIT_DIR = ROOT / "data" / "output" / "audit_logs"
STAMP = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")

# Updates-column value -> what it means in Salesforce. Values are exact picklist entries,
# verified against the live Opportunity describe on 2026-07-16.
ACTIONS = {
    "Move to Propspects": {           # typo is the source's, matched verbatim on purpose
        "StageName": "Prospects",
        "Loss_Reason__c": "",         # an open stage must not keep a loss reason
        "expect": 73,
    },
    "Change to Closed Lost - Existing BULK": {
        "StageName": "Closed Lost",
        "Loss_Reason__c": "Existing Contract",
        "expect": 14,
    },
}
PARKED = "Mark as Prospects - Assign to ?"

# Notes on these rows undercut the action. Staged but held back so a human rules on them first.
HOLD = {
    "Killeen_MDU_Morgan Manor Apts": "note says \"this could be wrong. Feel free to change @Chuck "
                                     "McNeely\" - author flagged his own note as unreliable",
    "Killeen_MDU_Lakeview Apts": "note \"Will not sign Bulk again\" is ambiguous - if they will not "
                                 "re-sign WITH SPECTRUM that is an opening, not a loss",
}


def sf(args):
    p = subprocess.run(args, capture_output=True, text=True, shell=True)
    i = p.stdout.find("{")
    if i < 0:
        raise SystemExit(f"sf cli gave no json:\n{p.stdout[:400]}\n{p.stderr[:400]}")
    d = json.loads(p.stdout[i:])
    if d.get("status") != 0:
        raise SystemExit(f"sf error: {str(d)[:400]}")
    return d["result"]


def soql(q):
    return sf(["sf", "data", "query", "--query", q, "--json"])["records"]


def esc(s):
    return s.replace("\\", "\\\\").replace("'", "\\'")


def read_workbook():
    ws = openpyxl.load_workbook(WORKBOOK)["ON Net MDUs"]
    hdr = [c.value for c in ws[1]]
    ci = {}
    for i, h in enumerate(hdr, 1):
        ci.setdefault(h, i)
    rows = []
    for r in range(2, ws.max_row + 1):
        g = lambda k: (str(ws.cell(row=r, column=ci[k]).value).strip()
                       if ws.cell(row=r, column=ci[k]).value is not None else "")
        if g("Updates"):
            rows.append({"site": g("Site Name"), "sf_name": g("Property Name (SF)"),
                         "update": g("Updates"), "state": g("Opportunity State"),
                         "loss": g("Loss Reason"), "owner": g("Owner"), "units": g("Total Units")})
    return rows


FIELDS = ("Id, Name, Agreement_Name__c, StageName, Loss_Reason__c, Probability, OwnerId, "
          "Owner.Name, RecordType.Name, Property_Address__c, Property_City__c, "
          "Property_State__c, Units__c")


def fetch(agree_names, names):
    """Key on Agreement_Name__c, not Name. Name is NOT unique.

    'River Oaks Apartments' is two unrelated buildings (Killeen TX, 228 units vs Tucson AZ,
    300 units); 'Woodgate Townhomes' is two more (Omaha 107th Plaza, 70 units vs Omaha 79th
    Plaza, 35 units). Matching on Name made those look like duplicates and would have pointed
    an update at whichever record came back first. Agreement_Name__c carries the market prefix
    ('Killeen_MDU_River Oaks Apartments') and matches the workbook's Site Name column exactly.
    """
    by_agree, by_name = {}, {}
    for i in range(0, len(agree_names), 40):
        chunk = ",".join(f"'{esc(n)}'" for n in agree_names[i:i + 40])
        for rec in soql(f"SELECT {FIELDS} FROM Opportunity WHERE Agreement_Name__c IN ({chunk})"):
            by_agree.setdefault(rec["Agreement_Name__c"], []).append(rec)
    for i in range(0, len(names), 40):
        chunk = ",".join(f"'{esc(n)}'" for n in names[i:i + 40])
        for rec in soql(f"SELECT {FIELDS} FROM Opportunity WHERE Name IN ({chunk})"):
            by_name.setdefault(rec["Name"], []).append(rec)
    return by_agree, by_name


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--execute", action="store_true", help="actually push (default: dry run)")
    args = ap.parse_args()

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_workbook()
    print(f"marked-up rows in workbook : {len(rows)}")
    print(f"  by action: {dict(Counter(r['update'] for r in rows))}")

    todo = [r for r in rows if r["update"] in ACTIONS]
    parked = [r for r in rows if r["update"] == PARKED]
    print(f"  actionable now: {len(todo)} | parked (owner unnamed): {len(parked)}")
    print()

    by_agree, by_name = fetch(sorted({r["site"] for r in todo if r["site"]}),
                              sorted({r["sf_name"] for r in todo if r["sf_name"]}))

    staged, problems = [], []
    matched_by = Counter()
    for r in todo:
        recs = by_agree.get(r["site"], [])
        how = "Agreement_Name__c"
        if not recs:                                  # fall back to Name, but only if unambiguous
            recs = by_name.get(r["sf_name"], [])
            how = "Name (fallback)"
        if not recs:
            problems.append((r, "NOT FOUND in SF by Agreement_Name__c or Name"))
            continue
        if len(recs) > 1:
            addrs = "; ".join(f"{x.get('Property_City__c')} {x.get('Units__c')}u" for x in recs)
            problems.append((r, f"{len(recs)} Opps match [{how}] - ambiguous ({addrs})"))
            continue
        rec = recs[0]
        matched_by[how] += 1
        tgt = ACTIONS[r["update"]]
        change = {}
        if rec["StageName"] != tgt["StageName"]:
            change["StageName"] = tgt["StageName"]
        if (rec.get("Loss_Reason__c") or "") != tgt["Loss_Reason__c"]:
            change["Loss_Reason__c"] = tgt["Loss_Reason__c"]
        staged.append({
            "site": r["site"], "id": rec["Id"], "name": rec["Name"], "units": r["units"],
            "action": r["update"],
            "from_stage": rec["StageName"], "to_stage": tgt["StageName"],
            "from_loss": rec.get("Loss_Reason__c") or "", "to_loss": tgt["Loss_Reason__c"],
            "owner": (rec.get("Owner") or {}).get("Name", ""),
            "record_type": (rec.get("RecordType") or {}).get("Name", ""),
            "probability": rec.get("Probability"),
            "change": change,
            "hold": HOLD.get(r["site"], ""),
        })

    # --- rollback snapshot: every field this script could touch, as it stands right now
    snap = AUDIT_DIR / f"txne-onnet-stage-push-SNAPSHOT-{STAMP}.csv"
    with open(snap, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "Name", "Agreement_Name__c", "StageName", "Loss_Reason__c",
                    "Probability", "OwnerId", "Owner_Name", "RecordType"])
        seen_ids = set()
        for recs in list(by_agree.values()) + list(by_name.values()):
            for rec in recs:
                if rec["Id"] in seen_ids:
                    continue
                seen_ids.add(rec["Id"])
                w.writerow([rec["Id"], rec["Name"], rec.get("Agreement_Name__c") or "",
                            rec["StageName"], rec.get("Loss_Reason__c") or "",
                            rec.get("Probability"), rec["OwnerId"],
                            (rec.get("Owner") or {}).get("Name", ""),
                            (rec.get("RecordType") or {}).get("Name", "")])

    # --- staged diff: exactly what would change, for review before anything is pushed
    diff = AUDIT_DIR / f"txne-onnet-stage-push-STAGED-{STAMP}.csv"
    with open(diff, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Site Name", "SF Id", "SF Name", "Units", "Action", "Owner", "Record Type",
                    "Stage: from", "Stage: to", "Loss Reason: from", "Loss Reason: to",
                    "Fields changing", "HOLD - needs a human"])
        for s in sorted(staged, key=lambda x: (x["action"], x["site"])):
            w.writerow([s["site"], s["id"], s["name"], s["units"], s["action"], s["owner"],
                        s["record_type"], s["from_stage"], s["to_stage"], s["from_loss"],
                        s["to_loss"], "; ".join(s["change"]) or "(already correct)", s["hold"]])

    # --- validate: staged counts must tie back to the workbook, or something is wrong
    print(f"matched by: {dict(matched_by)}")
    print()
    print("=== staged vs expected ===")
    ok = True
    for act, cfg in ACTIONS.items():
        got = sum(1 for s in staged if s["action"] == act)
        flag = "OK" if got == cfg["expect"] else "MISMATCH"
        if got != cfg["expect"]:
            ok = False
        print(f"  {flag:8s} {act:40s} staged={got:3d} expected={cfg['expect']:3d}")
    print()
    print("=== what actually changes ===")
    for act in ACTIONS:
        sub = [s for s in staged if s["action"] == act]
        print(f"  {act}")
        print(f"     stage moves : {dict(Counter(f'{s[chr(34)+chr(34)] if False else s['from_stage']} -> {s['to_stage']}' for s in sub))}")
        print(f"     loss reason : {dict(Counter(f'{s['from_loss'] or '(blank)'} -> {s['to_loss'] or '(cleared)'}' for s in sub))}")
        print(f"     no-op rows  : {sum(1 for s in sub if not s['change'])}")
    print()
    held = [s for s in staged if s["hold"]]
    if held:
        print(f"=== HELD BACK - {len(held)} rows a human must rule on first ===")
        for s in held:
            print(f"  {s['site']}\n      {s['hold']}")
        print()
    if problems:
        print(f"=== PROBLEMS - {len(problems)} rows not staged ===")
        for r, why in problems:
            print(f"  {r['site'][:44]:44s} | {why}")
        print()
    print(f"parked (owner is '?', needs Speaker 00): {len(parked)} rows, "
          f"{sum(int(r['units'] or 0) for r in parked):,} units")
    print()
    print(f"snapshot -> {snap}")
    print(f"staged   -> {diff}")

    if not args.execute:
        print("\nDRY RUN. Nothing was written to Salesforce. Review the staged CSV, then re-run "
              "with --execute.")
        return
    if not ok:
        raise SystemExit("refusing to execute: staged counts do not match the workbook")
    pushable = [s for s in staged if s["change"] and not s["hold"]]
    print(f"\nEXECUTE: {len(pushable)} records...")
    upd = AUDIT_DIR / f"txne-onnet-stage-push-APPLY-{STAMP}.csv"
    with open(upd, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "StageName", "Loss_Reason__c"])
        for s in pushable:
            # Bulk API 2.0 IGNORES an empty CSV value - it does not null the field. Clearing
            # requires the literal '#N/A'. Learned the hard way on 2026-07-16: the job reported
            # 85/85 successful while all 73 kept their 'No Contact Info' loss reason, leaving
            # open opportunities carrying a loss reason.
            w.writerow([s["id"], s["to_stage"], s["to_loss"] if s["to_loss"] else "#N/A"])
    res = sf(["sf", "data", "update", "bulk", "--sobject", "Opportunity", "--file", str(upd),
              "--wait", "10", "--json"])
    print(json.dumps(res)[:400])


if __name__ == "__main__":
    main()
