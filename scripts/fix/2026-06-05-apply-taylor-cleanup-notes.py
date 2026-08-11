"""Apply Taylor's 6/5 cleanup notes to Agreement__c records.

Source file: 'Cleanup need Ironclad IDs - Taylor's Notes 6.5.26.xlsx' (her notes
from the 6/5 Weekly Salesforce call). Two actions, keyed by the "Taylor's Note"
column on each Opportunity row:
  - "Delete PAL for this opportunity"            -> delete the Opp's PAL Agreement__c
  - "Change Agreement Type from PAL to ROE if not ROE already" -> PAL.Agreement_Type__c = ROE

Taylor Mauney is the MDU gatekeeper; her notes are authoritative (no per-owner re-consult).

Safety:
  - Dry-run by default. Pass --apply to write.
  - Before ANY write/delete, snapshots the FULL record of every affected Agreement__c
    to a JSON file (restorable) + a review CSV.
  - Audit CSV of every action (SF_Id / Name / Action / Before / After / Source / Timestamp).
  - Skips (does not guess) on: ambiguous duplicate Opp name, no PAL present for a
    delete row, or a retype row whose Opp already has an ACTIVE (non-Cancelled) ROE.
"""
import sys
import csv
import json
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


APPLY = "--apply" in sys.argv
FILE = sorted(Path(r"C:/Users/cass/OneDrive - Ubiquity Management/Desktop")
              .glob("Cleanup need Ironclad IDs*xlsx"))[-1]
OUT = Path("C:/Users/cass/Work_Projects/SalesForce/data/output/audit_logs")
OUT.mkdir(parents=True, exist_ok=True)
TS = datetime.now().strftime("%Y%m%d-%H%M%S")
SOURCE = FILE.name

sf = Salesforce(username=_SF["username"], password=_SF["password"],
                security_token=_SF["token"])

# Canonical MDU stage order (post 2026-04-29 restructure) for advance logic.
STAGE_ORDER = ["Closed Lost", "On Hold", "Prospects", "Prospecting", "Engaged",
               "Proposal Sent", "Contract Negotiations", "PAL/ROE Complete",
               "Marketing/Bulk In Progress", "Marketing/Bulk Complete"]
STAGE_RANK = {s: i for i, s in enumerate(STAGE_ORDER)}
COMPLETE_RANK = STAGE_RANK["PAL/ROE Complete"]


def categorize(note):
    n = (note or "").strip().lower()
    if n.startswith("change agreement type"):
        return "pal_to_roe"
    if n.startswith("delete pal"):
        return "delete_pal"
    return "skip"


# --- parse Taylor's file ---
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [r for r in ws.iter_rows(values_only=True)]
hdr = list(rows[0])
NI, OI = hdr.index("Taylor's Note"), hdr.index("Opportunity Name")
file_rows = [(str(r[OI]).strip(), categorize(r[NI])) for r in rows[1:]
             if r and r[OI] and categorize(r[NI]) != "skip"]
print(f"Source: {SOURCE}")
print(f"Actionable rows: {len(file_rows)}  {dict(Counter(c for _, c in file_rows))}")

# --- pull SF state ---
opps = sf.query_all("SELECT Id, Name, StageName FROM Opportunity")["records"]
by_name = defaultdict(list)
for o in opps:
    by_name[(o["Name"] or "").strip().lower()].append(o)

agr = sf.query_all("""
    SELECT Id, Name, Agreement_Type__c, Status__c, Opportunity__c
    FROM Agreement__c
""")["records"]
agr_by_opp = defaultdict(list)
for a in agr:
    agr_by_opp[a["Opportunity__c"]].append(a)

# --- build plan ---
deletes, retypes, skips = [], [], []

for nm, cat in file_rows:
    matches = by_name.get(nm.lower(), [])
    if not matches:
        skips.append((nm, cat, "opp_not_found"))
        continue

    if cat == "delete_pal":
        cand = [o for o in matches if any(
            (a.get("Agreement_Type__c") or "") == "PAL" for a in agr_by_opp.get(o["Id"], []))]
        if not cand:
            skips.append((nm, cat, "no_PAL_to_delete"))
            continue
        if len(cand) > 1:
            skips.append((nm, cat, f"ambiguous_{len(cand)}_opps_with_PAL"))
            continue
        opp = cand[0]
        for a in agr_by_opp.get(opp["Id"], []):
            if (a.get("Agreement_Type__c") or "") == "PAL":
                deletes.append((nm, opp, a))

    elif cat == "pal_to_roe":
        if len(matches) > 1:
            skips.append((nm, cat, f"dup_opp_name_{len(matches)}"))
            continue
        opp = matches[0]
        ags = agr_by_opp.get(opp["Id"], [])
        pals = [a for a in ags if (a.get("Agreement_Type__c") or "") == "PAL"]
        active_roe = [a for a in ags if (a.get("Agreement_Type__c") or "") == "ROE"
                      and (a.get("Status__c") or "") != "Cancelled"]
        if not pals:
            skips.append((nm, cat, "already_ROE_no_PAL"))   # the 63 no-ops
            continue
        if active_roe:
            skips.append((nm, cat, "has_active_ROE_already"))  # would duplicate; needs Taylor
            continue
        for a in pals:
            retypes.append((nm, opp, a))

# Stage advances: a retyped Opp keeps a Completed agreement, so if it sits below
# PAL/ROE Complete it should advance to match (approved 6/5: Pacifica HOA).
stage_advances = []
_seen_opp = set()
for nm, o, a in retypes:
    if o["Id"] in _seen_opp:
        continue
    _seen_opp.add(o["Id"])
    if STAGE_RANK.get(o["StageName"], 99) < COMPLETE_RANK:
        stage_advances.append((nm, o, o["StageName"], "PAL/ROE Complete"))

print(f"\nPLAN:")
print(f"  PAL agreements to DELETE: {len(deletes)}")
print(f"  PAL agreements to RETYPE -> ROE: {len(retypes)}")
print(f"  Opp STAGE advances -> PAL/ROE Complete: {len(stage_advances)}  {[s[0] for s in stage_advances]}")
print(f"  SKIPPED rows: {len(skips)}  {dict(Counter(s[2] for s in skips))}")

# --- snapshot full records of everything we will touch ---
touch_ids = [a["Id"] for _, _, a in deletes] + [a["Id"] for _, _, a in retypes]
snapshot = {}
for aid in touch_ids:
    snapshot[aid] = sf.Agreement__c.get(aid)  # full record for restore
for nm, o, frm, to in stage_advances:
    snapshot[o["Id"]] = sf.Opportunity.get(o["Id"])  # full Opp record for stage rollback
snap_path = OUT / f"taylor_cleanup_snapshot_{TS}.json"
snap_path.write_text(json.dumps(snapshot, indent=2, default=str), encoding="utf-8")
print(f"\nFull-record snapshot ({len(snapshot)} agreements): {snap_path}")

# --- review CSVs ---
del_csv = OUT / f"taylor_cleanup_DELETE_{TS}.csv"
with open(del_csv, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["opp_name", "opp_id", "opp_stage", "agr_id", "agr_name",
                "agr_type", "agr_status"])
    for nm, o, a in deletes:
        w.writerow([nm, o["Id"], o["StageName"], a["Id"], a["Name"],
                    a.get("Agreement_Type__c"), a.get("Status__c")])

rt_csv = OUT / f"taylor_cleanup_RETYPE_{TS}.csv"
with open(rt_csv, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["opp_name", "opp_id", "opp_stage", "agr_id", "agr_name",
                "from_type", "to_type", "agr_status"])
    for nm, o, a in retypes:
        w.writerow([nm, o["Id"], o["StageName"], a["Id"], a["Name"],
                    a.get("Agreement_Type__c"), "ROE", a.get("Status__c")])

skip_csv = OUT / f"taylor_cleanup_SKIPPED_{TS}.csv"
with open(skip_csv, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["opp_name", "category", "reason"])
    for nm, cat, reason in skips:
        w.writerow([nm, cat, reason])
print(f"Review CSVs: {del_csv.name}, {rt_csv.name}, {skip_csv.name}")

# Show skip detail (the non-no-op ones matter)
notable = [s for s in skips if s[2] not in ("already_ROE_no_PAL",)]
if notable:
    print(f"\nNotable skips (need a look, not plain no-ops):")
    for nm, cat, reason in notable:
        print(f"  [{cat}] {nm[:40]:40} {reason}")

# --- audit + apply ---
audit = OUT / f"taylor_cleanup_audit_{TS}.csv"
with open(audit, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["SF_Id", "Agreement_Name", "Action", "Before", "After",
                "Opp", "Source", "Result", "Timestamp"])
    action_label = "APPLIED" if APPLY else "PREVIEW"

    if not APPLY:
        for nm, o, a in retypes:
            w.writerow([a["Id"], a["Name"], "RETYPE_PAL_TO_ROE", "PAL", "ROE",
                        nm, SOURCE, "preview", datetime.now().isoformat()])
        for nm, o, a in deletes:
            w.writerow([a["Id"], a["Name"], "DELETE_PAL",
                        f"{a.get('Agreement_Type__c')}/{a.get('Status__c')}", "(deleted)",
                        nm, SOURCE, "preview", datetime.now().isoformat()])
        for nm, o, frm, to in stage_advances:
            w.writerow([o["Id"], o["Name"], "ADVANCE_STAGE", frm, to,
                        nm, SOURCE, "preview", datetime.now().isoformat()])
        print(f"\nAudit (preview): {audit}")
        print("\nPREVIEW ONLY. Re-run with --apply to execute.")
    else:
        rok = rfail = dok = dfail = 0
        for nm, o, a in retypes:
            try:
                sf.Agreement__c.update(a["Id"], {"Agreement_Type__c": "ROE"})
                w.writerow([a["Id"], a["Name"], "RETYPE_PAL_TO_ROE", "PAL", "ROE",
                            nm, SOURCE, "ok", datetime.now().isoformat()])
                rok += 1
            except Exception as e:
                w.writerow([a["Id"], a["Name"], "RETYPE_PAL_TO_ROE", "PAL", "ROE",
                            nm, SOURCE, f"error:{e}", datetime.now().isoformat()])
                rfail += 1
        for nm, o, a in deletes:
            try:
                sf.Agreement__c.delete(a["Id"])
                w.writerow([a["Id"], a["Name"], "DELETE_PAL",
                            f"{a.get('Agreement_Type__c')}/{a.get('Status__c')}", "(deleted)",
                            nm, SOURCE, "ok", datetime.now().isoformat()])
                dok += 1
            except Exception as e:
                w.writerow([a["Id"], a["Name"], "DELETE_PAL",
                            f"{a.get('Agreement_Type__c')}/{a.get('Status__c')}", "(deleted)",
                            nm, SOURCE, f"error:{e}", datetime.now().isoformat()])
                dfail += 1
        sok = sfail = 0
        for nm, o, frm, to in stage_advances:
            try:
                sf.Opportunity.update(o["Id"], {"StageName": to})
                w.writerow([o["Id"], o["Name"], "ADVANCE_STAGE", frm, to,
                            nm, SOURCE, "ok", datetime.now().isoformat()])
                sok += 1
            except Exception as e:
                w.writerow([o["Id"], o["Name"], "ADVANCE_STAGE", frm, to,
                            nm, SOURCE, f"error:{e}", datetime.now().isoformat()])
                sfail += 1
        print(f"\nRetype: ok={rok} fail={rfail}   Delete: ok={dok} fail={dfail}   Stage: ok={sok} fail={sfail}")
        print(f"Audit: {audit}")
        print(f"Restore from: {snap_path}")
