"""Apply Taylor's 6/10 bulk update for Brett Spivey's PAL/ROE-Complete opportunities.

Source file: 'Cleanup need Ironclad IDs - Taylor's Notes 6.10.26.xlsx' (SF report
export, sheet report1781120890475). Koa's instruction (6/11): every Opportunity
owned by Brett Spivey gets the SAME bulk values, already filled in the file:
  - PAL Agreement__c.IronClad_ID__c   = "IC-1152"
  - PAL Agreement__c.Expiration_Date__c = 2043-11-01
  - Opportunity.AccountId             -> "Jones & Jones Communities"

The IronClad ID + Expiration land on the Opp's PAL Agreement__c (per the 6/5
snapshot, PAL records carry IronClad_ID__c + Expiration_Date__c). The Account is
the Opportunity's parent Account (reparent / set).

Taylor Mauney is the MDU gatekeeper; her notes are authoritative (no per-owner
re-consult). Brett Spivey is the Jones & Jones Communities contact.

Safety (mirrors 2026-06-05-apply-taylor-cleanup-notes.py):
  - Dry-run by default. Pass --apply to write.
  - Pass --create-account to create the "Jones & Jones Communities" Account if it
    does not already exist (only honored with --apply).
  - Before ANY write, snapshots the FULL record of every affected Agreement__c and
    Opportunity to a JSON file (restorable) + review CSV + audit CSV.
  - Reads the target values FROM the file (does not hardcode) and asserts Brett's
    rows are internally consistent before using them.
  - Skips (does not guess) on: opp name not found, ambiguous duplicate opp name,
    or an opp with no PAL agreement (can't set IronClad/Expiration without one).
"""
import sys
import csv
import json
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


APPLY = "--apply" in sys.argv
CREATE_ACCOUNT = "--create-account" in sys.argv
FILE = sorted(Path(r"C:/Users/cass/OneDrive - Ubiquity Management/Desktop")
              .glob("Cleanup need Ironclad IDs*xlsx"))[-1]
OUT = Path("C:/Users/cass/Work_Projects/SalesForce/data/output/audit_logs")
OUT.mkdir(parents=True, exist_ok=True)
TS = datetime.now().strftime("%Y%m%d-%H%M%S")
SOURCE = FILE.name

OWNER = "Brett Spivey"

sf = Salesforce(username=_SF["username"], password=_SF["password"],
                security_token=_SF["token"])


def norm(v):
    return ("" if v is None else str(v)).strip()


def as_date(v):
    """Normalize a cell to YYYY-MM-DD for a SF Date field."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s[:10] if s else None


# --- parse the file: take only Brett's rows, pull target values from the file ---
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [r for r in ws.iter_rows(values_only=True)]
hdr = [norm(h) for h in rows[0]]
ci = {name: hdr.index(name) for name in
      ["Opportunity Name", "Opportunity Owner", "Ironclad ID#", "Expiration Date", "Account"]}

brett = []
for r in rows[1:]:
    if not r or not r[ci["Opportunity Name"]]:
        continue
    if norm(r[ci["Opportunity Owner"]]) != OWNER:
        continue
    brett.append({
        "name": norm(r[ci["Opportunity Name"]]),
        "ironclad": norm(r[ci["Ironclad ID#"]]),
        "expiration": as_date(r[ci["Expiration Date"]]),
        "account": norm(r[ci["Account"]]),
    })

print(f"Source: {SOURCE}")
print(f"{OWNER} rows in file: {len(brett)}")

# assert the bulk values are consistent across Brett's rows (this is a *bulk* update)
ic_set = {b["ironclad"] for b in brett}
ex_set = {b["expiration"] for b in brett}
ac_set = {b["account"] for b in brett}
print(f"  Ironclad ID#:    {ic_set}")
print(f"  Expiration Date: {ex_set}")
print(f"  Account:         {ac_set}")
if not (len(ic_set) == len(ex_set) == len(ac_set) == 1):
    sys.exit("ABORT: Brett's rows are not internally consistent — not a clean bulk update. Inspect the file.")
TARGET_IC = ic_set.pop()
TARGET_EX = ex_set.pop()
TARGET_ACCT = ac_set.pop()

# --- pull SF state ---
opps = sf.query_all(
    "SELECT Id, Name, StageName, AccountId, Account.Name, Owner.Name FROM Opportunity"
)["records"]
by_name = defaultdict(list)
for o in opps:
    by_name[norm(o["Name"]).lower()].append(o)

agr = sf.query_all("""
    SELECT Id, Name, Agreement_Type__c, Status__c, Opportunity__c,
           IronClad_ID__c, Expiration_Date__c
    FROM Agreement__c
""")["records"]
agr_by_opp = defaultdict(list)
for a in agr:
    agr_by_opp[a["Opportunity__c"]].append(a)

# --- resolve the target Account ---
acct = sf.query_all(
    f"SELECT Id, Name FROM Account WHERE Name = '{TARGET_ACCT}'"
)["records"]
acct_id = acct[0]["Id"] if acct else None
print(f"\nTarget Account '{TARGET_ACCT}': "
      f"{'EXISTS ' + acct_id if acct_id else 'NOT FOUND (use --create-account to create on apply)'}")

# --- build plan ---
agr_updates, opp_updates, skips = [], [], []

for b in brett:
    matches = by_name.get(b["name"].lower(), [])
    if not matches:
        skips.append((b["name"], "opp_not_found"))
        continue
    if len(matches) > 1:
        skips.append((b["name"], f"dup_opp_name_{len(matches)}"))
        continue
    opp = matches[0]

    # Account on the Opportunity
    cur_acct = norm(opp.get("Account", {}).get("Name") if opp.get("Account") else "")
    if cur_acct != TARGET_ACCT:
        opp_updates.append((b["name"], opp, cur_acct, TARGET_ACCT))

    # IronClad ID + Expiration on the PAL agreement
    pals = [a for a in agr_by_opp.get(opp["Id"], [])
            if norm(a.get("Agreement_Type__c")) == "PAL"]
    if not pals:
        skips.append((b["name"], "no_PAL_agreement"))
        continue
    for a in pals:
        chg = {}
        if norm(a.get("IronClad_ID__c")) != b["ironclad"]:
            chg["IronClad_ID__c"] = b["ironclad"]
        if as_date(a.get("Expiration_Date__c")) != b["expiration"]:
            chg["Expiration_Date__c"] = b["expiration"]
        if chg:
            agr_updates.append((b["name"], opp, a, chg))

print(f"\nPLAN:")
print(f"  Opp Account set -> '{TARGET_ACCT}': {len(opp_updates)}")
print(f"  PAL agreement IronClad/Expiration updates: {len(agr_updates)}")
print(f"  Skipped rows: {len(skips)}  {dict(Counter(s[1] for s in skips))}")
if skips:
    print("  Skip detail:")
    for nm, reason in skips:
        print(f"    {nm[:45]:45} {reason}")

# --- snapshot full records of everything we will touch ---
snapshot = {}
for _, opp, a, _ in agr_updates:
    snapshot[a["Id"]] = sf.Agreement__c.get(a["Id"])
for _, opp, _, _ in opp_updates:
    snapshot[opp["Id"]] = sf.Opportunity.get(opp["Id"])
snap_path = OUT / f"brett_jones_jones_snapshot_{TS}.json"
snap_path.write_text(json.dumps(snapshot, indent=2, default=str), encoding="utf-8")
print(f"\nFull-record snapshot ({len(snapshot)} records): {snap_path}")

# --- review CSV ---
rev = OUT / f"brett_jones_jones_PLAN_{TS}.csv"
with open(rev, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["opp_name", "opp_id", "opp_stage", "change_kind",
                "target", "field", "before", "after"])
    for nm, opp, cur, new in opp_updates:
        w.writerow([nm, opp["Id"], opp["StageName"], "OPP_ACCOUNT",
                    "Opportunity", "AccountId(Name)", cur, new])
    for nm, opp, a, chg in agr_updates:
        for fld, val in chg.items():
            before = norm(a.get(fld)) if fld != "Expiration_Date__c" else as_date(a.get(fld))
            w.writerow([nm, opp["Id"], opp["StageName"], "PAL_FIELD",
                        f"{a['Name']}", fld, before, val])
print(f"Review CSV: {rev}")

# --- audit + apply ---
audit = OUT / f"brett_jones_jones_audit_{TS}.csv"
with open(audit, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["SF_Id", "Name", "Action", "Field", "Before", "After",
                "Opp", "Source", "Result", "Timestamp"])

    if not APPLY:
        for nm, opp, cur, new in opp_updates:
            w.writerow([opp["Id"], opp["Name"], "SET_ACCOUNT", "AccountId",
                        cur, new, nm, SOURCE, "preview", datetime.now().isoformat()])
        for nm, opp, a, chg in agr_updates:
            for fld, val in chg.items():
                w.writerow([a["Id"], a["Name"], "SET_PAL_FIELD", fld,
                            norm(a.get(fld)), val, nm, SOURCE, "preview",
                            datetime.now().isoformat()])
        print(f"\nAudit (preview): {audit}")
        print("\nPREVIEW ONLY. Re-run with --apply (and --create-account if needed) to execute.")
    else:
        # create the account if requested and missing
        if acct_id is None:
            if CREATE_ACCOUNT:
                res = sf.Account.create({"Name": TARGET_ACCT})
                acct_id = res["id"]
                print(f"Created Account '{TARGET_ACCT}': {acct_id}")
                w.writerow([acct_id, TARGET_ACCT, "CREATE_ACCOUNT", "Name", "",
                            TARGET_ACCT, "", SOURCE, "ok", datetime.now().isoformat()])
            elif opp_updates:
                sys.exit(f"ABORT: Account '{TARGET_ACCT}' not found and --create-account not passed, "
                         f"but {len(opp_updates)} Opps need it. Re-run with --create-account.")

        ok = fail = 0
        for nm, opp, cur, new in opp_updates:
            try:
                sf.Opportunity.update(opp["Id"], {"AccountId": acct_id})
                w.writerow([opp["Id"], opp["Name"], "SET_ACCOUNT", "AccountId",
                            cur, new, nm, SOURCE, "ok", datetime.now().isoformat()])
                ok += 1
            except Exception as e:
                w.writerow([opp["Id"], opp["Name"], "SET_ACCOUNT", "AccountId",
                            cur, new, nm, SOURCE, f"error:{e}", datetime.now().isoformat()])
                fail += 1
        aok = afail = 0
        for nm, opp, a, chg in agr_updates:
            try:
                sf.Agreement__c.update(a["Id"], chg)
                for fld, val in chg.items():
                    w.writerow([a["Id"], a["Name"], "SET_PAL_FIELD", fld,
                                norm(a.get(fld)), val, nm, SOURCE, "ok",
                                datetime.now().isoformat()])
                aok += 1
            except Exception as e:
                w.writerow([a["Id"], a["Name"], "SET_PAL_FIELD", "(batch)", "", "",
                            nm, SOURCE, f"error:{e}", datetime.now().isoformat()])
                afail += 1
        print(f"\nOpp Account: ok={ok} fail={fail}   PAL fields: ok={aok} fail={afail}")
        print(f"Audit: {audit}")
        print(f"Restore from: {snap_path}")
