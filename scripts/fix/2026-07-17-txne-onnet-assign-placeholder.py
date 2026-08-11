"""Reassign the parked TX-NE On-Net 'Mark as Prospects - Assign to ?' orphans to the placeholder.

The 2026-07-15 call marked 47 rows 'Mark as Prospects - Assign to ?': move off the (now inactive)
owner, new owner left blank. Salesforce has no blank owner - OwnerId is required and Opportunities
can't be queue-owned - so per Koa (2026-07-17) they go to a dedicated placeholder user
'Unassigned MDU' (005WR00000J5pBpYAJ) that reads clearly as needing reassignment.

Scope corrected after checking live SF (2026-07-17): these are NOT Closed-Lost - most are already
Prospects, and a separate 'bills-list' reassign on 7/15 already moved several onto active Bill
Holick. So this touches ONLY rows still owned by an INACTIVE user (Chuck / Jeff), and changes
OWNER ONLY:
  * no stage change - several are past Prospects (On Hold, Marketing/Bulk Complete, PAL/ROE Complete)
    and forcing them to Prospects would regress real progress;
  * no loss-reason touch;
  * rows already on an active user are left alone.

REST per-record update. Snapshot -> rollback; every write is re-queried (never trust the count).
DRY RUN by default; --execute to write.
"""
import argparse
import csv
import datetime
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "analysis"))
from enrich_omaha_onnet_mdus import creds  # noqa: E402
from simple_salesforce import Salesforce   # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT.parent / "MDU" / "data" / "input" / "tx-ne-on-net-mdus-with-comments-2026-07-15.xlsx"
AUDIT = ROOT / "data" / "output" / "audit_logs"
PLACEHOLDER_ID = "005WR00000J5pBpYAJ"                 # Unassigned MDU
PARKED = "Mark as Prospects - Assign to ?"
STAMP = datetime.datetime.now().strftime("%Y-%m-%dT%H-%M-%S")


def parked_sites():
    ws = openpyxl.load_workbook(WORKBOOK)["ON Net MDUs"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i for i, h in enumerate(hdr, 1) if h}
    return [str(ws.cell(r, ci["Site Name"]).value).strip()
            for r in range(2, ws.max_row + 1)
            if str(ws.cell(r, ci["Updates"]).value or "").strip() == PARKED]


def esc(s):
    return s.replace("\\", "\\\\").replace("'", "\\'")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--execute", action="store_true", help="actually write (default: dry run)")
    args = ap.parse_args()
    AUDIT.mkdir(parents=True, exist_ok=True)

    sites = parked_sites()
    print(f"parked rows in workbook: {len(sites)}")

    sf = Salesforce(*creds())
    ph = sf.query(f"SELECT Id, Name, IsActive FROM User WHERE Id='{PLACEHOLDER_ID}'")["records"]
    assert ph and ph[0]["IsActive"], "placeholder user missing or inactive"
    print(f"placeholder owner: {ph[0]['Name']} ({ph[0]['Id']}) active={ph[0]['IsActive']}")

    fields = ("Id, Name, Agreement_Name__c, StageName, Loss_Reason__c, OwnerId, "
              "Owner.Name, Owner.IsActive")
    recs = []
    for i in range(0, len(sites), 40):
        chunk = ",".join(f"'{esc(s)}'" for s in sites[i:i + 40])
        recs += sf.query_all(f"SELECT {fields} FROM Opportunity "
                             f"WHERE Agreement_Name__c IN ({chunk})")["records"]
    by = defaultdict(list)
    for r in recs:
        by[r["Agreement_Name__c"]].append(r)

    staged, skip_active, ambiguous, notfound = [], [], [], []
    for s in sites:
        rs = by.get(s, [])
        if not rs:
            notfound.append(s)
        elif len(rs) > 1:
            ambiguous.append(s)
        elif rs[0]["Owner"]["IsActive"]:
            skip_active.append(rs[0])          # already on a live owner (Bill) - leave alone
        else:
            staged.append(rs[0])               # orphaned on an inactive user - reassign

    print(f"  reassign (inactive owner) : {len(staged)}")
    print(f"  skip (active owner, leave): {len(skip_active)} "
          f"{dict(Counter(r['Owner']['Name'] for r in skip_active))}")
    print(f"  not found by agree-name   : {len(notfound)}")
    print(f"  ambiguous (skipped)       : {len(ambiguous)} {ambiguous}")
    print(f"  reassigning FROM owners   : {dict(Counter(r['Owner']['Name'] for r in staged))}")
    print(f"  reassigning stages (kept) : {dict(Counter(r['StageName'] for r in staged))}")

    snap = AUDIT / f"txne-onnet-placeholder-assign-SNAPSHOT-{STAMP}.csv"
    with open(snap, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "Name", "Agreement_Name__c", "StageName", "Loss_Reason__c",
                    "OwnerId", "Owner_Name"])
        for r in staged:
            w.writerow([r["Id"], r["Name"], r["Agreement_Name__c"], r["StageName"],
                        r.get("Loss_Reason__c") or "", r["OwnerId"], r["Owner"]["Name"]])
    print(f"snapshot -> {snap}")

    if not args.execute:
        print("\nDRY RUN. Nothing written. Re-run with --execute.")
        return

    done, errs = 0, []
    for r in staged:
        try:
            sf.Opportunity.update(r["Id"], {"OwnerId": PLACEHOLDER_ID})   # OWNER ONLY
            done += 1
        except Exception as e:  # noqa: BLE001
            errs.append((r["Id"], str(e)[:200]))
    print(f"\nupdated {done}/{len(staged)} | errors {len(errs)}")
    for i, e in errs:
        print(f"  ERR {i}: {e}")

    ids = [r["Id"] for r in staged]
    ver = []
    for i in range(0, len(ids), 40):
        chunk = ",".join(f"'{x}'" for x in ids[i:i + 40])
        ver += sf.query_all(f"SELECT Id, StageName, OwnerId, Owner.Name "
                            f"FROM Opportunity WHERE Id IN ({chunk})")["records"]
    ok = sum(1 for v in ver if v["OwnerId"] == PLACEHOLDER_ID)
    print(f"VERIFIED {ok}/{len(ver)} now owned by the placeholder (stage untouched)")
    for v in ver:
        if v["OwnerId"] != PLACEHOLDER_ID:
            print(f"  NOT REASSIGNED {v['Id']} still owner={v['Owner']['Name']}")

    aud = AUDIT / f"txne-onnet-placeholder-assign-APPLIED-{STAMP}.csv"
    with open(aud, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Id", "old_owner", "new_owner", "StageName", "result"])
        vmap = {v["Id"]: v for v in ver}
        for r in staged:
            good = vmap.get(r["Id"], {}).get("OwnerId") == PLACEHOLDER_ID
            w.writerow([r["Id"], r["Owner"]["Name"], "Unassigned MDU", r["StageName"],
                        "OK" if good else "FAILED"])
    print(f"applied audit -> {aud}")


if __name__ == "__main__":
    main()
