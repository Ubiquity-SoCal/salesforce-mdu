"""Backfill Property_Street_Number__c / Property_Street__c / Property_County__c.

Fields deployed 2026-08-06 for the PAL/ROE acquisition refocus (Niraj 2026-07-29).

Sources:
  * Street number + street  -> parsed from the existing Property_Address__c, which
    is already populated on 3,630 of 3,646 MDU opportunities. Not re-keyed from the
    spreadsheet, so the whole MDU book gets populated rather than just the 528 rows
    on the TX/NE on-net list.
  * County -> workbook column J ("TX-NE ON Net MDUs - with Salesforce.xlsx"), which
    only covers the on-net list, so county lands on those rows only.

Validation: Niraj split street number and street himself in workbook columns F and
G. The script compares the parser against his split on every joinable row and
reports the disagreement rate BEFORE writing. If the parser disagrees badly, stop.

Snapshots every current value to an audit CSV first (rollback source of truth).
Dry-run by default.

Usage:
    python 2026-08-06-backfill-property-address-components.py            # dry run
    python 2026-08-06-backfill-property-address-components.py --apply
"""

import argparse
import collections
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO))

from _shared.sf_auth import get_sf  # noqa: E402

BOOK = REPO / "MDU" / "data" / "input" / "TX-NE ON Net MDUs - with Salesforce.xlsx"
AUDIT = REPO / "SalesForce" / "data" / "output" / "audit_logs"

# A leading house number may carry a range or a letter: 1015-1017, 4707, 12B.
NUM_RE = re.compile(r"^\s*([0-9][0-9A-Za-z\-/]*)\s+(.+)$")
UNIT_RE = re.compile(r"\s+(?:UNIT|STE|SUITE|APT|BLDG|#)\s*[A-Za-z0-9\-]+\s*$", re.I)
ZIP_RE = re.compile(r"\s+\d{5}(?:-\d{4})?\s*$")
STATE_NAMES = {"TX": "TEXAS", "NE": "NEBRASKA", "AZ": "ARIZONA", "CA": "CALIFORNIA"}


def split_street(addr, city=None, state=None, zipcode=None):
    """Return (street_number, street_name) from a full or partial address.

    Roughly a third of Property_Address__c values carry no comma, e.g.
    "1200 SOUTHEAST PKWY UNIT 1 AZLE TX 76020". Splitting on the comma alone
    dumps the city, state and zip into the street field, so for those we peel
    the record's own city / state / zip off the tail instead.
    """
    if not addr:
        return None, None
    head = str(addr).replace("\t", " ").strip()
    # A few records carry a stray control/encoding byte before the number.
    head = re.sub(r"^[^0-9A-Za-z]+", "", head).strip()

    if "," in head:
        head = head.split(",")[0].strip()
    else:
        # Peel trailing zip, then state (abbr or full), then city.
        head = ZIP_RE.sub("", head).strip()
        tails = []
        if state:
            tails += [str(state).strip(), STATE_NAMES.get(str(state).strip().upper(), "")]
        tails += ["USA", "US"]
        for _ in range(3):  # zip/state/country can stack
            for t in [x for x in tails if x]:
                if head.upper().endswith(" " + t.upper()):
                    head = head[: -(len(t) + 1)].strip()
            head = ZIP_RE.sub("", head).strip()
        if city and head.upper().endswith(" " + str(city).strip().upper()):
            head = head[: -(len(str(city).strip()) + 1)].strip()

    head = UNIT_RE.sub("", head).strip().rstrip(",").strip()
    if not head:
        return None, None
    m = NUM_RE.match(head)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, head


def norm(x):
    return (x or "").replace("_", " ").strip().upper()


# Niraj's sheet uses USPS abbreviations, Salesforce mostly spells them out.
# Canonicalise both sides so the comparison reports real differences, not style.
ABBR = {
    "avenue": "ave", "street": "st", "road": "rd", "drive": "dr",
    "plaza": "plz", "court": "ct", "boulevard": "blvd", "lane": "ln",
    "circle": "cir", "parkway": "pkwy", "place": "pl", "terrace": "ter",
    "highway": "hwy", "trail": "trl", "north": "n", "south": "s",
    "east": "e", "west": "w", "northeast": "ne", "northwest": "nw",
    "southeast": "se", "southwest": "sw",
}


def squash(x):
    """Compare street names ignoring case, punctuation and USPS abbreviation style."""
    toks = re.sub(r"[^a-z0-9 ]", " ", str(x or "").lower()).split()
    return " ".join(ABBR.get(t, t) for t in toks)


def load_workbook_rows():
    ws = openpyxl.load_workbook(BOOK, data_only=True)["ON Net MDUs"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        rec = dict(zip(hdr, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]))
        if rec.get("Site Name"):
            rec["_row"] = r
            out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    sf = get_sf("main")
    opps = sf.query_all(
        "SELECT Id, Name, Agreement_Name__c, Property_Address__c, "
        "Property_City__c, Property_State__c, Property_Zip__c, "
        "Property_Street_Number__c, Property_Street__c, Property_County__c "
        "FROM Opportunity WHERE RecordType.Name = 'MDU/SFU'"
    )["records"]
    print(f"MDU opportunities: {len(opps)}")

    # ---- county lookup from the workbook -----------------------------------
    rows = load_workbook_rows()
    county_by_key = {}
    wb_split = {}
    for r in rows:
        k = norm(str(r["Site Name"]).strip())
        if r.get("County"):
            county_by_key[k] = str(r["County"]).strip()
        wb_split[k] = (
            None if r.get("St Number") is None else str(r["St Number"]).strip(),
            None if r.get("Street") is None else str(r["Street"]).strip(),
        )
    print(f"workbook rows: {len(rows)}, with a county: {len(county_by_key)}")

    # ---- parse + compare against Niraj's own split -------------------------
    agree_num = dis_num = agree_st = dis_st = compared = 0
    dis_examples = []
    updates = []
    no_number = []
    for o in opps:
        num, street = split_street(
            o["Property_Address__c"], o["Property_City__c"],
            o["Property_State__c"], o["Property_Zip__c"],
        )
        key = norm(o["Agreement_Name__c"]) if o["Agreement_Name__c"] else None
        county = county_by_key.get(key) if key else None

        if o["Property_Address__c"] and num is None:
            no_number.append(o)

        if key and key in wb_split:
            wnum, wst = wb_split[key]
            if wnum is not None:
                compared += 1
                if squash(wnum) == squash(num):
                    agree_num += 1
                else:
                    dis_num += 1
                    if len(dis_examples) < 12:
                        dis_examples.append(
                            (o["Name"][:34], o["Property_Address__c"], num, wnum, street, wst)
                        )
                if squash(wst) == squash(street):
                    agree_st += 1
                else:
                    dis_st += 1

        payload = {}
        if num and num != o["Property_Street_Number__c"]:
            payload["Property_Street_Number__c"] = num
        if street and street != o["Property_Street__c"]:
            payload["Property_Street__c"] = street
        if county and county != o["Property_County__c"]:
            payload["Property_County__c"] = county
        if payload:
            updates.append((o, payload))

    print()
    print("=== PARSER VALIDATION vs Niraj's own split (workbook cols F and G) ===")
    if compared:
        print(f"   compared on {compared} rows")
        print(f"   street number: agree {agree_num} ({agree_num/compared:.1%}), differ {dis_num}")
        print(f"   street name  : agree {agree_st} ({agree_st/compared:.1%}), differ {dis_st}")
        for e in dis_examples:
            print(f"      {e[0]:<36} addr={str(e[1])[:38]:<40} mine={e[2]!r}/{e[4]!r} his={e[3]!r}/{e[5]!r}")
    else:
        print("   no overlap to compare")

    print()
    print(f"addresses with no leading house number (street only, number left blank): {len(no_number)}")
    for o in no_number[:8]:
        print(f"      {o['Name'][:34]:<36} {o['Property_Address__c']}")

    print()
    fieldcount = collections.Counter()
    for _, p in updates:
        for f in p:
            fieldcount[f] += 1
    print(f"=== PLANNED WRITES: {len(updates)} records ===")
    for f, n in fieldcount.most_common():
        print(f"   {f:<32} {n}")

    stamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    AUDIT.mkdir(parents=True, exist_ok=True)
    path = AUDIT / f"2026-08-06-property-address-components-{'apply' if args.apply else 'dryrun'}-{stamp}.csv"
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([
            "Id", "Name", "Agreement_Name__c", "Property_Address__c",
            "field", "old_value", "new_value",
        ])
        for o, p in updates:
            for f, v in p.items():
                w.writerow([
                    o["Id"], o["Name"], o["Agreement_Name__c"], o["Property_Address__c"],
                    f, o[f] or "", v,
                ])
    print(f"\naudit/rollback CSV -> {path.name}")

    if not args.apply:
        print("\nDRY RUN. Re-run with --apply to write.")
        return 0

    print()
    payload = [dict({"Id": o["Id"]}, **p) for o, p in updates]
    ok = fail = 0
    for i in range(0, len(payload), 200):
        chunk = payload[i:i + 200]
        res = sf.bulk.Opportunity.update(chunk)
        for r in res:
            if r.get("success"):
                ok += 1
            else:
                fail += 1
                print("   FAIL", r)
        print(f"   {min(i+200, len(payload))}/{len(payload)}")
    print(f"\napplied {ok}, failed {fail}")

    # ---- verify ------------------------------------------------------------
    after = sf.query_all(
        "SELECT Id, Property_Street_Number__c, Property_Street__c, Property_County__c "
        "FROM Opportunity WHERE RecordType.Name = 'MDU/SFU'"
    )["records"]
    idx = {r["Id"]: r for r in after}
    bad = 0
    for o, p in updates:
        live = idx.get(o["Id"], {})
        for f, v in p.items():
            if live.get(f) != v:
                bad += 1
    print(f"verify: {len(updates) - bad} of {len(updates)} records match their intended values, {bad} mismatched")
    filled = collections.Counter()
    for r in after:
        for f in ("Property_Street_Number__c", "Property_Street__c", "Property_County__c"):
            if r.get(f):
                filled[f] += 1
    print("populated org-wide now:")
    for f, n in filled.items():
        print(f"   {f:<32} {n} / {len(after)}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
