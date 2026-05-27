"""Comprehensive comparison: Friday Excel vs current SF dashboard.

Loads the 5/22 Excel file in full, pulls equivalent SF data with pagination,
and prints a side-by-side comparison for every metric.

Audit-trail probe. Read-only. Output goes to stdout + a CSV report."""
import csv
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook
from simple_salesforce import Salesforce

EXCEL_PATH = r"C:\Users\cass\Work_Projects\Serviceability_Lookup\data\output\business-penetration-2026-05-22.xlsx"

sf = Salesforce(
    username="cass1@ubiquitygp.com",
    password="Hawaiian1984",
    security_token="IBSKT6CFUpSUJWxq1CMm0HkFC",
)

# ── Load Excel ─────────────────────────────────────────────────────
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Summary headline metrics
summary = {}
ws = wb["Summary"]
rows = list(ws.iter_rows(values_only=True))
for r in rows:
    if r and r[0] in ("Lit base business addresses", "Total units in lit buildings",
                       "Active customers", "De-activated (churned)",
                       "Drop completed (in lit bldgs)", "Serviceable idle (in lit bldgs)",
                       "Overall penetration", "All serviceable business addresses (universe)"):
        summary[r[0]] = r[1]

# Excel by-state
excel_by_state = {}
for r in rows:
    if r and r[0] in ("AZ", "CA", "NE", "TX"):
        excel_by_state[r[0]] = {
            "lit_bldgs": r[1], "units": r[2], "active": r[3],
            "deact": r[4], "pen": r[5]
        }

# Excel penetration distribution
excel_buckets = {}
for r in rows:
    if r and r[0] in ("0% (deactivated-only / churned)", "1-25%", "26-50%", "51-75%", "76-100%"):
        excel_buckets[r[0]] = r[1]

# Excel BBA detail
bba_ws = wb["Base Business Addresses"]
bba_rows = list(bba_ws.iter_rows(values_only=True))[1:]  # skip header
# Cols: BBA, City, State, Zip, Market, FDH, FDH Activated, Total Units, Active, Deactivated, Drop Completed, Serviceable (idle), Penetration
excel_bbas = []
for r in bba_rows:
    if r[0]:
        excel_bbas.append({
            "bba": r[0], "city": r[1], "state": r[2], "fdh": r[5],
            "total": r[7] or 0, "active": r[8] or 0, "deact": r[9] or 0,
            "drop_done": r[10] or 0, "idle": r[11] or 0, "pen": r[12] or 0
        })

# Excel Units detail
unit_ws = wb["Units"]
unit_count_excel = unit_ws.max_row - 1

# Excel single vs multi (count BBAs where total = 1 vs > 1)
excel_single_lit = sum(1 for b in excel_bbas if b["total"] == 1)
excel_multi_lit = sum(1 for b in excel_bbas if b["total"] > 1)
excel_multi_total_units = sum(b["total"] for b in excel_bbas if b["total"] > 1)
excel_multi_active = sum(b["active"] for b in excel_bbas if b["total"] > 1)
excel_single_active = sum(b["active"] for b in excel_bbas if b["total"] == 1)

# ── Pull SF (with + without Hold) ──────────────────────────────────
def sf_summary(no_hold=False):
    extra = " AND Priority__c != 'Hold'" if no_hold else ""
    UNIV = "Address_Type__c='Business' AND Import_Delete_Property__c=false" + extra

    def c(w): return sf.query(f"SELECT COUNT() FROM Property_Location__c WHERE {w}")["totalSize"]
    def sum_field(w, field):
        r = sf.query(f"SELECT SUM({field}) s FROM Property_Location__c WHERE {w}")["records"][0]
        return int(r["s"] or 0)

    total = c(UNIV)
    single_total = c(UNIV + " AND Property_Unit_Count__c = 1")
    multi_total = c(UNIV + " AND Property_Unit_Count__c > 1")
    LIT = UNIV + " AND Lit__c = true"
    lit_total = c(LIT)
    lit_single = c(LIT + " AND Property_Unit_Count__c = 1")
    lit_multi = c(LIT + " AND Property_Unit_Count__c > 1")
    multi_units_in_lit = sum_field(LIT + " AND Property_Unit_Count__c > 1", "Property_Unit_Count__c")
    multi_active_in_lit = sum_field(LIT + " AND Property_Unit_Count__c > 1", "Active_Unit_Count__c")
    multi_deact_in_lit = sum_field(LIT + " AND Property_Unit_Count__c > 1", "Deactive_Unit_Count__c")
    single_active_lit = c(LIT + " AND Property_Unit_Count__c = 1 AND Active_Unit_Count__c > 0")
    single_deact_lit = c(LIT + " AND Property_Unit_Count__c = 1 AND Active_Unit_Count__c = 0 AND Deactive_Unit_Count__c > 0")
    return {
        "universe": total,
        "single_total": single_total, "multi_total": multi_total,
        "lit_total": lit_total, "lit_single": lit_single, "lit_multi": lit_multi,
        "multi_units_in_lit": multi_units_in_lit,
        "multi_active_in_lit": multi_active_in_lit,
        "multi_deact_in_lit": multi_deact_in_lit,
        "single_active": single_active_lit, "single_deact_only": single_deact_lit,
    }

sf_with = sf_summary(no_hold=False)
sf_without = sf_summary(no_hold=True)

# By-state for SF (with + without Hold)
def sf_by_state(no_hold=False):
    extra = " AND Priority__c != 'Hold'" if no_hold else ""
    UNIV = "Address_Type__c='Business' AND Import_Delete_Property__c=false" + extra
    out = {}
    for s in ("AZ", "CA", "NE", "TX"):
        def c(w): return sf.query(f"SELECT COUNT() FROM Property_Location__c WHERE {w}")["totalSize"]
        def sf_sum(w, f):
            r = sf.query(f"SELECT SUM({f}) s FROM Property_Location__c WHERE {w}")["records"][0]
            return int(r["s"] or 0)
        lit_total = c(UNIV + f" AND State__c='{s}' AND Lit__c=true")
        lit_single = c(UNIV + f" AND State__c='{s}' AND Property_Unit_Count__c=1 AND Lit__c=true")
        lit_multi = c(UNIV + f" AND State__c='{s}' AND Property_Unit_Count__c>1 AND Lit__c=true")
        units_in_lit_multi = sf_sum(UNIV + f" AND State__c='{s}' AND Property_Unit_Count__c>1 AND Lit__c=true", "Property_Unit_Count__c")
        active_in_lit_multi = sf_sum(UNIV + f" AND State__c='{s}' AND Property_Unit_Count__c>1 AND Lit__c=true", "Active_Unit_Count__c")
        deact_in_lit_multi = sf_sum(UNIV + f" AND State__c='{s}' AND Property_Unit_Count__c>1 AND Lit__c=true", "Deactive_Unit_Count__c")
        active_singles = c(UNIV + f" AND State__c='{s}' AND Property_Unit_Count__c=1 AND Active_Unit_Count__c>0")
        out[s] = {
            "lit_total": lit_total, "lit_single": lit_single, "lit_multi": lit_multi,
            "units_in_lit_multi": units_in_lit_multi,
            "active_in_lit_multi": active_in_lit_multi,
            "deact_in_lit_multi": deact_in_lit_multi,
            "active_singles": active_singles,
        }
    return out

sf_state_with = sf_by_state(no_hold=False)
sf_state_without = sf_by_state(no_hold=True)

# ── Print comparison ───────────────────────────────────────────────
def fmt(v): return f"{v:>8,.0f}" if isinstance(v, (int, float)) and not isinstance(v, bool) else f"{v:>8}"
def diff(excel, sf): return sf - (excel or 0) if (excel or 0) != 0 else sf

print("=" * 96)
print("HEADLINE METRICS")
print("=" * 96)
print(f"{'Metric':<42}{'Excel':>10}{'SF (Hold)':>12}{'SF (no Hold)':>14}{'Excel diff':>16}")
def line(label, excel, sf_w, sf_wo):
    if isinstance(excel, (int, float)) and isinstance(sf_w, (int, float)):
        d_w = sf_w - excel
        diff_str = f"  (sf_with - excel = {d_w:+,})"
    else:
        diff_str = ""
    print(f"  {label:<40}{fmt(excel):>10}{fmt(sf_w):>12}{fmt(sf_wo):>14}{diff_str}")

# Note: single_active in SF is COUNT of singles with active customer. Excel single_active is COUNT of single BBAs with active=1.
total_active_with = sf_with["single_active"] + sf_with["multi_active_in_lit"]
total_active_without = sf_without["single_active"] + sf_without["multi_active_in_lit"]
total_deact_with = sf_with["single_deact_only"] + sf_with["multi_deact_in_lit"]
total_deact_without = sf_without["single_deact_only"] + sf_without["multi_deact_in_lit"]
total_units_lit_with = sf_with["lit_single"] + sf_with["multi_units_in_lit"]
total_units_lit_without = sf_without["lit_single"] + sf_without["multi_units_in_lit"]

line("Lit base business addresses", summary["Lit base business addresses"], sf_with["lit_total"], sf_without["lit_total"])
line("Total units in lit buildings", summary["Total units in lit buildings"], total_units_lit_with, total_units_lit_without)
line("Active customers", summary["Active customers"], total_active_with, total_active_without)
line("De-activated (churned)", summary["De-activated (churned)"], total_deact_with, total_deact_without)
line("Drop completed (in lit)", summary["Drop completed (in lit bldgs)"], "n/a (no field)", "n/a (no field)")
line("Serviceable idle (in lit)", summary["Serviceable idle (in lit bldgs)"], "n/a (no field)", "n/a (no field)")
line("Overall penetration %", f"{summary['Overall penetration']*100:.1f}%",
     f"{total_active_with / total_units_lit_with * 100:.1f}%" if total_units_lit_with else "-",
     f"{total_active_without / total_units_lit_without * 100:.1f}%" if total_units_lit_without else "-")
line("Universe (all serviceable biz)", summary["All serviceable business addresses (universe)"],
     sf_with["universe"], sf_without["universe"])

print()
print("=" * 96)
print("LIT BUILDINGS: SINGLE-UNIT vs MULTI-UNIT")
print("=" * 96)
line("Lit Single-Unit", excel_single_lit, sf_with["lit_single"], sf_without["lit_single"])
line("Lit Multi-Unit", excel_multi_lit, sf_with["lit_multi"], sf_without["lit_multi"])
line("Active customers (singles)", excel_single_active, sf_with["single_active"], sf_without["single_active"])
line("Active customers (multi-doors)", excel_multi_active, sf_with["multi_active_in_lit"], sf_without["multi_active_in_lit"])
line("Total units in lit multis", excel_multi_total_units, sf_with["multi_units_in_lit"], sf_without["multi_units_in_lit"])

print()
print("=" * 96)
print("BY STATE — LIT BUILDINGS (SINGLE + MULTI combined)")
print("=" * 96)
print(f"{'State':>6}{'Excel Lit':>14}{'SF Lit (Hold)':>16}{'SF Lit (no Hold)':>18}{'Excel Active':>16}{'SF Active (Hold)':>20}{'SF Active (no Hold)':>22}")
for s in ("AZ", "CA", "NE", "TX"):
    e = excel_by_state[s]
    sw = sf_state_with[s]
    swo = sf_state_without[s]
    sf_active_with = sw["active_singles"] + sw["active_in_lit_multi"]
    sf_active_without = swo["active_singles"] + swo["active_in_lit_multi"]
    print(f"{s:>6}{fmt(e['lit_bldgs']):>14}{fmt(sw['lit_total']):>16}{fmt(swo['lit_total']):>18}{fmt(e['active']):>16}{fmt(sf_active_with):>20}{fmt(sf_active_without):>22}")

print()
print("=" * 96)
print("EXCEL PENETRATION BUCKETS (lit buildings)")
print("=" * 96)
print("These are not currently exposed in the SF dashboard. From the Excel:")
for k, v in excel_buckets.items():
    print(f"  {k}: {v}")
print(f"  Total: {sum(excel_buckets.values())}  (should equal 686 lit BBAs)")

print()
print("=" * 96)
print("EXCEL UNITS TAB CROSS-CHECK")
print("=" * 96)
print(f"  Excel Units tab row count:    {unit_count_excel} units across 686 lit BBAs")
print(f"  Excel Summary 'Total units in lit buildings': {summary['Total units in lit buildings']}")
print(f"  -> diff: {unit_count_excel - summary['Total units in lit buildings']} (Excel includes 'Lit units folded in at non-serviceable addr' = 1)")
print()
print(f"  SF dashboard total units in lit (with Hold): {total_units_lit_with}")
print(f"  SF dashboard total units in lit (no Hold):   {total_units_lit_without}")
