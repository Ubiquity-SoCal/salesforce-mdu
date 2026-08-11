"""
Compare SF Property_Location__c.Priority__c (labeled "Priority", values
"Category 1/2/3"/"All Active"/"Hold") for business vs the Vetro penetration pull.

SF Priority formula:
  Hold                                  -> Hold
  Active == Total                       -> All Active
  Total==1 & Active==0                  -> Category 3
  Total>1 & (Active>0 OR Deactive>0)    -> Category 1   (multi-unit LIT)
  Total>1 & Active==0                   -> Category 2

My "lit" = Active>=1 OR Deactivated>=1 (any unit count). So my lit set should map to
SF (Category 1 + All Active), with single-unit-deactivated as an edge case SF mislabels
Category 3.

Read-only.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


XLSX = Path(r'C:\Users\cass\Work_Projects\Serviceability_Lookup\data\output\business-penetration-2026-05-22.xlsx')

sf = Salesforce(username=_SF["username"], password=_SF["password"],
                security_token=_SF["token"])

print('=' * 64)
print('SF Property_Location__c business by Priority__c (non-stale)')
print('=' * 64)
# Priority__c is a formula -> can't GROUP BY in SOQL; pull rows and tally in Python.
recs = sf.query_all(
    "SELECT Priority__c, Property_Unit_Count__c, Active_Unit_Count__c, "
    "Deactive_Unit_Count__c, State__c FROM Property_Location__c "
    "WHERE Address_Type__c='Business' AND Import_Delete_Property__c=false"
)['records']
prio = Counter()
cat1_c = cat1_u = cat1_a = cat1_d = 0
cat1_state = Counter()
allact_c = allact_u = allact_a = 0
for r in recs:
    p = r.get('Priority__c') or '(blank)'
    prio[p] += 1
    u = r.get('Property_Unit_Count__c') or 0
    a = r.get('Active_Unit_Count__c') or 0
    d = r.get('Deactive_Unit_Count__c') or 0
    if p == 'Category 1':
        cat1_c += 1; cat1_u += u; cat1_a += a; cat1_d += d
        cat1_state[r.get('State__c')] += 1
    elif p == 'All Active':
        allact_c += 1; allact_u += u; allact_a += a
for k, v in prio.most_common():
    print(f"   {k:<14} {v:,}")
print(f"\n  Category 1 : {cat1_c:,} bldgs | units={cat1_u:,.0f} active={cat1_a:,.0f} "
      f"deact={cat1_d:,.0f} pen={cat1_a/(cat1_u or 1):.1%}")
print(f"  All Active : {allact_c:,} bldgs | units={allact_u:,.0f} active={allact_a:,.0f}")
print(f"  -> Cat1 + All Active buildings (= 'has customers') = {cat1_c + allact_c:,}")
print('\n  Category 1 business by State:')
for s in sorted(cat1_state, key=lambda x: (x is None, x)):
    print(f"   {str(s):<4} {cat1_state[s]:,}")

# ── My side: classify the lit buildings the same way SF would ────────────────
print('\n' + '=' * 64)
print('MY Vetro pull: classify lit buildings by SF priority rule')
print('=' * 64)
wb = load_workbook(XLSX, read_only=True)
ws = wb['Base Business Addresses']
hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
ix = {h: i for i, h in enumerate(hdr)}
buckets = Counter()
cat1_units = cat1_active = 0
cat1_by_state = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    total = row[ix['Total Units']]; active = row[ix['Active']]; deact = row[ix['Deactivated']]
    state = row[ix['State']]
    if active == total:
        buckets['All Active'] += 1
    elif total > 1 and (active > 0 or deact > 0):
        buckets['Category 1'] += 1
        cat1_units += total; cat1_active += active
        cat1_by_state[state] += 1
    elif total == 1 and active == 0:
        buckets['Category 3 (single deactivated - SF mislabels)'] += 1
    else:
        buckets['other'] += 1

for k, v in buckets.most_common():
    print(f"   {k:<48} {v:,}")
print(f"\n  My Category-1 equivalent: {buckets['Category 1']:,} bldgs | "
      f"units={cat1_units:,} active={cat1_active:,} pen={cat1_active/(cat1_units or 1):.1%}")
print('  My Category-1 by State:')
for s in sorted(cat1_by_state):
    print(f"   {str(s):<4} {cat1_by_state[s]:,}")
