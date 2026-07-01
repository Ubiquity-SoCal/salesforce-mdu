"""
Coverage gap analysis: which ROE / PAL (and other) agreements in the CURRENT
IronClad export are NOT tracked as an Agreement__c in Salesforce?

Goal (Koa): every ROE and PAL should have an Agreement__c in SF. This finds the
ones that don't so we can decide whether to create them.

Method (respects the reverse-link gotcha):
  - The IronClad__c.Agreement__c reverse lookup is NOT maintained -> querying it
    gives massive false positives. Instead we build the "covered" set from the
    Agreement side: every Agreement__c.IronClad_Record__r.IronClad_Id__c and every
    Agreement__c.IronClad_ID__c (text). One IronClad record can back many agreements.
  - Then diff the export's Ironclad Ids against that covered set.

Read-only. Writes a CSV of gaps to data/output.

Run: python ironclad_roe_pal_coverage_gap.py
"""

import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from simple_salesforce import Salesforce

USERNAME = "cass1@ubiquitygp.com"
PASSWORD = "Hawaiian1984"
SECURITY_TOKEN = "IBSKT6CFUpSUJWxq1CMm0HkFC"

EXPORT = Path("C:/Users/cass/Work_Projects/IronClad/data/input/exports/ironclad_export_2026-07-01_151529_all.xlsx")
OUT = Path("C:/Users/cass/Work_Projects/SalesForce/data/output")
OUT.mkdir(parents=True, exist_ok=True)

# The agreement types Koa wants fully tracked in SF.
ROE_PAL = {"Right of Entry Agreement", "Premises Access License"}
# Stages that represent a live/real agreement worth having in SF. Cancelled is
# reported separately (a cancelled workflow usually doesn't need an SF record).
ACTIVE_STAGES = {"create", "review", "sign", "completed", "archive", "paused"}


def load_export():
    wb = load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb["export"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    hi = {h: i for i, h in enumerate(headers) if h}

    def g(row, name):
        i = hi.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in rows:
        ic_id = g(row, "Ironclad Id")
        if not ic_id:
            continue
        out.append({
            "ic_id": str(ic_id).strip(),
            "record_type": g(row, "Record Type"),
            "stage": (g(row, "Stage") or "").lower(),
            "mdu_bus": g(row, "MDU or BUS"),
            "record_name": g(row, "Record Name"),
            "property_name": g(row, "Property Name"),
            "property_city": g(row, "Property Address Locality"),
            "property_state": g(row, "Property Address Region"),
            "counterparty": g(row, "Counterparty Name"),
            "agreement_date": g(row, "Agreement Date"),
        })
    return out


def main():
    sf = Salesforce(username=USERNAME, password=PASSWORD, security_token=SECURITY_TOKEN)
    export = load_export()
    print(f"Export rows: {len(export)}")

    # Build the covered set from the Agreement side (both link fields).
    agrs = sf.query_all("""
        SELECT Id, Name, Status__c,
               IronClad_ID__c,
               IronClad_Record__r.IronClad_Id__c,
               Opportunity__r.Name,
               Opportunity__r.RecordType.DeveloperName
        FROM Agreement__c
    """)["records"]
    print(f"Agreement__c total: {len(agrs)}")

    covered = set()
    # ic_id -> list of agreements covering it (for reporting duplicates / who has it)
    cover_map = defaultdict(list)
    for a in agrs:
        ids = set()
        lk = (a.get("IronClad_Record__r") or {}).get("IronClad_Id__c")
        if lk:
            ids.add(str(lk).strip())
        txt = a.get("IronClad_ID__c")
        if txt:
            ids.add(str(txt).strip())
        for i in ids:
            covered.add(i)
            cover_map[i].append(a)
    print(f"Distinct IronClad Ids covered by an Agreement__c: {len(covered)}")

    # Diff
    rows_out = []
    summary = Counter()          # (type_bucket, stage_bucket) -> count of gaps
    active_roe_pal_gaps = []
    for r in export:
        if r["ic_id"] in covered:
            continue  # tracked
        # It's a gap.
        is_roe_pal = r["record_type"] in ROE_PAL
        type_bucket = r["record_type"] or "(blank)"
        active = r["stage"] in ACTIVE_STAGES
        summary[(type_bucket, "active" if active else r["stage"])] += 1
        rows_out.append(r)
        if is_roe_pal and active:
            active_roe_pal_gaps.append(r)

    # Report
    print("\n" + "=" * 70)
    print("GAP SUMMARY (in export, NOT tracked as an Agreement__c)")
    print("=" * 70)
    total_gaps = len(rows_out)
    print(f"Total export records with no SF Agreement: {total_gaps}")

    # ROE/PAL focus
    roe_pal_gap = [r for r in rows_out if r["record_type"] in ROE_PAL]
    roe_pal_active = [r for r in roe_pal_gap if r["stage"] in ACTIVE_STAGES]
    roe_pal_cancelled = [r for r in roe_pal_gap if r["stage"] == "cancelled"]
    print(f"\n  ROE/PAL gaps total:        {len(roe_pal_gap)}")
    print(f"    - ACTIVE (real gaps):    {len(roe_pal_active)}   <-- these matter")
    print(f"    - cancelled:             {len(roe_pal_cancelled)}")

    print("\n  ROE/PAL active gaps by stage:")
    for st, c in Counter(r["stage"] for r in roe_pal_active).most_common():
        print(f"      {st:<12} {c}")
    print("\n  ROE/PAL active gaps by MDU/BUS:")
    for mb, c in Counter(r["mdu_bus"] for r in roe_pal_active).most_common():
        print(f"      {str(mb):<12} {c}")

    print("\n  All gap types (type -> active / cancelled / other counts):")
    by_type = defaultdict(Counter)
    for r in rows_out:
        st = "active" if r["stage"] in ACTIVE_STAGES else r["stage"]
        by_type[r["record_type"] or "(blank)"][st] += 1
    for t, c in sorted(by_type.items(), key=lambda kv: -sum(kv[1].values())):
        print(f"      {str(t):<42} {dict(c)}")

    # Sample the active ROE/PAL gaps
    print("\n  Sample ACTIVE ROE/PAL gaps (first 25):")
    for r in active_roe_pal_gaps[:25]:
        print(f"    {r['ic_id']:<10} {str(r['record_type'])[:24]:<24} {r['stage']:<10} "
              f"{str(r['property_name'] or r['record_name'])[:34]:<34} {str(r['counterparty'])[:24]}")

    # Write full CSV of gaps
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = OUT / f"ironclad_roe_pal_coverage_gap_{ts}.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["IronClad_Id", "Record_Type", "Stage", "MDU_or_BUS", "Is_ROE_PAL",
                    "Active", "Record_Name", "Property_Name", "City", "State",
                    "Counterparty", "Agreement_Date"])
        # ROE/PAL active first (most actionable), then the rest
        ordered = (
            [r for r in rows_out if r["record_type"] in ROE_PAL and r["stage"] in ACTIVE_STAGES]
            + [r for r in rows_out if r["record_type"] in ROE_PAL and r["stage"] not in ACTIVE_STAGES]
            + [r for r in rows_out if r["record_type"] not in ROE_PAL]
        )
        for r in ordered:
            w.writerow([
                r["ic_id"], r["record_type"], r["stage"], r["mdu_bus"],
                r["record_type"] in ROE_PAL, r["stage"] in ACTIVE_STAGES,
                r["record_name"], r["property_name"], r["property_city"],
                r["property_state"], r["counterparty"],
                r["agreement_date"].strftime("%Y-%m-%d") if hasattr(r["agreement_date"], "strftime") else r["agreement_date"],
            ])
    print(f"\nFull gap CSV: {path}")


if __name__ == "__main__":
    main()
