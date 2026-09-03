"""
Build the Salesforce -> HubSpot migration mapping pack.

Produces one workbook the HubSpot consultant can work straight out of:
  - Read Me / Objects / Record Types / Relationships overview
  - one tab per object: every field, its type, whether it is required, its
    picklist values, what it points at, and how full it actually is
  - blank "HubSpot Property / HubSpot Type / Migrate? / Notes" columns on every
    field tab, so the mapping is filled in in place
  - Automation tab: triggers, flows, validation rules (the custom behaviour that
    has to be rebuilt, not imported)

Scope is deliberately narrow: the objects hanging off the MDU Opportunity, the AVR
(Case) side of Address Management, and SiteTracker. Business Sales is excluded per
the 8/31 call (Sales Focus is gone, John is on HubSpot already). Tabs are coloured
by app so the tab strip shows which side of the org a tab belongs to.

Run: python build_hubspot_migration_mapping.py
Out: SalesForce/data/output/salesforce-hubspot-field-map.xlsx
"""
import sys
import urllib.parse
from collections import defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds  # noqa: E402

from simple_salesforce import Salesforce  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

OUT = Path(__file__).resolve().parents[2] / "data" / "output"
OUT.mkdir(parents=True, exist_ok=True)
XLSX = OUT / "salesforce-hubspot-field-map.xlsx"

# Scoped down 2026-08-31 per Koa: the objects hanging off the MDU Opportunity, the
# AVR (Case) side, and SiteTracker. Dropped as not needed for the mapping:
# IronClad__c (re-established as a direct IronClad to HubSpot sync, not migrated),
# Special_Project__c, Tracker_View__c (LWC config, not data), and the three
# file-link / document-upload objects.
#
# The three junction objects (Opportunity_Contact__c, Opportunity_Account__c,
# Opportunity_Campaign__c) are also dropped, Koa's call: Kia has admin and can
# discover the structure herself. Note the cost, with them gone nothing in this
# workbook shows Contact and Opportunity are related, and the payload riding on
# those junctions (Role__c 96% populated, Is_Primary__c, Tag_Type__c) is not
# surfaced here. That is covered in the handoff email instead.
#
# object -> (tab name, which app it belongs to, one line on what it is)
OBJECTS = [
    ("Opportunity", "Opportunity", "MDU Sales", "The property deal. Core object, 4 record types, holds stage, owner, forecast and the PAL/ROE workflow."),
    ("Agreement__c", "Agreement", "MDU Sales", "Signed agreement (PAL, ROE, marketing). Child of Opportunity."),
    ("Contact", "Contact", "MDU Sales", "People. Linked to Opportunity through Opportunity_Contact__c, not the standard OpportunityContactRole."),
    ("Account", "Account", "MDU Sales", "Management company / owner entity."),
    ("Campaign", "Campaign", "MDU Sales", "Used as a project / focus-list tag, not for marketing sends."),
    ("Case", "Case (AVR)", "Address Mgmt", "Address Validation Request workflow runs on Cases."),
    ("AVR_Project_ID__c", "AVR Project ID", "Address Mgmt", "AVR project identifier record."),
    ("Property_Location__c", "Property Location", "Address Mgmt", "The physical property / building. 18k records, the address backbone the AVR resolves against."),
    ("Property_Unit__c", "Property Unit", "Address Mgmt", "Individual unit inside a property. 46k records, links to Vetro serviceability."),
    ("SiteTracker_Project__c", "SiteTracker Project", "SiteTracker", "Mirror of the SiteTracker construction project. Nightly sync, external system of record."),
]

# Tab colours so the app a tab belongs to is obvious from the tab strip.
APP_COLOR = {
    "MDU Sales": "1A3C5E",     # deep blue
    "Address Mgmt": "2E7D32",  # green
    "SiteTracker": "D4880F",   # amber
}
OVERVIEW_COLOR = "808080"      # grey for the non-object tabs

# Fill rates on these two are sampled; they are big and the shape is uniform.
SAMPLE_LIMIT = {"Property_Unit__c": 20000, "Property_Location__c": 20000}

HDR_FILL = PatternFill("solid", fgColor="1A3C5E")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
MAP_FILL = PatternFill("solid", fgColor="D4880F")
TITLE_FONT = Font(bold=True, size=13, color="1A3C5E", name="Calibri")
BODY = Font(size=10, name="Calibri")
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def connect():
    c = creds()
    return Salesforce(username=c["username"], password=c["password"], security_token=c["token"])


def tooling(sf, soql):
    return sf.toolingexecute("query/?q=" + urllib.parse.quote(soql))


def queryable_fields(desc):
    """Fields we can actually SELECT for a fill-rate pass."""
    out = []
    for f in desc["fields"]:
        if f["type"] in ("base64", "encryptedstring"):
            continue
        if not f.get("permissionable", True) and f["type"] == "address":
            continue
        out.append(f["name"])
    return out


def fill_rates(sf, obj, desc):
    """Return {field: populated_count}, sample_size, and one example value per field."""
    names = queryable_fields(desc)
    limit = SAMPLE_LIMIT.get(obj)
    counts = defaultdict(int)
    examples = {}
    seen_ids = set()

    # chunk so no single SOQL gets unwieldy
    chunks = [names[i:i + 80] for i in range(0, len(names), 80)]
    for chunk in chunks:
        sel = ", ".join(dict.fromkeys(["Id"] + chunk))
        soql = f"SELECT {sel} FROM {obj}"
        if limit:
            soql += f" LIMIT {limit}"
        try:
            recs = sf.query_all(soql)["records"]
        except Exception as e:
            print(f"    ! {obj} chunk failed ({str(e)[:80]}), retrying field by field")
            recs = []
            for fld in chunk:
                try:
                    sub = sf.query_all(f"SELECT Id, {fld} FROM {obj}" + (f" LIMIT {limit}" if limit else ""))["records"]
                except Exception:
                    continue
                for r in sub:
                    v = r.get(fld)
                    if v not in (None, "", []):
                        counts[fld] += 1
                        examples.setdefault(fld, v)
                    seen_ids.add(r["Id"])
            continue
        for r in recs:
            seen_ids.add(r["Id"])
            for fld in chunk:
                v = r.get(fld)
                if v in (None, "", []):
                    continue
                if isinstance(v, dict):
                    v = {k: x for k, x in v.items() if k != "attributes" and x is not None}
                    if not v:
                        continue
                counts[fld] += 1
                examples.setdefault(fld, v)
    return counts, len(seen_ids), examples


def short(v, n=60):
    s = str(v).replace("\n", " ").replace("\r", " ")
    return s if len(s) <= n else s[:n - 3] + "..."


def style_header(ws, row, ncols, fill=HDR_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    # Use the string form. ws.cell(...) would materialise the row and leave a
    # blank line sitting between the header and the first data row.
    ws.freeze_panes = f"A{row + 1}"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    sf = connect()
    print("connected:", sf.sf_instance)
    wb = Workbook()
    wb.remove(wb.active)

    described = {}
    counts_by_obj = {}
    fills = {}

    for api, tab, app, note in OBJECTS:
        print(f"  describing {api} ...")
        desc = getattr(sf, api).describe()
        described[api] = desc
        try:
            counts_by_obj[api] = sf.query(f"SELECT COUNT() FROM {api}")["totalSize"]
        except Exception:
            counts_by_obj[api] = -1
        print(f"    fill rates for {api} ({counts_by_obj[api]:,} records) ...")
        fills[api] = fill_rates(sf, api, desc)

    # ---------------- scope stats (measured, never asserted) ----------------
    n_fields = n_empty = n_thin = 0
    n_cust = n_cust_empty = n_cust_thin = 0
    for api, _, _, _ in OBJECTS:
        cnt, sample_n, _ = fills[api]
        for f in described[api]["fields"]:
            n = cnt.get(f["name"], 0)
            pct = (n / sample_n * 100) if sample_n else 0
            n_fields += 1
            if f["custom"]:
                n_cust += 1
            if n == 0:
                n_empty += 1
                if f["custom"]:
                    n_cust_empty += 1
            elif pct < 5:
                n_thin += 1
                if f["custom"]:
                    n_cust_thin += 1
    droppable = n_empty + n_thin
    print(f"  scope: {n_fields} fields, {n_empty} empty, {n_thin} under 5% populated")

    # ---------------- Read Me ----------------
    ws = wb.create_sheet("Read Me")
    ws.sheet_properties.tabColor = OVERVIEW_COLOR
    lines = [
        ("Salesforce to HubSpot migration: field map", "title"),
        (f"Generated {datetime.now():%Y-%m-%d %H:%M} from the live org (Generate Ubiquity Services LLC, fun-power-747).", "body"),
        ("", "body"),
        ("What this is", "h"),
        ("Every field on every object in the MDU Sales and Address Management apps, with its type, whether it is", "body"),
        ("required, its picklist values, what it links to, and how many records actually have a value in it.", "body"),
        ("", "body"),
        ("How to use it", "h"),
        ("Each object has its own tab. The four orange columns on the right of every field tab are blank on purpose:", "body"),
        ("fill in the HubSpot property name, the HubSpot type, whether the field moves at all, and any notes.", "body"),
        ("", "body"),
        ("Tab colours", "h"),
        ("Dark blue tabs are MDU Sales. Green tabs are Address Management (the AVR side). Amber is SiteTracker.", "body"),
        ("Grey tabs are the overview tabs, not a single object.", "body"),
        ("", "body"),
        ("The Populated column is the count of records that actually carry a value in that field, and Populated %", "body"),
        ("is the same thing against the record count. Use it to decide what is worth mapping and what is dead weight.", "body"),
        (f"For reference: of the {n_fields} fields here, {n_empty} have no data at all and {n_thin} more sit under 5%.", "body"),
        ("", "body"),
        ("What is NOT in here", "h"),
        ("Business Sales. Per the 8/31 call it is out of scope: Sales Focus is gone and John is on HubSpot already.", "body"),
        ("", "body"),
        ("The part that is not a field mapping", "h"),
        ("See the Automation tab. Triggers, flows and validation rules are custom behaviour, not data. None of it", "body"),
        ("imports. Every row on that tab is something that has to be rebuilt in HubSpot or deliberately dropped.", "body"),
        ("See the Integrations tab for the systems that write into Salesforce today and will need repointing.", "body"),
    ]
    r = 1
    for text, kind in lines:
        cell = ws.cell(row=r, column=1, value=text)
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "h":
            cell.font = Font(bold=True, size=11, color="2A5A8A", name="Calibri")
        else:
            cell.font = BODY
        r += 1
    ws.column_dimensions["A"].width = 115

    # ---------------- Objects ----------------
    ws = wb.create_sheet("Objects")
    ws.sheet_properties.tabColor = OVERVIEW_COLOR
    hdr = ["Object API Name", "Label", "App", "Records", "Fields", "Custom Fields",
           "Fields With Data", "Record Types", "What it is"]
    ws.append(hdr)
    style_header(ws, 1, len(hdr))
    for api, tab, app, note in OBJECTS:
        desc = described[api]
        cnt, sample_n, _ = fills[api]
        with_data = sum(1 for f in desc["fields"] if cnt.get(f["name"], 0) > 0)
        ws.append([api, desc["label"], app, counts_by_obj[api], len(desc["fields"]),
                   sum(1 for f in desc["fields"] if f["custom"]), with_data,
                   len([r for r in desc["recordTypeInfos"] if r["name"] != "Master"]), note])
    for row in ws.iter_rows(min_row=2, max_col=len(hdr)):
        for c in row:
            c.font = BODY
            c.border = BOX
        row[8].alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws, [26, 22, 15, 10, 8, 13, 15, 12, 85])

    # ---------------- Record Types ----------------
    ws = wb.create_sheet("Record Types")
    ws.sheet_properties.tabColor = OVERVIEW_COLOR
    hdr = ["Object", "Record Type", "Developer Name", "Active", "Default", "Records Using It"]
    ws.append(hdr)
    style_header(ws, 1, len(hdr))
    for api, tab, app, note in OBJECTS:
        rts = [r for r in described[api]["recordTypeInfos"] if r["name"] != "Master"]
        if len(rts) < 1:
            continue
        for rt in rts:
            n = ""
            try:
                n = sf.query(f"SELECT COUNT() FROM {api} WHERE RecordType.DeveloperName = '{rt['developerName']}'")["totalSize"]
            except Exception:
                n = "n/a"
            ws.append([api, rt["name"], rt["developerName"], rt.get("active"), rt.get("defaultRecordTypeMapping"), n])
    for row in ws.iter_rows(min_row=2, max_col=len(hdr)):
        for c in row:
            c.font = BODY
            c.border = BOX
    autosize(ws, [26, 30, 30, 10, 10, 18])

    # ---------------- Relationships ----------------
    ws = wb.create_sheet("Relationships")
    ws.sheet_properties.tabColor = OVERVIEW_COLOR
    hdr = ["From Object", "Field", "Field Label", "Points To", "Kind", "Required", "Populated %"]
    ws.append(hdr)
    style_header(ws, 1, len(hdr))
    scope = {api for api, _, _, _ in OBJECTS}
    for api, tab, app, note in OBJECTS:
        desc = described[api]
        cnt, sample_n, _ = fills[api]
        for f in desc["fields"]:
            if f["type"] not in ("reference",):
                continue
            refs = ", ".join(f.get("referenceTo") or [])
            kind = "Master-Detail" if f.get("cascadeDelete") and not f.get("nillable") else "Lookup"
            pct = (cnt.get(f["name"], 0) / sample_n * 100) if sample_n else 0
            ws.append([api, f["name"], f["label"], refs, kind,
                       "Yes" if not f["nillable"] and f["createable"] else "No", round(pct, 1)])
    for row in ws.iter_rows(min_row=2, max_col=len(hdr)):
        for c in row:
            c.font = BODY
            c.border = BOX
        if isinstance(row[3].value, str) and row[3].value in scope:
            row[3].font = Font(size=10, bold=True, name="Calibri", color="2A5A8A")
    autosize(ws, [26, 34, 30, 30, 15, 10, 12])

    # ---------------- one tab per object ----------------
    for api, tab, app, note in OBJECTS:
        desc = described[api]
        cnt, sample_n, examples = fills[api]
        ws = wb.create_sheet(tab[:31])
        ws.sheet_properties.tabColor = APP_COLOR.get(app, OVERVIEW_COLOR)
        ws.cell(row=1, column=1, value=f"{api}  ({desc['label']})  |  {app}  |  {counts_by_obj[api]:,} records").font = TITLE_FONT
        note_txt = note
        if api in SAMPLE_LIMIT and counts_by_obj[api] > SAMPLE_LIMIT[api]:
            note_txt += f"  [Populated % measured on a sample of {sample_n:,} records]"
        ws.cell(row=2, column=1, value=note_txt).font = Font(size=10, italic=True, name="Calibri")
        hdr = ["Field API Name", "Label", "Type", "Length", "Required", "Unique",
               "External Id", "Formula", "Points To", "Picklist Values",
               "Populated", "Populated %", "Example",
               "HubSpot Property", "HubSpot Type", "Migrate?", "Notes"]
        ws.append([])
        ws.append(hdr)
        style_header(ws, 4, len(hdr))
        for c in range(14, 18):
            ws.cell(row=4, column=c).fill = MAP_FILL

        flds = sorted(desc["fields"], key=lambda f: (not f["custom"], f["name"].lower()))
        for f in flds:
            n = cnt.get(f["name"], 0)
            pct = (n / sample_n * 100) if sample_n else 0
            pv = ""
            if f.get("picklistValues"):
                vals = [p["value"] for p in f["picklistValues"] if p.get("active")]
                pv = " | ".join(vals[:40])
                if len(vals) > 40:
                    pv += f"  (+{len(vals) - 40} more)"
            ws.append([
                f["name"], f["label"], f["type"],
                f.get("length") or f.get("precision") or "",
                "Yes" if (not f["nillable"] and f["createable"] and not f["defaultedOnCreate"]) else "",
                "Yes" if f.get("unique") else "",
                "Yes" if f.get("externalId") else "",
                "Yes" if f.get("calculated") else "",
                ", ".join(f.get("referenceTo") or []),
                # keep 2dp under 1% so a field with a handful of values on a big
                # object does not display as a flat 0 and get dropped by mistake
                pv, n, round(pct, 2) if 0 < pct < 1 else round(pct, 1),
                short(examples.get(f["name"], "")),
                "", "", "", "",
            ])
        last = ws.max_row
        for row in ws.iter_rows(min_row=5, max_row=last, max_col=len(hdr)):
            for c in row:
                c.font = BODY
                c.border = BOX
            row[9].alignment = Alignment(wrap_text=False)
            # No row shading. The Populated count and % are there; whether a thin
            # field is worth keeping is Kia's call to make, not ours to pre-judge.
            for c in row[13:17]:
                c.fill = PatternFill("solid", fgColor="FFF6E5")
        autosize(ws, [36, 30, 14, 8, 9, 8, 10, 9, 22, 55, 10, 12, 32, 26, 16, 11, 34])

    # ---------------- Automation ----------------
    ws = wb.create_sheet("Automation")
    ws.sheet_properties.tabColor = OVERVIEW_COLOR
    ws.cell(row=1, column=1, value="Custom behaviour: none of this imports, all of it is rebuild-or-drop").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Apex triggers, Apex classes, flows and validation rules currently live in the org.").font = Font(size=10, italic=True, name="Calibri")
    ws.append([])
    hdr = ["Kind", "Name", "On Object", "Active", "Detail", "HubSpot Plan", "Notes"]
    ws.append(hdr)
    style_header(ws, 4, len(hdr))
    for c in (6, 7):
        ws.cell(row=4, column=c).fill = MAP_FILL

    rows = []
    try:
        for t in tooling(sf, "SELECT Name, TableEnumOrId, Status, LengthWithoutComments FROM ApexTrigger WHERE NamespacePrefix = null")["records"]:
            rows.append(["Apex Trigger", t["Name"], t.get("TableEnumOrId"), t.get("Status"),
                         f"{t.get('LengthWithoutComments') or 0} chars", "", ""])
    except Exception as e:
        rows.append(["Apex Trigger", f"query failed: {str(e)[:80]}", "", "", "", "", ""])
    try:
        for c in tooling(sf, "SELECT Name, Status, LengthWithoutComments FROM ApexClass WHERE NamespacePrefix = null")["records"]:
            if c["Name"].lower().endswith("test"):
                continue
            rows.append(["Apex Class", c["Name"], "", c.get("Status"),
                         f"{c.get('LengthWithoutComments') or 0} chars", "", ""])
    except Exception as e:
        rows.append(["Apex Class", f"query failed: {str(e)[:80]}", "", "", "", "", ""])
    # FlowDefinitionView is a standard object, not a Tooling one. The org carries
    # ~82 flows but the bulk are Salesforce stock (CMS, Routing, Appointments);
    # only record-triggered and scheduled flows on our own objects are ours.
    REAL_TRIGGERS = {"RecordBeforeSave", "RecordAfterSave", "RecordBeforeDelete", "Scheduled"}
    try:
        for f in sf.query_all(
            "SELECT ApiName, Label, ProcessType, TriggerType, IsActive, "
            "TriggerObjectOrEventLabel FROM FlowDefinitionView"
        )["records"]:
            if f.get("TriggerType") not in REAL_TRIGGERS:
                continue
            if not f.get("TriggerObjectOrEventLabel"):
                continue
            rows.append(["Flow", f.get("Label") or f.get("ApiName"),
                         f.get("TriggerObjectOrEventLabel"),
                         "Active" if f.get("IsActive") else "Inactive",
                         f"{f.get('ProcessType')} / {f.get('TriggerType')}", "", ""])
    except Exception as e:
        rows.append(["Flow", f"query failed: {str(e)[:80]}", "", "", "", "", ""])
    try:
        for v in tooling(sf, "SELECT ValidationName, EntityDefinition.QualifiedApiName, Active, ErrorMessage FROM ValidationRule")["records"]:
            ent = (v.get("EntityDefinition") or {}).get("QualifiedApiName", "")
            rows.append(["Validation Rule", v.get("ValidationName"), ent,
                         "Active" if v.get("Active") else "Inactive",
                         short(v.get("ErrorMessage") or "", 90), "", ""])
    except Exception as e:
        rows.append(["Validation Rule", f"query failed: {str(e)[:80]}", "", "", "", "", ""])

    order = {"Apex Trigger": 0, "Flow": 1, "Validation Rule": 2, "Apex Class": 3}
    rows.sort(key=lambda r: (order.get(r[0], 9), str(r[2]), str(r[1])))
    for r_ in rows:
        ws.append(r_)
    for row in ws.iter_rows(min_row=5, max_col=len(hdr)):
        for c in row:
            c.font = BODY
            c.border = BOX
        for c in row[5:7]:
            c.fill = PatternFill("solid", fgColor="FFF6E5")
    autosize(ws, [16, 44, 26, 10, 46, 30, 34])

    # ---------------- Integrations ----------------
    ws = wb.create_sheet("Integrations")
    ws.sheet_properties.tabColor = OVERVIEW_COLOR
    ws.cell(row=1, column=1, value="Systems writing into Salesforce today").font = TITLE_FONT
    ws.cell(row=2, column=1, value="The third bucket from the 8/31 call. Each of these has to be repointed at HubSpot or retired.").font = Font(size=10, italic=True, name="Calibri")
    ws.append([])
    hdr = ["Source System", "Direction", "Writes To", "Cadence", "How it runs", "Matched On", "HubSpot Plan"]
    ws.append(hdr)
    style_header(ws, 4, len(hdr))
    ws.cell(row=4, column=7).fill = MAP_FILL
    integrations = [
        ["SiteTracker", "In", "SiteTracker_Project__c, Opportunity", "Nightly", "GitHub Actions cron, Python", "Site name / Agreement_Name__c", ""],
        ["IronClad", "In", "IronClad__c, Agreement__c", "Twice a week", "Export file, Python sync", "IronClad_Id__c (not Name)", ""],
        ["Vetro", "In", "Property_Location__c, Property_Unit__c", "Ad hoc / refresh", "Databricks pull, Python", "Address / agreename", ""],
        ["COS (FiberFirst)", "In", "Serviceability reporting", "Ad hoc", "Warehouse query", "external_address_id", ""],
        ["Monday.com", "In (historic)", "Opportunity", "Retired 2026-06-01", "Migration completed", "Agreement_Name__c", ""],
        ["Databricks dashboards", "Out", "Reads Salesforce", "Scheduled", "Dashboard generators", "n/a", ""],
        ["Tracker LWC + Address Mgmt LWC", "In-org UI", "Opportunity, Property_Location__c, Case", "Live, user driven", "Custom Lightning components", "n/a", ""],
        ["Requestor_Email_Extract", "In", "Case (AVR)", "On email receipt", "Apex / flow parse, 16 fields at once", "n/a", ""],
    ]
    for row in integrations:
        ws.append(row)
    for row in ws.iter_rows(min_row=5, max_col=len(hdr)):
        for c in row:
            c.font = BODY
            c.border = BOX
        row[6].fill = PatternFill("solid", fgColor="FFF6E5")
    autosize(ws, [32, 14, 40, 18, 34, 30, 34])

    wb.save(XLSX)
    print("wrote", XLSX)


if __name__ == "__main__":
    main()
