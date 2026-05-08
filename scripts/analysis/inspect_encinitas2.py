"""
Focused inspection of the Encinitas PBI file - cleaner output.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

FILE_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build\CompleteEncinitasView_PBI.xlsx"

wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False)

print("=" * 100)
print("1. POWERBI_DATA SHEET - COMPLETE COLUMN MAP")
print("=" * 100)

ws_dv = wb_values["PowerBI_Data"]
ws_df = wb_formulas["PowerBI_Data"]

print(f"Total rows: {ws_dv.max_row}, Total cols: {ws_dv.max_column}")
print()

# All headers with their column letters
print("ALL COLUMN HEADERS:")
print(f"{'Col':<5} {'Letter':<5} {'Header':<45} {'Hidden':<8} {'Width':<10}")
print("-" * 75)
for col in range(1, ws_dv.max_column + 1):
    letter = get_column_letter(col)
    header = ws_dv.cell(row=1, column=col).value or "(empty)"
    hidden = ws_dv.column_dimensions.get(letter, None)
    is_hidden = hidden.hidden if hidden else False
    width = hidden.width if hidden else "default"
    print(f"{col:<5} {letter:<5} {str(header):<45} {str(is_hidden):<8} {width}")

# Check formulas in rows 2-5 for ALL columns
print("\n\nFORMULA CHECK - Row 2 (all columns):")
for col in range(1, ws_dv.max_column + 1):
    letter = get_column_letter(col)
    header = ws_dv.cell(row=1, column=col).value
    form = ws_df.cell(row=2, column=col).value
    val = ws_dv.cell(row=2, column=col).value
    is_formula = isinstance(form, str) and form.startswith("=")
    if is_formula:
        print(f"  {letter} ({header}): FORMULA = {form}")
        print(f"     -> Evaluated value: {val}")

# Check a few more rows for formula patterns
print("\nFORMULA CHECK - Row 3:")
for col in range(1, ws_dv.max_column + 1):
    letter = get_column_letter(col)
    header = ws_dv.cell(row=1, column=col).value
    form = ws_df.cell(row=3, column=col).value
    val = ws_dv.cell(row=3, column=col).value
    is_formula = isinstance(form, str) and form.startswith("=")
    if is_formula:
        print(f"  {letter} ({header}): FORMULA = {form}")

# Specifically look at Address Type, Phase, Serviceable columns
print("\n\nENRICHED COLUMNS DETAIL:")
for col in range(1, ws_dv.max_column + 1):
    header = str(ws_dv.cell(row=1, column=col).value or "").lower()
    if any(k in header for k in ["address type", "phase", "serviceable", "fdh"]):
        letter = get_column_letter(col)
        actual_header = ws_dv.cell(row=1, column=col).value
        print(f"\n  Column {letter}: '{actual_header}'")
        # Check first 5 data rows
        for row in range(2, 7):
            form = ws_df.cell(row=row, column=col).value
            val = ws_dv.cell(row=row, column=col).value
            is_formula = isinstance(form, str) and form.startswith("=")
            if is_formula:
                print(f"    Row {row}: FORMULA={form} -> value={val}")
            else:
                print(f"    Row {row}: STATIC value={val}")


print("\n\n" + "=" * 100)
print("2. SUMMARY SHEET - COMPLETE STRUCTURE")
print("=" * 100)

ws_sv = wb_values["Summary"]
ws_sf = wb_formulas["Summary"]

print(f"Dimensions: {ws_sv.dimensions}")
print(f"Total rows: {ws_sv.max_row}, Total cols: {ws_sv.max_column}")
print()

# Column widths
print("COLUMN DIMENSIONS:")
for letter in "ABCDEFGHIJ":
    cd = ws_sv.column_dimensions.get(letter)
    if cd:
        print(f"  {letter}: width={cd.width}")
    else:
        print(f"  {letter}: (default)")
print()

# Merged cells
print("MERGED CELLS:", list(ws_sv.merged_cells.ranges))
print()

# Conditional formatting
print("CONDITIONAL FORMATTING:")
for cf in ws_sv.conditional_formatting:
    for rule in cf.rules:
        print(f"  Range: {cf.cells}")
        print(f"    Type: {rule.type}")
        if rule.formula:
            print(f"    Formula: {rule.formula}")
        if rule.dxf:
            dxf = rule.dxf
            if dxf.font:
                print(f"    Font: bold={dxf.font.bold}, color={dxf.font.color}")
            if dxf.fill:
                print(f"    Fill: {dxf.fill.fgColor}, pattern={dxf.fill.patternType}")
print()

# Row by row - first 30 rows with formatting summary
print("ROW-BY-ROW CONTENT (showing B-G columns):")
print("-" * 100)

for row in range(1, min(35, ws_sv.max_row + 1)):
    row_data = {}
    for col in range(1, ws_sv.max_column + 1):
        val_cell = ws_sv.cell(row=row, column=col)
        form_cell = ws_sf.cell(row=row, column=col)
        letter = get_column_letter(col)
        val = val_cell.value
        form = form_cell.value
        is_formula = isinstance(form, str) and form.startswith("=")

        if val is not None or is_formula:
            indent = val_cell.alignment.indent if val_cell.alignment.indent else 0
            bold = val_cell.font.bold
            h_align = val_cell.alignment.horizontal
            numfmt = val_cell.number_format if val_cell.number_format != "General" else None

            entry = f"{repr(val)}"
            tags = []
            if is_formula:
                tags.append(f"FORMULA:{form}")
            if bold:
                tags.append("BOLD")
            if indent > 0:
                tags.append(f"indent={indent}")
            if h_align:
                tags.append(f"align={h_align}")
            if numfmt:
                tags.append(f"fmt={numfmt}")

            # Border
            border = val_cell.border
            bsides = []
            for side_name in ["top", "bottom", "left", "right"]:
                side = getattr(border, side_name)
                if side and side.style:
                    bsides.append(f"{side_name}={side.style}")
            if bsides:
                tags.append(f"border=[{','.join(bsides)}]")

            if tags:
                entry += f"  [{', '.join(tags)}]"
            row_data[letter] = entry

    if row_data:
        print(f"\nRow {row}:")
        for letter, entry in sorted(row_data.items()):
            print(f"  {letter}: {entry}")


# Now check the LEFT side and RIGHT side patterns (I,J columns)
print("\n\n" + "=" * 100)
print("3. SUMMARY SHEET - RIGHT SIDE (Columns I, J) - First 30 rows")
print("=" * 100)
for row in range(1, min(35, ws_sv.max_row + 1)):
    for col in [9, 10]:  # I and J
        val_cell = ws_sv.cell(row=row, column=col)
        form_cell = ws_sf.cell(row=row, column=col)
        letter = get_column_letter(col)
        val = val_cell.value
        form = form_cell.value
        is_formula = isinstance(form, str) and form.startswith("=")
        if val is not None or is_formula:
            indent = val_cell.alignment.indent if val_cell.alignment.indent else 0
            bold = val_cell.font.bold
            entry = f"Row {row} {letter}: {repr(val)}"
            if is_formula:
                entry += f"  FORMULA:{form}"
            if bold:
                entry += " [BOLD]"
            if indent:
                entry += f" [indent={indent}]"
            print(entry)


# Understand the grouping pattern - analyze deeper
print("\n\n" + "=" * 100)
print("4. SUMMARY SHEET - GROUPING PATTERN ANALYSIS (B column levels)")
print("=" * 100)
print("Analyzing indent levels and data types to understand hierarchy:")
for row in range(2, min(50, ws_sv.max_row + 1)):
    cell = ws_sv.cell(row=row, column=2)  # Column B
    if cell.value is not None:
        indent = cell.alignment.indent if cell.alignment.indent else 0
        bold = cell.font.bold
        val = cell.value
        val_type = type(val).__name__

        # Determine what kind of row this is
        if bold and indent == 0:
            row_type = "PHASE HEADER"
        elif indent == 1 and isinstance(val, str) and "FDH" in str(val):
            row_type = "FDH NAME"
        elif indent == 2 and isinstance(val, datetime):
            row_type = "SERVICEABLE DATE"
        elif indent == 0 and not bold:
            row_type = "PHASE NUMBER?"
        else:
            row_type = "UNKNOWN"

        print(f"  Row {row}: indent={indent}, bold={bold}, type={val_type}, "
              f"value={repr(val)}, -> {row_type}")


# Distinct values in key columns
print("\n\n" + "=" * 100)
print("5. SUMMARY - COLUMN D and E HEADERS AND FORMULAS")
print("=" * 100)
for col in range(2, ws_sv.max_column + 1):
    letter = get_column_letter(col)
    # Check row 2 (likely header area) and row 3
    for r in [2, 3, 4, 5]:
        val = ws_sv.cell(row=r, column=col).value
        form = ws_sf.cell(row=r, column=col).value
        is_formula = isinstance(form, str) and form.startswith("=")
        if val is not None or is_formula:
            print(f"  Row {r} Col {letter}: value={repr(val)}", end="")
            if is_formula:
                print(f"  FORMULA={form}", end="")
            print()


# Check formulas in D and E columns for data rows
print("\n\nFORMULAS IN D, E, G COLUMNS (rows 5-20):")
for row in range(5, 21):
    for col_letter, col_idx in [("D", 4), ("E", 5), ("G", 7)]:
        form = ws_sf.cell(row=row, column=col_idx).value
        val = ws_sv.cell(row=row, column=col_idx).value
        is_formula = isinstance(form, str) and form.startswith("=")
        if val is not None or is_formula:
            line = f"  Row {row} {col_letter}: value={repr(val)}"
            if is_formula:
                line += f"  FORMULA={form}"
            print(line)

wb_values.close()
wb_formulas.close()
print("\n\nDone.")
