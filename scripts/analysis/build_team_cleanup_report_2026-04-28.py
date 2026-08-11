"""
MDU Sales team cleanup report.
Outputs a multi-tab Excel workbook the team can use to drive cleanup.

Tabs:
  1. Summary
  2. Stage requires Agreement (Opps with stage that should have certain agreements)
  3. Agreements missing IronClad ID (focus on Signed/Completed first)
  4. Opps missing required links/fields
  5. Stale active Opps (no recent notes)
"""

from datetime import datetime, timezone, timedelta
from pathlib import Path
from collections import defaultdict
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


USERNAME = _SF["username"]
PASSWORD = _SF["password"]
SECURITY_TOKEN = _SF["token"]
LIGHTNING = "https://fun-power-747.lightning.force.com/lightning/r"
OUT = Path(__file__).parent / f"MDU_Cleanup_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

NAVY = "1F3A68"
ACCENT = "2E7D32"
WHITE = "FFFFFF"
LIGHT = "F0F4FA"
WARN = "FFE0B2"
RED = "FCE4EC"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER


def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autofit(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        m = 10
        for c in col:
            v = c.value
            if v is None:
                continue
            l = len(str(v))
            if l > m:
                m = l
        ws.column_dimensions[letter].width = min(m + 2, max_width)


def opp_link(opp_id):
    return f"{LIGHTNING}/Opportunity/{opp_id}/view"


def agr_link(agr_id):
    return f"{LIGHTNING}/Agreement__c/{agr_id}/view"


def main():
    sf = Salesforce(username=USERNAME, password=PASSWORD, security_token=SECURITY_TOKEN)
    print("Connected to Salesforce")

    # ---------- Pull active MDU Opps ----------
    soql_opps = """
        SELECT Id, Name, StageName, Sales_Status__c, Hold_Reason__c, Loss_Reason__c,
               Property_City__c, Property_State__c, Property_Zip__c, Units__c,
               Property_Type__c, Property_Location__c,
               Agreement_Name__c, SiteTracker_Project_ID__c, Projected_Close_Date__c,
               Agreement_Count__c, Notes_Count__c,
               Owner.Name, Owner.IsActive,
               RE_Assigned__c, RE_Assigned__r.Name,
               LastModifiedDate
        FROM Opportunity
        WHERE RecordType.DeveloperName = 'MDU'
          AND IsClosed = FALSE
    """
    opps = sf.query_all(soql_opps)["records"]
    print(f"Active MDU Opps: {len(opps)}")

    # ---------- Pull all Agreements (active Opps only) ----------
    opp_ids = [o["Id"] for o in opps]
    agreements_by_opp = defaultdict(list)
    # SOQL has 4000-element IN limit, chunk if needed
    chunk = 500
    for i in range(0, len(opp_ids), chunk):
        ids_str = ",".join(f"'{x}'" for x in opp_ids[i:i + chunk])
        soql_a = f"""
            SELECT Id, Name, Agreement_Type__c, Status__c, IronClad_ID__c,
                   Signed_Date__c, Requested_Date__c, Opportunity__c
            FROM Agreement__c
            WHERE Opportunity__c IN ({ids_str})
        """
        for a in sf.query_all(soql_a)["records"]:
            agreements_by_opp[a["Opportunity__c"]].append(a)
    total_agr = sum(len(v) for v in agreements_by_opp.values())
    print(f"Agreements on active MDU Opps: {total_agr}")

    # ---------- Pull recent ContentNote dates per Opp ----------
    last_note_by_opp = {}
    for i in range(0, len(opp_ids), chunk):
        ids_str = ",".join(f"'{x}'" for x in opp_ids[i:i + chunk])
        soql_cdl = f"""
            SELECT LinkedEntityId,
                   ContentDocument.LatestPublishedVersion.CreatedDate
            FROM ContentDocumentLink
            WHERE LinkedEntityId IN ({ids_str})
              AND ContentDocument.FileType = 'SNOTE'
        """
        for r in sf.query_all(soql_cdl)["records"]:
            opp_id = r["LinkedEntityId"]
            cd = (r.get("ContentDocument") or {}).get("LatestPublishedVersion") or {}
            created = cd.get("CreatedDate")
            if not created:
                continue
            prev = last_note_by_opp.get(opp_id)
            if prev is None or created > prev:
                last_note_by_opp[opp_id] = created
    print(f"Opps with at least one note: {len(last_note_by_opp)}")

    # ---------- Build workbook ----------
    wb = Workbook()
    wb.remove(wb.active)

    # === Tab 2: Stage requires Agreement ===
    ws = wb.create_sheet("Stage missing Agreement")
    headers = ["Property", "Stage", "Sales Status", "What's missing",
               "Opp Owner", "RE Assigned",
               "City", "State", "Units", "Agreement_Name",
               "Projected Close", "Last Modified", "Open Opp"]
    ws.append(headers)
    style_header_row(ws)

    stage_missing = []
    for o in opps:
        stage = o.get("StageName")
        agrs = agreements_by_opp.get(o["Id"], [])
        signed_pal = any(a["Agreement_Type__c"] == "PAL"
                         and a["Status__c"] in ("Completed", "Sign", "Archive")
                         for a in agrs)
        signed_roe = any(a["Agreement_Type__c"] == "ROE"
                         and a["Status__c"] in ("Completed", "Sign", "Archive")
                         for a in agrs)
        active_agrs = [a for a in agrs if a["Status__c"] in ("Create", "Review", "Sign", "Completed", "Archive")]

        missing = None
        if stage == "Under Contract" and not signed_pal:
            missing = "Under Contract but no signed PAL"
        elif stage == "ROE Secured" and not signed_roe:
            missing = "ROE Secured but no signed ROE Agreement"
        elif stage == "Contract Negotiations" and not active_agrs:
            missing = "Contract Negotiations but no active Agreement"
        elif stage == "Engaged" and not active_agrs:
            missing = "Engaged with no Agreement record (consider creating draft)"
        if not missing:
            continue
        stage_missing.append((o, missing))
        owner = (o.get("Owner") or {}).get("Name") or ""
        re_assigned = (o.get("RE_Assigned__r") or {}).get("Name") or ""
        ws.append([
            o["Name"], stage, o.get("Sales_Status__c") or "", missing,
            owner, re_assigned,
            o.get("Property_City__c") or "", o.get("Property_State__c") or "",
            int(o.get("Units__c") or 0),
            o.get("Agreement_Name__c") or "",
            (o.get("Projected_Close_Date__c") or "")[:10] if o.get("Projected_Close_Date__c") else "",
            (o.get("LastModifiedDate") or "")[:10],
            "Open",
        ])
        link_cell = ws.cell(row=ws.max_row, column=len(headers))
        link_cell.hyperlink = opp_link(o["Id"])
        link_cell.font = LINK_FONT
    freeze_and_filter(ws)
    autofit(ws)

    # === Tab 3: Agreements missing IronClad ID ===
    ws = wb.create_sheet("Agreements missing IronClad ID")
    headers = ["Agreement", "Type", "Status", "Signed Date",
               "Property", "Stage", "Opp Owner", "RE Assigned",
               "Agreement_Name", "Priority", "Open Agreement", "Open Opp"]
    ws.append(headers)
    style_header_row(ws)
    missing_ic = []
    for o in opps:
        for a in agreements_by_opp.get(o["Id"], []):
            if a.get("IronClad_ID__c"):
                continue
            status = a.get("Status__c") or ""
            # Skip Cancelled / blank — those aren't real
            if status in ("Cancelled", "", None):
                continue
            # Priority: Signed/Completed first, then Sign, then Review, then Create/Paused
            prio = {"Completed": "1 - High", "Archive": "1 - High", "Sign": "2 - Medium",
                    "Review": "3 - Low", "Create": "4 - Low", "Paused": "5 - Defer"}.get(status, "5 - Defer")
            missing_ic.append((o, a, prio))
            owner = (o.get("Owner") or {}).get("Name") or ""
            re_assigned = (o.get("RE_Assigned__r") or {}).get("Name") or ""
            ws.append([
                a["Name"], a["Agreement_Type__c"], status,
                (a.get("Signed_Date__c") or "")[:10] if a.get("Signed_Date__c") else "",
                o["Name"], o.get("StageName") or "", owner, re_assigned,
                o.get("Agreement_Name__c") or "", prio, "Open Agreement", "Open Opp",
            ])
            row = ws.max_row
            ws.cell(row=row, column=11).hyperlink = agr_link(a["Id"])
            ws.cell(row=row, column=11).font = LINK_FONT
            ws.cell(row=row, column=12).hyperlink = opp_link(o["Id"])
            ws.cell(row=row, column=12).font = LINK_FONT
    freeze_and_filter(ws)
    autofit(ws)

    # === Tab 4: Opps missing required links/fields ===
    ws = wb.create_sheet("Opps missing data")
    headers = ["Property", "Stage", "What's missing", "Opp Owner", "RE Assigned",
               "City", "State", "Zip", "Property_Location linked?",
               "Agreement_Name", "Last Modified", "Open Opp"]
    ws.append(headers)
    style_header_row(ws)
    missing_data = []
    for o in opps:
        flags = []
        if not o.get("Property_Location__c"):
            flags.append("No Property Location")
        if not (o.get("Owner") or {}).get("IsActive", True):
            flags.append("Owner inactive")
        if not o.get("RE_Assigned__c") and o.get("StageName") in ("ROE Secured", "Contract Negotiations", "Under Contract"):
            flags.append("No RE Assigned for active stage")
        if not o.get("Agreement_Name__c"):
            flags.append("No Agreement_Name (cross-system key)")
        if not o.get("Projected_Close_Date__c") and o.get("StageName") in ("Contract Negotiations", "ROE Secured", "Under Contract"):
            flags.append("No Projected Close Date")
        if not flags:
            continue
        missing_data.append((o, flags))
        owner = (o.get("Owner") or {}).get("Name") or ""
        re_assigned = (o.get("RE_Assigned__r") or {}).get("Name") or ""
        ws.append([
            o["Name"], o.get("StageName") or "",
            "; ".join(flags), owner, re_assigned,
            o.get("Property_City__c") or "", o.get("Property_State__c") or "",
            o.get("Property_Zip__c") or "",
            "Yes" if o.get("Property_Location__c") else "No",
            o.get("Agreement_Name__c") or "",
            (o.get("LastModifiedDate") or "")[:10],
            "Open",
        ])
        link_cell = ws.cell(row=ws.max_row, column=len(headers))
        link_cell.hyperlink = opp_link(o["Id"])
        link_cell.font = LINK_FONT
    freeze_and_filter(ws)
    autofit(ws)

    # === Tab 5: Stale active Opps ===
    ws = wb.create_sheet("Stale active Opps")
    headers = ["Property", "Stage", "Days since last note", "Opp Owner", "RE Assigned",
               "City", "State", "Last Note Date", "Last Modified", "Open Opp"]
    ws.append(headers)
    style_header_row(ws)
    now = datetime.now(timezone.utc)
    stale = []
    for o in opps:
        stage = o.get("StageName")
        if stage not in ("Engaged", "Contract Negotiations", "ROE Secured", "Under Contract"):
            continue
        last_note = last_note_by_opp.get(o["Id"])
        if not last_note:
            days_since = None
            label = "(no notes ever)"
        else:
            try:
                dt = datetime.fromisoformat(last_note.replace("Z", "+00:00"))
                days_since = (now - dt).days
            except Exception:
                continue
            if days_since < 60:
                continue
            label = (last_note or "")[:10]
        stale.append((o, days_since, label))
        owner = (o.get("Owner") or {}).get("Name") or ""
        re_assigned = (o.get("RE_Assigned__r") or {}).get("Name") or ""
        ws.append([
            o["Name"], stage,
            days_since if days_since is not None else "n/a",
            owner, re_assigned,
            o.get("Property_City__c") or "", o.get("Property_State__c") or "",
            label, (o.get("LastModifiedDate") or "")[:10],
            "Open",
        ])
        link_cell = ws.cell(row=ws.max_row, column=len(headers))
        link_cell.hyperlink = opp_link(o["Id"])
        link_cell.font = LINK_FONT
    freeze_and_filter(ws)
    autofit(ws)

    # === Tab 1: Summary (insert at front) ===
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "MDU Sales Cleanup Report"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    ws["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A3"] = f"Active MDU Opportunities scanned: {len(opps)}"
    ws["A3"].font = Font(name="Calibri", size=10, color="555555")

    ws.append([])
    rows = [
        ["Cleanup category", "Count", "Tab", "What to do"],
        ["Stage missing required Agreement", len(stage_missing), "Stage missing Agreement",
         "Open each Opp, create or update the Agreement that the stage requires."],
        ["Agreements missing IronClad ID", len(missing_ic), "Agreements missing IronClad ID",
         "RE/Sales: paste the IronClad ID into IronClad ID field. Sort by Priority column."],
        ["Opps missing required data", len(missing_data), "Opps missing data",
         "Fill the field listed in 'What's missing'. Property Location, RE Assigned, Projected Close, etc."],
        ["Stale active Opps (no notes 60+ days)", len(stale), "Stale active Opps",
         "Either log a note with current status, move to On Hold with a reason, or close as Lost."],
    ]
    start = ws.max_row + 1
    for r in rows:
        ws.append(r)
    style_header_row(ws, row=start)

    # Bold the count column
    for r in range(start + 1, start + len(rows)):
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 70

    # Add a "How to use this report" section
    ws.append([])
    ws.append(["How to use this report"])
    ws.cell(row=ws.max_row, column=1).font = Font(name="Calibri", size=12, bold=True, color=NAVY)
    notes = [
        "1. Open the tab matching what you own. Each row links directly to the Salesforce record (rightmost column).",
        "2. Sort or filter by Owner / RE Assigned to see only your work.",
        "3. The IronClad ID tab is sorted by Priority. Tackle '1 - High' rows first (Completed agreements that need the ID for audit trail).",
        "4. If an Opp should not be in this report (already cleaned, no longer relevant), update Salesforce; the next refresh of this report will drop it.",
        "5. Reach out to Cass with questions or if a record looks wrong.",
    ]
    for n in notes:
        ws.append([n])

    wb.save(OUT)
    print(f"\nWrote: {OUT}")
    print(f"\nSummary:")
    print(f"  Stage missing Agreement:      {len(stage_missing)}")
    print(f"  Agreements missing IronClad:  {len(missing_ic)}")
    print(f"  Opps missing data:            {len(missing_data)}")
    print(f"  Stale active Opps (60+ days): {len(stale)}")


if __name__ == "__main__":
    main()
