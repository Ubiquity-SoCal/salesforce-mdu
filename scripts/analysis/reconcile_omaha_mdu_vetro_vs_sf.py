"""
Reconcile the 65 Omaha MDUs from "OMAHA MDUs from Vetro.xlsx" against Salesforce.
Question: which are already loaded (matched by Agreement_Name OR address), which are
missing and need loading. Match by name may miss them, so address is the fallback.

Read-only. Output: console buckets + CSV.
NOTE (per Koa 2026-07-08 correction): do NOT treat these as the ONLY Cat 1 MDUs and
do NOT hunt SF Cat1-not-on-list for re-bucketing - that premise is retracted.
"""
import sys, csv
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl
from rapidfuzz import fuzz
from simple_salesforce import Salesforce

sys.path.insert(0, str(Path(r"C:\Users\cass\Work_Projects\SalesForce\scripts\analysis")))
from lookup_agree_names_for_unlinked import house, st_tokens, norm_name, numset

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


USERNAME = _SF["username"]; PASSWORD = _SF["password"]; SECURITY_TOKEN = _SF["token"]
VETRO = Path(r"C:\Users\cass\AppData\Local\Temp\claude\C--Users-cass-Work-Projects\eb63bd25-744f-4859-9c80-939403c2cb8e\scratchpad\omaha_mdus.xlsx")
OUT = Path(r"C:\Users\cass\Work_Projects\SalesForce\data\output\2026-07-08-omaha-mdu-vetro-vs-sf.csv")

def load_vetro_list():
    ws = openpyxl.load_workbook(VETRO, read_only=True, data_only=True).active
    rows = [r for r in list(ws.iter_rows(values_only=True))[1:] if any(c not in (None, "") for c in r)]
    return [dict(site=str(r[0]).strip(), monday=str(r[1] or "").strip(),
                 cat=str(r[2] or "").strip(), units=r[3], addr=str(r[4] or "").strip())
            for r in rows]

def main():
    sf = Salesforce(username=USERNAME, password=PASSWORD, security_token=SECURITY_TOKEN)
    vlist = load_vetro_list()
    print(f"Vetro Omaha MDUs: {len(vlist)}")

    q = """SELECT Id, Name, Agreement_Name__c, Property_Address__c, Property_City__c,
           Property_State__c, Property_Category__c, StageName, RecordType.Name,
           Loss_Reason__c, CloseDate
           FROM Opportunity
           WHERE Name LIKE 'Omaha%' OR Agreement_Name__c LIKE 'Omaha%'
              OR Property_City__c LIKE 'Omaha%'
              OR (Property_State__c IN ('NE','Nebraska') AND RecordType.Name='MDU/SFU')"""
    opps = sf.query_all(q)["records"]
    print(f"SF Omaha-universe Opps: {len(opps)}")
    SOFT = {"No Contact Info", "No Decision / Non-Responsive", "Other", "None", None}

    # indexes
    by_agree = defaultdict(list)
    by_house = defaultdict(list)
    for o in opps:
        agn = (o["Agreement_Name__c"] or "").strip().lower()
        if agn:
            by_agree[agn].append(o)
        h = house(o["Property_Address__c"])
        if h:
            by_house[h].append(o)

    def gather(v):
        """Union of all candidate Opps: agree-name exact + address(house+street)."""
        hits = {}
        basis = set()
        for o in by_agree.get(v["site"].lower(), []):
            hits[o["Id"]] = o; basis.add("agree-name")
        h = house(v["addr"]); gt = st_tokens(v["addr"])
        for o in by_house.get(h, []):
            ov = gt & st_tokens(o["Property_Address__c"])
            if ov:
                hits[o["Id"]] = o; basis.add("address")
        return list(hits.values()), basis

    def alpha_toks(a):
        return {t for t in st_tokens(a) if not t.isdigit()}

    def near_addr(v):
        """Same distinctive street (>=2 shared alpha tokens) + house number within 25.
        Catches e.g. 8515 vs 8509 Indian Hills Dr; won't merge numbered streets."""
        h = house(v["addr"]); ga = alpha_toks(v["addr"])
        if not h or len(ga) < 2:
            return None
        hi = int(h); out = []
        for o in opps:
            oh = house(o["Property_Address__c"])
            if not oh:
                continue
            if len(ga & alpha_toks(o["Property_Address__c"])) >= 2 and abs(int(oh) - hi) <= 25:
                out.append((abs(int(oh) - hi), o))
        if not out:
            return None
        out.sort(key=lambda x: x[0])
        detail = f"{house(out[0][1]['Property_Address__c'])} vs {h}"
        return [o for _, o in out], detail

    def fuzzy(v):
        q = norm_name(v["site"]); bf = None
        for o in opps:
            cand = norm_name(o["Agreement_Name__c"] or o["Name"])
            if not cand:
                continue
            score = min(fuzz.token_set_ratio(q, cand), fuzz.ratio(q, cand))
            if bf is None or score > bf[0]:
                bf = (score, o)
        return bf if (bf and bf[0] >= 90) else None

    def pick(hits):  # representative Opp: prefer active, then most-advanced-ish
        return sorted(hits, key=lambda o: (o["StageName"] == "Closed Lost",))[0]

    ACTIVE = lambda o: o["StageName"] != "Closed Lost"
    buckets = defaultdict(list)
    rows_out = []
    for v in vlist:
        hits, basis = gather(v)
        how = "+".join(sorted(basis)) if basis else ""
        if not hits:
            n = near_addr(v)
            if n:
                hits = n[0]; how = f"address (near: {n[1]})"
        if not hits:
            f = fuzzy(v)
            if f:
                hits = [f[1]]; how = f"name-fuzzy({f[0]})"
        active = [o for o in hits if ACTIVE(o)]
        if not hits:
            bucket = "MISSING"; rep = None
        elif active:
            bucket = "loaded-ACTIVE"; rep = pick(active)
        else:
            bucket = "loaded-CLOSED-LOST"; rep = pick(hits)
        buckets[bucket].append((v, rep, how, len(hits), len(active)))

        # helpful note
        if rep is None:
            note = "Not in Salesforce - needs loading"
        elif bucket == "loaded-CLOSED-LOST":
            r = rep["Loss_Reason__c"]; tag = "soft, re-approachable" if r in SOFT else "hard - fiber won't change"
            note = f"Closed Lost: {r} ({str(rep['CloseDate'])[:7]}) - {tag}"
        elif rep["StageName"] in ("PAL/ROE Complete", "Marketing/Bulk Complete", "Marketing/Bulk In Progress"):
            note = f"Active - secured/in build ({rep['StageName']})"
        else:
            note = f"Active in pipeline ({rep['StageName']})"
        if "near" in how:
            note = f"{note}  [address off by a few - verify same property]"
        if len(hits) > 1:
            note = f"{note}  [{len(hits)} SF Opps on this property]"

        # readable match basis
        basis_label = {"agree-name": "Agree-name", "address": "Address",
                       "address+agree-name": "Agree-name + Address"}.get(how, how or "MISSING")

        rows_out.append(dict(
            vetro_site=v["site"], vetro_addr=v["addr"], vetro_units=v["units"], vetro_cat=v["cat"],
            in_sf=("No" if rep is None else "Yes"),
            current_status=(rep["StageName"] if rep else ""),
            sf_category=(rep["Property_Category__c"] or "" if rep else ""),
            helpful_note=note, match_basis=basis_label, bucket=bucket,
            n_opps=len(hits), sf_name=(rep["Name"] if rep else ""),
            sf_agree=(rep["Agreement_Name__c"] or "" if rep else ""),
            sf_id=(rep["Id"] if rep else "")))

    order = ["loaded-ACTIVE", "loaded-CLOSED-LOST", "MISSING"]
    print("\n=== SUMMARY ===")
    for k in order:
        u = sum(int(v['units'] or 0) for v, *_ in buckets.get(k, []))
        print(f"  {k:20} {len(buckets.get(k, [])):3}  ({u:,} units)")
    print(f"  {'TOTAL':20} {len(vlist):3}")

    for k in order:
        if not buckets.get(k):
            continue
        print(f"\n--- {k} ({len(buckets[k])}) ---")
        for v, o, how, n, na in sorted(buckets[k], key=lambda x: -int(x[0]['units'] or 0)):
            if o:
                dup = f" x{n}" if n > 1 else ""
                extra = f" reason={o['Loss_Reason__c']} ({str(o['CloseDate'])[:7]})" if k == "loaded-CLOSED-LOST" else ""
                print(f"  {v['site'][:30]:30} u={str(v['units']):4} -> {o['Name'][:22]:22} "
                      f"[{o['Property_Category__c'] or '-'}/{o['StageName'][:14]}]{dup}{extra}")
            else:
                print(f"  {v['site'][:30]:30} u={str(v['units']):4} -> (LOAD) {v['addr'][:38]}")

    # loss-reason breakdown for the matched closed-lost + soft/hard split
    cl = buckets.get("loaded-CLOSED-LOST", [])
    reasons = Counter(o["Loss_Reason__c"] for v, o, *_ in cl)
    soft = [(v, o) for v, o, *_ in cl if o["Loss_Reason__c"] in SOFT]
    print(f"\n=== Closed-Lost matched to fiber list: loss reasons ({len(cl)}) ===")
    for r, c in reasons.most_common():
        tag = "  <- soft, re-approachable" if r in SOFT else ""
        print(f"  {str(r):32} {c}{tag}")
    su = sum(int(v['units'] or 0) for v, o in soft)
    print(f"  SOFT (re-approachable): {len(soft)} MDUs / {su:,} units")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader(); w.writerows(rows_out)
    print(f"\nwrote {OUT}")

    # ---- formatted Excel deliverable ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    COLS = [("Vetro Site Name", "vetro_site", 34), ("Full Address", "vetro_addr", 34),
            ("Units", "vetro_units", 7), ("Vetro Category", "vetro_cat", 13),
            ("In Salesforce", "in_sf", 12), ("Current Status", "current_status", 18),
            ("SF Category", "sf_category", 11), ("Helpful Note", "helpful_note", 60),
            ("Match Basis", "match_basis", 22), ("# SF Opps", "n_opps", 9),
            ("SF Opp Name", "sf_name", 28), ("SF Agreement Name", "sf_agree", 34),
            ("SF Opp Id", "sf_id", 20)]
    prio = {"MISSING": 0, "loaded-CLOSED-LOST": 1, "loaded-ACTIVE": 2}
    def sortkey(r):
        soft = 0 if "soft" in r["helpful_note"] else 1
        return (prio.get(r["bucket"], 3), soft if r["bucket"] == "loaded-CLOSED-LOST" else 0,
                -int(r["vetro_units"] or 0))
    rows_sorted = sorted(rows_out, key=sortkey)

    wb = Workbook(); ws = wb.active; ws.title = "Omaha MDU vs SF"
    hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D0D0D0"); border = Border(*(thin,) * 4)
    for c, (label, _, w_) in enumerate(COLS, 1):
        cell = ws.cell(1, c, label); cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[chr(64 + c) if c <= 26 else "A" + chr(64 + c - 26)].width = w_
    fills = {"MISSING": "FFC7CE", "ACTIVE": "C6EFCE", "SOFT": "FFF2CC", "HARD": "E7E6E6"}
    for ridx, r in enumerate(rows_sorted, 2):
        if r["bucket"] == "MISSING": key = "MISSING"
        elif r["bucket"] == "loaded-ACTIVE": key = "ACTIVE"
        else: key = "SOFT" if "soft" in r["helpful_note"] else "HARD"
        fill = PatternFill("solid", fgColor=fills[key])
        for c, (_, field, _w) in enumerate(COLS, 1):
            cell = ws.cell(ridx, c, r[field]); cell.fill = fill; cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "helpful_note"))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64+len(COLS))}{len(rows_sorted)+1}"
    xlsx = OUT.with_name("omaha-mdu-vetro-vs-sf-reconciliation.xlsx")
    wb.save(xlsx)
    print(f"wrote {xlsx}")

if __name__ == "__main__":
    main()
