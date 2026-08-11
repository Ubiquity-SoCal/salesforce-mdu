"""
Resolve the 'agree name' for Opportunities that are NOT linked to SiteTracker.

Given a tracker workbook whose columns include an Opportunity Name, a Property
Address, and a Property State, this looks each address up in BOTH:
  - SiteTracker  (the real ST org)  -> sitetracker__Site__c.Name  (PRIMARY)
  - Vetro        (local snapshot)   -> agreename                  (fallback)

WHY ST-primary (learned 2026-07-02): Vetro's `agreename` is sparse and noisy at
the building level for unlinked Opps (usually 'nan' or junk like 'private st',
'Moratorium 2026', or a neighbor-complex bleed). SiteTracker's Site.Name is
address-exact and equals the Ubiquity agree-name / Agreement_Name__c linking key.
See memories: vetro-agreename-gap-matching, taylor-sitename-sitetracker-linking.

Matching per row:
  1. ST exact-address : house-number == + street-token overlap (state-gated).
  2. ST name-fuzzy    : normalize(Opp Name) vs normalize(Site.Name),
                        min(token_set, ratio) >= 80, number-set + state gates.
  3. Vetro agreename  : house-number == + street-token overlap, cleaned.
Recommend ST-address > ST-name(>=90 HIGH / >=80 MED) > clean Vetro > none.

Usage:
    python lookup_agree_names_for_unlinked.py "<tracker.xlsx>"          # preview
    python lookup_agree_names_for_unlinked.py "<tracker.xlsx>" --write  # backup + fill
      (cols written: D=agree name, +Source, +Confidence, +Lookup Notes)

Read-only against SF/Vetro. --write backs the workbook up before touching it.
Column indexes below match Koa's "Not Linked to Site Tracker.xlsx" export.
"""
import re, sys, csv, shutil, argparse
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz
from simple_salesforce import Salesforce

sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "Vetro" / "scripts" / "lib"))
from load_vetro import load_vetro, service_locations

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_ST = _sf_creds("st")


# tracker column indexes (1-based) — adjust here if the export layout changes
COL_OPP, COL_ADDR, COL_STATE, COL_AGREE = 1, 3, 5, 4

ST = dict(username=_ST["username"], password=_ST["password"],
          security_token=_ST["token"])  # real SiteTracker org

DIRSUF = (r"\b(n|s|e|w|no|so|north|south|east|west|st|street|ave|avenue|rd|road|blvd|dr|drive|"
          r"ln|lane|way|cir|circle|ct|court|pl|place|plz|plaza|trl|trail|hwy|highway|us|unit|"
          r"apt|bldg|building|ste|suite|pkwy|parkway|fm|cr|rdt)\b")
STATE_FULL = {'CA':'California','TX':'Texas','AZ':'Arizona','NE':'Nebraska','CO':'Colorado',
              'OK':'Oklahoma','UT':'Utah','NM':'New Mexico'}
BAD_AGN = {"nan","null","none","","0","mdu","sfu","bus","mtu","private st","hoa",
           "not built","unknown-not built","unknown"}

def house(a):
    m = re.match(r"\s*(\d+)", str(a or "")); return m.group(1) if m else None
def st_tokens(a):
    line = (str(a or "").split(",")[0].splitlines() or [""])[0].lower()
    line = re.sub(r"^\s*\d+[-\d]*\s*", " ", line)
    line = re.sub(r"\b(unit|apt|bldg)\b.*$", " ", line)
    line = re.sub(r"[.,#]", " ", line); line = re.sub(r"\b\d{5}\b", " ", line)
    line = re.sub(DIRSUF, " ", line)
    return {t for t in re.findall(r"[a-z0-9]+", line) if len(t) >= 2}
def numset(s): return set(re.findall(r"\d+", s or ""))
def norm_state(s):
    s = str(s or "").strip().upper()
    return {v.upper(): k for k, v in STATE_FULL.items()}.get(s, s[:2]) if len(s) > 2 else s
def norm_name(n):
    if not n: return ""
    s = str(n).strip()
    if " - " in s: s = s.split(" - ")[-1].strip()
    s = re.sub(r"^[A-Za-z ]+_MDU_", "", s, flags=re.I)
    s = re.sub(r"^[A-Z]{2}_[A-Za-z0-9 ]+_(S[A-Z0-9]+|FDH[0-9]+|FB[0-9]+|EFC)[_A-Z0-9]*\s*", "", s, flags=re.I)
    s = s.lower(); s = re.sub(r"[_\-/]+", " ", s); s = re.sub(r"[^a-z0-9& ]", " ", s)
    s = re.sub(r"\b(apartments|apartment|apts|community|hoa|the|llc|lp|inc|condos|condo|"
               r"townhomes|townhouse|pvt|private|road|st)\b", " ", s)
    return re.sub(r"\s+", " ", s).strip()
def clean_agn(a):
    a = str(a or "").strip(); al = a.lower()
    return "" if (al in BAD_AGN or "moratorium" in al or al.startswith("unknown")) else a

def build_indexes():
    dv = service_locations(load_vetro(warn_if_stale_days=999)).copy()
    dv["agn"] = dv.agreename.map(clean_agn); dv = dv[dv.agn != ""]
    vidx = defaultdict(list)
    for hn, sn, ss, stt, agn in zip(dv.housenum, dv.streetname, dv.streetsuff, dv.state, dv.agn):
        vidx[(str(hn).strip(), str(stt).strip().upper())].append((st_tokens(f"{sn} {ss or ''}"), agn))
    sf = Salesforce(**ST)
    res = sf.query("SELECT Name, sitetracker__Street_Address__c, sitetracker__State__c "
                   "FROM sitetracker__Site__c WHERE sitetracker__Street_Address__c!=null")
    sites = res['records']
    while not res['done']:
        res = sf.query_more(res['nextRecordsUrl'], True); sites.extend(res['records'])
    by_house = defaultdict(list); by_state = defaultdict(list)
    for s in sites:
        stt = norm_state(s.get('sitetracker__State__c')); h = house(s['sitetracker__Street_Address__c'])
        rec = dict(name=s['Name'], toks=st_tokens(s['sitetracker__Street_Address__c']),
                   norm=norm_name(s['Name']), numset=numset(s['Name']))
        if h: by_house[(h, stt)].append(rec)
        if rec['norm']: by_state[stt].append(rec)
    return vidx, by_house, by_state

def resolve(opp, addr, state, vidx, by_house, by_state):
    h = house(addr); gt = st_tokens(addr)
    # 1. ST exact address
    best = None
    for rec in by_house.get((h, state), []):
        ov = gt & rec['toks']
        if ov and (best is None or len(ov) > best[0]): best = (len(ov), rec)
    if best: return best[1]['name'], "SiteTracker (address match)", "HIGH"
    # 2. ST name fuzzy
    q = norm_name(opp); qn = numset(opp); bf = None
    for rec in by_state.get(state, []):
        if qn and rec['numset'] and not (qn & rec['numset']): continue
        dual = min(fuzz.token_set_ratio(q, rec['norm']), fuzz.ratio(q, rec['norm']))
        if bf is None or dual > bf[0]: bf = (dual, rec)
    if bf and bf[0] >= 90: return bf[1]['name'], "SiteTracker (name match)", "HIGH"
    if bf and bf[0] >= 80: return bf[1]['name'], "SiteTracker (name match)", "MED"
    # 3. Vetro fallback
    best = None
    for toks, agn in vidx.get((h, state), []):
        ov = gt & toks
        if ov and (best is None or len(ov) > best[0]): best = (len(ov), agn)
    if best: return best[1], "Vetro agreename", "MED"
    return "", "", "none"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("tracker"); ap.add_argument("--write", action="store_true")
    a = ap.parse_args()
    path = Path(a.tracker)
    print("building ST + Vetro indexes...")
    vidx, by_house, by_state = build_indexes()
    wb = openpyxl.load_workbook(path); ws = wb.active
    NONE_NOTE = "Checked SiteTracker + Vetro - no matching site yet (likely new / unbuilt)"
    out = []
    for r in range(2, ws.max_row + 1):
        opp = ws.cell(r, COL_OPP).value; addr = ws.cell(r, COL_ADDR).value
        state = norm_state(ws.cell(r, COL_STATE).value)
        agree, src, conf = resolve(opp, addr, state, vidx, by_house, by_state)
        out.append((r, opp, addr, state, agree, src, conf))
        print(f"{r:>3} [{conf:4}] {str(addr)[:30]:30} => {agree[:44]}")
    n = sum(1 for x in out if x[6] != "none")
    print(f"\nmatched {n}/{len(out)}")
    if not a.write:
        print("preview only; pass --write to fill the workbook."); return
    bak = path.with_name(path.stem + " (backup before agree-name fill).xlsx")
    shutil.copy2(path, bak); print("backup ->", bak)
    S, T, U = ws.max_column + 1, ws.max_column + 2, ws.max_column + 3
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="305496")
    for c, lbl in ((S, "Agree Name Source"), (T, "Match Confidence"), (U, "Lookup Notes")):
        ws.cell(1, c, lbl).font = hf; ws.cell(1, c).fill = hfill
    for r, opp, addr, state, agree, src, conf in out:
        if conf == "none":
            ws.cell(r, T).value = "none"; ws.cell(r, U).value = NONE_NOTE
        else:
            ws.cell(r, COL_AGREE).value = agree
            ws.cell(r, S).value = src; ws.cell(r, T).value = conf
    for col, w in ((COL_AGREE, 40), (S, 26), (T, 16), (U, 52)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    wb.save(path); print("wrote ->", path)

if __name__ == "__main__":
    main()
