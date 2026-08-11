"""
Enrich Niraj's "Texas -NE - ON Net MDUs.xlsx" (552 rows) with the 7 Salesforce
columns he asked for (2026-07-15). Same reconciliation as the Omaha on-net rounds,
scaled to TX+NE. Preserves his row order and his source columns; only APPENDS.

Niraj's 7 asks  ->  appended column:
  1 In Salesforce (YES/NO)   -> Agreement_Name__c exact join, then address/name-fuzzy
  2 Property name (SF)        -> Opportunity.Name
  3 Category (SF)             -> Property_Category__c
  4 Opportunity State         -> StageName
  5 Reason for lost (if lost) -> Loss_Reason__c
  6 Owner                     -> Owner.Name (+ deactivated flag)
  7 Any other relevant info   -> derived (deactivated-owner / weak-match / twin / cat-blank)
Plus "Latest Note (SF)": the concrete last touch on each Opp, so the higher-priority
rows can be picked out for momentum (Niraj's stated goal).

Matching ladder reused from Omaha (do not regress the guards):
  agree-name exact -> address(house# + shared street token) -> near-address -> name-fuzzy.
PLACE_TOKENS is built PER RUN here: TX+NE span many cities, so the city name is a
discriminating token, NOT a free shared one like "omaha" was. We strip only the
state/country tokens plus every city that actually appears (from the Site Name prefix
and SF Property_City__c), so two different buildings in one city can't share {fort,worth}.

Read-only against Salesforce. Output: xlsx (data/output + Desktop) + CSV audit.
"""
import sys
import csv
import re
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz
from simple_salesforce import Salesforce

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))
from lookup_agree_names_for_unlinked import house, st_tokens, norm_name, numset  # noqa: E402
from enrich_omaha_onnet_mdus import creds, fetch_latest_notes  # noqa: E402  (both SF-only, PLACE_TOKENS-independent)

DESKTOP = Path(r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop")
SRC = DESKTOP / "Texas -NE - ON Net MDUs.xlsx"
OUT_DIR = Path(r"C:\Users\cass\Work_Projects\SalesForce\data\output")
OUT_XLSX = OUT_DIR / "tx-ne-onnet-mdus-with-salesforce.xlsx"
OUT_DESKTOP = DESKTOP / "TX-NE ON Net MDUs - with Salesforce.xlsx"
OUT_CSV = OUT_DIR / "2026-07-15-tx-ne-onnet-mdus-sf-enrichment.csv"

STATE_TOKENS = {"tx", "texas", "ne", "nebraska", "usa", "us"}
HOUSE_WINDOW = 12
SOFT_REASONS = {"No Contact Info", "No Decision / Non-Responsive", "Other", "None", None}
# "Existing Fiber" is deliberately NOT soft: same open question as Omaha (our fiber or a
# competitor's?). Flagged in Other Info, treated as hard until Taylor confirms the meaning.

# Niraj's exact source columns, in his order (col B and col F are both "Site Name").
SRC_COLS = ["Sate", "Site Name", "MDU Site Category", "Total Units",
            "Full Address", "Site Name ", "Monday.com name"]
NEW_COLS = ["In Salesforce", "Property Name (SF)", "Category (SF)", "Opportunity State",
            "Loss Reason", "Owner", "Other Relevant Info", "Latest Note (SF)"]

FILLS = {"ACTIVE": "C6EFCE", "CLOSED": "E7E6E6", "REVIEW": "FFEB9C", "MISSING": "FFC7CE"}

# generic words that carry no identity: they must not be the token two rows "share"
GENERIC = {"apartments", "apartment", "apts", "apt", "homes", "home", "estates", "estate",
           "village", "court", "courts", "place", "park", "mobile", "manor", "flats",
           "tower", "towers", "ranch", "condominium", "condominiums", "condos", "plex",
           "mhp", "mdu", "sfu", "bus", "mtu", "the", "at", "of", "and", "on", "net"}


def prop_part(s):
    """Property-name portion of an agree-name: text after _MDU_/_SFU_/_BUS_/_MTU_."""
    parts = re.split(r"_(?:MDU|SFU|BUS|MTU)_", s or "", flags=re.I, maxsplit=1)
    return parts[1] if len(parts) > 1 else (s or "")


def name_tokens(s):
    return {t for t in re.findall(r"[a-z0-9]+", (s or "").lower())
            if t not in GENERIC and len(t) >= 4}


def is_mdu(o):
    rt = (o.get("RecordType") or {}).get("Name") or ""
    return "MDU" in rt or "SFU" in rt


def load_source():
    """Return the 552 rows IN Niraj's original order (list of dicts)."""
    ws = openpyxl.load_workbook(SRC, read_only=True, data_only=True)["ON Net MDUs"]
    rows = []
    for state, sn, cat, units, addr, sn2, mon in ws.iter_rows(min_row=2, values_only=True):
        if sn in (None, ""):
            continue
        rows.append({
            "Sate": state,
            "Site Name": str(sn).strip(),
            "MDU Site Category": cat,
            "Total Units": int(units or 0),
            "Full Address": str(addr or "").strip(),
            "Site Name ": str(sn2 or "").strip(),
            "Monday.com name": str(mon or "").strip(),
        })
    return rows


def city_prefix(site):
    """The city portion of an agree-name: text before _MDU_/_SFU_/_BUS_/_MTU_."""
    parts = re.split(r"_(?:MDU|SFU|BUS|MTU)_", site, flags=re.I, maxsplit=1)
    return parts[0] if len(parts) > 1 else ""


def build_place_tokens(rows, opps):
    place = set(STATE_TOKENS)
    for r in rows:
        for t in re.findall(r"[a-z]+", city_prefix(r["Site Name"]).lower()):
            place.add(t)
    for o in opps:
        for t in re.findall(r"[a-z]+", (o.get("Property_City__c") or "").lower()):
            place.add(t)
    return place


def load_vetro_footprints(names):
    """agreename(lower) -> [(housenum, 'house street suff', city_norm)] from the Vetro snapshot.
    A Vetro agreename can span several street faces; SF may key off any one of them, so this
    lets us match a row via a face Niraj did not list. Graceful no-op if the snapshot is absent."""
    try:
        sys.path.insert(0, r"C:\Users\cass\Work_Projects\Vetro\scripts\lib")
        import warnings
        warnings.filterwarnings("ignore")
        from load_vetro import load_vetro, service_locations
    except Exception as e:
        print(f"(Vetro footprints unavailable - skipping that stage: {e})")
        return {}
    v = service_locations(load_vetro())
    v = v[v.state.isin(["TX", "NE"])]
    nameset = {n.lower() for n in names}
    foot = defaultdict(set)
    for agn, hn, sn, sfx, city in zip(v.agreename, v.housenum, v.streetname, v.streetsuff, v.city):
        a = str(agn).lower()
        if a in nameset and str(hn) not in ("", "nan", "None"):
            foot[a].add((str(hn), f"{hn} {sn} {sfx}".strip(), ncity(city)))
    return {k: list(val) for k, val in foot.items()}


def fetch_opps(sf, names):
    fields = ("Id, Name, Agreement_Name__c, Property_Address__c, Property_City__c, "
              "Property_State__c, Property_Category__c, StageName, RecordType.Name, "
              "Loss_Reason__c, CloseDate, Owner.Name, Owner.IsActive")
    got = {}
    names = sorted({n for n in names if n})
    for i in range(0, len(names), 200):
        vals = "','".join(n.replace("'", "\\'") for n in names[i:i + 200])
        for r in sf.query_all(f"SELECT {fields} FROM Opportunity "
                              f"WHERE Agreement_Name__c IN ('{vals}')")["records"]:
            got[r["Id"]] = r
    # TX/NE universe for the address / name-fuzzy fallback on rows that miss the exact join
    for r in sf.query_all(f"SELECT {fields} FROM Opportunity WHERE "
                          f"Property_State__c IN ('TX','Texas','NE','Nebraska') "
                          f"AND Property_Address__c != null")["records"]:
        got.setdefault(r["Id"], r)
    return list(got.values())


def ncity(s):
    return re.sub(r"[^a-z]", "", (s or "").lower())


JUNK_CITY = {"", "ne", "tx", "texas", "nebraska", "usa", "us"}  # uninformative Property_City__c values


class Matcher:
    """agree-name exact -> address -> Vetro-footprint -> near-address -> name-fuzzy.
    Guards from the Omaha rounds; the Vetro stage is new for TX+NE (2026-07-15)."""

    def __init__(self, opps, place, footprints=None):
        self.opps = opps
        self.mdu = [o for o in opps if is_mdu(o)]  # fallback only trusts MDU/SFU record types
        self.place = place
        self.footprints = footprints or {}  # agreename(lower) -> [(housenum, street_full, city_norm)]
        self.by_agree = defaultdict(list)
        self.by_house = defaultdict(list)
        self.by_hc = defaultdict(list)     # (housenum, city_norm) -> [mdu opp]
        for o in opps:
            agn = (o["Agreement_Name__c"] or "").strip().lower()
            if agn:
                self.by_agree[agn].append(o)
        for o in self.mdu:
            h = house(o["Property_Address__c"])
            if h:
                self.by_house[h].append(o)
                self.by_hc[(h, ncity(o.get("Property_City__c")))].append(o)

    def _street(self, a):
        return {t for t in st_tokens(a) if not t.isdigit()} - self.place

    def _city_ok(self, row_city, o):
        """Address-basis matches must be same-city (an agree-name match is exempt - it's the
        authoritative key). Junk/state Property_City__c is allowed so it doesn't drop legit
        rows; a real different city (Turkey Creek/Bridgeport vs Cedar Park) is blocked."""
        sc = ncity(o.get("Property_City__c"))
        return (not row_city) or sc in JUNK_CITY or sc == row_city

    def _exact(self, site, addr):
        hits, basis = {}, set()
        for o in self.by_agree.get(site.lower(), []):
            hits[o["Id"]] = o
            basis.add("agree-name")
        gt = self._street(addr)
        row_city = ncity(city_prefix(site))
        for o in self.by_house.get(house(addr), []):
            if gt & self._street(o["Property_Address__c"]) and self._city_ok(row_city, o):
                hits[o["Id"]] = o
                basis.add("address")
        return list(hits.values()), basis

    def _near(self, addr, row_city):
        h, ga = house(addr), self._street(addr)
        if not h or len(ga) < 2:
            return None
        hi, out = int(h), []
        for o in self.mdu:
            oh = house(o["Property_Address__c"])
            if (oh and len(ga & self._street(o["Property_Address__c"])) >= 2
                    and abs(int(oh) - hi) <= HOUSE_WINDOW and self._city_ok(row_city, o)):
                out.append((abs(int(oh) - hi), o))
        if not out:
            return None
        out.sort(key=lambda x: x[0])
        best = out[0][1]
        return [o for _, o in out], f"{house(best['Property_Address__c'])} vs {h}"

    def _addr_compatible(self, addr, o):
        oa = o["Property_Address__c"]
        if self._street(addr) & self._street(oa):
            return True
        h, oh = house(addr), house(oa)
        return bool(h and oh and abs(int(h) - int(oh)) <= HOUSE_WINDOW)

    def _fuzzy(self, site, addr, monday):
        """Name match on agree-name OR the Monday.com display name; digits must agree."""
        cands = [norm_name(site)]
        if monday:
            cands.append(norm_name(monday))
        best = None
        for q in cands:
            qn = numset(q)
            for o in self.mdu:
                cand = norm_name(o["Agreement_Name__c"] or o["Name"])
                if not cand or numset(cand) != qn or not self._addr_compatible(addr, o):
                    continue
                score = min(fuzz.token_set_ratio(q, cand), fuzz.ratio(q, cand))
                if best is None or score > best[0]:
                    best = (score, o)
        return best if (best and best[0] >= 90) else None

    def _vetro(self, site):
        """A Vetro agreename can cover several street faces (Bell Presidio = 7); Niraj listed
        one, SF may key off another. Test EVERY Vetro face for this agreename against SF by
        (house# + city + shared street token). Authoritative building footprint, so high trust."""
        hits, face = {}, {}
        for hn, full, city in self.footprints.get(site.lower(), ()):
            for o in self.by_hc.get((hn, city), []):
                if self._street(full) & self._street(o["Property_Address__c"]):
                    hits[o["Id"]] = o
                    face.setdefault(o["Id"], full)
        return (list(hits.values()), face) if hits else (None, None)

    def match(self, site, addr, monday):
        row_city = ncity(city_prefix(site))
        hits, basis = self._exact(site, addr)
        how = "+".join(sorted(basis)) if basis else ""
        if not hits:
            vf, face = self._vetro(site)
            if vf:
                hits, how = vf, f"vetro-footprint ({face[vf[0]['Id']]})"
        if not hits:
            n = self._near(addr, row_city)
            if n:
                hits, how = n[0], f"address (near: {n[1]})"
        if not hits:
            f = self._fuzzy(site, addr, monday)
            if f:
                hits, how = [f[1]], f"name-fuzzy({f[0]:.0f})"
        return hits, how

    def suggest(self, site, addr, monday):
        """For a row the confident ladder could NOT match: surface plausible existing Opps
        so a human verifies, rather than us asserting MISSING (which risks a duplicate load).
        A suggestion is NOT a confirmed match. Two signals:
          name  - shares a distinctive (>=4 char, non-generic) token AND token_set_ratio>=80
          addr  - same house# w/ shared street token, or house within window w/ shared street
        """
        norm_city = lambda s: re.sub(r"[^a-z]", "", (s or "").lower())
        row_city = norm_city(city_prefix(site))
        src_names = [n for n in (prop_part(site), monday) if n]
        # place tokens (city/state) are not identity: a shared "bridgeport"/"jarrell" is geography
        src_tok = (set().union(*(name_tokens(n) for n in src_names)) if src_names else set()) - self.place
        cands = {}
        for o in self.mdu:
            # name suggestions must be same-city; else "station"/"authority" bleed across TX<->NE
            if row_city and norm_city(o.get("Property_City__c")) and row_city != norm_city(o["Property_City__c"]):
                continue
            op = prop_part(o["Agreement_Name__c"] or o["Name"])
            shared = src_tok & ((name_tokens(op) | name_tokens(o["Name"])) - self.place)
            if not shared:
                continue
            score = max((fuzz.token_set_ratio(n, op) for n in src_names), default=0)
            if score >= 80 or max((len(t) for t in shared), default=0) >= 6:
                cands[o["Id"]] = (score + 2 * len(shared), o,
                                  f"name~'{'/'.join(sorted(shared))}'")
        h, ga = house(addr), self._street(addr)
        if h:
            hi = int(h)
            for o in self.mdu:
                oh = house(o["Property_Address__c"])
                if not oh:
                    continue
                shared_st = ga & self._street(o["Property_Address__c"])
                if not shared_st:
                    continue
                if oh == h:
                    cands[o["Id"]] = max(cands.get(o["Id"], (0,)), (96, o, f"addr {h} same street"),
                                         key=lambda x: x[0])
                elif abs(int(oh) - hi) <= HOUSE_WINDOW:
                    cand = (92 - abs(int(oh) - hi), o, f"addr {oh} vs {h} same street")
                    if cand[0] > cands.get(o["Id"], (0,))[0]:
                        cands[o["Id"]] = cand
        return sorted(cands.values(), key=lambda x: -x[0])[:3]


def pick(hits):
    return sorted(hits, key=lambda o: (o["StageName"] == "Closed Lost",))[0]


def other_info(rep, bucket, soft, how, n_opps, owner_gone=False):
    bits = []
    if bucket == "ACTIVE":
        name = rep["Owner"]["Name"] if rep.get("Owner") else ""
        if owner_gone:
            bits.append(f"ACTIVE but owner ({name}) is a DEACTIVATED user with no active record - "
                        f"reassign to a live owner to work it")
        elif rep.get("Owner") and not rep["Owner"]["IsActive"]:
            bits.append(f"ACTIVE - assigned to a stale duplicate '{name}' record; re-point to the "
                        f"active {name} (owner is still here)")
        else:
            bits.append("Active in pipeline")
    elif bucket == "CLOSED":
        reason = rep["Loss_Reason__c"] or "(blank)"
        if reason == "Existing Fiber":
            bits.append("Closed Lost 'Existing Fiber' - confirm ours vs competitor before re-approaching")
        elif soft:
            bits.append(f"Closed Lost ({reason}) - re-approachable now fiber is on-net")
        else:
            bits.append(f"Closed Lost ({reason})")
    if rep is not None and not (rep.get("Property_Category__c") or ""):
        bits.append("SF category blank")
    if "vetro-footprint" in how:
        face = how.split("(", 1)[1].rstrip(")")
        bits.append(f"Matched via Vetro building footprint - SF Opp is at another face of the same "
                    f"building ({face}), not the address on this row")
    if "near" in how:
        bits.append(f"Address off by a few ({how.split('near: ')[1].rstrip(')')}) - verify same property")
    if "fuzzy" in how:
        bits.append("Matched on name not exact agree-name - verify same property")
    if n_opps > 1:
        bits.append(f"{n_opps} SF Opps on this property (showing the representative)")
    return "; ".join(bits)


def main():
    sf = Salesforce(*creds())
    rows = load_source()
    print(f"source rows: {len(rows)}  "
          f"(TX {sum(r['Sate']=='TX' for r in rows)}, NE {sum(r['Sate']=='NE' for r in rows)})")

    opps = fetch_opps(sf, [r["Site Name"] for r in rows])
    place = build_place_tokens(rows, opps)
    foot = load_vetro_footprints([r["Site Name"] for r in rows])
    rtd = Counter((o.get("RecordType") or {}).get("Name") for o in opps)
    print(f"SF TX/NE universe: {len(opps)} Opps  |  MDU/SFU: {sum(is_mdu(o) for o in opps)}")
    print(f"  record types: {dict(rtd)}")
    print(f"  place-token stoplist: {len(place)} tokens")
    print(f"  Vetro footprints for {len(foot)}/{len(rows)} agreenames "
          f"({sum(len(v) for v in foot.values())} address faces)")
    m = Matcher(opps, place, foot)

    # A name is only "gone" if it has NO active user record anywhere. Some names (Brett Spivey,
    # Melissa Baker) have BOTH an active and a stale inactive user; an opp on the dead twin is a
    # re-point, not an orphan. Chuck McNeely / Jeff Chao have only the inactive record = truly gone.
    active_owner_names = {o["Owner"]["Name"] for o in opps
                          if o.get("Owner") and o["Owner"]["IsActive"]}

    for r in rows:
        hits, how = m.match(r["Site Name"], r["Full Address"], r["Monday.com name"])
        active = [o for o in hits if o["StageName"] != "Closed Lost"]
        rep = pick(active) if active else (pick(hits) if hits else None)
        if rep is None:
            # no confident match: look for likely-existing Opps before calling it MISSING
            sugg = m.suggest(r["Site Name"], r["Full Address"], r["Monday.com name"])
            bucket = "REVIEW" if sugg else "MISSING"
        elif active:
            bucket, sugg = "ACTIVE", []
        else:
            bucket, sugg = "CLOSED", []
        soft = bool(rep and bucket == "CLOSED" and rep["Loss_Reason__c"] in SOFT_REASONS)
        owner_gone = bool(rep and rep.get("Owner") and not rep["Owner"]["IsActive"]
                          and rep["Owner"]["Name"] not in active_owner_names)

        if bucket == "REVIEW":
            info = "Possible SF match(es) - VERIFY (not auto-matched): " + " | ".join(
                f"{o['Name']} @ {o['Property_Address__c']} [{o['StageName']}"
                + (f", {o['Loss_Reason__c']}" if o['StageName'] == 'Closed Lost' and o['Loss_Reason__c'] else "")
                + f"] ({why})" for _sc, o, why in sugg[:2])
        else:
            info = other_info(rep, bucket, soft, how, len(hits), owner_gone) if rep else ""

        r.update({
            "In Salesforce": {"MISSING": "NO", "REVIEW": "REVIEW"}.get(bucket, "YES"),
            "Property Name (SF)": rep["Name"] if rep else "",
            "Category (SF)": (rep["Property_Category__c"] or "") if rep else "",
            "Opportunity State": rep["StageName"] if rep else "",
            "Loss Reason": (rep["Loss_Reason__c"] or "") if rep and bucket == "CLOSED" else "",
            "Owner": rep["Owner"]["Name"] if rep and rep.get("Owner") else "",
            "Other Relevant Info": info,
            "Latest Note (SF)": "",
            "_bucket": bucket, "_soft": soft, "_match": how,
            "_sf_id": rep["Id"] if rep else "",
            "_sugg": sugg,
            # actual owner record's status (NOT keyed on name - two "Brett Spivey" / "Melissa
            # Baker" users exist, one active one inactive; only the OwnerId tells them apart)
            "_owner_active": (rep["Owner"]["IsActive"] if rep and rep.get("Owner") else None),
            "_owner_gone": owner_gone,  # inactive AND no active twin -> truly orphaned
        })

    # ---- twin / second-feed post-pass (Vetro emits one row per fiber feed) ----
    by_addr = defaultdict(list)
    for r in rows:
        by_addr[re.sub(r"[^a-z0-9]", "", r["Full Address"].lower())].append(r)
    inherited = 0
    for group in by_addr.values():
        if len(group) < 2:
            continue
        donor = next((g for g in group if g["_bucket"] in ("ACTIVE", "CLOSED")), None)
        twins = ", ".join(sorted({g["Site Name"] for g in group}))
        for r in group:
            if r["_bucket"] in ("MISSING", "REVIEW") and donor:
                for f in NEW_COLS[:6]:
                    r[f] = donor[f]
                r["_bucket"], r["_soft"], r["_sf_id"] = donor["_bucket"], donor["_soft"], donor["_sf_id"]
                inherited += 1
            note = f"Same building as its twin (second fiber feed): {twins}. Do not double-load/count."
            r["Other Relevant Info"] = f"{r['Other Relevant Info']}; {note}".lstrip("; ")
    twin_rows = sum(len(g) for g in by_addr.values() if len(g) > 1)
    print(f"\n{len(by_addr)} distinct addresses; {twin_rows} row(s) share an address; "
          f"{inherited} twin(s) inherited a sibling match")

    # ---- latest note per Opp ----
    sf_ids = sorted({r["_sf_id"] for r in rows if r["_sf_id"]})
    latest = fetch_latest_notes(sf, sf_ids)
    for r in rows:
        r["Latest Note (SF)"] = latest.get(r["_sf_id"], "")
    print(f"latest note resolved for {sum(1 for r in rows if r['Latest Note (SF)'])} rows "
          f"({len(latest)}/{len(sf_ids)} distinct Opps have a note)")

    # ---- summary (parts must sum to whole) ----
    b = Counter(r["_bucket"] for r in rows)
    print("\n=== SUMMARY (552 in scope) ===")
    for k in ("ACTIVE", "CLOSED", "REVIEW", "MISSING"):
        u = sum(r["Total Units"] for r in rows if r["_bucket"] == k)
        print(f"  {k:8} {b[k]:3}  ({u:,} units)")
    assert sum(b.values()) == len(rows) == 552, "bucket counts must sum to 552"
    print(f"  {'TOTAL':8} {sum(b.values()):3}  ({sum(r['Total Units'] for r in rows):,} units)")
    for st in ("TX", "NE"):
        sub = [r for r in rows if r["Sate"] == st]
        print(f"    {st}: {len(sub)} rows, {sum(r['Total Units'] for r in sub):,} units, "
              f"{sum(r['In Salesforce']=='YES' for r in sub)} in SF")

    cl = [r for r in rows if r["_bucket"] == "CLOSED"]
    soft = [r for r in cl if r["_soft"]]
    print(f"\n=== Closed Lost ({len(cl)}) reasons ===")
    for reason, n in Counter(r["Loss Reason"] or "(blank)" for r in cl).most_common():
        print(f"  {reason:34} {n}")
    print(f"  SOFT (re-approachable): {len(soft)} / {sum(r['Total Units'] for r in soft):,} units | "
          f"HARD: {len(cl)-len(soft)} / {sum(r['Total Units'] for r in cl)-sum(r['Total Units'] for r in soft):,} units")

    deact = [r for r in rows if r["_bucket"] == "ACTIVE" and r["_owner_gone"]]
    print(f"\n=== MOMENTUM: {len(deact)} ACTIVE Opps on truly-gone owners ({sum(r['Total Units'] for r in deact):,} units) ===")
    for r in sorted(deact, key=lambda r: -r["Total Units"])[:15]:
        print(f"  {r['Site Name'][:44]:44} u={r['Total Units']:<5} {r['Owner']}")

    review = [r for r in rows if r["_bucket"] == "REVIEW"]
    print(f"\n=== REVIEW ({len(review)}, {sum(r['Total Units'] for r in review):,} units) - likely in SF under a variant, verify ===")
    for r in sorted(review, key=lambda r: -r["Total Units"]):
        print(f"  {r['Sate']} {r['Site Name'][:42]:42} u={r['Total Units']:<5} -> {r['Other Relevant Info'][60:150]}")

    miss = [r for r in rows if r["_bucket"] == "MISSING"]
    print(f"\n=== NOT in Salesforce ({len(miss)}, {sum(r['Total Units'] for r in miss):,} units) - load candidates ===")
    for r in sorted(miss, key=lambda r: -r["Total Units"]):
        print(f"  {r['Sate']} {r['Site Name'][:44]:44} u={r['Total Units']:<5} {r['Full Address'][:44]}")

    vfp = [r for r in rows if "vetro-footprint" in r["_match"]]
    print(f"\n=== {len(vfp)} matched via Vetro building footprint (SF keys off a different address face) ===")
    for r in sorted(vfp, key=lambda r: -r["Total Units"]):
        print(f"  {r['Site Name'][:40]:40} u={r['Total Units']:<5} -> {r['Property Name (SF)'][:28]:28} [{r['Opportunity State']}]")

    weak = [r for r in rows if "near" in r["_match"] or "fuzzy" in r["_match"]]
    print(f"\n=== {len(weak)} weak matches to verify ===")
    for r in weak:
        print(f"  {r['Site Name'][:40]:40} [{r['_match']}] -> {r['Property Name (SF)'][:34]}")
    print(f"\nCategory (SF): " + ", ".join(f"{c or '(blank)'}={n}"
          for c, n in Counter(r["Category (SF)"] for r in rows if r["In Salesforce"] == "YES").most_common()))

    # ---- CSV audit ----
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"\nwrote {OUT_CSV}")

    # ---- deliverable: Niraj's columns + order preserved, 8 columns appended ----
    write_workbook(rows)


def write_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ON Net MDUs"
    cols = SRC_COLS + NEW_COLS
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E78")
    nfill = PatternFill("solid", fgColor="2E75B6")
    border = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    widths = [6, 42, 15, 9, 40, 42, 26, 12, 34, 12, 20, 22, 14, 52, 90]
    for i, name in enumerate(cols, 1):
        c = ws.cell(1, i, name.strip())
        c.font, c.fill = hfont, (nfill if name in NEW_COLS else hfill)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = widths[i - 1]
    # Only the rows that need a human action get a fill: yellow=verify, red=not in SF.
    # Matched rows (active / closed lost) stay unshaded - the Opportunity State column carries that.
    HILITE = {"REVIEW": "FFEB9C", "MISSING": "FFC7CE"}
    for i, r in enumerate(rows, 2):  # Niraj's original order, no re-sort
        hl = HILITE.get(r["_bucket"])
        fill = PatternFill("solid", fgColor=hl) if hl else None
        for j, name in enumerate(cols, 1):
            c = ws.cell(i, j, r[name])
            c.border = border
            if fill:
                c.fill = fill
            c.alignment = Alignment(vertical="top",
                                    wrap_text=name in ("Full Address", "Other Relevant Info", "Latest Note (SF)"))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(cols)).column_letter}{len(rows) + 1}"

    build_summary_tab(wb, rows)  # lands first so status is the first thing he sees

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    try:
        wb.save(OUT_DESKTOP)
        print(f"wrote {OUT_DESKTOP}")
    except Exception as e:
        print(f"(desktop copy skipped: {e})")


def build_summary_tab(wb, rows):
    """A one-glance status tab: In-SF split, every Opportunity stage, and the momentum
    rollup (active-but-orphaned, soft vs hard closed lost). Placed as the first sheet."""
    U = lambda rs: sum(r["Total Units"] for r in rs)
    tot_u = U(rows)
    active = [r for r in rows if r["_bucket"] == "ACTIVE"]
    closed = [r for r in rows if r["_bucket"] == "CLOSED"]
    review = [r for r in rows if r["_bucket"] == "REVIEW"]
    missing = [r for r in rows if r["_bucket"] == "MISSING"]
    matched = active + closed
    # orphaned = active Opp whose owner is truly gone (inactive AND no active user record).
    # Excludes stale-duplicate owners (Brett Spivey / Melissa Baker are still active users).
    deact = [r for r in active if r["_owner_gone"]]
    stale = [r for r in active if r["_owner_active"] is False and not r["_owner_gone"]]
    soft = [r for r in closed if r["_soft"]]
    hard = [r for r in closed if not r["_soft"]]

    ws = wb.create_sheet("Summary", 0)
    wb.active = 0
    for col, w_ in zip("ABCD", (46, 13, 13, 44)):
        ws.column_dimensions[col].width = w_
    white = Font(bold=True, color="FFFFFF")
    blue = PatternFill("solid", fgColor="1F4E78")
    grey = PatternFill("solid", fgColor="D9E1F2")
    R = 1

    def section(title):
        nonlocal R
        for col in range(1, 5):
            ws.cell(R, col).fill = blue
        ws.cell(R, 1, title).font = white
        R += 1

    def cols(a, b_="Properties", c="Units", d=""):
        nonlocal R
        for col, val in enumerate((a, b_, c, d), 1):
            cell = ws.cell(R, col, val)
            cell.font = Font(bold=True)
            cell.fill = grey
        R += 1

    def line(label, props, units, note="", fill=None, bold=False):
        nonlocal R
        vals = [label, props, f"{units:,}" if isinstance(units, int) else units, note]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(R, col, val)
            if bold:
                cell.font = Font(bold=True)
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if col == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        R += 1

    def gap():
        nonlocal R
        R += 1

    # ---- title ----
    ws.cell(R, 1, "TX + NE On-Net MDUs - Salesforce Status").font = Font(bold=True, size=14)
    R += 1
    ws.cell(R, 1, f"{len(rows)} properties / {tot_u:,} units from Niraj's Vetro list (CX complete, "
                  f"MDU not active). Full detail on the 'ON Net MDUs' tab - one row per property; "
                  f"yellow = needs review, red = not in Salesforce.").alignment = \
        Alignment(wrap_text=True)
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
    ws.row_dimensions[R].height = 30
    R += 2

    # ---- by state ----
    section("BY STATE")
    cols("State")
    for st in ("TX", "NE"):
        sub = [r for r in rows if r["Sate"] == st]
        line(st, len(sub), U(sub))
    line("Total", len(rows), tot_u, bold=True)
    gap()

    # ---- in salesforce ----
    section("IN SALESFORCE")
    cols("Status", "Properties", "Units", "")
    line("In Salesforce (matched)", len(matched), U(matched),
         "Active or Closed Lost - see next section")
    line("Needs review", len(review), U(review),
         "Similar Opp exists under a variant - verify", fill="FFEB9C")
    line("Not in Salesforce", len(missing), U(missing),
         "Load candidates - verify first", fill="FFC7CE")
    line("Total", len(rows), tot_u, bold=True)
    gap()

    # ---- opportunity status (the core view) ----
    section("OPPORTUNITY STATUS (matched properties)")
    cols("Opportunity State")
    for stage, n in Counter(r["Opportunity State"] for r in active).most_common():
        line(stage, n, U([r for r in active if r["Opportunity State"] == stage]))
    line("Active - subtotal", len(active), U(active), bold=True)
    line("Closed Lost", len(closed), U(closed))
    line("Matched - total", len(matched), U(matched), bold=True)
    gap()

    # ---- momentum ----
    section("WHERE THE MOMENTUM IS")
    cols("Opportunity group", "Properties", "Units", "Action")
    deact_owners = ", ".join(sorted({r["Owner"] for r in deact})) or "n/a"
    line("Active, owner no longer here", len(deact), U(deact),
         f"Reassign to a live owner - fastest wins (owner: {deact_owners})")
    if stale:
        line("Active, on a stale duplicate user", len(stale), U(stale),
             "Owner is still active - just re-point to their live record (data hygiene)")
    line("Closed Lost - soft reason", len(soft), U(soft),
         "Re-approachable now fiber is on-net")
    line("Closed Lost - hard reason", len(hard), U(hard),
         "Fiber won't change these (incl. Existing Fiber)")
    gap()

    # ---- loss reasons ----
    section("CLOSED LOST - REASONS")
    cols("Reason", "Properties", "Units", "")
    for reason, n in Counter(r["Loss Reason"] or "(blank)" for r in closed).most_common():
        sub = [r for r in closed if (r["Loss Reason"] or "(blank)") == reason]
        tag = "soft - re-approachable" if reason in SOFT_REASONS or reason == "(blank)" else "hard"
        line(reason, n, U(sub), tag)
    gap()

    # ---- owners ----
    section("OWNERS (matched properties)")
    cols("Owner", "Total", "Active", "Closed Lost / note")
    own_tot = Counter(r["Owner"] for r in matched if r["Owner"])
    for owner, n in own_tot.most_common():
        na = sum(1 for r in active if r["Owner"] == owner)
        nc = sum(1 for r in closed if r["Owner"] == owner)
        # by the actual owner RECORD: "gone" = no active twin; "stale" = active user, dead dup record
        gone = sum(1 for r in matched if r["Owner"] == owner and r["_owner_gone"])
        stl = sum(1 for r in matched if r["Owner"] == owner
                  and r["_owner_active"] is False and not r["_owner_gone"])
        flag = (f"  |  INACTIVE USER - reassign {gone}" if gone
                else f"  |  {stl} on a stale duplicate record (owner active)" if stl else "")
        line(owner, n, na, f"{nc} closed lost{flag}", fill=("FFEB9C" if gone else None))
    gap()

    # ---- category ----
    section("CATEGORY (Salesforce)")
    cols("Category", "Properties", "Units", "")
    for cat, n in Counter(r["Category (SF)"] or "(blank)" for r in matched).most_common():
        sub = [r for r in matched if (r["Category (SF)"] or "(blank)") == cat]
        line(cat, n, U(sub))
    gap()

    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    main()
