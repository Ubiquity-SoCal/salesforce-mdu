"""Inspect Property_Location__c for parcel-related fields + compare xlsx Parcel # vs Parcel columns."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


sf = Salesforce(username=_SF["username"], password=_SF["password"], security_token=_SF["token"])

# 1. PL fields — any "parcel" or "apn" mentions?
print("=== Property_Location__c — parcel/APN-related fields ===")
d = sf.Property_Location__c.describe()
KEYS = ['parcel', 'apn', 'tax', 'lot', 'plat', 'assess']
for f in d['fields']:
    n, label, t = f['name'], f['label'], f['type']
    if any(k in n.lower() or k in label.lower() for k in KEYS):
        extra = f"  len={f.get('length')}" if f.get('length') else ''
        print(f"  {n:45s} {label[:35]:35s} {t:10s}{extra}")

# Also dump all custom long-text/text fields in case there's an unfortunately-named one
print("\n=== All custom text fields on Property_Location__c (in case) ===")
for f in d['fields']:
    if f['name'].endswith('__c') and f['type'] in ('string', 'textarea'):
        print(f"  {f['name']:45s} {f['label'][:35]:35s} {f['type']:10s} len={f.get('length')}")

# 2. Compare xlsx columns
XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb['ROE_Tracking']
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]

print(f"\n=== xlsx Parcel column comparison ({len(data)} rows) ===")
filled_p1 = sum(1 for r in data if r.get('Parcel #') not in (None, ''))
filled_p2 = sum(1 for r in data if r.get('Parcel') not in (None, ''))
both = sum(1 for r in data if r.get('Parcel #') not in (None, '') and r.get('Parcel') not in (None, ''))
only1 = sum(1 for r in data if r.get('Parcel #') not in (None, '') and r.get('Parcel') in (None, ''))
only2 = sum(1 for r in data if r.get('Parcel #') in (None, '') and r.get('Parcel') not in (None, ''))
neither = sum(1 for r in data if r.get('Parcel #') in (None, '') and r.get('Parcel') in (None, ''))
print(f"  Parcel # filled:           {filled_p1}")
print(f"  Parcel   filled:           {filled_p2}")
print(f"  Both filled:               {both}")
print(f"  Only Parcel # filled:      {only1}")
print(f"  Only Parcel   filled:      {only2}")
print(f"  Neither filled:            {neither}")

# Where both are filled, are they equal? Different?
print("\n=== Rows where BOTH filled — agreement check ===")
disagree = []
agree = 0
for r in data:
    p1 = str(r.get('Parcel #') or '').strip()
    p2 = str(r.get('Parcel') or '').strip()
    if p1 and p2:
        # normalize: drop "common parcel" prefix, lowercase, strip whitespace
        p1n = re.sub(r'\s+', ' ', p1.lower())
        p2n = re.sub(r'\s+', ' ', re.sub(r'common parcel', '', p2, flags=re.I).strip().lower())
        if p1n == p2n:
            agree += 1
        else:
            disagree.append((r.get('Business Buildings'), p1, p2))
print(f"  Both equal (after normalization): {agree}")
print(f"  Both filled, DIFFERENT:           {len(disagree)}")
print(f"\n  Sample disagreements (first 15):")
for bn, p1, p2 in disagree[:15]:
    print(f"    [{(bn or '')[:35]:35s}] Parcel#={p1[:50]}  Parcel={p2[:50]}")

# Format diversity in each column
print("\n=== Format samples ===")
print("\n  Parcel # (first 25 unique non-blank):")
seen = []
for r in data:
    v = r.get('Parcel #')
    if v and str(v).strip() not in seen:
        seen.append(str(v).strip())
        if len(seen) >= 25:
            break
for s in seen:
    print(f"    {s}")

print("\n  Parcel (first 25 unique non-blank):")
seen = []
for r in data:
    v = r.get('Parcel')
    if v and str(v).strip() not in seen:
        seen.append(str(v).strip())
        if len(seen) >= 25:
            break
for s in seen:
    print(f"    {s}")
