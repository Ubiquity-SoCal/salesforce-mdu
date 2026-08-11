"""Generate the CAT1 triage workbook.

Pulls every open CAT1 opportunity + its ContentNotes from Salesforce, classifies
why each isn't being worked (via scripts/lib/cat1_notes), and writes an Excel
workbook with read-only context, the note-derived story/evidence, and editable
columns (Proposed Owner / Proposed Pursuit Status / Decision) for the cleanup
meeting. Bulk update back to SF is a separate, later step.

Re-runnable. Output: data/output/2026-06-10-cat1-triage.xlsx

Usage:  python cat1_triage_workbook.py
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from simple_salesforce import Salesforce
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPTS = Path(__file__).resolve().parents[1]   # SalesForce/scripts
SF_ROOT = Path(__file__).resolve().parents[2]   # SalesForce
sys.path.insert(0, str(SCRIPTS / "lib"))
import cat1_notes as C  # noqa: E402

CREDS_PATH = SF_ROOT / "api" / "Salesforce_Credentials.txt"
OUT = SF_ROOT / "data" / "output" / "2026-06-10-cat1-triage.xlsx"

GROUP_FILL = {
    "A. Done (won/live)":      "C6E0B4",  # green
    "B. Dead/blocked":         "F4B6B6",  # red
    "C. Route/re-verify":      "FFE699",  # amber
    "D. Workable (orphaned)":  "BDD7EE",  # blue
}
HEADER_FILL = PatternFill("solid", fgColor="44546A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def connect():
    creds = {}
    for line in open(CREDS_PATH, encoding="utf-8"):
        if ":" in line:
            k, v = line.split(":", 1)
            creds[k.strip().lower()] = v.strip()
    sf = Salesforce(username=creds["username"], password=creds["password"],
                    security_token=creds["security token"])
    lightning = creds.get("lightning url", "").rstrip("/")
    return sf, lightning


def chunked_in(sf, ids, tmpl):
    rows = []
    for i in range(0, len(ids), 200):
        inlist = ",".join(f"'{x}'" for x in ids[i:i + 200])
        rows += sf.query_all(tmpl.format(inlist=inlist))["records"]
    return rows


def fetch(sf):
    opps = sf.query_all("""
        SELECT Id, Name, Owner.Name, Owner.IsActive, StageName, Substatus__c,
               Property_State__c, Property_City__c, Units__c, Agreement_Count__c,
               IronClad_URL__c, CreatedDate, LastModifiedDate, Description
        FROM Opportunity
        WHERE Property_Category__c='Cat 1' AND IsClosed=false
        ORDER BY Owner.Name, Name
    """)["records"]
    ids = [o["Id"] for o in opps]

    links = chunked_in(sf, ids,
        "SELECT LinkedEntityId, ContentDocumentId FROM ContentDocumentLink "
        "WHERE LinkedEntityId IN ({inlist})")
    opp_docs = defaultdict(list)
    for l in links:
        opp_docs[l["LinkedEntityId"]].append(l["ContentDocumentId"])
    all_docs = sorted({l["ContentDocumentId"] for l in links})

    vers = chunked_in(sf, all_docs,
        "SELECT ContentDocumentId, TextPreview, CreatedDate FROM ContentVersion "
        "WHERE ContentDocumentId IN ({inlist}) AND IsLatest=true")
    doc_ver = {v["ContentDocumentId"]: v for v in vers}
    return opps, opp_docs, doc_ver


def active_reps(sf):
    rows = sf.query_all("""
        SELECT Owner.Name FROM Opportunity
        WHERE IsClosed=false AND Owner.IsActive=true
        GROUP BY Owner.Name ORDER BY Owner.Name
    """)["records"]
    return [r["Name"] for r in rows if r.get("Name")]


GROUP_ORDER = {"D. Workable (orphaned)": 0, "C. Route/re-verify": 1,
               "B. Dead/blocked": 2, "A. Done (won/live)": 3}


def build_rows(opps, opp_docs, doc_ver, lightning):
    out = []
    for o in opps:
        notes = sorted((doc_ver[d] for d in opp_docs.get(o["Id"], []) if d in doc_ver),
                       key=lambda v: v.get("CreatedDate") or "", reverse=True)
        bodies = [C.clean(v.get("TextPreview")) for v in notes]
        cls = C.classify_opp(bodies, o.get("Description"))
        out.append({
            "name": o["Name"],
            "link": f"{lightning}/lightning/r/Opportunity/{o['Id']}/view" if lightning else "",
            "owner": o["Owner"]["Name"],
            "active": "Yes" if o["Owner"]["IsActive"] else "No",
            "stage": o["StageName"],
            "cur_pursuit": o.get("Substatus__c") or "",
            "state": o.get("Property_State__c") or "",
            "city": o.get("Property_City__c") or "",
            "units": o.get("Units__c"),
            "agreements": o.get("Agreement_Count__c"),
            "ironclad": "Yes" if o.get("IronClad_URL__c") else "No",
            "created": (o.get("CreatedDate") or "")[:10],
            "modified": (o.get("LastModifiedDate") or "")[:10],
            "group": cls["group"], "story": cls["story"],
            "snippet": cls["snippet"][:300], "suggested": cls["suggested_pursuit_status"],
        })
    # inactive-first -> group (workable first) -> state
    out.sort(key=lambda r: (r["active"] == "Yes", GROUP_ORDER.get(r["group"], 9),
                            r["state"], r["name"]))
    return out


COLS = [
    ("Opportunity", "name", 34), ("Link", "link", 12), ("Owner", "owner", 18),
    ("Owner Active?", "active", 11), ("Stage", "stage", 20),
    ("Current Pursuit Status", "cur_pursuit", 20), ("State", "state", 7),
    ("City", "city", 16), ("Units", "units", 8), ("Agreements", "agreements", 10),
    ("IronClad?", "ironclad", 9), ("Created", "created", 11),
    ("Last Modified", "modified", 12), ("Action Group", "group", 22),
    ("Note Story", "story", 34), ("Note Snippet (evidence)", "snippet", 60),
    ("Suggested Pursuit Status", "suggested", 22),
    ("Proposed Owner", None, 18), ("Proposed Pursuit Status", None, 22),
    ("Decision / notes", None, 30),
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def write_triage(wb, rows, reps):
    ws = wb.active
    ws.title = "Triage"
    group_col = next(i for i, (h, k, w) in enumerate(COLS, 1) if k == "group")
    edit_start = next(i for i, (h, k, w) in enumerate(COLS, 1) if h == "Proposed Owner")
    edit_fill = PatternFill("solid", fgColor="FFF2CC")

    for i, (h, k, w) in enumerate(COLS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(COLS))

    for r, row in enumerate(rows, start=2):
        for i, (h, k, w) in enumerate(COLS, 1):
            cell = ws.cell(row=r, column=i)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(h in (
                "Note Snippet (evidence)", "Note Story", "Opportunity")))
            if k is None:  # editable
                cell.fill = edit_fill
                continue
            if k == "link":
                if row["link"]:
                    cell.value = "open"
                    cell.hyperlink = row["link"]
                    cell.font = Font(color="0563C1", underline="single")
                continue
            cell.value = row[k]
        ws.cell(row=r, column=group_col).fill = PatternFill(
            "solid", fgColor=GROUP_FILL.get(row["group"], "FFFFFF"))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"

    # ---- data-validation dropdowns sourced from a hidden Lists sheet
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    lists["A1"] = "Active reps"
    for i, name in enumerate(reps, start=2):
        lists.cell(row=i, column=1, value=name)
    lists["B1"] = "Pursuit statuses"
    for i, val in enumerate(C.PURSUIT_STATUS_VALUES, start=2):
        lists.cell(row=i, column=2, value=val)

    last = len(rows) + 1
    dv_owner = DataValidation(
        type="list", formula1=f"Lists!$A$2:$A${len(reps) + 1}", allow_blank=True)
    dv_ps = DataValidation(
        type="list",
        formula1=f"Lists!$B$2:$B${len(C.PURSUIT_STATUS_VALUES) + 1}", allow_blank=True)
    ws.add_data_validation(dv_owner)
    ws.add_data_validation(dv_ps)
    owner_l = get_column_letter(edit_start)
    ps_l = get_column_letter(edit_start + 1)
    dv_owner.add(f"{owner_l}2:{owner_l}{last}")
    dv_ps.add(f"{ps_l}2:{ps_l}{last}")


def write_summary(wb, rows):
    ws = wb.create_sheet("Summary")
    groups = ["D. Workable (orphaned)", "C. Route/re-verify",
              "B. Dead/blocked", "A. Done (won/live)"]
    # Action Group x Owner-Active
    ws["A1"] = "Action Group"
    ws["B1"] = "Inactive-owned"
    ws["C1"] = "Active-owned"
    ws["D1"] = "Total"
    style_header(ws, 4)
    for r, g in enumerate(groups, start=2):
        ina = sum(1 for x in rows if x["group"] == g and x["active"] == "No")
        act = sum(1 for x in rows if x["group"] == g and x["active"] == "Yes")
        ws.cell(r, 1, g).fill = PatternFill("solid", fgColor=GROUP_FILL[g])
        ws.cell(r, 2, ina)
        ws.cell(r, 3, act)
        ws.cell(r, 4, ina + act)
    tr = len(groups) + 2
    ws.cell(tr, 1, "TOTAL").font = Font(bold=True)
    for col in (2, 3, 4):
        ws.cell(tr, col, f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{tr-1})").font = Font(bold=True)

    # Note Story x State
    base = tr + 3
    states = sorted({x["state"] for x in rows if x["state"]})
    stories = [lbl for (lbl, a, p) in C.RULES] + [C.NO_STORY[0]]
    stories = [s for s in stories if any(x["story"] == s for x in rows)]
    ws.cell(base, 1, "Note Story")
    for j, st in enumerate(states, start=2):
        ws.cell(base, j, st)
    ws.cell(base, len(states) + 2, "Total")
    style_header(ws, len(states) + 2)
    # re-apply header style only to the second header row cells
    for c in range(1, len(states) + 3):
        cell = ws.cell(row=base, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, story in enumerate(stories, start=base + 1):
        ws.cell(r, 1, story)
        total = 0
        for j, st in enumerate(states, start=2):
            n = sum(1 for x in rows if x["story"] == story and x["state"] == st)
            if n:
                ws.cell(r, j, n)
            total += n
        ws.cell(r, len(states) + 2, total)
    ws.column_dimensions["A"].width = 40
    for j in range(2, len(states) + 3):
        ws.column_dimensions[get_column_letter(j)].width = 8


def write_legend(wb):
    ws = wb.create_sheet("Legend")
    ws["A1"] = "Action Group"
    ws["B1"] = "Meaning"
    style_header(ws, 2)
    meanings = [
        ("A. Done (won/live)", "Activated/built or agreement secured. Reassign owner or close-won; nothing to 'work'."),
        ("B. Dead/blocked", "Real reason not to pursue (existing competitor Bulk, denial, DQ, moratorium). Leave or close-lost."),
        ("C. Route/re-verify", "Sold or belongs to another team. Re-verify owner / hand off."),
        ("D. Workable (orphaned)", "No blocking reason in the notes. Stalled when the owner left -> reassign to an active rep and work."),
    ]
    for r, (g, m) in enumerate(meanings, start=2):
        ws.cell(r, 1, g).fill = PatternFill("solid", fgColor=GROUP_FILL[g])
        ws.cell(r, 2, m).alignment = Alignment(wrap_text=True)

    base = len(meanings) + 4
    ws.cell(base, 1, "Note Story").font = Font(bold=True)
    ws.cell(base, 2, "Suggested Pursuit Status (advisory)").font = Font(bold=True)
    for r, (story, ps) in enumerate(C.SUGGESTED_PURSUIT_STATUS.items(), start=base + 1):
        ws.cell(r, 1, story)
        ws.cell(r, 2, ps)
    note_r = base + len(C.SUGGESTED_PURSUIT_STATUS) + 2
    ws.cell(note_r, 1,
            "Note: Story/Action are keyword-derived from the opp's notes and are "
            "advisory. Always check the Note Snippet before acting. Suggested "
            "Pursuit Status never auto-fills the editable column.").alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 70


def main():
    sf, lightning = connect()
    print("Fetching CAT1 opps + notes...")
    opps, opp_docs, doc_ver = fetch(sf)
    reps = active_reps(sf)
    print(f"  opps: {len(opps)}  active reps: {len(reps)}")

    rows = build_rows(opps, opp_docs, doc_ver, lightning)

    wb = openpyxl.Workbook()
    write_triage(wb, rows, reps)
    write_summary(wb, rows)
    write_legend(wb)
    # move Lists sheet to end (already hidden)
    wb.move_sheet("Lists", offset=len(wb.sheetnames))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")

    # console summary
    g = defaultdict(int)
    for r in rows:
        g[r["group"]] += 1
    print("\nAction groups:")
    for k in ["A. Done (won/live)", "B. Dead/blocked", "C. Route/re-verify",
              "D. Workable (orphaned)"]:
        print(f"  {k:<26} {g[k]}")
    ina = sum(1 for r in rows if r["active"] == "No")
    print(f"\ninactive-owned: {ina}   active-owned: {len(rows) - ina}   total: {len(rows)}")


if __name__ == "__main__":
    main()
