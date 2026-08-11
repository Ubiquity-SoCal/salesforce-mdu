"""Enrich the 'ROEs_Missing IC#' worklist (SF report export) with property-locator
fields so the team can find each ROE in IronClad. Joins each AGR-XXXX back to its
Agreement__c + parent Opportunity in Salesforce and appends address/units/etc.
Writes a NEW file next to the original (does not overwrite). Re-runnable as the
list regenerates."""
import sys, shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "deploy"))
from _md_deploy import connect

SRC = Path(r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\ROEs_Missing IC#.xlsx")
OUT = SRC.with_name("ROEs_Missing IC#_enriched.xlsx")
AGR_COL_HEADER = "Agreement: Agreement Number"

# 1. Read the original worklist.
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
header, data = rows[0], rows[1:]
agr_idx = header.index(AGR_COL_HEADER)
agrs = [r[agr_idx] for r in data if r[agr_idx]]
print(f"{len(data)} rows; {len(agrs)} AGR numbers")

# 2. Pull enrichment from SF (single IN query; 75 < 200 ok).
sf = connect()
in_list = "'" + "','".join(agrs) + "'"
q = (f"SELECT Name, Requested_Date__c, Signer__r.Name, "
     f"Opportunity__r.Property_Address__c, Opportunity__r.Property_City__c, "
     f"Opportunity__r.Property_Zip__c, Opportunity__r.Units__c, "
     f"Opportunity__r.Management_Company__c, Opportunity__r.Account.Name "
     f"FROM Agreement__c WHERE Name IN ({in_list})")
enr = {}
for a in sf.query_all(q)["records"]:
    o = a.get("Opportunity__r") or {}
    enr[a["Name"]] = {
        "Property Address": o.get("Property_Address__c"),
        "City": o.get("Property_City__c"),
        "Zip": o.get("Property_Zip__c"),
        "Units": o.get("Units__c"),
        "Management Company": o.get("Management_Company__c"),
        "Account": (o.get("Account") or {}).get("Name") if o.get("Account") else None,
        "Signer": (a.get("Signer__r") or {}).get("Name") if a.get("Signer__r") else None,
        "Requested Date": a.get("Requested_Date__c"),
    }

# 3. Fill rates -> keep address/city/zip/units always; keep optional cols only if any data.
ALWAYS = ["Property Address", "City", "Zip", "Units"]
OPTIONAL = ["Management Company", "Account", "Signer", "Requested Date"]
n = len(agrs)
def filled(col): return sum(1 for a in agrs if enr.get(a, {}).get(col) not in (None, ""))
print("fill rates:")
for col in ALWAYS + OPTIONAL:
    print(f"  {col:<20} {filled(col)}/{n}")
add_cols = ALWAYS + [c for c in OPTIONAL if filled(c) > 0]
print("adding columns:", add_cols)

# 4. Write enriched copy (original columns + appended enrichment).
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "ROEs Missing IC#"
out_ws.append(list(header) + add_cols)
for r in data:
    agr = r[agr_idx]
    e = enr.get(agr, {})
    out_ws.append(list(r) + [e.get(c) for c in add_cols])
# light formatting: bold header, reasonable widths
for cell in out_ws[1]:
    cell.font = Font(bold=True)
for i, col in enumerate(list(header) + add_cols, start=1):
    out_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(len(str(col)) + 2, 12), 42)
out_wb.save(OUT)
print(f"\nWrote {OUT}")
