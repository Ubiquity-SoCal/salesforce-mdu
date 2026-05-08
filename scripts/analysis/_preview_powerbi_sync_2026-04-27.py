"""
Preview the PowerBI -> Salesforce sync diff WITHOUT making any changes.

Builds the same Location + Unit records the production sync script would,
compares to current SF state, and reports:
- New Property_Locations to be created
- Existing PLs that would be updated (and what fields change)
- SF PLs not in this export (potential stale)
- Same for Units

Usage: python _preview_powerbi_sync_2026-04-27.py
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from simple_salesforce import Salesforce

XLSX = Path(r'C:\Users\cass\Work_Projects\SalesForce\PowerBI_Report\data - 2026-04-27T103900.191.xlsx')
HEADER_ROW = 3
DATA_START_ROW = 4

sf = Salesforce(username='cass1@ubiquitygp.com', password='Hawaiian1984', security_token='IBSKT6CFUpSUJWxq1CMm0HkFC')


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def normalize_address(v):
    s = to_str(v)
    return re.sub(r'\s+', ' ', s) if s else None


def to_date_str(v):
    if v is None:
        return None
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s if s else None


# ── Load xlsx ──
print(f'[Load] {XLSX.name}')
wb = load_workbook(XLSX, data_only=True)
ws = wb['Sheet1']
headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
rows = []
for row_num in range(DATA_START_ROW, ws.max_row + 1):
    row = {h: ws.cell(row=row_num, column=i + 1).value for i, h in enumerate(headers) if h}
    rows.append(row)
wb.close()
print(f'  Read {len(rows)} data rows')

# Print filters from row 1
filt_cell = ws.cell(row=1, column=1).value
print(f'  Filters applied (from row 1):\n    {filt_cell.replace(chr(10), chr(10)+"    ") if filt_cell else "(none)"}')

# ── Build PL + Unit records (same logic as sync script) ──
seen_bbas = set()
loc_records = []
loc_bad = 0
unit_records = []
unit_bad = 0

for r in rows:
    bba = normalize_address(r.get('Business Base Address'))
    cid = to_str(r.get('Circuit ID'))
    if not bba or not cid:
        loc_bad += 1
        unit_bad += 1
        continue

    if bba not in seen_bbas:
        seen_bbas.add(bba)
        state = to_str(r.get('State'))
        market = to_str(r.get('Market'))
        loc_records.append({
            'Business_Base_Address__c': bba,
            'Name': bba[:80],
            'Market__c': market,
            'State__c': state,
            'FDH_Activated_Date__c': to_date_str(r.get('FDH Activated Date')),
            'FDH_Name__c': to_str(r.get('FDH Name')),
            'Serving_Area__c': to_str(r.get('Serving Area')),
            'City__c': to_str(r.get('City')),
            'Business_Building_Id__c': to_str(r.get('Business Building Id')),
            'Circuit_ID__c': to_str(r.get('Circuit ID')),
        })

    unit_records.append({
        'Circuit_ID__c': cid,
        'Name': (to_str(r.get('Address')) or '')[:80],
        'Unit__c': to_str(r.get('Unit #')),
        'Activated__c': to_str(r.get('Activated')),
        'Address_Activation_Date__c': to_date_str(r.get('Address Activation Date')),
        'ValidForFF__c': to_str(r.get('ValidForFF')),
        'Address_De_activation_Date__c': to_date_str(r.get('Address De-activation Date')),
        'Address_Deactivated__c': to_str(r.get('Address Deactivated')),
        'Ordered_Product__c': to_str(r.get('Ordered Product')),
        'AreaId__c': to_str(r.get('AreaId')),
        '_parent_bba': bba,
    })

print(f'\n[Build] {len(loc_records)} unique PLs, {len(unit_records)} units (bad rows skipped: {loc_bad})')

# ── Pull SF state ──
print('\n[SF] Querying current state...')
sf_pls = sf.query_all("""
  SELECT Id, Business_Base_Address__c, Name, Market__c, State__c, City__c,
         Circuit_ID__c, FDH_Name__c, Serving_Area__c, FDH_Activated_Date__c,
         Business_Building_Id__c, Property_Status__c, Type_of_Property__c
  FROM Property_Location__c
""")['records']
print(f'  PLs in SF: {len(sf_pls)}')
sf_units = sf.query_all("""
  SELECT Id, Circuit_ID__c, Unit__c, Activated__c,
         Property_Location__r.Business_Base_Address__c
  FROM Property_Unit__c
""")['records']
print(f'  Units in SF: {len(sf_units)}')

# Index SF
sf_pl_by_bba = {p['Business_Base_Address__c']: p for p in sf_pls if p.get('Business_Base_Address__c')}
sf_unit_by_cid = {u['Circuit_ID__c']: u for u in sf_units if u.get('Circuit_ID__c')}

# ── Diff: PLs ──
export_bbas = {r['Business_Base_Address__c'] for r in loc_records}
sf_bbas = set(sf_pl_by_bba.keys())

new_pls = export_bbas - sf_bbas
matched_pls = export_bbas & sf_bbas
sf_only_pls = sf_bbas - export_bbas

# Updates: among matched, count how many would actually change
updates = 0
unchanged = 0
for r in loc_records:
    if r['Business_Base_Address__c'] not in matched_pls:
        continue
    sf_p = sf_pl_by_bba[r['Business_Base_Address__c']]
    diffs = []
    for k in ['Market__c', 'State__c', 'City__c', 'Circuit_ID__c', 'FDH_Name__c',
              'Serving_Area__c', 'FDH_Activated_Date__c', 'Business_Building_Id__c']:
        old = sf_p.get(k)
        new = r.get(k)
        if old != new and (old or new):
            diffs.append((k, old, new))
    if diffs:
        updates += 1
    else:
        unchanged += 1

# ── Diff: Units ──
export_cids = {r['Circuit_ID__c'] for r in unit_records}
sf_cids = set(sf_unit_by_cid.keys())

new_units = export_cids - sf_cids
matched_units = export_cids & sf_cids
sf_only_units = sf_cids - export_cids

# State-by-state breakdown for new PLs
state_dist_new = Counter()
for r in loc_records:
    if r['Business_Base_Address__c'] in new_pls:
        state_dist_new[r.get('State__c')] += 1

# ── Report ──
print('\n' + '=' * 70)
print('PREVIEW SUMMARY')
print('=' * 70)
print(f'\nProperty Locations:')
print(f'  Export rows (unique BBA):        {len(export_bbas):>6}')
print(f'  SF total today:                  {len(sf_bbas):>6}')
print(f'  → NEW (in export, not in SF):    {len(new_pls):>6}')
print(f'  → UPDATE (matched, with diffs):  {updates:>6}')
print(f'  → UNCHANGED:                     {unchanged:>6}')
print(f'  → SF-ONLY (not in this export):  {len(sf_only_pls):>6}')

print(f'\n  NEW PLs by state:')
for s, c in state_dist_new.most_common():
    print(f'    {c:>5}  {s}')

print(f'\nProperty Units:')
print(f'  Export units (unique CID):       {len(export_cids):>6}')
print(f'  SF total today:                  {len(sf_cids):>6}')
print(f'  → NEW (in export, not in SF):    {len(new_units):>6}')
print(f'  → MATCHED (will be updated):     {len(matched_units):>6}')
print(f'  → SF-ONLY (not in this export):  {len(sf_only_units):>6}')

# Sample new PLs
print(f'\n  Sample NEW PLs (first 15):')
new_pl_records = [r for r in loc_records if r['Business_Base_Address__c'] in new_pls][:15]
for r in new_pl_records:
    print(f'    {r["State__c"]!s:3s} {r["City__c"]!s:20s} {r["Business_Base_Address__c"][:55]:55s}')

# Sample sample updates
print(f'\n  Sample updates with diffs (first 8):')
shown = 0
for r in loc_records:
    if shown >= 8:
        break
    bba = r['Business_Base_Address__c']
    if bba not in matched_pls:
        continue
    sf_p = sf_pl_by_bba[bba]
    diffs = []
    for k in ['Market__c', 'State__c', 'City__c', 'Circuit_ID__c', 'FDH_Name__c',
              'Serving_Area__c', 'FDH_Activated_Date__c', 'Business_Building_Id__c']:
        old = sf_p.get(k)
        new = r.get(k)
        if old != new and (old or new):
            diffs.append((k, old, new))
    if diffs:
        shown += 1
        print(f'\n    {bba[:55]}:')
        for k, old, new in diffs:
            print(f'      {k}: {old!r} -> {new!r}')
