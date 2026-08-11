"""
Match the 247 ROE_Tracking buildings against Property_Location__c.
Reports: matched / unmatched / ambiguous, plus a CSV of unmatched for review.
"""
import sys, io, re, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


SHEET = r'C:/Users/cass/Work_Projects/SalesForce/scripts/analysis/_smb_roe_copy.xlsx'

sf = Salesforce(
    username=_SF["username"],
    password=_SF["password"],
    security_token=_SF["token"],
)


def norm(s):
    """Strip ZIP, 'UNIT XXX', extra whitespace; uppercase; collapse spaces."""
    if not s:
        return ''
    s = str(s).upper()
    s = re.sub(r'\bUNIT\s+\S+', '', s)
    s = re.sub(r'\bSTE\s+\S+', '', s)
    s = re.sub(r'\bSUITE\s+\S+', '', s)
    s = re.sub(r'\b\d{5}(-\d{4})?\b', '', s)  # ZIP
    s = re.sub(r'[#,]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# 1. Read sheet
wb = openpyxl.load_workbook(SHEET, data_only=True)
ws = wb['ROE_Tracking']
hdr = [c.value for c in ws[1]]
H = {h: i for i, h in enumerate(hdr) if h is not None}
buildings = []
for row in ws.iter_rows(min_row=2, values_only=True):
    bb = row[H['Business Buildings']]
    addr = row[H.get('Updated Business Address', H['Address'])]
    state = row[H['State']]
    units = row[H['Units']]
    roe_status = row[H['ROE Status']]
    re_status = row[H['RE Status']]
    if not bb and not addr:
        continue
    buildings.append({
        'business_building': bb,
        'address': addr,
        'state': state,
        'units': units,
        'roe_status': roe_status,
        're_status': re_status,
        'norm_bb': norm(bb),
        'norm_addr': norm(addr),
    })
print(f'ROE_Tracking rows: {len(buildings)}')

# 2. Pull all Property_Location__c (limit fields, normalize on our side)
print('Pulling Property_Location__c...')
pl = sf.query_all("""
    SELECT Id, Name, Business_Base_Address__c, City__c, State__c,
           Property_Type__c, Property_Unit_Count__c, ROE_Status__c
    FROM Property_Location__c
""")['records']
print(f'Property_Location__c records: {len(pl)}')

# Build lookup maps
by_name = {}
by_bba = {}
for p in pl:
    n = norm(p.get('Name'))
    b = norm(p.get('Business_Base_Address__c'))
    by_name.setdefault(n, []).append(p)
    if b:
        by_bba.setdefault(b, []).append(p)

# 3. Match
matched = []
unmatched = []
ambiguous = []
for b in buildings:
    keys = [b['norm_addr'], b['norm_bb']]
    hits = []
    for k in keys:
        if not k:
            continue
        hits += by_name.get(k, [])
        hits += by_bba.get(k, [])
    # de-dup by Id
    uniq = {}
    for h in hits:
        uniq[h['Id']] = h
    hits = list(uniq.values())
    if len(hits) == 1:
        matched.append((b, hits[0]))
    elif len(hits) > 1:
        ambiguous.append((b, hits))
    else:
        unmatched.append(b)

print(f'\n=== RESULTS ===')
print(f'Matched: {len(matched)}')
print(f'Ambiguous (>1 PL hit): {len(ambiguous)}')
print(f'Unmatched: {len(unmatched)}')

if ambiguous:
    print(f'\n--- Ambiguous samples ---')
    for b, hits in ambiguous[:5]:
        print(f'  Sheet: {b["business_building"]}  ({b["state"]})')
        for h in hits:
            print(f'     ↔ {h["Name"]}  | bba={h.get("Business_Base_Address__c")}  | {h.get("City__c")}, {h.get("State__c")}')

if unmatched:
    print(f'\n--- Unmatched samples (first 10) ---')
    for b in unmatched[:10]:
        print(f'  · {b["business_building"]}  |  state={b["state"]}  |  units={b["units"]}')

# Write unmatched to CSV
out_csv = r'C:/Users/cass/Work_Projects/SalesForce/scripts/analysis/_smb_roe_unmatched.csv'
with open(out_csv, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['business_building', 'address', 'state', 'units', 'roe_status', 're_status'])
    for b in unmatched:
        w.writerow([b['business_building'], b['address'], b['state'], b['units'], b['roe_status'], b['re_status']])
print(f'\nUnmatched CSV: {out_csv}')

# Sanity: % match
total = len(buildings)
print(f'\nMatch rate: {len(matched)/total*100:.1f}%')
