"""
Quick inspector: dump columns + samples + value frequencies from SMB_ROE_Project.xlsx
for owner / management company / contact related fields.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
SHEET = 'ROE_Tracking'

wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[SHEET]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = []
for r in rows[1:]:
    if not any(c is not None for c in r):
        continue
    data.append(dict(zip(header, r)))

print(f"Rows: {len(data)}")
print(f"\nALL COLUMNS ({len(header)}):")
for i, h in enumerate(header):
    fill = sum(1 for r in data if r.get(h) not in (None, ''))
    print(f"  [{i:2d}] {h!r:55s} filled={fill}/{len(data)}")

# Likely owner/contact columns — look for any that contain related keywords
KEYWORDS = ['owner', 'manage', 'contact', 'phone', 'email', 'company',
            'name', 'llc', 'mgmt', 'mgr', 'tenant', 'attn', 'c/o', 'parcel']
candidates = [h for h in header if h and any(k in str(h).lower() for k in KEYWORDS)]
print(f"\n=== CANDIDATE COLUMNS for owner/contact parsing ({len(candidates)}) ===")
for col in candidates:
    print(f"\n--- {col!r} ---")
    vals = [r.get(col) for r in data if r.get(col) not in (None, '')]
    print(f"  Filled: {len(vals)}/{len(data)}")
    samples = vals[:8]
    for s in samples:
        s_str = str(s).replace('\n', ' | ')[:200]
        print(f"    {s_str}")
    if len(vals) > 8:
        # Show 3 random later samples
        import random
        random.seed(42)
        more = random.sample(vals[8:], min(3, len(vals)-8))
        for s in more:
            s_str = str(s).replace('\n', ' | ')[:200]
            print(f"    ...{s_str}")

# For Management Company specifically — full distribution if it exists
for col in candidates:
    if 'manage' in str(col).lower() or 'mgmt' in str(col).lower():
        print(f"\n=== {col} — FULL VALUE DISTRIBUTION (top 30) ===")
        c = Counter(str(r.get(col)).strip() for r in data if r.get(col) not in (None, ''))
        for v, n in c.most_common(30):
            print(f"  {n:4d}  {v[:90]}")

# Same for Owner / Owner Contact
for col in candidates:
    cl = str(col).lower()
    if 'owner' in cl and 'manage' not in cl:
        print(f"\n=== {col} — sample raw values (first 25 unique) ===")
        seen = []
        for r in data:
            v = r.get(col)
            if v and str(v).strip() not in [s for s in seen]:
                seen.append(str(v).strip())
                if len(seen) >= 25:
                    break
        for s in seen:
            print(f"  {s[:200]}")

# Detect c/o, Attn:, llc patterns in any owner-ish column
print(f"\n=== PATTERN PROBE — c/o, Attn, LLC across candidate cols ===")
patterns = {'c/o': r'c/o', 'Attn': r'(?i)attn', 'LLC': r'(?i)\bllc\b',
            'Inc': r'(?i)\binc\b', 'phone': r'\d{3}[-.\s]?\d{3}[-.\s]?\d{4}',
            'email': r'[\w.-]+@[\w.-]+'}
for col in candidates:
    counts = Counter()
    for r in data:
        v = r.get(col)
        if not v:
            continue
        s = str(v)
        for label, pat in patterns.items():
            if re.search(pat, s):
                counts[label] += 1
    if counts:
        print(f"  {col!r}:")
        for label, n in counts.most_common():
            print(f"    {label:8s} {n}")
