import openpyxl
import os

BASE = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build"
fpath = os.path.join(BASE, "CompleteEncinitasView_PBI.xlsx")

# ============================================================
# Check all formula patterns in Summary
# ============================================================
print("=" * 120)
print("ALL UNIQUE FORMULA PATTERNS in Summary sheet")
print("=" * 120)

wb = openpyxl.load_workbook(fpath, data_only=False)
ws = wb['Summary']

# Collect all formulas by column
formulas_by_col = {}
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            col_letter = cell.column_letter
            if col_letter not in formulas_by_col:
                formulas_by_col[col_letter] = []
            formulas_by_col[col_letter].append((cell.coordinate, cell.value))

for col in sorted(formulas_by_col.keys()):
    entries = formulas_by_col[col]
    print(f"\nColumn {col}: {len(entries)} formulas")
    # Show unique patterns (replace row numbers with #)
    import re
    patterns = set()
    for coord, formula in entries:
        # Generalize the formula
        pattern = re.sub(r'\d+', '#', formula)
        patterns.add(pattern)
    for p in sorted(patterns):
        print(f"  Pattern: {p}")
    # Show a few examples
    for coord, formula in entries[:3]:
        print(f"  Example: {coord} = {formula}")

wb.close()

# ============================================================
# Summary sheet structure: what's in each column region
# ============================================================
print("\n" + "=" * 120)
print("SUMMARY SHEET LAYOUT ANALYSIS")
print("=" * 120)

wb2 = openpyxl.load_workbook(fpath, data_only=True)
ws2 = wb2['Summary']

# Columns B-E seem to be a pivot table of Phase/FDH/Serviceable Date vs Address Status
# Columns G-H seem to be a lookup
# Columns I-J seem to be FDH Name vs Count

print("\nColumn B content samples (first 20 non-empty):")
count = 0
for row_idx in range(1, ws2.max_row + 1):
    val = ws2.cell(row=row_idx, column=2).value
    if val is not None and count < 20:
        print(f"  Row {row_idx}: {val}")
        count += 1

print("\nColumn I content samples (first 20 non-empty):")
count = 0
for row_idx in range(1, ws2.max_row + 1):
    val = ws2.cell(row=row_idx, column=9).value
    if val is not None and count < 20:
        print(f"  Row {row_idx}: {val}")
        count += 1

print("\nColumn J content samples (first 20 non-empty):")
count = 0
for row_idx in range(1, ws2.max_row + 1):
    val = ws2.cell(row=row_idx, column=10).value
    if val is not None and count < 20:
        print(f"  Row {row_idx}: {val}")
        count += 1

# Check columns H and beyond
print("\nAll columns with data in row 4 (header context row):")
for col_idx in range(1, ws2.max_column + 1):
    val = ws2.cell(row=4, column=col_idx).value
    if val is not None:
        print(f"  Col {col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): {val}")

print("\nAll columns with data in row 5 (sub-header row):")
for col_idx in range(1, ws2.max_column + 1):
    val = ws2.cell(row=5, column=col_idx).value
    if val is not None:
        print(f"  Col {col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): {val}")

# Check the XLOOKUP context: col G is XLOOKUP(B,PowerBI_Data!D:D,PowerBI_Data!C:C)
# PowerBI_Data col D = FDH Name, col C = FDH Activation Date
# So G = lookup the FDH Name (from Summary col B) in PowerBI_Data to get FDH Activation Date
print("\n--- XLOOKUP interpretation ---")
print("Column G formula: XLOOKUP(B_value, PowerBI_Data!D:D, PowerBI_Data!C:C)")
print("  B_value = FDH Name (from the pivot table rows)")
print("  PowerBI_Data col D = FDH Name")
print("  PowerBI_Data col C = FDH Activation Date")
print("  Result: FDH Activation Date for each FDH Name in the summary pivot")

wb2.close()

# ============================================================
# Understand the second table (cols I-J)
# ============================================================
print("\n" + "=" * 120)
print("SECOND TABLE IN SUMMARY (Columns I-J)")
print("=" * 120)

wb3 = openpyxl.load_workbook(fpath, data_only=True)
ws3 = wb3['Summary']

print("\nRows 4-6 in columns I-J:")
for row_idx in range(4, 7):
    for col_idx in range(9, 11):
        val = ws3.cell(row=row_idx, column=col_idx).value
        print(f"  Row {row_idx}, Col {openpyxl.utils.get_column_letter(col_idx)}: {val}")

# This appears to be: FDH Name / FDH Activation Date / Count of Address
# grouped differently from the left table

print("\nFull second table (I-J), rows 5-136:")
for row_idx in range(5, 137):
    i_val = ws3.cell(row=row_idx, column=9).value
    j_val = ws3.cell(row=row_idx, column=10).value
    if i_val is not None or j_val is not None:
        print(f"  Row {row_idx}: I={i_val}, J={j_val}")

wb3.close()
