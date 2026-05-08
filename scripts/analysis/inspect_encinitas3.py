"""
Final details: tail of Summary, border patterns, last rows, right-side full extent.
"""
import openpyxl
from openpyxl.utils import get_column_letter

FILE_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build\CompleteEncinitasView_PBI.xlsx"

wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False)

ws_sv = wb_values["Summary"]
ws_sf = wb_formulas["Summary"]

# 1. Last rows of the Summary sheet
print("=" * 80)
print("SUMMARY SHEET - LAST 20 ROWS")
print("=" * 80)
start = max(1, ws_sv.max_row - 20)
for row in range(start, ws_sv.max_row + 1):
    row_data = {}
    for col in range(1, ws_sv.max_column + 1):
        val = ws_sv.cell(row=row, column=col).value
        form = ws_sf.cell(row=row, column=col).value
        letter = get_column_letter(col)
        is_formula = isinstance(form, str) and form.startswith("=")
        if val is not None or is_formula:
            indent = ws_sv.cell(row=row, column=col).alignment.indent or 0
            bold = ws_sv.cell(row=row, column=col).font.bold
            entry = repr(val)
            if is_formula:
                entry += f" FORMULA:{form}"
            if bold:
                entry += " [BOLD]"
            if indent:
                entry += f" [indent={indent}]"
            row_data[letter] = entry
    if row_data:
        print(f"Row {row}: {row_data}")

# 2. Count hierarchy levels
print("\n\n" + "=" * 80)
print("SUMMARY LEFT SIDE (B-G) - HIERARCHY STATISTICS")
print("=" * 80)

phase_count = 0
fdh_count = 0
date_count = 0
phases = []

for row in range(6, ws_sv.max_row + 1):
    cell = ws_sv.cell(row=row, column=2)
    val = cell.value
    indent = cell.alignment.indent if cell.alignment.indent else 0

    if val is None:
        continue

    if indent == 0 and isinstance(val, int):
        phase_count += 1
        phases.append(val)
    elif indent == 1:
        fdh_count += 1
    elif indent == 2:
        date_count += 1

print(f"Phase rows (indent=0, int): {phase_count}")
print(f"FDH Name rows (indent=1): {fdh_count}")
print(f"Serviceable Date rows (indent=2): {date_count}")
print(f"Phase values: {sorted(phases)}")

# 3. Right side (I, J) - hierarchy statistics
print("\n\n" + "=" * 80)
print("SUMMARY RIGHT SIDE (I, J) - HIERARCHY STATISTICS")
print("=" * 80)

fdh_right = 0
date_right = 0
right_side_extent = 0

for row in range(5, ws_sv.max_row + 1):
    cell_i = ws_sv.cell(row=row, column=9)
    val_i = cell_i.value
    if val_i is not None:
        right_side_extent = row
        indent = cell_i.alignment.indent if cell_i.alignment.indent else 0
        if indent == 0:
            fdh_right += 1
        elif indent == 1:
            date_right += 1

print(f"FDH Name rows (indent=0): {fdh_right}")
print(f"FDH Activation Date rows (indent=1): {date_right}")
print(f"Right side extends to row: {right_side_extent}")

# 4. Verify: does left side and right side have same extent?
left_extent = 0
for row in range(6, ws_sv.max_row + 1):
    if ws_sv.cell(row=row, column=2).value is not None:
        left_extent = row
print(f"Left side extends to row: {left_extent}")

# 5. Check for Grand Total row
print("\n\n" + "=" * 80)
print("LOOKING FOR GRAND TOTAL / SUMMARY ROWS")
print("=" * 80)
for row in range(ws_sv.max_row - 5, ws_sv.max_row + 1):
    for col in range(1, ws_sv.max_column + 1):
        val = ws_sv.cell(row=row, column=col).value
        if val is not None:
            letter = get_column_letter(col)
            bold = ws_sv.cell(row=row, column=col).font.bold
            print(f"  Row {row} {letter}: {repr(val)} {'[BOLD]' if bold else ''}")

# 6. Check C column (future_serviceable) - when does it have values vs not?
print("\n\n" + "=" * 80)
print("COLUMN C (future_serviceable) - ROWS WITH VALUES")
print("=" * 80)
c_rows = 0
c_rows_list = []
for row in range(6, ws_sv.max_row + 1):
    val = ws_sv.cell(row=row, column=3).value
    if val is not None:
        c_rows += 1
        if c_rows <= 30:
            b_val = ws_sv.cell(row=row, column=2).value
            b_indent = ws_sv.cell(row=row, column=2).alignment.indent or 0
            c_rows_list.append((row, val, b_val, b_indent))
print(f"Total rows with C values: {c_rows}")
print("First 30:")
for r, c, b, bi in c_rows_list:
    print(f"  Row {r}: C={c}, B={repr(b)} indent={bi}")

# 7. G column (XLOOKUP) - check behavior pattern
print("\n\n" + "=" * 80)
print("COLUMN G (XLOOKUP) - RETURN VALUE PATTERN")
print("=" * 80)
g_has_value = 0
g_empty = 0
g_value_at_indent = {0: 0, 1: 0, 2: 0}
g_empty_at_indent = {0: 0, 1: 0, 2: 0}

for row in range(6, ws_sv.max_row + 1):
    g_val = ws_sv.cell(row=row, column=7).value
    b_indent = ws_sv.cell(row=row, column=2).alignment.indent or 0
    b_indent = int(b_indent)
    if b_indent not in g_value_at_indent:
        g_value_at_indent[b_indent] = 0
        g_empty_at_indent[b_indent] = 0

    if g_val is not None and g_val != "":
        g_has_value += 1
        g_value_at_indent[b_indent] = g_value_at_indent.get(b_indent, 0) + 1
    else:
        g_empty += 1
        g_empty_at_indent[b_indent] = g_empty_at_indent.get(b_indent, 0) + 1

print(f"G column has value: {g_has_value}")
print(f"G column empty/None: {g_empty}")
print(f"G has value at indent levels: {g_value_at_indent}")
print(f"G empty at indent levels: {g_empty_at_indent}")
print("(XLOOKUP looks up B column value in PowerBI_Data D:D, returns C:C i.e. FDH Activation Date)")
print("So it returns a date only when B contains an FDH Name (indent=1), not for Phase nums or dates")


# 8. D and E columns - are they formulas or static?
print("\n\n" + "=" * 80)
print("D AND E COLUMNS - FORMULA CHECK")
print("=" * 80)
d_formula_count = 0
d_static_count = 0
e_formula_count = 0
e_static_count = 0
for row in range(6, ws_sv.max_row + 1):
    for col_idx, col_name in [(4, "D"), (5, "E")]:
        form = ws_sf.cell(row=row, column=col_idx).value
        val = ws_sv.cell(row=row, column=col_idx).value
        is_formula = isinstance(form, str) and form.startswith("=")
        if col_idx == 4:
            if is_formula:
                d_formula_count += 1
            elif val is not None:
                d_static_count += 1
        else:
            if is_formula:
                e_formula_count += 1
            elif val is not None:
                e_static_count += 1

print(f"Column D: {d_formula_count} formulas, {d_static_count} static values")
print(f"Column E: {e_formula_count} formulas, {e_static_count} static values")

# 9. Check C column formulas
c_formula_count = 0
c_static_count = 0
for row in range(6, ws_sv.max_row + 1):
    form = ws_sf.cell(row=row, column=3).value
    val = ws_sv.cell(row=row, column=3).value
    is_formula = isinstance(form, str) and form.startswith("=")
    if is_formula:
        c_formula_count += 1
    elif val is not None:
        c_static_count += 1
print(f"Column C: {c_formula_count} formulas, {c_static_count} static values")

# 10. J column formulas?
j_formula_count = 0
j_static_count = 0
for row in range(5, ws_sv.max_row + 1):
    form = ws_sf.cell(row=row, column=10).value
    val = ws_sv.cell(row=row, column=10).value
    is_formula = isinstance(form, str) and form.startswith("=")
    if is_formula:
        j_formula_count += 1
    elif val is not None:
        j_static_count += 1
print(f"Column J: {j_formula_count} formulas, {j_static_count} static values")

# 11. Border patterns on key rows
print("\n\n" + "=" * 80)
print("BORDER PATTERNS - rows 2-10")
print("=" * 80)
for row in range(2, 11):
    borders = []
    for col in range(1, 11):
        cell = ws_sv.cell(row=row, column=col)
        letter = get_column_letter(col)
        b = cell.border
        sides = {}
        for s in ["top", "bottom", "left", "right"]:
            side = getattr(b, s)
            if side and side.style:
                sides[s] = side.style
        if sides:
            borders.append(f"{letter}:{sides}")
    if borders:
        print(f"  Row {row}: {borders}")

# 12. Fill/background colors on key rows
print("\n\n" + "=" * 80)
print("FILL COLORS - rows 2-10")
print("=" * 80)
for row in range(2, 11):
    fills = []
    for col in range(1, 11):
        cell = ws_sv.cell(row=row, column=col)
        letter = get_column_letter(col)
        fill = cell.fill
        if fill.patternType and fill.patternType != "none":
            fills.append(f"{letter}: pattern={fill.patternType}, fg={fill.fgColor.rgb if fill.fgColor else None}, "
                        f"bg={fill.bgColor.rgb if fill.bgColor else None}")
    if fills:
        print(f"  Row {row}: {fills}")

wb_values.close()
wb_formulas.close()
print("\n\nDone.")
