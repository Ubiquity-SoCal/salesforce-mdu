# -*- coding: utf-8 -*-
"""
Analyze which columns in the Encinitas PBI file were enriched via XLOOKUP
from Invoice Support files, by comparing against the base Carlsbad PBI columns.
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from pathlib import Path

BASE = Path(r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build")


def get_header_row(ws, max_scan=10):
    """Find the first row where most cells have string values (the header row).
    Returns (row_number, headers_list)."""
    best_row = 1
    best_headers = []
    best_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), 1):
        vals = [cell.value for cell in row]
        str_count = sum(1 for v in vals if isinstance(v, str) and v.strip())
        if str_count > best_count:
            best_count = str_count
            best_row = row_idx
            best_headers = vals
    return best_row, best_headers


# ========================================================================
# 1. Carlsbad PBI headers (BASE columns)
# ========================================================================
print("=" * 80)
print("1. CARLSBAD PBI (ComprehensiveCarlsbadView_PBI.xlsx  -  Sheet1)")
print("=" * 80)

wb = openpyxl.load_workbook(BASE / "ComprehensiveCarlsbadView_PBI.xlsx", read_only=True, data_only=True)
ws = wb["Sheet1"]
carlsbad_pbi_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
wb.close()

print(f"Column count: {len(carlsbad_pbi_headers)}")
for i, h in enumerate(carlsbad_pbi_headers, 1):
    print(f"  {i:>3}. {h}")

# ========================================================================
# 2. Encinitas PBI headers (ENRICHED columns)
# ========================================================================
print("\n" + "=" * 80)
print("2. ENCINITAS PBI (CompleteEncinitasView_PBI.xlsx  -  PowerBI_Data)")
print("=" * 80)

wb = openpyxl.load_workbook(BASE / "CompleteEncinitasView_PBI.xlsx", read_only=True, data_only=True)
ws = wb["PowerBI_Data"]
encinitas_pbi_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
wb.close()

print(f"Column count: {len(encinitas_pbi_headers)}")
for i, h in enumerate(encinitas_pbi_headers, 1):
    print(f"  {i:>3}. {h}")

# ========================================================================
# 3. Columns in Encinitas but NOT in Carlsbad
# ========================================================================
print("\n" + "=" * 80)
print("3. ENRICHED COLUMNS (in Encinitas PBI but NOT in Carlsbad PBI)")
print("   --> These are the columns added via XLOOKUP from Invoice Support")
print("=" * 80)

carlsbad_set = set(carlsbad_pbi_headers)
enriched = [(i, h) for i, h in enumerate(encinitas_pbi_headers, 1) if h not in carlsbad_set]

print(f"Count: {len(enriched)}")
for pos, h in enriched:
    print(f"  Col {pos:>2} in Encinitas PBI: {h}")

# Sanity check: Carlsbad columns missing from Encinitas
encinitas_set = set(encinitas_pbi_headers)
missing_from_enc = [h for h in carlsbad_pbi_headers if h not in encinitas_set]
if missing_from_enc:
    print(f"\n  [Note] Carlsbad columns NOT present in Encinitas ({len(missing_from_enc)}):")
    for h in missing_from_enc:
        print(f"    - {h}")

# ========================================================================
# 4. Encinitas Invoice Support headers
# ========================================================================
print("\n" + "=" * 80)
print("4. ENCINITAS INVOICE SUPPORT (ENC_InvoiceSupport_February_MarchReview.xlsx)")
print("   Sheet: 'ServiceableDoors in Jan'")
print("=" * 80)

wb = openpyxl.load_workbook(BASE / "ENC_InvoiceSupport_February_MarchReview.xlsx", read_only=True, data_only=True)
print(f"  Available sheets: {wb.sheetnames}")
ws = wb["ServiceableDoors in Jan"]
enc_inv_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
wb.close()

print(f"Column count: {len(enc_inv_headers)}")
for i, h in enumerate(enc_inv_headers, 1):
    print(f"  {i:>3}. {h}")

# ========================================================================
# 5. Carlsbad Invoice Support headers
# ========================================================================
print("\n" + "=" * 80)
print("5. CARLSBAD INVOICE SUPPORT")
print("   (Carlsbad Invoice Support_February - March ViewCopy.xlsm)")
print("   Sheet: 'Serviceable Doors in Jan'")
print("=" * 80)

wb = openpyxl.load_workbook(BASE / "Carlsbad Invoice Support_February - March ViewCopy.xlsm",
                            read_only=True, data_only=True)
print(f"  Available sheets: {wb.sheetnames}")
target_sheet = None
for sn in wb.sheetnames:
    if "serviceable" in sn.lower() and "jan" in sn.lower():
        target_sheet = sn
        break

carl_inv_headers = []
if target_sheet is None:
    print("  WARNING: Could not find 'Serviceable Doors in Jan' sheet.")
else:
    ws = wb[target_sheet]
    print(f"  Using sheet: '{target_sheet}'")
    # Scan first 10 rows to find the real header row
    row_num, carl_inv_headers = get_header_row(ws, max_scan=10)
    print(f"  Header row detected at row: {row_num}")
    print(f"  Column count: {len(carl_inv_headers)}")
    for i, h in enumerate(carl_inv_headers, 1):
        print(f"  {i:>3}. {h}")
wb.close()

# ========================================================================
# 6. Solana Beach Invoice Support headers
# ========================================================================
print("\n" + "=" * 80)
print("6. SOLANA BEACH INVOICE SUPPORT")
print("   (Solana Beach Invoice Support_February_MarchView.xlsm)")
print("   Sheet: 'Serviceable Doors - in Jan'")
print("=" * 80)

wb = openpyxl.load_workbook(BASE / "Solana Beach Invoice Support_February_MarchView.xlsm",
                            read_only=True, data_only=True)
print(f"  Available sheets: {wb.sheetnames}")
target_sheet = None
for sn in wb.sheetnames:
    if "serviceable" in sn.lower() and "jan" in sn.lower():
        target_sheet = sn
        break

sb_inv_headers = []
if target_sheet is None:
    print("  WARNING: Could not find matching sheet.")
else:
    ws = wb[target_sheet]
    print(f"  Using sheet: '{target_sheet}'")
    row_num, sb_inv_headers = get_header_row(ws, max_scan=10)
    print(f"  Header row detected at row: {row_num}")
    print(f"  Column count: {len(sb_inv_headers)}")
    for i, h in enumerate(sb_inv_headers, 1):
        print(f"  {i:>3}. {h}")
wb.close()

# ========================================================================
# 7. Mapping: enriched PBI columns --> invoice support columns
# ========================================================================
print("\n" + "=" * 80)
print("7. COLUMN MAPPING: Enriched Encinitas PBI cols --> Invoice Support sources")
print("=" * 80)

enriched_names = [h for _, h in enriched]


def norm(s):
    """Normalise a column name for fuzzy comparison."""
    if s is None:
        return ""
    return s.strip().lower().replace("_", " ").replace("-", " ")


enc_inv_norm = {norm(h): h for h in enc_inv_headers if h}
carl_inv_norm = {norm(h): h for h in carl_inv_headers if h}
sb_inv_norm = {norm(h): h for h in sb_inv_headers if h}

print(f"\n{'Enriched PBI Column':<45} | {'ENC Invoice?':<35} | {'CARL Invoice?':<35} | {'SB Invoice?'}")
print("-" * 160)

for col in enriched_names:
    cn = norm(col)
    enc_match = enc_inv_norm.get(cn, "")
    carl_match = carl_inv_norm.get(cn, "")
    sb_match = sb_inv_norm.get(cn, "")

    # If exact normalised match didn't work, try substring / contains
    if not enc_match:
        for k, v in enc_inv_norm.items():
            if cn and k and (cn in k or k in cn):
                enc_match = f"~{v}"
                break
    if not carl_match:
        for k, v in carl_inv_norm.items():
            if cn and k and (cn in k or k in cn):
                carl_match = f"~{v}"
                break
    if not sb_match:
        for k, v in sb_inv_norm.items():
            if cn and k and (cn in k or k in cn):
                sb_match = f"~{v}"
                break

    e_flag = "YES" if enc_match and not enc_match.startswith("~") else ("PARTIAL" if enc_match else "---")
    c_flag = "YES" if carl_match and not carl_match.startswith("~") else ("PARTIAL" if carl_match else "---")
    s_flag = "YES" if sb_match and not sb_match.startswith("~") else ("PARTIAL" if sb_match else "---")

    enc_display = enc_match.lstrip("~") if enc_match else ""
    carl_display = carl_match.lstrip("~") if carl_match else ""
    sb_display = sb_match.lstrip("~") if sb_match else ""

    print(f"  {str(col):<43} | {e_flag:<6} {enc_display:<28} | {c_flag:<6} {carl_display:<28} | {s_flag:<6} {sb_display}")

# ========================================================================
# 8. All Invoice Support columns NOT used in PBI enrichment
# ========================================================================
print("\n" + "=" * 80)
print("8. INVOICE SUPPORT COLUMNS *NOT* AMONG THE ENRICHED PBI COLUMNS")
print("   (These exist in Invoice but were NOT pulled into PBI)")
print("=" * 80)

enriched_norm = {norm(h) for h in enriched_names}

print("\n  Encinitas Invoice Support unused columns:")
for h in enc_inv_headers:
    if h and norm(h) not in enriched_norm:
        # Check partial too
        found = any(norm(h) in en or en in norm(h) for en in enriched_norm if en)
        tag = " (partial match with enriched)" if found else ""
        print(f"    - {h}{tag}")

if carl_inv_headers:
    print("\n  Carlsbad Invoice Support unused columns:")
    for h in carl_inv_headers:
        if h and norm(h) not in enriched_norm:
            found = any(norm(h) in en or en in norm(h) for en in enriched_norm if en)
            tag = " (partial match with enriched)" if found else ""
            print(f"    - {h}{tag}")

if sb_inv_headers:
    print("\n  Solana Beach Invoice Support unused columns:")
    for h in sb_inv_headers:
        if h and norm(h) not in enriched_norm:
            found = any(norm(h) in en or en in norm(h) for en in enriched_norm if en)
            tag = " (partial match with enriched)" if found else ""
            print(f"    - {h}{tag}")

# ========================================================================
# SUMMARY
# ========================================================================
print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"  Carlsbad PBI base columns:            {len(carlsbad_pbi_headers)}")
print(f"  Encinitas PBI total columns:           {len(encinitas_pbi_headers)}")
print(f"  Enriched (XLOOKUP'd) columns:          {len(enriched)}")
print(f"  Encinitas Invoice Support columns:     {len(enc_inv_headers)}")
print(f"  Carlsbad Invoice Support columns:      {len(carl_inv_headers)}")
print(f"  Solana Beach Invoice Support columns:  {len(sb_inv_headers)}")
