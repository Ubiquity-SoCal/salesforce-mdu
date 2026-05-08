"""
Generate a combined PBI vs Build Match & Date Discrepancy HTML report.

Reads from all 3 enriched PBI files:
  - Match Report sheet: Circuit ID match stats (PBI vs Invoice)
  - PowerBI_Data sheet: Per-address data to compute date discrepancies

Groups by Build Phase to show:
  - Which FDH names PowerBI assigned to each phase
  - The single PowerBI FDH Activation Date vs multiple Build serviceable dates
  - Addresses with no phase go into a single "Unassigned" bucket
"""

import openpyxl
from datetime import datetime, date, time as dtime
from pathlib import Path
from collections import defaultdict, OrderedDict

BASE = Path(r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build")

FILES = {
    "Encinitas":    BASE / "CompleteEncinitasView_PBI.xlsx",
    "Carlsbad":     BASE / "CompleteCarlsbadView_PBI.xlsx",
    "Solana Beach": BASE / "CompleteSolanaBeachView_PBI.xlsx",
}


def fmt_date(val):
    """Format a date value to M/D/YYYY string."""
    if isinstance(val, dtime) and not isinstance(val, (datetime, date)):
        return None
    if isinstance(val, datetime):
        return f"{val.month}/{val.day}/{val.year}"
    if isinstance(val, date):
        return f"{val.month}/{val.day}/{val.year}"
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None


def date_sort_key(d):
    """Sort key for date strings in M/D/YYYY format."""
    if not d or d == "(no date)":
        return "9999"
    try:
        parts = d.split("/")
        return f"{int(parts[2]):04d}{int(parts[0]):02d}{int(parts[1]):02d}"
    except (ValueError, IndexError):
        return d


def phase_sort_key(phase):
    """Sort phases numerically where possible."""
    if not phase:
        return (999999, "")
    s = str(phase).strip()
    # Pure number
    try:
        return (int(float(s)), "")
    except (ValueError, TypeError):
        pass
    # CV11-41 -> sort by 41
    if "-" in s:
        num = s.split("-")[-1]
        try:
            return (int(num), s)
        except ValueError:
            pass
    # SA01 FDH05 -> sort by 5
    if "FDH" in s:
        for part in s.split():
            if part.startswith("FDH"):
                try:
                    return (int(part[3:]), s)
                except ValueError:
                    pass
    return (99999, s)


def read_match_report(path):
    """Read the Match Report sheet -- summary metrics + unmatched circuits."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Match Report"]

    metrics = {}
    for row in ws.iter_rows(min_row=4, max_row=8, values_only=True):
        label, value = row[0], row[1]
        if label:
            metrics[label] = value

    unmatched = []
    for row in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True):
        if row[0] is not None:
            unmatched.append({
                "circuit_id": row[0],
                "address": row[1] or "",
                "fdh_name": row[2] or "",
                "status": row[3] or "",
            })

    wb.close()
    return metrics, unmatched


def read_pbi_data(path):
    """Read the PowerBI_Data sheet. Returns list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["PowerBI_Data"]

    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for j, v in enumerate(row):
            if v:
                headers[v.strip()] = j
        break

    col_addr = headers.get("Address")
    col_fdh_date = headers.get("FDH Activation Date")
    col_fdh_name = headers.get("FDH Name")
    col_phase = headers.get("Phase")
    col_svc_date = headers.get("Serviceable from date")
    col_circuit = headers.get("Circuit ID")
    col_status = headers.get("Address Status")

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        address = row[col_addr] if col_addr is not None else None
        if not address:
            continue

        phase = row[col_phase] if col_phase is not None else None
        if phase is not None:
            phase = str(phase).strip()
            if phase == "" or phase == "None":
                phase = None

        records.append({
            "address": address,
            "fdh_activation_date": fmt_date(row[col_fdh_date]) if col_fdh_date is not None else None,
            "fdh_name": row[col_fdh_name] if col_fdh_name is not None else "",
            "phase": phase,
            "serviceable_from_date": fmt_date(row[col_svc_date]) if col_svc_date is not None else None,
            "circuit_id": row[col_circuit] if col_circuit is not None else "",
            "status": row[col_status] if col_status is not None else "",
        })

    wb.close()
    return records


def analyze_by_phase(records):
    """
    Group records by Build Phase. For each phase, show:
    - Which FDH names are in it (and their PowerBI dates)
    - The Build serviceable dates per FDH
    """
    phase_groups = defaultdict(list)
    no_phase = []

    for r in records:
        if r["phase"]:
            phase_groups[r["phase"]].append(r)
        else:
            no_phase.append(r)

    phase_analysis = []
    for phase in sorted(phase_groups.keys(), key=phase_sort_key):
        addrs = phase_groups[phase]

        # Group by FDH within this phase
        fdh_sub = defaultdict(list)
        for a in addrs:
            fdh_sub[a["fdh_name"] or "(no FDH)"].append(a)

        fdh_details = []
        for fdh_name in sorted(fdh_sub.keys()):
            fdh_addrs = fdh_sub[fdh_name]
            pbi_date = fdh_addrs[0]["fdh_activation_date"] or "N/A"

            # Status counts
            svc = sum(1 for a in fdh_addrs if a["status"] == "serviceable")
            future = sum(1 for a in fdh_addrs if a["status"] == "future_serviceable")

            # Build date breakdown
            build_dates = defaultdict(int)
            for a in fdh_addrs:
                d = a["serviceable_from_date"] or "(no date)"
                build_dates[d] += 1
            sorted_dates = sorted(build_dates.items(), key=lambda x: date_sort_key(x[0]))

            fdh_details.append({
                "fdh_name": fdh_name,
                "pbi_date": pbi_date,
                "count": len(fdh_addrs),
                "serviceable": svc,
                "future_serviceable": future,
                "build_dates": sorted_dates,
                "has_multi_dates": len(build_dates) > 1,
            })

        # Flags
        has_multi_fdhs = len(fdh_details) > 1
        has_date_issues = any(f["has_multi_dates"] for f in fdh_details)

        # Phase-level status totals
        phase_svc = sum(1 for a in addrs if a["status"] == "serviceable")
        phase_future = sum(1 for a in addrs if a["status"] == "future_serviceable")

        phase_analysis.append({
            "phase": phase,
            "total": len(addrs),
            "serviceable": phase_svc,
            "future_serviceable": phase_future,
            "fdh_details": fdh_details,
            "has_multi_fdhs": has_multi_fdhs,
            "has_date_issues": has_date_issues,
        })

    # Bucket for no-phase addresses, grouped by FDH
    no_phase_fdhs = defaultdict(list)
    for a in no_phase:
        no_phase_fdhs[a["fdh_name"] or "(no FDH)"].append(a)

    no_phase_details = []
    for fdh_name in sorted(no_phase_fdhs.keys()):
        fdh_addrs = no_phase_fdhs[fdh_name]
        pbi_date = fdh_addrs[0]["fdh_activation_date"] or "N/A"
        svc = sum(1 for a in fdh_addrs if a["status"] == "serviceable")
        future = sum(1 for a in fdh_addrs if a["status"] == "future_serviceable")
        no_phase_details.append({
            "fdh_name": fdh_name,
            "pbi_date": pbi_date,
            "count": len(fdh_addrs),
            "serviceable": svc,
            "future_serviceable": future,
        })

    return phase_analysis, no_phase_details, len(no_phase)


def generate_html(all_data):
    """Generate the combined HTML report. Top-level tabs = cities."""
    now = datetime.now().strftime("%B %d, %Y at %I:%M %p")

    total_pbi = sum(d["metrics"].get("PBI records", 0) for d in all_data.values())
    total_matched = sum(d["metrics"].get("Matched to Invoice", 0) for d in all_data.values())
    total_unmatched = sum(d["metrics"].get("Not in Invoice", 0) for d in all_data.values())

    # Pre-compute per-city stats for the summary table
    city_rows = []
    for city, data in all_data.items():
        m = data["metrics"]
        phases = data["phases"]
        issue_count = sum(1 for p in phases if p["has_date_issues"] or p["has_multi_fdhs"])
        issue_addrs = sum(p["total"] for p in phases if p["has_date_issues"] or p["has_multi_fdhs"])
        total_svc = sum(p["serviceable"] for p in phases) + sum(np["serviceable"] for np in data["no_phase_details"])
        total_future = sum(p["future_serviceable"] for p in phases) + sum(np["future_serviceable"] for np in data["no_phase_details"])
        city_rows.append({
            "city": city,
            "pbi": m.get("PBI records", 0),
            "serviceable": total_svc,
            "future_serviceable": total_future,
            "matched": m.get("Matched to Invoice", 0),
            "unmatched": m.get("Not in Invoice", 0),
            "phases": len(phases),
            "issues": issue_count,
            "issue_addrs": issue_addrs,
            "no_phase": data["no_phase_count"],
        })

    total_phases_issues = sum(r["issues"] for r in city_rows)
    total_issue_addrs = sum(r["issue_addrs"] for r in city_rows)
    total_no_phase = sum(r["no_phase"] for r in city_rows)

    # Build city tab buttons
    city_tab_buttons = ""
    city_ids = []
    for i, (city, _) in enumerate(all_data.items()):
        cid = city.lower().replace(" ", "")
        city_ids.append(cid)
        active = " active" if i == 0 else ""
        city_tab_buttons += f'    <button class="city-tab{active}" onclick="switchCity(\'{cid}\')">{city}</button>\n'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PBI vs Build &mdash; Phase &amp; Date Summary</title>
<style>
  :root {{
    --blue: #1565C0;
    --blue-light: #E3F2FD;
    --blue-dark: #0D47A1;
    --green: #2E7D32;
    --green-light: #E8F5E9;
    --orange: #E65100;
    --orange-light: #FFF3E0;
    --red: #C62828;
    --red-light: #FFEBEE;
    --gray: #546E7A;
    --gray-light: #ECEFF1;
    --purple: #6A1B9A;
    --purple-light: #F3E5F5;
    --bg: #F5F7FA;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: var(--bg);
    color: #263238;
    line-height: 1.5;
    padding: 20px;
  }}
  .container {{ max-width: 1400px; margin: 0 auto; }}
  header {{
    background: linear-gradient(135deg, var(--blue-dark), var(--blue));
    color: white;
    padding: 24px 32px;
    border-radius: 12px;
    margin-bottom: 24px;
  }}
  header h1 {{ font-size: 1.6em; font-weight: 600; }}
  header .subtitle {{ opacity: 0.85; margin-top: 4px; font-size: 0.9em; }}

  .billing-callout {{
    background: #FFF8E1;
    border: 1px solid #FFD54F;
    border-left: 4px solid #F9A825;
    border-radius: 8px;
    padding: 14px 16px;
    margin-bottom: 20px;
    font-size: 0.88em;
  }}
  .billing-callout strong {{ color: #E65100; }}

  /* Summary table */
  .summary-table-wrap {{
    background: white;
    border-radius: 10px;
    padding: 20px;
    margin-bottom: 24px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
  }}
  .summary-table-wrap h3 {{ font-size: 1.05em; color: var(--blue-dark); margin-bottom: 12px; }}
  .summary-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.88em;
  }}
  .summary-table th {{
    background: var(--blue-dark);
    color: white;
    padding: 10px 14px;
    text-align: right;
    font-weight: 600;
  }}
  .summary-table th:first-child {{ text-align: left; }}
  .summary-table td {{
    padding: 10px 14px;
    border-bottom: 1px solid #e0e0e0;
    text-align: right;
    font-variant-numeric: tabular-nums;
  }}
  .summary-table td:first-child {{ text-align: left; font-weight: 600; color: var(--blue-dark); }}
  .summary-table tr:last-child td {{
    border-top: 2px solid var(--blue-dark);
    font-weight: 700;
    background: var(--blue-light);
  }}
  .summary-table tr:hover td {{ background: #F5F5F5; }}
  .summary-table tr:last-child:hover td {{ background: var(--blue-light); }}
  /* removed rate styling - no longer used */

  /* City tabs (top-level) */
  .city-tabs {{
    display: flex;
    gap: 0;
    margin-bottom: 0;
  }}
  .city-tab {{
    padding: 12px 28px;
    background: var(--gray-light);
    border: none;
    cursor: pointer;
    font-size: 1em;
    font-weight: 500;
    border-radius: 10px 10px 0 0;
    color: var(--gray);
    transition: all 0.2s;
  }}
  .city-tab:hover {{ background: #CFD8DC; }}
  .city-tab.active {{
    background: white;
    color: var(--blue-dark);
    font-weight: 700;
    box-shadow: 0 -2px 6px rgba(0,0,0,0.06);
  }}
  .city-content {{
    display: none;
    background: white;
    border-radius: 0 10px 10px 10px;
    padding: 24px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    margin-bottom: 24px;
  }}
  .city-content.active {{ display: block; }}

  /* Sub-section headers (collapsible) */
  .section-toggle {{
    cursor: pointer;
    user-select: none;
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 12px 0;
    font-weight: 700;
    font-size: 1em;
    color: var(--blue-dark);
    border-bottom: 2px solid var(--blue-light);
    margin-bottom: 12px;
  }}
  .section-toggle .arrow-icon {{ transition: transform 0.2s; display: inline-block; font-size: 0.8em; }}
  .section-toggle.open .arrow-icon {{ transform: rotate(90deg); }}
  .section-toggle .section-count {{
    font-size: 0.75em;
    font-weight: 500;
    color: var(--gray);
    margin-left: auto;
  }}
  .section-body {{ display: none; margin-bottom: 24px; }}
  .section-body.open {{ display: block; }}

  /* Phase grid & cards */
  /* Phase table */
  .phase-table {{ font-size: 0.82em; }}
  .phase-table th {{ padding: 6px 8px; white-space: nowrap; }}
  .phase-table td {{ padding: 5px 8px; }}
  .phase-row {{ cursor: default; }}
  .phase-row:hover td {{ background: rgba(0,0,0,0.03); }}
  .expand-arrow {{ display: inline-block; transition: transform 0.2s; font-size: 0.75em; }}
  .expand-arrow.open {{ transform: rotate(90deg); }}

  .detail-row td {{ padding: 0 8px 8px 8px !important; background: #FAFAFA; }}
  .detail-fdh {{
    margin: 4px 0;
    padding: 6px 10px;
    background: white;
    border: 1px solid #e0e0e0;
    border-radius: 5px;
  }}
  .detail-fdh-header {{ font-size: 0.88em; margin-bottom: 3px; }}
  .detail-date {{
    display: flex;
    justify-content: space-between;
    font-size: 0.82em;
    padding: 1px 0;
    border-bottom: 1px dotted #eee;
  }}
  .detail-date:last-child {{ border-bottom: none; }}
  .detail-cnt {{ font-weight: 600; color: var(--gray); }}

  .badge {{
    font-size: 0.62em;
    padding: 2px 7px;
    border-radius: 10px;
    font-weight: 600;
    text-transform: uppercase;
    margin-left: 2px;
  }}
  .badge-warn {{ background: #FFE0B2; color: var(--orange); }}
  .badge-purple {{ background: #E1BEE7; color: var(--purple); }}
  .badge-ok {{ background: #C8E6C9; color: var(--green); }}

  .date-row {{
    display: flex;
    justify-content: space-between;
    padding: 2px 0;
    font-size: 0.8em;
    border-bottom: 1px dotted #ddd;
  }}
  .date-row:last-child {{ border-bottom: none; }}
  .date-row .d {{ color: #333; }}
  .date-row .c {{ font-weight: 600; color: var(--gray); }}
  .date-mismatch {{ color: var(--red) !important; font-weight: 600; }}

  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.82em;
  }}
  th {{
    background: var(--blue-light);
    color: var(--blue-dark);
    padding: 8px 10px;
    text-align: left;
    font-weight: 600;
    position: sticky;
    top: 0;
  }}
  td {{ padding: 5px 10px; border-bottom: 1px solid #eee; }}
  tr:hover td {{ background: #F5F5F5; }}

  .legend {{
    display: flex;
    gap: 14px;
    flex-wrap: wrap;
    margin-bottom: 12px;
    padding: 8px 12px;
    background: var(--gray-light);
    border-radius: 8px;
    font-size: 0.76em;
  }}
  .legend-item {{ display: flex; align-items: center; gap: 5px; }}
  .legend-swatch {{ width: 14px; height: 14px; border-radius: 3px; border: 1px solid #ccc; }}

  .scrollable {{ max-height: 500px; overflow-y: auto; }}

  @media (max-width: 768px) {{
    .city-tab {{ padding: 10px 16px; font-size: 0.9em; }}
  }}
</style>
</head>
<body>
<div class="container">

<header>
  <h1>PBI vs Build &mdash; Phase &amp; Date Discrepancy Report</h1>
  <div class="subtitle">Generated {now} &bull; Encinitas, Carlsbad, Solana Beach</div>
</header>

<div class="billing-callout">
  <strong>Billing Impact:</strong> Billing is based on when an address becomes serviceable.
  PowerBI assigns a single FDH Activation Date to all addresses under an FDH, but the Build
  shows addresses actually became serviceable on different dates. This report is organized
  by <strong>Build Phase</strong> to show which FDHs and dates PowerBI assigned to each phase,
  and where those dates don't match.
</div>

<!-- Per-city summary table -->
<div class="summary-table-wrap">
  <h3>Summary by City</h3>
  <p style="font-size:0.82em; color:var(--gray); margin-bottom:10px;">
    PBI records are <strong>serviceable addresses from the Build</strong> only.
    Circuit ID Match = PBI circuits found in Invoice Support data.
  </p>
  <table class="summary-table">
    <thead>
      <tr>
        <th>City</th>
        <th>PBI Records</th>
        <th>Serviceable</th>
        <th>Future Serviceable</th>
        <th>Circuit ID Match</th>
        <th>Not in Invoice</th>
        <th>Phases</th>
        <th>Phases w/ Issues</th>
        <th>Affected Addrs</th>
        <th>No Phase</th>
      </tr>
    </thead>
    <tbody>
"""

    for r in city_rows:
        html += f"""
      <tr>
        <td>{r['city']}</td>
        <td>{r['pbi']:,}</td>
        <td>{r['serviceable']:,}</td>
        <td>{r['future_serviceable']:,}</td>
        <td>{r['matched']:,}</td>
        <td>{r['unmatched']:,}</td>
        <td>{r['phases']}</td>
        <td>{r['issues']}</td>
        <td>{r['issue_addrs']:,}</td>
        <td>{r['no_phase']:,}</td>
      </tr>
"""

    total_svc = sum(r['serviceable'] for r in city_rows)
    total_future = sum(r['future_serviceable'] for r in city_rows)
    html += f"""
      <tr>
        <td>Total</td>
        <td>{total_pbi:,}</td>
        <td>{total_svc:,}</td>
        <td>{total_future:,}</td>
        <td>{total_matched:,}</td>
        <td>{total_unmatched:,}</td>
        <td>{sum(r['phases'] for r in city_rows)}</td>
        <td>{total_phases_issues}</td>
        <td>{total_issue_addrs:,}</td>
        <td>{total_no_phase:,}</td>
      </tr>
    </tbody>
  </table>
</div>

<!-- City tabs -->
<div class="city-tabs">
{city_tab_buttons}</div>
"""

    # Generate each city's content
    for i, (city, data) in enumerate(all_data.items()):
        phases = data["phases"]
        no_phase_details = data["no_phase_details"]
        no_phase_count = data["no_phase_count"]
        unmatched_circuits = data["unmatched"]

        city_id = city.lower().replace(" ", "")
        active = " active" if i == 0 else ""

        html += f"""
<div id="city-{city_id}" class="city-content{active}">
"""

        # --- All Phases table ---
        issue_count = sum(1 for p in phases if p["has_date_issues"] or p["has_multi_fdhs"])
        html += f"""
  <div class="section-toggle open" onclick="toggleSection(this)">
    <span class="arrow-icon">&#9654;</span>
    All Phases
    <span class="section-count">{len(phases)} phases &bull; {issue_count} with issues &bull; {sum(p['total'] for p in phases):,} addresses</span>
  </div>
  <div class="section-body open">
    <div style="display:flex; align-items:center; gap:12px; margin-bottom:8px; flex-wrap:wrap;">
      <div class="legend" style="margin-bottom:0; flex:1;">
        <div class="legend-item"><div class="legend-swatch" style="background:var(--orange-light);"></div> Date split</div>
        <div class="legend-item"><div class="legend-swatch" style="background:var(--purple-light);"></div> Multi-FDH</div>
        <div class="legend-item"><div class="legend-swatch" style="background:var(--red-light);"></div> Both</div>
        <div class="legend-item"><span style="color:var(--red); font-weight:600;">Red date</span> = differs from PBI</div>
      </div>
      <button onclick="toggleAllDetails('{city_id}')" id="expand-btn-{city_id}"
        style="padding:5px 14px; font-size:0.78em; border:1px solid var(--blue); color:var(--blue); background:white; border-radius:5px; cursor:pointer; white-space:nowrap;">
        Expand All
      </button>
    </div>
    <table class="phase-table">
      <thead>
        <tr>
          <th style="width:30px;"></th>
          <th>Phase</th>
          <th>FDH Name</th>
          <th>PBI Date</th>
          <th>Build Date(s)</th>
          <th>Total</th>
          <th>Svc</th>
          <th>Future</th>
          <th>Flags</th>
        </tr>
      </thead>
      <tbody>
"""

        for p in phases:
            has_issues = p["has_date_issues"] or p["has_multi_fdhs"]
            has_detail = has_issues or len(p["fdh_details"]) > 1

            # Row background
            if p["has_multi_fdhs"] and p["has_date_issues"]:
                row_bg = "background:var(--red-light);"
            elif p["has_multi_fdhs"]:
                row_bg = "background:var(--purple-light);"
            elif p["has_date_issues"]:
                row_bg = "background:var(--orange-light);"
            else:
                row_bg = ""

            # Badges
            badges = ""
            if p["has_multi_fdhs"]:
                badges += f'<span class="badge badge-purple">{len(p["fdh_details"])} FDHs</span> '
            if p["has_date_issues"]:
                badges += '<span class="badge badge-warn">DATE SPLIT</span> '
            if not has_issues:
                badges = '<span class="badge badge-ok">OK</span>'

            # Summary FDH info for main row
            fdh0 = p["fdh_details"][0]
            fdh_name_display = fdh0["fdh_name"]
            if len(p["fdh_details"]) > 1:
                fdh_name_display += f' <span style="color:var(--gray);">+{len(p["fdh_details"])-1} more</span>'

            pbi_date_display = fdh0["pbi_date"]

            # Build dates summary
            all_dates = []
            for fdh in p["fdh_details"]:
                for d, cnt in fdh["build_dates"]:
                    all_dates.append((d, cnt))
            if len(all_dates) == 1:
                bd_display = all_dates[0][0]
                if bd_display != fdh0["pbi_date"] and bd_display != "(no date)":
                    bd_display = f'<span class="date-mismatch">{bd_display}</span>'
            else:
                bd_display = f'{len(all_dates)} dates'

            # Expand arrow
            expand = ""
            row_id = f"{city_id}-ph-{p['phase']}"
            if has_detail and (len(all_dates) > 1 or len(p["fdh_details"]) > 1):
                expand = f'<span class="expand-arrow" style="cursor:pointer;" onclick="toggleDetail(\'{row_id}\')">&#9654;</span>'
            else:
                expand = '<span style="width:10px; display:inline-block;"></span>'

            html += f"""
        <tr style="{row_bg}" class="phase-row">
          <td>{expand}</td>
          <td><strong>{p['phase']}</strong></td>
          <td>{fdh_name_display}</td>
          <td>{pbi_date_display}</td>
          <td>{bd_display}</td>
          <td>{p['total']:,}</td>
          <td>{p['serviceable']:,}</td>
          <td>{p['future_serviceable']:,}</td>
          <td>{badges}</td>
        </tr>
"""

            # Detail rows (hidden by default)
            if has_detail and (len(all_dates) > 1 or len(p["fdh_details"]) > 1):
                html += f'        <tr id="{row_id}" class="detail-row" style="display:none;">\n'
                html += '          <td></td>\n'
                html += '          <td colspan="8">\n'

                for fdh in p["fdh_details"]:
                    svc_str = f'{fdh["serviceable"]:,} svc'
                    if fdh["future_serviceable"]:
                        svc_str += f' / {fdh["future_serviceable"]:,} future'

                    html += f'            <div class="detail-fdh">\n'
                    html += f'              <div class="detail-fdh-header"><strong>{fdh["fdh_name"]}</strong> <span style="color:var(--blue);">PBI: {fdh["pbi_date"]}</span> &bull; {fdh["count"]:,} ({svc_str})</div>\n'

                    for d, cnt in fdh["build_dates"]:
                        mc = ""
                        if d != fdh["pbi_date"] and d != "(no date)":
                            mc = " date-mismatch"
                        html += f'              <div class="detail-date"><span class="{mc}">{d}</span><span class="detail-cnt">{cnt:,}</span></div>\n'

                    html += '            </div>\n'

                html += '          </td>\n        </tr>\n'

        html += """
      </tbody>
    </table>
  </div>
"""

        # --- Section 3: No Phase ---
        html += f"""
  <div class="section-toggle" onclick="toggleSection(this)">
    <span class="arrow-icon">&#9654;</span>
    No Phase Assigned
    <span class="section-count">{no_phase_count:,} addresses &bull; {len(no_phase_details)} FDHs</span>
  </div>
  <div class="section-body">
"""
        if no_phase_details:
            html += """
    <table>
      <thead><tr><th>FDH Name</th><th>PowerBI Date</th><th>Addresses</th><th>Serviceable</th><th>Future</th></tr></thead>
      <tbody>
"""
            for np in no_phase_details:
                html += f'        <tr><td>{np["fdh_name"]}</td><td>{np["pbi_date"]}</td><td>{np["count"]:,}</td><td>{np["serviceable"]:,}</td><td>{np["future_serviceable"]:,}</td></tr>\n'
            html += "      </tbody>\n    </table>\n"
        else:
            html += '    <p style="color:var(--green);">All addresses have a phase assigned.</p>\n'

        html += "  </div>\n"

        # --- Section 4: Unmatched ---
        html += f"""
  <div class="section-toggle" onclick="toggleSection(this)">
    <span class="arrow-icon">&#9654;</span>
    PBI Circuits Not in Invoice
    <span class="section-count">{len(unmatched_circuits):,} circuits</span>
  </div>
  <div class="section-body">
"""
        if unmatched_circuits:
            html += """
    <div class="scrollable">
    <table>
      <thead><tr><th>Circuit ID</th><th>Address</th><th>FDH Name</th><th>Status</th></tr></thead>
      <tbody>
"""
            for uc in unmatched_circuits:
                html += f'        <tr><td>{uc["circuit_id"]}</td><td>{uc["address"]}</td><td>{uc["fdh_name"]}</td><td>{uc["status"]}</td></tr>\n'
            html += "      </tbody>\n    </table>\n    </div>\n"
        else:
            html += '    <p style="color:var(--green);">All PBI circuits matched.</p>\n'

        html += """
  </div>

</div>
"""

    # JavaScript
    city_ids_js = str(city_ids).replace("'", '"')
    html += f"""
<script>
var cityIds = {city_ids_js};

function switchCity(cityId) {{
  cityIds.forEach(function(cid) {{
    var el = document.getElementById('city-' + cid);
    if (el) el.classList.remove('active');
  }});
  document.querySelectorAll('.city-tab').forEach(function(t) {{ t.classList.remove('active'); }});

  var target = document.getElementById('city-' + cityId);
  if (target) target.classList.add('active');

  var tabs = document.querySelectorAll('.city-tab');
  var idx = cityIds.indexOf(cityId);
  if (idx >= 0 && tabs[idx]) tabs[idx].classList.add('active');
}}

function toggleSection(el) {{
  el.classList.toggle('open');
  el.nextElementSibling.classList.toggle('open');
}}

function toggleDetail(rowId) {{
  var row = document.getElementById(rowId);
  if (!row) return;
  var visible = row.style.display !== 'none';
  row.style.display = visible ? 'none' : 'table-row';
  var prev = row.previousElementSibling;
  if (prev) {{
    var arrow = prev.querySelector('.expand-arrow');
    if (arrow) arrow.classList.toggle('open', !visible);
  }}
}}

function toggleAllDetails(cityId) {{
  var container = document.getElementById('city-' + cityId);
  if (!container) return;
  var btn = document.getElementById('expand-btn-' + cityId);
  var rows = container.querySelectorAll('.detail-row');
  var expanding = btn && btn.textContent.trim() === 'Expand All';

  rows.forEach(function(row) {{
    row.style.display = expanding ? 'table-row' : 'none';
    var prev = row.previousElementSibling;
    if (prev) {{
      var arrow = prev.querySelector('.expand-arrow');
      if (arrow) arrow.classList.toggle('open', expanding);
    }}
  }});

  if (btn) btn.textContent = expanding ? 'Collapse All' : 'Expand All';
}}
</script>

</div>
</body>
</html>
"""
    return html


def main():
    print("Reading data from all 3 cities...")
    all_data = OrderedDict()

    for city, path in FILES.items():
        print(f"\n  {city}: {path.name}")

        metrics, unmatched = read_match_report(path)
        print(f"    Match Rate: {metrics.get('Match Rate', 'N/A')} | Unmatched: {len(unmatched):,}")

        records = read_pbi_data(path)
        print(f"    PowerBI records: {len(records):,}")

        phases, no_phase_details, no_phase_count = analyze_by_phase(records)
        issue_count = sum(1 for p in phases if p["has_date_issues"] or p["has_multi_fdhs"])
        print(f"    Phases: {len(phases)} | With issues: {issue_count} | No phase: {no_phase_count:,}")

        all_data[city] = {
            "metrics": metrics,
            "unmatched": unmatched,
            "phases": phases,
            "no_phase_details": no_phase_details,
            "no_phase_count": no_phase_count,
        }

    print("\nGenerating HTML report...")
    html = generate_html(all_data)

    out_path = BASE / "Match_Summary.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"\nSaved to: {out_path}")
    print(f"File size: {out_path.stat().st_size / 1024:.0f} KB")


if __name__ == "__main__":
    main()
