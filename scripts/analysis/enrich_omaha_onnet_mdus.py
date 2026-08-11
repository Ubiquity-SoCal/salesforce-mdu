"""
Enrich Niraj's "OMAHA ON net - Future MDUs.xlsx" (232 rows) with Salesforce columns.

Ask (Niraj, 2026-07-09): the 227 On-Net Omaha MDUs where Vetro HP CX is complete and the
MDU is not active. Add: In Salesforce (YES/NO), Property name from SF, Category from SF,
Opportunity State (= StageName), Owner. Loss Reason added per Koa - ~70% of these land in
Closed Lost and the reason is the only thing that makes the row actionable.

Scope per Koa: only the 227 rows where `SFU CX Completed` == "Yes". The other 5 rows
(4x #N/A, 1x "No") are dropped from the deliverable and listed on the console.

Matching reuses the ladder proven on the 65-row list (2026-07-08):
  agree-name exact -> address (house# + shared street tokens) -> near-address -> name-fuzzy.

Read-only against Salesforce. Output: xlsx deliverable + CSV audit.
"""
import sys
import csv
import re
import html
from pathlib import Path
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz
from simple_salesforce import Salesforce

sys.path.insert(0, str(Path(__file__).parent))
from lookup_agree_names_for_unlinked import house, st_tokens, norm_name, numset  # noqa: E402

# Both sides carry the city/state in the address string ("4760 LAFAYETTE AVE Omaha, NE 68132"),
# so `omaha` is a free shared token. Left in, it silently halves the ">=2 shared street tokens"
# gate and merges neighbouring buildings on dense blocks (4750 vs 4760 Lafayette).
PLACE_TOKENS = {"omaha", "ne", "nebraska", "usa", "us"}
HOUSE_WINDOW = 12  # blocks here are one-MDU-per-address; 25 spans several buildings

API = Path(r"C:\Users\cass\Work_Projects\SalesForce\api")
DESKTOP = Path(r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop")
SRC = DESKTOP / "OMAHA ON net - Future MDUs.xlsx"
OUT_DIR = Path(r"C:\Users\cass\Work_Projects\SalesForce\data\output")
OUT_XLSX = OUT_DIR / "omaha-onnet-future-mdus-with-salesforce.xlsx"
OUT_CSV = OUT_DIR / "2026-07-09-omaha-onnet-mdus-sf-enrichment.csv"

SOFT_REASONS = {"No Contact Info", "No Decision / Non-Responsive", "Other", "None", None}
SRC_COLS = ["Site Name", "Site Type", "Site Status", "Monday.com name", "MDU Site Category",
            "Parent Site", "P# from FB", "Total Units", "Full Address", "SFU CX Completed"]
NEW_COLS = ["In Salesforce", "Status", "Property Name (SF)", "Category (SF)",
            "Opportunity State", "Loss Reason", "Owner", "Match Note", "Latest Note"]
# "Status" says in words what the row colour says in colour. Colour alone is a bad carrier:
# it does not survive printing, filtering, or copy-paste, and it forces a legend lookup.
# "Match Note" is blank for the ~97% of rows matched on an exact agree-name or address.
# It only fires where a human should look: an inexact match, or several Opps on one property.

# The 8 Opportunities created 2026-07-09 by scripts/fix/2026-07-09-load-missing-omaha-onnet-mdus.py.
# Without this they render identically to the 69 that were already in the pipeline, so the
# reader cannot tell what this exercise actually changed.
LOADED_2026_07_09 = {
    "Omaha_MDU_4750 LAFAYETTE AVE", "Omaha_MDU_Chalet Apartments", "Omaha_MDU_5016 California St",
    "Omaha_MDU_4314 N 65th St", "Omaha_MDU_814 N 50th Ave", "Omaha_MDU_914 Mercer Blvd",
    "Omaha_MDU_Maple Villa Condominium", "Omaha_MDU_5711 N 24th St 6Plex",
}
# Two colours only, per Koa: Active vs Not Active. The soft/hard closed-lost split is still
# in the data (Loss Reason column) and still drives the console summary; it just does not
# get its own colour, because the reader's first question is "is anyone working this or not".
STATUS_TEXT = {
    "NEW":    "Active - added to Salesforce 7/9 (new Prospect)",
    "ACTIVE": "Active in pipeline",
    "SOFT":   "Not active - Closed Lost (worth re-approaching)",
    "HARD":   "Not active - Closed Lost",
    "MISSING": "NOT in Salesforce - needs loading",
}
COLOR_OF = {"NEW": "ACTIVE", "ACTIVE": "ACTIVE", "SOFT": "NOT_ACTIVE",
            "HARD": "NOT_ACTIVE", "MISSING": "MISSING"}
FILLS = {"ACTIVE": "C6EFCE", "NOT_ACTIVE": "E7E6E6", "MISSING": "FFC7CE"}


def creds():
    kv = {}
    for line in (API / "Salesforce_Credentials.txt").read_text().splitlines():
        if ":" in line:
            k, v = line.split(":", 1)
            kv[k.strip().lower()] = v.strip()
    return kv["username"], kv["password"], kv["security token"]


def load_source():
    """Return the 227 CX-complete rows, plus the excluded ones for reporting."""
    ws = openpyxl.load_workbook(SRC, read_only=True, data_only=True).active
    rows = [r for r in list(ws.iter_rows(values_only=True))[1:] if any(c not in (None, "") for c in r)]
    keep, drop = [], []
    for r in rows:
        rec = dict(zip(SRC_COLS, r))
        rec["Total Units"] = int(rec["Total Units"] or 0)
        rec["Site Name"] = str(rec["Site Name"] or "").strip()
        rec["Full Address"] = str(rec["Full Address"] or "").strip()
        (keep if rec["SFU CX Completed"] == "Yes" else drop).append(rec)
    return keep, drop


def fetch_opps(sf):
    q = """SELECT Id, Name, Agreement_Name__c, Property_Address__c, Property_City__c,
           Property_State__c, Property_Category__c, StageName, RecordType.Name,
           Loss_Reason__c, CloseDate, Owner.Name
           FROM Opportunity
           WHERE Name LIKE 'Omaha%' OR Agreement_Name__c LIKE 'Omaha%'
              OR Property_City__c LIKE 'Omaha%'
              OR (Property_State__c IN ('NE','Nebraska') AND RecordType.Name='MDU/SFU')"""
    return sf.query_all(q)["records"]


# Notes on these Opps come from two places and must be dated differently:
#  1. Bulk-migrated notes (loaded 2026-03-24 / 03-31 as "Cass Parker"). The real authored
#     date is embedded in the body as "Jeff Chao | 2024-12-11 18:00:52"; CreatedDate is only
#     when the migration ran, so it must NOT be used as the note's date.
#  2. Notes written natively in Salesforce since (Melissa Baker, Bill Holick, Taylor Mauney,
#     titles like "6/25 Update:"). No header, and CreatedDate IS the authored date.
# The Opportunity field Latest_Note_Snippet__c is NOT usable here: it was computed at
# migration time, so it structurally cannot see any note written after 2026-03-31.
NOTE_HEADER = re.compile(r"^\s*(.{2,40}?)\s*\|\s*(\d{4}-\d{2}-\d{2})[ T](\d{2}:\d{2}:\d{2})\s*")
MIGRATION_DAYS = {"2026-03-24", "2026-03-31"}
# On a migrated note the CreatedBy is whoever ran the load (Cass Parker), not the person who
# wrote it. For these the real author is in the title, e.g. "Rosemarie Shortino - ROE Outreach
# Notes". Crediting the migrator would misattribute 82 rows of somebody else's outreach.
TITLE_AUTHOR = re.compile(r"^\s*([A-Z][a-z]+(?: [A-Z][a-z'\-]+)+)\s+-\s+.*Notes?\s*$")
TAG_RE = re.compile(r"<[^>]+>")
WS_RE = re.compile(r"\s+")


def fetch_latest_notes(sf, opp_ids):
    """opp_id -> 'YYYY-MM-DD - Author - text'. Latest by best-known authored date."""
    best = {}
    for i in range(0, len(opp_ids), 150):
        chunk = "','".join(opp_ids[i:i + 150])
        for link in sf.query_all(
                f"SELECT LinkedEntityId, ContentDocument.FileType, ContentDocument.Title, "
                f"ContentDocument.LatestPublishedVersionId, "
                f"ContentDocument.LatestPublishedVersion.TextPreview, "
                f"ContentDocument.LatestPublishedVersion.CreatedDate, "
                f"ContentDocument.LatestPublishedVersion.CreatedBy.Name "
                f"FROM ContentDocumentLink WHERE LinkedEntityId IN ('{chunk}')")["records"]:
            doc = link["ContentDocument"]
            if doc["FileType"] != "SNOTE":
                continue
            v = doc["LatestPublishedVersion"]
            preview = v["TextPreview"] or ""
            created, creator = v["CreatedDate"][:10], v["CreatedBy"]["Name"]

            m = NOTE_HEADER.match(preview)
            if m:
                author, date, approx = m.group(1), m.group(2), False
                body = preview[m.end():]
            else:
                author, date, body = creator, created, preview
                # migrated-but-undated (e.g. the 85 "ROE Outreach Notes"): the date we have
                # is when it was attached, not when the outreach happened. Say so.
                approx = created in MIGRATION_DAYS and creator == "Cass Parker"
                if approx:
                    t = TITLE_AUTHOR.match(doc["Title"] or "")
                    if t:
                        author = t.group(1)

            key = link["LinkedEntityId"]
            if key not in best or date > best[key][0]:
                best[key] = (date, author, body, doc["Title"], approx,
                             doc["LatestPublishedVersionId"], len(preview))

    # only the winning notes, and only where the 255-char preview was actually truncated
    out = {}
    for oid, (date, author, body, title, approx, vid, plen) in best.items():
        if plen >= 250:
            try:
                raw = sf.session.get(
                    f"{sf.base_url}sobjects/ContentVersion/{vid}/VersionData",
                    headers=sf.headers).content.decode("utf-8", errors="replace")
                full = html.unescape(TAG_RE.sub(" ", raw))
                m = NOTE_HEADER.match(full)
                body = full[m.end():] if m else full
            except Exception:
                pass  # keep the preview
        body = WS_RE.sub(" ", body).strip()
        if not body:
            body = f"(note: {title})"
        tail = "  [migrated note, date approximate]" if approx else ""
        out[oid] = f"{date} - {author} - {body}{tail}"
    return out


class Matcher:
    def __init__(self, opps):
        self.opps = opps
        self.by_agree = defaultdict(list)
        self.by_house = defaultdict(list)
        for o in opps:
            agn = (o["Agreement_Name__c"] or "").strip().lower()
            if agn:
                self.by_agree[agn].append(o)
            h = house(o["Property_Address__c"])
            if h:
                self.by_house[h].append(o)

    @staticmethod
    def _street(a):
        """Street-name tokens only: no bare house numbers, no city/state."""
        return {t for t in st_tokens(a) if not t.isdigit()} - PLACE_TOKENS

    def _exact(self, site, addr):
        """Same house number AND at least one shared street-name token."""
        hits, basis = {}, set()
        for o in self.by_agree.get(site.lower(), []):
            hits[o["Id"]] = o
            basis.add("agree-name")
        gt = self._street(addr)
        for o in self.by_house.get(house(addr), []):
            if gt & self._street(o["Property_Address__c"]):
                hits[o["Id"]] = o
                basis.add("address")
        return list(hits.values()), basis

    def _near(self, addr):
        """Same distinctive street + house# within HOUSE_WINDOW.

        Requires >=2 shared street-name tokens, so it fires on a real multi-word street
        ("Indian Hills Dr" -> {indian, hills}, catching 8509 vs 8515) but never on a
        single-token street ({lafayette}, {65th}) where each address is its own building.
        """
        h, ga = house(addr), self._street(addr)
        if not h or len(ga) < 2:
            return None
        hi, out = int(h), []
        for o in self.opps:
            oh = house(o["Property_Address__c"])
            if (oh and len(ga & self._street(o["Property_Address__c"])) >= 2
                    and abs(int(oh) - hi) <= HOUSE_WINDOW):
                out.append((abs(int(oh) - hi), o))
        if not out:
            return None
        out.sort(key=lambda x: x[0])
        best = out[0][1]
        return [o for _, o in out], f"{house(best['Property_Address__c'])} vs {h}"

    def _addr_compatible(self, addr, o):
        """Guard for name matches: the two addresses must not actively contradict.
        'Chalet Apartments' @ 4728 Seward vs SF 'Chalet Apartments' @ 4858 Izard is a
        name collision between two different buildings, not a spelling variant."""
        oa = o["Property_Address__c"]
        if self._street(addr) & self._street(oa):
            return True
        h, oh = house(addr), house(oa)
        return bool(h and oh and abs(int(h) - int(oh)) <= HOUSE_WINDOW)

    def _fuzzy(self, site, addr):
        """Name match, but digits must agree (4313 != 4314) and the address must not contradict."""
        q, qn, best = norm_name(site), numset(norm_name(site)), None
        for o in self.opps:
            cand = norm_name(o["Agreement_Name__c"] or o["Name"])
            if not cand or numset(cand) != qn or not self._addr_compatible(addr, o):
                continue
            score = min(fuzz.token_set_ratio(q, cand), fuzz.ratio(q, cand))
            if best is None or score > best[0]:
                best = (score, o)
        return best if (best and best[0] >= 90) else None

    def match(self, site, addr):
        hits, basis = self._exact(site, addr)
        how = "+".join(sorted(basis)) if basis else ""
        if not hits:
            n = self._near(addr)
            if n:
                hits, how = n[0], f"address (near: {n[1]})"
        if not hits:
            f = self._fuzzy(site, addr)
            if f:
                hits, how = [f[1]], f"name-fuzzy({f[0]:.0f})"
        return hits, how


def pick(hits):
    """Representative Opp: an active one if any exists, else the closed-lost."""
    return sorted(hits, key=lambda o: (o["StageName"] == "Closed Lost",))[0]


def main():
    sf = Salesforce(*creds())
    keep, drop = load_source()
    print(f"source rows: {len(keep) + len(drop)}  ->  CX-complete (in scope): {len(keep)}")
    if drop:
        print(f"excluded ({len(drop)}, SFU CX Completed != 'Yes'):")
        for r in drop:
            print(f"   {str(r['SFU CX Completed']):6} {r['Site Name'][:44]:44} u={r['Total Units']:<4} {r['Full Address']}")

    opps = fetch_opps(sf)
    print(f"\nSF Omaha-universe Opportunities: {len(opps)}")
    m = Matcher(opps)

    out = []
    for r in keep:
        hits, how = m.match(r["Site Name"], r["Full Address"])
        active = [o for o in hits if o["StageName"] != "Closed Lost"]
        rep = pick(active) if active else (pick(hits) if hits else None)
        notes = []
        if "near" in how:
            notes.append(f"Address off by a few ({how.split('near: ')[1].rstrip(')')}) - verify same property")
        if "fuzzy" in how:
            notes.append("Matched on property name, not address - verify same property")
        if len(hits) > 1:
            notes.append(f"{len(hits)} SF Opps on this property; showing the "
                         f"{'active' if active else 'closed-lost'} one")

        row = dict(r)
        row.update({
            "In Salesforce": "NO" if rep is None else "YES",
            "Property Name (SF)": rep["Name"] if rep else "",
            "Category (SF)": (rep["Property_Category__c"] or "") if rep else "",
            "Opportunity State": rep["StageName"] if rep else "",
            "Loss Reason": (rep["Loss_Reason__c"] or "") if rep and rep["StageName"] == "Closed Lost" else "",
            "Owner": rep["Owner"]["Name"] if rep and rep.get("Owner") else "",
            "Status": "",       # filled below, once the display key is known
            "Latest Note": "",  # filled below, keyed on the resolved Opp id
            "Match Note": "; ".join(notes),
            "_bucket": ("MISSING" if rep is None else ("ACTIVE" if active else "CLOSED-LOST")),
            "_soft": bool(rep and not active and rep["Loss_Reason__c"] in SOFT_REASONS),
            "_match": how,
            "_n_opps": len(hits),
            "_sf_id": rep["Id"] if rep else "",
        })
        out.append(row)

    # ---- duplicate-address post-pass ----
    # Three buildings appear on two rows each: Vetro emits one row per fiber feed (FB), so
    # "<name>-2" is a second feed into the same building, not a second property. If one row
    # matched and its twin did not, the twin is not "missing" - it is the same building, and
    # loading it would create a duplicate Opportunity.
    import re as _re
    by_addr = defaultdict(list)
    for r in out:
        by_addr[_re.sub(r"[^a-z0-9]", "", r["Full Address"].lower())].append(r)
    inherited = 0
    for group in by_addr.values():
        if len(group) < 2:
            continue
        donor = next((g for g in group if g["_bucket"] != "MISSING"), None)
        twins = ", ".join(sorted({g["Site Name"] for g in group}))
        for r in group:
            if r["_bucket"] == "MISSING" and donor:
                for f in ("In Salesforce", "Property Name (SF)", "Category (SF)",
                          "Opportunity State", "Loss Reason", "Owner"):
                    r[f] = donor[f]
                r["_bucket"], r["_soft"], r["_sf_id"] = donor["_bucket"], donor["_soft"], donor["_sf_id"]
                r["_match"] = "duplicate-address twin"
                inherited += 1
            note = (f"Same building as its twin row (second fiber feed). "
                    f"Rows: {twins}. Do not load twice or double-count units.")
            r["Match Note"] = f"{r['Match Note']}; {note}" if r["Match Note"] else note
    print(f"\n{len(by_addr)} distinct addresses across {len(out)} rows; "
          f"{inherited} twin row(s) inherited a sibling's Salesforce match")

    # ---- display key: one value driving BOTH the Status text and the row colour ----
    # Runs after the twin post-pass so an inherited row gets its donor's key, not "MISSING".
    for r in out:
        if r["_bucket"] == "MISSING":
            key = "MISSING"
        elif r["Site Name"] in LOADED_2026_07_09:
            key = "NEW"
        elif r["_bucket"] == "ACTIVE":
            key = "ACTIVE"
        else:
            key = "SOFT" if r["_soft"] else "HARD"
        r["_key"] = key
        r["Status"] = STATUS_TEXT[key]

    # ---- latest note per Opportunity (runs after the twin pass, so twins share a note) ----
    sf_ids = sorted({r["_sf_id"] for r in out if r["_sf_id"]})
    latest = fetch_latest_notes(sf, sf_ids)
    for r in out:
        r["Latest Note"] = latest.get(r["_sf_id"], "")
    have = sum(1 for r in out if r["Latest Note"])
    print(f"\nlatest note resolved for {have} of {len(out)} rows "
          f"({len(latest)} of {len(sf_ids)} distinct Opps have any note)")

    # ---- summary (parts must sum to whole) ----
    b = Counter(r["_bucket"] for r in out)
    print("\n=== SUMMARY (227 in scope) ===")
    for k in ("ACTIVE", "CLOSED-LOST", "MISSING"):
        u = sum(r["Total Units"] for r in out if r["_bucket"] == k)
        print(f"  {k:14} {b[k]:3}  ({u:,} units)")
    assert sum(b.values()) == len(out) == len(keep), "bucket counts must sum to scope"
    print(f"  {'TOTAL':14} {sum(b.values()):3}  ({sum(r['Total Units'] for r in out):,} units)")
    dedup = {_re.sub(r"[^a-z0-9]", "", r["Full Address"].lower()): r for r in out}
    print(f"  {'distinct bldgs':14} {len(dedup):3}  ({sum(r['Total Units'] for r in dedup.values()):,} units)")

    print(f"\n=== NOT in Salesforce ({b['MISSING']}) - candidates to load ===")
    for r in sorted((r for r in out if r["_bucket"] == "MISSING"), key=lambda r: -r["Total Units"]):
        print(f"  {r['Site Name'][:44]:44} u={r['Total Units']:<4} {r['Full Address']}")

    cl = [r for r in out if r["_bucket"] == "CLOSED-LOST"]
    soft = [r for r in cl if r["_soft"]]
    print(f"\n=== Closed Lost ({len(cl)}) loss reasons ===")
    for reason, n in Counter(r["Loss Reason"] or "(blank)" for r in cl).most_common():
        tag = "  <- soft, re-approachable" if reason.strip("()blank ") in {""} or reason in SOFT_REASONS else ""
        print(f"  {reason:34} {n}{tag}")
    print(f"  SOFT: {len(soft)} MDUs / {sum(r['Total Units'] for r in soft):,} units"
          f"   |   HARD: {len(cl) - len(soft)} / {sum(r['Total Units'] for r in cl) - sum(r['Total Units'] for r in soft):,} units")
    assert len(soft) + (len(cl) - len(soft)) == len(cl)

    print(f"\n=== Owners on matched rows ===")
    for own, n in Counter(r["Owner"] for r in out if r["Owner"]).most_common():
        print(f"  {own:28} {n}")
    print(f"\n=== Category (SF) ===")
    for c, n in Counter(r["Category (SF)"] or "(blank)" for r in out).most_common():
        print(f"  {c:14} {n}")

    weak = [r for r in out if "near" in r["_match"] or "fuzzy" in r["_match"]]
    if weak:
        print(f"\n=== {len(weak)} rows matched on a weak basis - verify these ===")
        for r in weak:
            print(f"  {r['Site Name'][:38]:38} [{r['_match']}] -> {r['Property Name (SF)'][:32]}")
    multi = [r for r in out if r["_n_opps"] > 1]
    print(f"\n{len(multi)} rows have >1 SF Opp on the property (representative = active if one exists)")

    # ---- CSV audit ----
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(out[0].keys()))
        w.writeheader()
        w.writerows(out)
    print(f"\nwrote {OUT_CSV}")

    # ---- deliverable: Niraj's sheet + 6 appended columns ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OMAHA ON net - Future MDUs"
    cols = SRC_COLS + NEW_COLS
    hfont = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1F4E78")
    nfill = PatternFill("solid", fgColor="2E75B6")  # new columns get a lighter header
    thin = Side(style="thin", color="D9D9D9")
    border = Border(thin, thin, thin, thin)
    widths = [40, 9, 10, 26, 14, 22, 12, 10, 40, 15, 13, 40, 34, 12, 22, 28, 20, 46, 100]
    for i, (name, w_) in enumerate(zip(cols, widths), 1):
        c = ws.cell(1, i, name)
        c.font = hfont
        c.fill = nfill if name in NEW_COLS else hfill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w_

    # Not-active first: those are the rows that need a decision. Within them, the
    # re-approachable (soft) ones lead, then by door count.
    prio = {"MISSING": 0, "NOT_ACTIVE": 1, "ACTIVE": 2}
    rows_sorted = sorted(out, key=lambda r: (prio[COLOR_OF[r["_key"]]],
                                             0 if r["_key"] == "SOFT" else 1,
                                             -r["Total Units"]))
    for i, r in enumerate(rows_sorted, 2):
        fill = PatternFill("solid", fgColor=FILLS[COLOR_OF[r["_key"]]])
        for j, name in enumerate(cols, 1):
            c = ws.cell(i, j, r[name])
            c.fill = fill
            c.border = border
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(name in ("Full Address", "Match Note", "Status", "Latest Note")))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(cols)).column_letter}{len(rows_sorted) + 1}"

    # legend: only describes colours that actually occur, with live counts, so it can never
    # drift out of sync with the sheet the way the old 4-colour legend did.
    present = Counter(COLOR_OF[r["_key"]] for r in out)
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 16
    lg.column_dimensions["B"].width = 10
    lg.column_dimensions["C"].width = 96
    for j, h in enumerate(("Row color", "Rows", "Meaning"), 1):
        lg.cell(1, j, h).font = hfont
        lg.cell(1, j).fill = hfill
    legend = [
        ("ACTIVE", "Green", "Active Opportunity in the Salesforce pipeline. Includes the 8 MDUs added "
                            "on 7/9 - see the Status column to tell those apart."),
        ("NOT_ACTIVE", "Grey", "Not active: the Opportunity is Closed Lost. See the Loss Reason column. "
                               "Reasons like No Contact Info / Non-Responsive / Other are worth "
                               "re-approaching now that fiber is on-net; Not Interested / Existing Fiber "
                               "are not changed by fiber."),
        ("MISSING", "Red", "Not in Salesforce at all - needs loading."),
    ]
    row = 2
    for key, label, meaning in legend:
        if not present.get(key):
            continue  # do not describe a colour that appears nowhere
        lg.cell(row, 1, label).fill = PatternFill("solid", fgColor=FILLS[key])
        lg.cell(row, 1).border = border
        lg.cell(row, 2, present[key])
        lg.cell(row, 3, meaning).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    lg.cell(row + 1, 1, "Status column").font = Font(bold=True)
    lg.cell(row + 1, 3, "Says the same thing in words, so you never need this legend: "
                        + " | ".join(sorted({r["Status"] for r in out})))
    lg.cell(row + 1, 3).alignment = Alignment(wrap_text=True, vertical="top")
    lg.cell(row + 3, 1, "Scope").font = Font(bold=True)
    lg.cell(row + 3, 3, f"227 On Net MDUs where SFU CX Completed = Yes. {len(drop)} rows from the source "
                        f"file were excluded (CX not complete). Note 3 addresses appear on two rows each "
                        f"(360 Skyview, Indian Hills Village, 7805 Harney St), so 227 rows = 224 distinct "
                        f"properties.")
    lg.cell(row + 3, 3).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
