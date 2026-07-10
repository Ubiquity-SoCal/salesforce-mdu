"""
Contact hygiene review workbook for Rosemary.

Koa (2026-07-09) chose to FLAG these rather than auto-delete, so this script is read-only
against Salesforce. It writes one workbook with the three problems that corrupt the
"Primary Contact and Contact Count" columns on Niraj's report:

  1. Duplicate links  - same Contact linked to the same Opportunity more than once.
  2. Orphan links     - Opportunity_Contact__c rows with no Contact at all.
  3. Junk primaries   - the chosen Primary Contact is a company name, not a person
                        (e.g. "HARVEST DEVELOPMENT LLC"), or has no Account set.

Re-runnable. Output: data/output/contact-hygiene-review-<date>.xlsx
"""
import sys
import io
import re
from pathlib import Path
from datetime import date
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts" / "analysis"))

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from simple_salesforce import Salesforce  # noqa: E402
from enrich_omaha_onnet_mdus import creds  # noqa: E402

OUT = ROOT / "data" / "output" / f"contact-hygiene-review-{date.today():%Y-%m-%d}.xlsx"

COMPANY = re.compile(
    r"\b(LLC|L\.L\.C|INC|LP|LTD|CORP|PROPERTIES|PROPERTY|MANAGEMENT|MGMT|HOLDINGS|"
    r"DEVELOPMENT|REALTY|REAL ESTATE|CAPITAL|PARTNERS|GROUP|ENTERPRISES|INVESTMENTS|"
    r"APARTMENTS|TRUST|COMPANY|ASSOCIATES|RENTALS|INVESTMENT)\b",
    re.I,
)

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")

sf = Salesforce(*creds())

links = sf.query_all(
    "SELECT Id, CreatedDate, Opportunity__c, Opportunity__r.Name, "
    "Opportunity__r.Property_State__c, Opportunity__r.StageName, "
    "Contact__c, Contact__r.Name, Contact__r.AccountId, Contact__r.Account.Name, "
    "Contact__r.Phone, Contact__r.Email, Role__c "
    "FROM Opportunity_Contact__c WHERE Opportunity__c != null "
    "ORDER BY Opportunity__c, CreatedDate"
)["records"]
print(f"junction rows: {len(links)}")


def g(rec, path, default=""):
    cur = rec
    for p in path.split("."):
        cur = (cur or {}).get(p)
    return cur if cur is not None else default


# ---- 1. duplicate (opp, contact) pairs -----------------------------------
pairs = defaultdict(list)
for r in links:
    if r["Contact__c"]:
        pairs[(r["Opportunity__c"], r["Contact__c"])].append(r)
dupes = {k: v for k, v in pairs.items() if len(v) > 1}
dup_rows = []
for (oid, cid), rows in dupes.items():
    keep = rows[0]  # oldest
    for r in rows[1:]:
        dup_rows.append([
            g(r, "Opportunity__r.Name"), g(r, "Opportunity__r.Property_State__c"),
            g(r, "Opportunity__r.StageName"), g(r, "Contact__r.Name"),
            r.get("Role__c", ""), g(r, "Contact__r.Phone"),
            r["Id"], keep["Id"], r["CreatedDate"][:10],
        ])
print(f"duplicate pairs: {len(dupes)}  -> redundant rows to remove: {len(dup_rows)}")

# ---- 2. orphan links ------------------------------------------------------
orphans = [r for r in links if not r["Contact__c"]]
orphan_rows = [[
    g(r, "Opportunity__r.Name"), g(r, "Opportunity__r.Property_State__c"),
    g(r, "Opportunity__r.StageName"), r.get("Role__c", ""), r["Id"], r["CreatedDate"][:10],
] for r in orphans]
print(f"orphan links (no Contact): {len(orphans)}")

# ---- 3. junk primaries ----------------------------------------------------
opps = sf.query_all(
    "SELECT Id, Name, Property_State__c, StageName, Contact_Count__c, "
    "Primary_Contact__c, Primary_Contact__r.Name, Primary_Contact__r.AccountId, "
    "Primary_Contact__r.Account.Name, Primary_Contact_Role__c "
    "FROM Opportunity WHERE Primary_Contact__c != null"
)["records"]
junk_rows = []
for o in opps:
    name = g(o, "Primary_Contact__r.Name")
    looks_company = bool(COMPANY.search(name or ""))
    no_account = not g(o, "Primary_Contact__r.AccountId", None)
    if not (looks_company or no_account):
        continue
    issues = []
    if looks_company:
        issues.append("company name in person field")
    if no_account:
        issues.append("no property mgmt company set")
    junk_rows.append([
        o["Name"], o.get("Property_State__c", ""), o.get("StageName", ""),
        name, o.get("Primary_Contact_Role__c", ""),
        g(o, "Primary_Contact__r.Account.Name"),
        int(o.get("Contact_Count__c") or 0), "; ".join(issues), o["Id"],
    ])
print(f"opps with a primary contact: {len(opps)}  -> flagged primaries: {len(junk_rows)}")

# ---- write ----------------------------------------------------------------
wb = openpyxl.Workbook()

SHEETS = [
    ("Duplicate Links", dup_rows,
     ["Opportunity", "State", "Stage", "Contact", "Role", "Phone",
      "Duplicate Link Id (delete)", "Keep This Link Id", "Created"],
     "Same contact linked to the same opportunity more than once. "
     "Delete the 'Duplicate Link Id' rows; the 'Keep' row is the oldest."),
    ("Orphan Links", orphan_rows,
     ["Opportunity", "State", "Stage", "Role", "Link Id (delete)", "Created"],
     "Junction rows with a Role but no Contact attached. They render as a blank name."),
    ("Junk Primary Contacts", junk_rows,
     ["Opportunity", "State", "Stage", "Primary Contact", "Role",
      "Property Mgmt Company", "Contact Count", "Issue", "Opportunity Id"],
     "The contact now showing on Niraj's report is a company name, or has no management "
     "company set. Fix the Contact record, not the Opportunity."),
]

first = True
for title, rows, headers, note in SHEETS:
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    first = False
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=9)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 3):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A3"
    for c, h in enumerate(headers, 1):
        width = max([len(str(h))] + [len(str(r[c - 1])) for r in rows[:200]] or [10])
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 12), 46)
    print(f"  sheet '{title}': {len(rows)} rows")

wb.save(OUT)
print(f"\nwrote {OUT}")
print(f"TOTAL rows needing a human: {len(dup_rows) + len(orphan_rows) + len(junk_rows)}")
