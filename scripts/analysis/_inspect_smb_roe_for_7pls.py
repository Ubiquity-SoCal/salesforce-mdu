"""
Inspect SMB_ROE_Project.xlsx ROE_Tracking sheet for the 7 PLs we just flagged.
Pull Build Effort, RE Assigned, RE Notes for them.

Read-only.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
SHEET = 'ROE_Tracking'

# Addresses of the 7 PLs we just flagged (from SF Name field)
TARGETS = [
    "1179 SOUTHEAST PKWY AZLE TX",
    "1855 W BASELINE RD MESA AZ",
    "6237 N 89TH CIR OMAHA NE",
    "303 E SOUTHERN AVE MESA AZ",
    "320 E 10TH DR MESA AZ",
    "347 E SOUTHERN AVE MESA AZ",
    "5139 LEAVENWORTH ST OMAHA NE",
]

def norm(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s).strip().upper())

target_set = {norm(t) for t in TARGETS}

wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[SHEET]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
print("HEADER COLUMNS:")
for i, h in enumerate(header):
    print(f"  [{i}] {h!r}")
print()

# Try to find the relevant column names
def find_col(*candidates):
    for cand in candidates:
        cand_norm = cand.lower().replace(' ', '').replace('_', '')
        for i, h in enumerate(header):
            if h and cand_norm in str(h).lower().replace(' ', '').replace('_', ''):
                return i, h
    return None, None

idx_building, name_building = find_col('Business Buildings', 'Building')
idx_state,    name_state    = find_col('State')
idx_re,       name_re       = find_col('RE_Assigned', 'RE Assigned', 'REAssigned')
idx_status,   name_status   = find_col('RE Status', 'RE_Status')
idx_effort,   name_effort   = find_col('Build Effort', 'Build_Effort', 'BuildEffort', 'Effort')
idx_notes,    name_notes    = find_col('RE Notes', 'RE_Notes', 'Notes')

print(f"Resolved columns:")
print(f"  Building     [{idx_building}] {name_building!r}")
print(f"  State        [{idx_state}] {name_state!r}")
print(f"  RE Assigned  [{idx_re}] {name_re!r}")
print(f"  RE Status    [{idx_status}] {name_status!r}")
print(f"  Build Effort [{idx_effort}] {name_effort!r}")
print(f"  RE Notes     [{idx_notes}] {name_notes!r}")
print()

print("MATCHES:")
for r in rows[1:]:
    if not any(c is not None for c in r):
        continue
    bn = r[idx_building] if idx_building is not None else ''
    bn_norm = norm(bn)
    if bn_norm in target_set:
        print(f"\n--- {bn} ---")
        print(f"  RE Status:    {r[idx_status] if idx_status is not None else 'N/A'}")
        print(f"  RE Assigned:  {r[idx_re] if idx_re is not None else 'N/A'}")
        print(f"  Build Effort: {r[idx_effort] if idx_effort is not None else 'N/A'}")
        print(f"  RE Notes:     {r[idx_notes] if idx_notes is not None else 'N/A'}")
        target_set.discard(bn_norm)

if target_set:
    print(f"\nUNMATCHED — fuzzy searching for street numbers in any column:")
    # Extract street number prefixes from unmatched addresses
    target_numbers = {}
    for t in target_set:
        m = re.match(r'^(\d+)\s', t)
        if m:
            target_numbers[m.group(1)] = t
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue
        for col_idx in [0, 1, 17]:  # Building, Address, Updated Business Address
            if col_idx >= len(r):
                continue
            cell = r[col_idx]
            if not cell:
                continue
            cell_str = str(cell).upper()
            for num, t in target_numbers.items():
                if cell_str.startswith(num + ' ') or f' {num} ' in cell_str:
                    if num in t.split()[0]:
                        print(f"\n--- Possible match for {t} ---")
                        print(f"  Building (col 0): {r[0]}")
                        print(f"  Address  (col 1): {r[1]}")
                        print(f"  Updated  (col 17): {r[17]}")
                        print(f"  RE Status:    {r[idx_status] if idx_status is not None else 'N/A'}")
                        print(f"  RE Assigned:  {r[idx_re] if idx_re is not None else 'N/A'}")
                        print(f"  Build Effort: {r[idx_effort] if idx_effort is not None else 'N/A'}")
                        print(f"  RE Notes:     {r[idx_notes] if idx_notes is not None else 'N/A'}")
                        break
