"""
Investor PAL Pipeline Report — v3.

Sheet 1: Investor Pipeline (3 fixed-layout sections, always all 6 buckets shown)
Sheet 2: Action Items — data-hygiene punch-list for Taylor & team

Bucket logic (priority — most advanced wins; each opp appears in exactly ONE bucket):
  ON Air Serviceable      — ST Site_Status='In Service' OR Activation_Actual populated
  On Net - AAC            — (signed Agreement OR ST PAL_Signed_Date) + Property_Category='Cat 1'
  Near Net - AAC          — (signed Agreement OR ST PAL_Signed_Date) + Property_Category='Cat 2'
  Needs Classification    — signed Agreement + Cat 3 or null
  Proposal Sent           — Agreement Status IN ('Sign','Review')
  Prospects               — StageName='Contract Negotiations' OR Agreement Status='Create'

Column dedup: if opp qualifies for both ROE and PAL in the same bucket, count under PAL.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sf = Salesforce(username='cass1@ubiquitygp.com', password='Hawaiian1984', security_token='IBSKT6CFUpSUJWxq1CMm0HkFC')

BUCKETS_ADVANCED_FIRST = ['ON Air Serviceable','On Net - AAC','Near Net - AAC','Needs Classification','Proposal Sent','Prospects']
BUCKETS_DISPLAY = ['Prospects','Proposal Sent','On Net - AAC','Near Net - AAC','ON Air Serviceable','Needs Classification']
FRANCHISES = ['In-Franchise','National','Total']
AG_TYPES = ['ROE','PAL']


def bucket_for_type(opp, ag_type):
    cat = opp.get('Property_Category__c')
    ags = (opp.get('Agreements__r') or {}).get('records') or []
    type_ags = [a for a in ags if a.get('Agreement_Type__c') == ag_type]
    sts = (opp.get('SiteTracker_Projects__r') or {}).get('records') or []
    signed = any(a.get('Signed_Date__c') for a in type_ags)
    in_flight = any((a.get('Status__c') in ('Sign','Review')) and not a.get('Signed_Date__c') for a in type_ags)
    create = any(a.get('Status__c') == 'Create' for a in type_ags)
    st_in_service = any(s.get('Site_Status__c') == 'In Service' or s.get('Activation_Actual__c') for s in sts)
    st_pal_signed = any(s.get('PAL_Signed_Date__c') for s in sts)
    cat12 = cat in ('Cat 1','Cat 2')
    stage_neg = opp.get('StageName') == 'Contract Negotiations'

    if st_in_service and (signed or ag_type == 'PAL'):
        return 'ON Air Serviceable'
    agreement_effective = signed or (ag_type == 'PAL' and st_pal_signed)
    if agreement_effective and cat == 'Cat 1': return 'On Net - AAC'
    if agreement_effective and cat == 'Cat 2': return 'Near Net - AAC'
    if agreement_effective and not cat12:      return 'Needs Classification'
    if in_flight:                              return 'Proposal Sent'
    if stage_neg or create:                    return 'Prospects'
    return None


def classify(opp):
    pal_b = bucket_for_type(opp, 'PAL')
    roe_b = bucket_for_type(opp, 'ROE')
    candidates = []
    if pal_b: candidates.append(('PAL', pal_b))
    if roe_b: candidates.append(('ROE', roe_b))
    if not candidates:
        return None, None
    chosen = min(candidates, key=lambda c: BUCKETS_ADVANCED_FIRST.index(c[1]))
    col, bucket = chosen
    if pal_b == roe_b == bucket:
        col = 'PAL'
    return bucket, col


# ── Pull Opportunities (with agreements + ST) ──────────────────────────────
q = sf.query_all("""
  SELECT Id, Name, StageName, Property_Category__c, Franchise_Type__c, Units__c,
         Monday_Item_ID__c, Agreement_Name__c, In_SiteTracker__c,
         (SELECT Agreement_Type__c, Status__c, Signed_Date__c, IronClad_Record__c FROM Agreements__r),
         (SELECT Build_Status__c, Site_Status__c, PAL_Signed_Date__c, Activation_Actual__c FROM SiteTracker_Projects__r)
  FROM Opportunity
""")
opps = q['records']

# IronClad / Agreement — scope to ROE + PAL types only (per 2026-04-24 instruction:
# ignore FiberFirst Enterprise, Easement, Pole Attachment, Marketing agreements, etc.)
IC_ROE_PAL_TYPES = "('Right of Entry Agreement','Premises Access License')"
SF_ROE_PAL_TYPES = "('ROE','PAL')"

iron_no_ag_total = sf.query(f"SELECT Count() FROM IronClad__c WHERE Agreement__c = null AND Record_Type_IC__c IN {IC_ROE_PAL_TYPES}")['totalSize']
iron_no_ag = sf.query_all(f"""
  SELECT Id, Record_Name__c, Record_Type_IC__c, Counterparty_Name__c, Agree_Name__c, Agreement_Date__c
  FROM IronClad__c WHERE Agreement__c = null AND Record_Type_IC__c IN {IC_ROE_PAL_TYPES}
  ORDER BY CreatedDate DESC LIMIT 200
""")['records']

ag_no_ic_total = sf.query(f"SELECT Count() FROM Agreement__c WHERE Signed_Date__c != null AND IronClad_Record__c = null AND Agreement_Type__c IN {SF_ROE_PAL_TYPES}")['totalSize']
ag_no_ic = sf.query_all(f"""
  SELECT Id, Opportunity__c, Opportunity__r.Name, Agreement_Type__c, Status__c, Signed_Date__c
  FROM Agreement__c
  WHERE Signed_Date__c != null AND IronClad_Record__c = null AND Agreement_Type__c IN {SF_ROE_PAL_TYPES}
  ORDER BY Signed_Date__c DESC LIMIT 200
""")['records']

st_no_ag_total = sf.query("""
  SELECT Count() FROM SiteTracker_Project__c
  WHERE Opportunity__c != null
    AND Opportunity__c NOT IN (SELECT Opportunity__c FROM Agreement__c WHERE Signed_Date__c != null)
""")['totalSize']
st_no_ag = sf.query_all("""
  SELECT Id, Opportunity__c, Opportunity__r.Name, Build_Status__c
  FROM SiteTracker_Project__c
  WHERE Opportunity__c != null
    AND Opportunity__c NOT IN (SELECT Opportunity__c FROM Agreement__c WHERE Signed_Date__c != null)
  LIMIT 200
""")['records']


# ── Classify + aggregate ───────────────────────────────────────────────────
counts = {f: {b: {a: {'n': 0, 'u': 0.0} for a in AG_TYPES} for b in BUCKETS_DISPLAY} for f in FRANCHISES}
included, excl_cat3, excl_null = 0, 0, 0
action_items = {
    'A_needs_cat': [],            # signed agreement + Cat 3 or null
    'B_uc_no_ag': [],             # Under Contract + no signed agreement
    'C_roes_no_ag': [],           # ROE Secured + no signed ROE
    'D_ic_no_ag_link': [],        # IronClad__c without Agreement__c link
    'E_ag_no_ic': [],             # Signed Agreement__c without IronClad link
    'F_st_no_ag': [],             # ST project without signed Agreement on Opp
    'G_uc_pal_no_st': [],         # Under Contract + signed PAL + no ST project
}

for o in opps:
    cat = o.get('Property_Category__c')
    ags = (o.get('Agreements__r') or {}).get('records') or []
    has_signed = any(a.get('Signed_Date__c') for a in ags)
    has_signed_roe = any(a.get('Agreement_Type__c')=='ROE' and a.get('Signed_Date__c') for a in ags)
    has_signed_pal = any(a.get('Agreement_Type__c')=='PAL' and a.get('Signed_Date__c') for a in ags)
    in_st = o.get('In_SiteTracker__c')

    # Inclusion
    included_this = cat in ('Cat 1','Cat 2') or has_signed
    if not included_this:
        if cat == 'Cat 3':
            excl_cat3 += 1
        elif not cat:
            excl_null += 1
    else:
        bucket, col = classify(o)
        if bucket:
            included += 1
            units = float(o.get('Units__c') or 0)
            fr = o.get('Franchise_Type__c')
            for fk in ([fr] if fr in ('In-Franchise','National') else []):
                counts[fk][bucket][col]['n'] += 1; counts[fk][bucket][col]['u'] += units
            counts['Total'][bucket][col]['n'] += 1; counts['Total'][bucket][col]['u'] += units

    # Action-item categorization (independent of inclusion)
    if has_signed and (cat == 'Cat 3' or not cat):
        action_items['A_needs_cat'].append({
            'Name': o['Name'], 'Id': o['Id'], 'Stage': o.get('StageName'),
            'Current Category': cat or '(null)', 'Franchise': o.get('Franchise_Type__c'),
            'Agreement_Name__c': o.get('Agreement_Name__c'),
        })
    if o.get('StageName') == 'Under Contract' and not has_signed:
        action_items['B_uc_no_ag'].append({
            'Name': o['Name'], 'Id': o['Id'], 'Franchise': o.get('Franchise_Type__c'),
            'Monday_Item_ID': o.get('Monday_Item_ID__c'),
        })
    if o.get('StageName') == 'ROE Secured' and not has_signed_roe:
        action_items['C_roes_no_ag'].append({
            'Name': o['Name'], 'Id': o['Id'], 'Franchise': o.get('Franchise_Type__c'),
        })
    if o.get('StageName') == 'Under Contract' and has_signed_pal and not in_st:
        action_items['G_uc_pal_no_st'].append({
            'Name': o['Name'], 'Id': o['Id'], 'Franchise': o.get('Franchise_Type__c'),
            'Agreement_Name__c': o.get('Agreement_Name__c'),
        })

for ic in iron_no_ag:
    action_items['D_ic_no_ag_link'].append({
        'Record_Name': ic.get('Record_Name__c'), 'Id': ic['Id'],
        'Type': ic.get('Record_Type_IC__c'), 'Counterparty': ic.get('Counterparty_Name__c'),
        'Agree_Name': ic.get('Agree_Name__c') or '(blank — needs population in IronClad)',
        'Agreement_Date': ic.get('Agreement_Date__c'),
    })

for a in ag_no_ic:
    action_items['E_ag_no_ic'].append({
        'Opp': (a.get('Opportunity__r') or {}).get('Name'), 'Opp_Id': a.get('Opportunity__c'),
        'Agreement_Id': a['Id'], 'Type': a.get('Agreement_Type__c'),
        'Status': a.get('Status__c'), 'Signed_Date': a.get('Signed_Date__c'),
    })

for s in st_no_ag:
    action_items['F_st_no_ag'].append({
        'Opp': (s.get('Opportunity__r') or {}).get('Name'), 'Opp_Id': s.get('Opportunity__c'),
        'ST_Id': s['Id'], 'Build_Status': s.get('Build_Status__c'),
    })


# ── Render workbook ─────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Investor Pipeline ──
ws = wb.active
ws.title = 'Investor Pipeline'

HDR_FONT = Font(bold=True, size=11, color='FFFFFF')
BUCKET_FONT = Font(bold=True, size=10)
SEC_FILL = {'In-Franchise': PatternFill('solid', fgColor='4F81BD'),
            'National':     PatternFill('solid', fgColor='E46C0A'),
            'Total':        PatternFill('solid', fgColor='70AD47')}
SEC_ROW = {'In-Franchise': PatternFill('solid', fgColor='DCE6F1'),
           'National':     PatternFill('solid', fgColor='FBD4B4'),
           'Total':        PatternFill('solid', fgColor='E2EFDA')}
SUB = PatternFill('solid', fgColor='D9D9D9')
NEED = PatternFill('solid', fgColor='FFE699')
CTR = Alignment(horizontal='center', vertical='center')
LFT = Alignment(horizontal='left', vertical='center', indent=1)
THIN = Side(style='thin', color='A6A6A6')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

row = 1
ws.cell(row=row, column=1, value='Investor PAL Pipeline').font = Font(bold=True, size=14); row += 1
ws.cell(row=row, column=1, value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} from Salesforce").font = Font(italic=True, size=9); row += 2

for franchise in FRANCHISES:
    ws.cell(row=row, column=1, value=franchise).font = HDR_FONT
    ws.cell(row=row, column=1).fill = SEC_FILL[franchise]
    ws.cell(row=row, column=1).alignment = LFT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value='ROE'); c.font = HDR_FONT; c.alignment = CTR; c.fill = SEC_FILL[franchise]; c.border = BORDER
    ws.cell(row=row, column=3).fill = SEC_FILL[franchise]; ws.cell(row=row, column=3).border = BORDER
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    c = ws.cell(row=row, column=4, value='PAL'); c.font = HDR_FONT; c.alignment = CTR; c.fill = SEC_FILL[franchise]; c.border = BORDER
    ws.cell(row=row, column=5).fill = SEC_FILL[franchise]; ws.cell(row=row, column=5).border = BORDER
    row += 1

    ws.cell(row=row, column=1, value='').fill = SUB
    for i, label in enumerate(['Properties','Units','Properties','Units']):
        c = ws.cell(row=row, column=i+2, value=label)
        c.font = Font(bold=True, size=9, italic=True); c.alignment = CTR; c.fill = SUB; c.border = BORDER
    row += 1

    for bucket in BUCKETS_DISPLAY:
        c = ws.cell(row=row, column=1, value=bucket)
        c.alignment = LFT; c.font = BUCKET_FONT; c.border = BORDER
        c.fill = NEED if bucket == 'Needs Classification' else SEC_ROW[franchise]
        if bucket == 'Needs Classification':
            c.font = Font(bold=True, italic=True, size=10)
        roe = counts[franchise][bucket]['ROE']; pal = counts[franchise][bucket]['PAL']
        # ALWAYS render cells (0 shown as "0" not blank, to make the structure fixed)
        for i, v in enumerate([roe['n'], int(roe['u']), pal['n'], int(pal['u'])]):
            cc = ws.cell(row=row, column=i+2, value=v)
            cc.alignment = CTR; cc.border = BORDER
            if bucket == 'Needs Classification': cc.fill = NEED
            if i in (1, 3): cc.number_format = '#,##0'
        row += 1
    row += 1

ws.column_dimensions['A'].width = 30
for col in ('B','C','D','E'):
    ws.column_dimensions[col].width = 13

ws.cell(row=row, column=1, value='Excluded from report (no signed agreement + Cat 3 or null Category):').font = Font(bold=True, italic=True); row += 1
ws.cell(row=row, column=1, value=f'  Cat 3 (Off Net): {excl_cat3}').font = Font(italic=True, size=9); row += 1
ws.cell(row=row, column=1, value=f'  Null Category: {excl_null}').font = Font(italic=True, size=9); row += 1

# ── Sheet 2: Action Items ──
ws2 = wb.create_sheet('Action Items')
row = 1
ws2.cell(row=row, column=1, value='Data Hygiene Action Items').font = Font(bold=True, size=14); row += 1
ws2.cell(row=row, column=1, value='Suggested tasks to bridge gaps so the investor pipeline is accurate.').font = Font(italic=True, size=9); row += 2

SECTION_HDR = PatternFill('solid', fgColor='4F81BD')
TBL_HDR = PatternFill('solid', fgColor='D9D9D9')

def write_section(title, desc, owner, items, columns, section_id, total_count=None):
    global row
    shown_count = total_count if total_count is not None else len(items)
    c = ws2.cell(row=row, column=1, value=f"{section_id}. {title}  ({shown_count} opps)")
    c.font = HDR_FONT; c.fill = SECTION_HDR; c.alignment = LFT
    for col in range(2, len(columns)+2):
        ws2.cell(row=row, column=col).fill = SECTION_HDR
    row += 1
    ws2.cell(row=row, column=1, value=f"Fix: {desc}").font = Font(italic=True, size=10); row += 1
    ws2.cell(row=row, column=1, value=f"Suggested owner: {owner}").font = Font(italic=True, size=10); row += 1
    if items:
        for i, col in enumerate(columns):
            c = ws2.cell(row=row, column=i+1, value=col)
            c.font = Font(bold=True, size=9); c.fill = TBL_HDR; c.alignment = CTR; c.border = BORDER
        row += 1
        for it in items[:15]:
            for i, col in enumerate(columns):
                c = ws2.cell(row=row, column=i+1, value=str(it.get(col) or ''))
                c.alignment = LFT; c.border = BORDER
                c.font = Font(size=9)
            row += 1
        full_total = total_count if total_count is not None else len(items)
        if full_total > 15:
            ws2.cell(row=row, column=1, value=f"  ... + {full_total-15} more (full list available on request)").font = Font(italic=True, size=9); row += 1
    else:
        ws2.cell(row=row, column=1, value='  (no items — good!)').font = Font(italic=True, size=9); row += 1
    row += 2

write_section(
    title="Opps with signed agreement but no Property_Category",
    desc="Set Property_Category to Cat 1 (On Net), Cat 2 (Near Net), or Cat 3 (Off Net) on each of these opps. Investor report currently surfaces them in the 'Needs Classification' row.",
    owner="Taylor Mauney (MDU) / Justin (CA) / market owners",
    items=action_items['A_needs_cat'],
    columns=['Name','Id','Stage','Current Category','Franchise','Agreement_Name__c'],
    section_id='A',
)
write_section(
    title="Opps in 'Under Contract' stage but no signed Agreement in SF",
    desc="Either (a) move the Opportunity stage back to Contract Negotiations until a PAL is signed, or (b) create the Agreement__c record matching the signed PAL.",
    owner="Taylor Mauney",
    items=action_items['B_uc_no_ag'],
    columns=['Name','Id','Franchise','Monday_Item_ID'],
    section_id='B',
)
write_section(
    title="Opps in 'ROE Secured' stage but no signed ROE Agreement",
    desc="Either move the stage back, or create a signed ROE Agreement__c for each. Opp stage implies ROE was signed.",
    owner="Taylor Mauney",
    items=action_items['C_roes_no_ag'],
    columns=['Name','Id','Franchise'],
    section_id='C',
)
write_section(
    title="IronClad records with no Agreement__c / Opportunity link",
    desc="These IronClad contracts exist in SF but are not linked to an Opportunity. Populate 'AgreeName' in IronClad to enable auto-linking, or manually create the Agreement__c record.",
    owner="Legal / Taylor Mauney",
    items=action_items['D_ic_no_ag_link'],
    columns=['Record_Name','Id','Type','Counterparty','Agree_Name','Agreement_Date'],
    section_id='D',
    total_count=iron_no_ag_total,
)
write_section(
    title="Signed Agreement__c records with no IronClad link",
    desc="Agreement is signed in SF but not traceable to an IronClad document. Add the IronClad Record ID or AgreeName cross-reference.",
    owner="Legal / Taylor Mauney",
    items=action_items['E_ag_no_ic'],
    columns=['Opp','Opp_Id','Agreement_Id','Type','Status','Signed_Date'],
    section_id='E',
    total_count=ag_no_ic_total,
)
write_section(
    title="SiteTracker projects without a signed Agreement on linked Opp",
    desc="ST project created but no signed Agreement__c on the linked Opp. Either the Agreement wasn't imported or the ST was created prematurely.",
    owner="Eric / Ops / Taylor",
    items=action_items['F_st_no_ag'],
    columns=['Opp','Opp_Id','ST_Id','Build_Status'],
    section_id='F',
    total_count=st_no_ag_total,
)
write_section(
    title="Under Contract + signed PAL but no SiteTracker project",
    desc="Per business rule 'signed PAL/ROE should trigger ST project creation'. These opps signed but the ST project never got created.",
    owner="Eric / SiteTracker ops",
    items=action_items['G_uc_pal_no_st'],
    columns=['Name','Id','Franchise','Agreement_Name__c'],
    section_id='G',
)

for col, w in zip(['A','B','C','D','E','F'], [45, 22, 18, 25, 30, 15]):
    ws2.column_dimensions[col].width = w

out_path = rf"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Investor_PAL_Pipeline_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
wb.save(out_path)
print(f"Wrote: {out_path}")

# Console summary
print(f"\n=== Summary ===")
print(f"Included in report: {included}  |  Excluded Cat3: {excl_cat3}  |  Excluded null: {excl_null}")
print(f"\nAction Items:")
for key, items in action_items.items():
    print(f"  {key}: {len(items)}")