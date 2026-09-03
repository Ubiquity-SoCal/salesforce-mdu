"""
Build the Salesforce -> HubSpot sample record pack.

The 8/31 ask was: "pull in a couple records just with all the fields, one that
has a lot of contacts and a lot of agreements, send that over so you can look at
it and map it out, and then do the same thing for address management."

So: pick the richest real records automatically, then dump each one end to end,
parent record plus every related child, field by field with labels and values.
One tab per sample record so the consultant can read a whole record vertically.

Samples are chosen by data, not hardcoded, so this stays correct as the org moves:
  - the MDU Sales opportunities carrying the most contacts + agreements, across
    different record types and stages
  - the Address Management property location carrying the most units, plus its
    units and any AVR case

Run: python build_hubspot_sample_records.py
Out: SalesForce/data/output/salesforce-hubspot-sample-records.xlsx
"""
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds  # noqa: E402

from simple_salesforce import Salesforce  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

OUT = Path(__file__).resolve().parents[2] / "data" / "output"
OUT.mkdir(parents=True, exist_ok=True)
XLSX = OUT / "salesforce-hubspot-sample-records.xlsx"

TITLE_FONT = Font(bold=True, size=13, color="1A3C5E", name="Calibri")
SEC_FILL = PatternFill("solid", fgColor="2A5A8A")
SEC_FONT = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
HDR_FILL = PatternFill("solid", fgColor="D9E2EC")
HDR_FONT = Font(bold=True, size=10, name="Calibri")
BODY = Font(size=10, name="Calibri")
BLANK_FONT = Font(size=10, name="Calibri", color="A0A0A0", italic=True)
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def connect():
    c = creds()
    return Salesforce(username=c["username"], password=c["password"], security_token=c["token"])


def fieldinfo(sf, obj, cache={}):
    """{api_name: (label, type)} plus ordered field list."""
    if obj not in cache:
        d = getattr(sf, obj).describe()
        cache[obj] = (
            {f["name"]: (f["label"], f["type"]) for f in d["fields"]},
            [f["name"] for f in d["fields"] if f["type"] not in ("base64", "encryptedstring")],
        )
    return cache[obj]


def get_record(sf, obj, rec_id):
    info, names = fieldinfo(sf, obj)
    out = {}
    for i in range(0, len(names), 80):
        chunk = list(dict.fromkeys(["Id"] + names[i:i + 80]))
        q = f"SELECT {', '.join(chunk)} FROM {obj} WHERE Id = '{rec_id}'"
        r = sf.query(q)["records"]
        if r:
            out.update({k: v for k, v in r[0].items() if k != "attributes"})
    return out


def get_children(sf, obj, where, limit=None):
    """Fields are fetched in chunks and merged by Id, so any LIMIT must sit on a
    deterministic ORDER BY. Without it each chunk could return a different set of
    rows and the merged record would be a blend of several."""
    info, names = fieldinfo(sf, obj)
    tail = f" ORDER BY Id LIMIT {limit}" if limit else ""
    rows = {}
    for i in range(0, len(names), 80):
        chunk = list(dict.fromkeys(["Id"] + names[i:i + 80]))
        q = f"SELECT {', '.join(chunk)} FROM {obj} WHERE {where}{tail}"
        for r in sf.query_all(q)["records"]:
            rows.setdefault(r["Id"], {}).update({k: v for k, v in r.items() if k != "attributes"})
    return list(rows.values())


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, dict):
        v = {k: x for k, x in v.items() if k != "attributes" and x is not None}
        return "; ".join(f"{k}={x}" for k, x in v.items())
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v)


class Sheet:
    """Vertical record writer: sections, then Field / Label / Type / Value rows."""

    def __init__(self, wb, name, title, subtitle):
        self.ws = wb.create_sheet(name[:31])
        self.ws.cell(row=1, column=1, value=title).font = TITLE_FONT
        self.ws.cell(row=2, column=1, value=subtitle).font = Font(size=10, italic=True, name="Calibri")
        self.r = 4

    def section(self, text):
        for c in range(1, 5):
            cell = self.ws.cell(row=self.r, column=c)
            cell.fill = SEC_FILL
            cell.font = SEC_FONT
            cell.border = BOX
        self.ws.cell(row=self.r, column=1, value=text)
        self.r += 1
        for c, h in enumerate(["Field API Name", "Label", "Type", "Value"], start=1):
            cell = self.ws.cell(row=self.r, column=c, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = BOX
        self.r += 1

    def note(self, text):
        self.ws.cell(row=self.r, column=1, value=text).font = Font(size=10, italic=True, name="Calibri", color="808080")
        self.r += 1

    def record(self, sf, obj, rec, skip_blank=False):
        info, names = fieldinfo(sf, obj)
        for n in names:
            if n not in rec:
                continue
            val = fmt(rec.get(n))
            if skip_blank and val == "":
                continue
            label, typ = info.get(n, (n, ""))
            self.ws.cell(row=self.r, column=1, value=n).font = BODY
            self.ws.cell(row=self.r, column=2, value=label).font = BODY
            self.ws.cell(row=self.r, column=3, value=typ).font = BODY
            cell = self.ws.cell(row=self.r, column=4, value=val)
            cell.font = BLANK_FONT if val == "" else BODY
            if val == "":
                cell.value = "(blank)"
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            for c in range(1, 5):
                self.ws.cell(row=self.r, column=c).border = BOX
            self.r += 1
        self.r += 1

    def finish(self):
        for i, w in enumerate([38, 34, 16, 78], start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        self.ws.freeze_panes = "A4"


def pick_opportunities(sf, n=3):
    """Richest opps by contacts + agreements, spread across record types."""
    oc = Counter(r["Opportunity__c"] for r in sf.query_all(
        "SELECT Opportunity__c FROM Opportunity_Contact__c WHERE Opportunity__c != null")["records"])
    ag = Counter(r["Opportunity__c"] for r in sf.query_all(
        "SELECT Opportunity__c FROM Agreement__c WHERE Opportunity__c != null")["records"])
    ranked = sorted(
        set(oc) | set(ag),
        key=lambda k: (-(oc.get(k, 0) + ag.get(k, 0)), -min(oc.get(k, 0), ag.get(k, 0))),
    )[:40]
    ids = ",".join(f"'{k}'" for k in ranked)
    meta = {r["Id"]: r for r in sf.query_all(
        f"SELECT Id, Name, StageName, RecordType.Name FROM Opportunity WHERE Id IN ({ids})")["records"]}
    picked, seen_rt = [], set()
    # first pass: one per record type, preferring records with BOTH children
    for k in sorted(ranked, key=lambda k: -min(oc.get(k, 0), ag.get(k, 0))):
        rt = (meta.get(k, {}).get("RecordType") or {}).get("Name", "?")
        if rt in seen_rt:
            continue
        seen_rt.add(rt)
        picked.append((k, meta[k], oc.get(k, 0), ag.get(k, 0)))
        if len(picked) >= n:
            break
    # top up on raw volume if we ran out of record types
    for k in ranked:
        if len(picked) >= n:
            break
        if any(k == p[0] for p in picked):
            continue
        picked.append((k, meta[k], oc.get(k, 0), ag.get(k, 0)))
    return picked


def main():
    sf = connect()
    print("connected:", sf.sf_instance)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Read Me")
    txt = [
        ("Salesforce to HubSpot: sample records", "t"),
        (f"Generated {datetime.now():%Y-%m-%d %H:%M} from the live org.", "b"),
        ("", "b"),
        ("Each tab is one real record read top to bottom: the parent record with every field, then every", "b"),
        ("related record hanging off it. Blank fields are shown as (blank) on purpose, so the shape of a real", "b"),
        ("record is visible, not just the fields that happen to be filled in.", "b"),
        ("", "b"),
        ("The one thing to notice on the MDU tabs", "h"),
        ("Contacts are linked through Opportunity_Contact__c, a custom junction object. The standard", "b"),
        ("Salesforce OpportunityContactRole is NOT used and is empty. Any contact mapping has to follow", "b"),
        ("the junction, or it will come back with nothing.", "b"),
        ("", "b"),
        ("Pair this with salesforce-hubspot-field-map.xlsx, which has the full field list per object.", "b"),
    ]
    r = 1
    for t, k in txt:
        cell = ws.cell(row=r, column=1, value=t)
        cell.font = TITLE_FONT if k == "t" else Font(bold=True, size=11, color="2A5A8A", name="Calibri") if k == "h" else BODY
        r += 1
    ws.column_dimensions["A"].width = 110

    # ---------- MDU Sales samples ----------
    for oid, meta, n_con, n_ag in pick_opportunities(sf, n=3):
        name = meta["Name"]
        rt = (meta.get("RecordType") or {}).get("Name", "?")
        print(f"  building {name} ({rt}, {n_con} contacts, {n_ag} agreements)")
        tabname = "".join(ch for ch in name if ch not in "[]:*?/\\")[:28]
        sh = Sheet(wb, tabname, f"{name}", f"MDU Sales | Record Type: {rt} | Stage: {meta['StageName']} | {n_con} contacts, {n_ag} agreements | Id {oid}")

        sh.section("OPPORTUNITY  (the property deal)")
        sh.record(sf, "Opportunity", get_record(sf, "Opportunity", oid))

        links = get_children(sf, "Opportunity_Contact__c", f"Opportunity__c = '{oid}'")
        sh.section(f"OPPORTUNITY_CONTACT__c  (junction rows: {len(links)})")
        sh.note("This is how contacts attach to a deal. Standard OpportunityContactRole is unused in this org.")
        for i, lnk in enumerate(links, 1):
            sh.note(f"-- junction row {i} of {len(links)} --")
            sh.record(sf, "Opportunity_Contact__c", lnk)
        cids = [lnk.get("Contact__c") for lnk in links if lnk.get("Contact__c")]
        if cids:
            contacts = get_children(sf, "Contact", "Id IN (" + ",".join(f"'{c}'" for c in cids) + ")")
            sh.section(f"CONTACT  (the people behind those junction rows: {len(contacts)})")
            for i, con in enumerate(contacts, 1):
                sh.note(f"-- contact {i} of {len(contacts)}: {con.get('Name')} --")
                sh.record(sf, "Contact", con)

        ags = get_children(sf, "Agreement__c", f"Opportunity__c = '{oid}'")
        sh.section(f"AGREEMENT__c  (signed agreements: {len(ags)})")
        for i, a in enumerate(ags, 1):
            sh.note(f"-- agreement {i} of {len(ags)}: {a.get('Name')} --")
            sh.record(sf, "Agreement__c", a)

        camps = get_children(sf, "Opportunity_Campaign__c", f"Opportunity__c = '{oid}'")
        if camps:
            sh.section(f"OPPORTUNITY_CAMPAIGN__c  (project tags: {len(camps)})")
            for c in camps:
                sh.record(sf, "Opportunity_Campaign__c", c)

        accs = get_children(sf, "Opportunity_Account__c", f"Opportunity__c = '{oid}'")
        if accs:
            sh.section(f"OPPORTUNITY_ACCOUNT__c  (management company links: {len(accs)})")
            for a in accs:
                sh.record(sf, "Opportunity_Account__c", a)

        tasks = get_children(sf, "Task", f"WhatId = '{oid}'")
        if tasks:
            sh.section(f"TASK  (activity history: {len(tasks)}, showing up to 3)")
            for t in tasks[:3]:
                sh.record(sf, "Task", t, skip_blank=True)
        sh.finish()

    # ---------- Address Management sample ----------
    print("  building Address Management sample")
    units = Counter(r["Property_Location__c"] for r in sf.query_all(
        "SELECT Property_Location__c FROM Property_Unit__c WHERE Property_Location__c != null")["records"])
    if units:
        plid, nunits = units.most_common(1)[0]
        pl = get_record(sf, "Property_Location__c", plid)
        sh = Sheet(wb, "Address Mgmt sample",
                   f"{pl.get('Name')}",
                   f"Address Management | Property_Location__c with the most units | {nunits} units | Id {plid}")
        sh.section("PROPERTY_LOCATION__c  (the physical property)")
        sh.record(sf, "Property_Location__c", pl)

        us = get_children(sf, "Property_Unit__c", f"Property_Location__c = '{plid}'", limit=3)
        sh.section(f"PROPERTY_UNIT__c  ({nunits} units on this property, showing 3 in full)")
        for i, u in enumerate(us, 1):
            sh.note(f"-- unit {i} of 3 shown ({nunits} total) --")
            sh.record(sf, "Property_Unit__c", u)

        cases = get_children(sf, "Case", "Id != null", limit=1)
        if cases:
            sh.section("CASE  (Address Validation Request workflow, sample case)")
            sh.note("AVR runs on Cases. Requestor_Email_Extract parses an inbound email into 16 fields in one update.")
            sh.record(sf, "Case", cases[0])

        avr = get_children(sf, "AVR_Project_ID__c", "Id != null", limit=1)
        if avr:
            sh.section("AVR_PROJECT_ID__c  (sample)")
            sh.record(sf, "AVR_Project_ID__c", avr[0])
        sh.finish()

    wb.save(XLSX)
    print("wrote", XLSX)


if __name__ == "__main__":
    main()
