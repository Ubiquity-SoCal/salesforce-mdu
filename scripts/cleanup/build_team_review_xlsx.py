"""Build stage_cleanup_team_review.xlsx — owner-grouped action-required workbook.
Pulls live data from SF for every Opp currently flagged in stage_cleanup_team_review.md
plus the EMA/Bulk In Progress Brett cluster. One row per Opp with explicit Action Required.

Sheet layout: one Summary tab + one tab per owner."""
from simple_salesforce import Salesforce
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sf = Salesforce(
    username=_SF["username"],
    password=_SF["password"],
    security_token=_SF["token"],
)

# ---- review items, hand-curated from stage_cleanup_team_review.md ----
# Each row: (stage, owner, sf_id, current_state, action_required)
REVIEW_ITEMS = [
    # ---- Engaged ----
    ('Engaged', 'Justin Barry', '006WR00000yur0lYAA',
     '1/27 voicemail to Stephanie. Three prior calls, no callback. "10/10 pending contact from Mason."',
     'Confirm still Engaged. If yes, set Next_Action. If unreachable, drop to Prospects or Closed Lost.'),
    ('Engaged', 'Justin Barry', '006WR00000ywTYXYA2',
     '3/11 followup with Simon. 2/11 sent Easement Agreement. 1/27/26 completed easement.',
     'Is Simon still moving? If yes, set Next_Action with target. If stalled, drop stage.'),
    ('Engaged', 'Justin Barry', '006WR00000xuFKAYA2',
     'Bridgeport_MDU_Federal Housing Administration. Moved into Engaged 2026-05-01. No Next_Action set.',
     'Set Next_Action describing where this stands.'),

    # ---- Contract Negotiations ----
    ('Contract Negotiations', 'Brett Spivey', '006WR00000wj8KSYAY',
     'PAL Paused (manual SF entry, no IronClad). No Next_Action, no Projected.',
     'Still in Contract Negotiations? Set Next_Action / Projected, or move to On Hold.'),
    ('Contract Negotiations', 'Brett Spivey', '006WR00000wkAW3YAM',
     'PAL Paused (manual). No Next_Action, no Projected.',
     'Still in negotiations or On Hold?'),
    ('Contract Negotiations', 'Jeff Chao', '006WR00000wjnlUYAQ',
     'PAL Paused + EMA Review (manual). No Next_Action, no Projected, no IronClad.',
     'Still alive or On Hold?'),
    ('Contract Negotiations', 'Jeff Chao', '006WR00000wjnmKYAQ',
     'PAL Paused + EMA Completed signed 2025-01-21 (manual, no IronClad backing).',
     'EMA shows Completed/signed 1/21/25 — if real, this is PAL/ROE Complete. Confirm: was EMA actually signed?'),
    ('Contract Negotiations', 'Tanya Friese', '006WR00000wk9SGYAY',
     'PAL Review (manual). No Next_Action, no Projected, no IronClad.',
     'Still active or On Hold?'),
    ('Contract Negotiations', 'Melissa Baker', '006WR000011cV9lYAE',
     'NONE — empty record. Zero data: no agreements, no notes, no Next_Action, no Projected.',
     'Was this created in error or as a placeholder? Delete or define purpose.'),
    ('Contract Negotiations', 'Chuck McNeely (inactive)', '006WR00000wk9SjYAI',
     'PAL Review + ROE Cancelled (manual).',
     'Reassign owner. Verify Contract Negotiations is correct.'),
    ('Contract Negotiations', 'Chuck McNeely (inactive)', '006WR00000wk9S0YAI',
     'PAL Review (manual).',
     'Reassign owner. Verify stage.'),
    ('Contract Negotiations', 'Chuck McNeely (inactive)', '006WR00000wk9SwYAI',
     'PAL Paused (manual).',
     'Reassign owner. Likely On Hold given Paused agreement + zero recent signals.'),
    ('Contract Negotiations', 'Chuck McNeely (inactive)', '006WR00000wk9S5YAI',
     'PAL+EMA Paused (manual).',
     'Reassign owner. Likely On Hold.'),

    # ---- PAL/ROE Complete ----
    ('PAL/ROE Complete', 'Justin Barry', '006WR00000wk9YTYAY',
     '"Confirmed Ting bulk signed at this property per Dave Putnam" (10/4/24). No Bulk Agreement record exists.',
     'Verify Bulk signed. Create Bulk Agreement__c record (auto-bumps to EMA/Bulk Complete) or confirm not active.'),
    ('PAL/ROE Complete', 'Justin Barry', '006WR00000wkCllYAE',
     '"Ting BULK Confirmed by SFU Team" (10/8/24). No Bulk Agreement record exists.',
     'Verify Bulk signed. Create Bulk Agreement__c record or confirm not active.'),
    ('PAL/ROE Complete', 'Justin Barry', '006WR00000wkClkYAE',
     '"Ting Bulk confirmed by SFU Team" (10/8/24). No Bulk Agreement record exists.',
     'Verify Bulk signed. Create Bulk Agreement__c record or confirm not active.'),
    ('PAL/ROE Complete', 'Justin Barry', '006WR00000wkClNYAU',
     '"Ting BULK Confirmed by SFU Team" (10/8/24). No Bulk Agreement record exists.',
     'Verify Bulk signed. Create Bulk Agreement__c record or confirm not active.'),
    ('PAL/ROE Complete', 'Tanya Friese', '006WR00000wkEbPYAU',
     '"PAL signed, waiting on PAL addendum and verbal on EMA" (4/24/26).',
     'Is the EMA active enough to bump to EMA/Bulk In Progress, or stay in PAL/ROE Complete?'),
    ('PAL/ROE Complete', 'Tanya Friese', '006WR00000xuzoQYAQ',
     '"verbal on EMA" (weekly tracker note). Possible dupe with Bradley Arms_Colt RE.',
     'Confirm EMA status. Confirm dupe with Bradley Arms_Colt RE (006WR00000wkCjuYAE).'),
    ('PAL/ROE Complete', 'Justin Barry', '006WR00000wkA6iYAE',
     'ROE AGR-1417 Status=Sign but Next_Action says "ROE signed - awaiting build orders".',
     'Update Status to Completed + populate Signed_Date.'),
    ('PAL/ROE Complete', 'Brett Spivey', '006WR00000wkEc3YAE',
     'Bulk AGR-1040 Status=Completed but Signed_Date=None and no IronClad. Next_Action says "Brett to contact owner re bulk requirement".',
     'Either Bulk really is Completed (fill Signed_Date + IronClad link) or Status is wrong.'),
    ('PAL/ROE Complete', 'Brett Spivey + Melissa Baker', '006WR00000yvY5dYAE',
     'Possible duplicate. Both empty Converge Justin records. Brett=PAL/ROE Complete, Melissa=CN (006WR000011cV9lYAE).',
     'Pick one to keep, close the other.'),
    ('PAL/ROE Complete', 'Chuck McNeely (inactive)', '006WR00000wkEcKYAU',
     'Lexington Place (Monterey). No Agreement, no Next_Action, no Projected, owner inactive.',
     'Reassign + verify stage.'),
    # Missing PAL/ROE Agreement records (data gap, stage OK per CX-tracking rule)
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Decatur_MDU_Smallwood Trailer Park — no PAL/ROE Agreement__c record (CX-tracking).',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Omaha_MDU_4612 Redman Ave — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Omaha_MDU_4760 LAFAYETTE AVE — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Omaha_MDU_5004 Davenport St — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Omaha_MDU_9208 Ohio St — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Omaha_MDU_Benson Crest Apartments 2 — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Melissa Baker', None,
     'Omaha_MDU_Farnam Flats — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Tanya Friese', None,
     'Paul Mark Apts — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
    ('PAL/ROE Complete', 'Justin Barry', None,
     '512-514 Via De La Valle — no PAL/ROE Agreement__c record.',
     'Backfill PAL Agreement record, or accept the gap?'),
]

# Brett's 31 EMA/Bulk In Progress cluster
BRETT_EMABULK_CLUSTER = [
    ('Baldwin Manor', '006WR00000wkEaRYAU', 'AGR-0793 Review', 'AGR-0794 Review', 'P-003350'),
    ('Bali Apartments', '006WR00000wkEaTYAU', 'AGR-0799 Review', 'AGR-0800 Review', 'P-003352'),
    ('Brody Terrace (Riverside)', '006WR00000wkEahYAE', 'AGR-0841 Review', 'AGR-0842 Review', 'P-003331'),
    ('Brookside Terrace Apartments', '006WR00000wkEbLYAU', 'AGR-0935 Review', 'AGR-0936 Review', 'P-003349'),
    ('Casa Linda', '006WR00000wkEaXYAU', 'AGR-0811 Review', 'AGR-0812 Review', 'P-003347'),
    ('Chatsworth Plaza', '006WR00000wkEaVYAU', 'AGR-0805 Review', 'AGR-0806 Review', 'P-003345'),
    ('Coliseum Apartments', '006WR00000wkEaZYAU', 'AGR-0817 Review', 'AGR-0818 Review', 'P-003343'),
    ('Colonial Manor Apartments', '006WR00000wkEaaYAE', 'AGR-0820 Review', 'AGR-0821 Review', 'P-003342'),
    ('Corbett Avenue Apartments', '006WR00000wkEaUYAU', 'AGR-0802 Review', 'AGR-0803 Review', 'P-003383'),
    ('First Casa De Marina', '006WR00000wkEbIYAU', 'AGR-0926 Review', 'AGR-0927 Review', 'P-003281'),
    ('Jaclyn Terrace', '006WR00000wkEagYAE', 'AGR-0838 Review', 'AGR-0839 Review', 'P-003338'),
    ('John Manor', '006WR00000wkEaWYAU', 'AGR-0808 Review', 'AGR-0809 Review', 'P-003384'),
    ('Kling Trio Apartments', '006WR00000wkEafYAE', 'AGR-0835 Review', 'AGR-0836 Review', 'P-003336'),
    ('Krystal Terrace', '006WR00000wkEadYAE', 'AGR-0829 Review', 'AGR-0830 Review', 'P-003387'),
    ('Lombardi Apartments', '006WR00000wkEbJYAU', 'AGR-0929 Review', 'AGR-0930 Review', 'P-003335'),
    ('Parkview Terrace', '006WR00000wkEaiYAE', 'AGR-0844 Review', 'AGR-0845 Review', 'P-003334'),
    ('Parkway Terrace', '006WR00000wkEaeYAE', 'AGR-0832 Review', 'AGR-0833 Review', 'P-003388'),
    ('Parthenia Terrace', '006WR00000wkEbNYAU', 'AGR-0941 Review', 'AGR-0942 Review', 'P-003332'),
    ('Riverside Villa Apartments', '006WR00000wkEakYAE', 'AGR-0850 Review', 'AGR-0851 Review', 'P-003330'),
    ('Roxanne Apartments', '006WR00000wkEajYAE', 'AGR-0847 Review', 'AGR-0848 Review', 'P-003353'),
    ('San Vicente Apartments', '006WR00000wkEabYAE', 'AGR-0823 Review', 'AGR-0824 Review', 'P-003385'),
    ('St. Andrews Manor', '006WR00000wkEacYAE', 'AGR-0826 Review', 'AGR-0827 Review', 'P-003386'),
    ('The Banyans', '006WR00000wkEalYAE', 'AGR-0853 Review', 'AGR-0854 Review', 'P-003351'),
    ('The Meadows at Westlake Village', '006WR00000wkEbKYAU', 'AGR-0932 Review', 'AGR-0933 Review', 'P-003348'),
    ('The Palms', '006WR00000wkEapYAE', 'AGR-0865 Review', 'AGR-0866 Review', 'P-003346'),
    ('Topanga Apartments', '006WR00000wkEamYAE', 'AGR-0856 Review', 'AGR-0857 Review', 'P-003344'),
    ('Topanga Terrace', '006WR00000wkEanYAE', 'AGR-0859 Review', 'AGR-0860 Review', 'P-003341'),
    ('Vista Apartments', '006WR00000wkEaoYAE', 'AGR-0862 Review', 'AGR-0863 Review', 'P-003340'),
    ('White Oak Terrace', '006WR00000wkEaSYAU', 'AGR-0796 Review', 'AGR-0797 Review', 'P-003337'),
    ('Windsor Manor', '006WR00000wkEaYYAU', 'AGR-0814 Review', 'AGR-0815 Review', 'P-003381'),
    ('Woodland Trio The Oaks', '006WR00000wkEaQYAU', 'AGR-0790 Review', 'AGR-0791 Review', 'P-003339'),
]
for name, sfid, ema, bulk, st in BRETT_EMABULK_CLUSTER:
    REVIEW_ITEMS.append((
        'EMA/Bulk In Progress', 'Brett Spivey', sfid,
        f'{name}. PAL Completed (signed 2023-11-01 — bulk import date). EMA={ema}, Bulk={bulk} (manual, no IronClad). ST={st} at "Project - PAL/ROE Signed" (no construction). No Next_Action / Projected.',
        '(A) IN NEGOTIATION — keep here, set Next_Action + Projected. (B) NOT NEGOTIATING — drop to PAL/ROE Complete, cancel placeholder EMA + Bulk. (C) OTHER — explain.',
    ))

# EMA/Bulk Complete review (Chuck inactive, no signed EMA/Bulk)
EMABULK_COMPLETE_REVIEW = [
    ('Arbors of Killeen', '006WR00000wkEcmYAE',
     'PAL AGR-1165 Completed (IC-1084), EMA AGR-1166 Archive (never signed), PAL Addendum AGR-1167 Completed. ST P-004231 Build Completed.'),
    ('The Bluffs of Brookside', '006WR00000wk1ElYAI',
     'PAL AGR-1298 Completed (IC-1499), EMA AGR-1299 Archive (never signed), PAL Addendum AGR-1300 Completed. ST P-006292 Build Completed.'),
    ('The Renaissance at Stoney Creek', '006WR00000wk1EjYAI',
     'PAL AGR-1291 Completed (IC-1504), EMA AGR-1292 Archive (never signed), PAL Addendum AGR-1293 Completed. ST P-006296 Build Completed.'),
]
for name, sfid, state in EMABULK_COMPLETE_REVIEW:
    REVIEW_ITEMS.append((
        'EMA/Bulk Complete', 'Chuck McNeely (inactive)', sfid,
        f'{name}. {state}',
        '(A) MOVE TO PAL/ROE COMPLETE — EMA never signed (Archive correct). (B) STAY HERE — EMA actually signed, fix Status + Signed_Date. PLUS reassign owner.',
    ))

# On Hold review — only flag the 3 records that show signs of active work (Koa: On Hold is lowest priority)
import csv as _csv

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()

on_hold_rows = []
try:
    with open('C:/Users/cass/Work_Projects/SalesForce/audit_logs/stage_audit_on_hold_2026-05-04.csv', encoding='utf-8') as fh:
        on_hold_rows = list(_csv.DictReader(fh))
except FileNotFoundError:
    print("On Hold CSV not found — skip On Hold integration")

# These 3 are the only On Hold records where Next_Action / Projected indicates real activity
# (the other 3 with Next_Action use it as status documentation, not active work)
ON_HOLD_ACTIVE_REVIEW = {
    # SF Id -> (opp_name, owner, signal, action)
    None: None,
}
# Identify them programmatically: has Projected, OR Next_Action that doesn't read like status doc
STATUS_DOC_PHRASES = ('stalled', 'on hold -', 'build not approved', 'no agreement', 'owner out of country', 'deal stuck')
for r in on_hold_rows:
    na = (r['Next_Action'] or '').strip()
    pj = r['Projected'] or ''
    if not (na or pj):
        continue
    # Skip status-doc-only Next_Actions
    na_lower = na.lower()
    is_status_doc = any(phrase in na_lower for phrase in STATUS_DOC_PHRASES) and not pj
    if is_status_doc:
        continue
    if 'TEST PROPERTY' in r['Name'].upper():
        action = 'DELETE — test record in production pipeline.'
        signal = f'Name="{r["Name"]}", Projected={pj}'
    elif 'cancel' in na_lower:
        action = 'Cancellation in progress — should this close out as Closed Lost rather than sit On Hold?'
        signal = f'Next_Action: "{na[:80]}"'
    else:
        action = 'Active work + forecast date. Why On Hold rather than Engaged / Prospecting / Contract Negotiations?'
        signal = f'Next_Action: "{na[:80]}", Projected={pj}'
    REVIEW_ITEMS.append(('On Hold', r['Owner'], r['Id'], signal, action))

# ---- enrich with live SF data ----
ids = [r[2] for r in REVIEW_ITEMS if r[2]]
ids_str = "','".join(ids)
opp_data = {}
for o in sf.query_all(f"""
    SELECT Id, Name, Owner.Name, Sales_Status__c, Hold_Reason__c,
           Projected_Close_Date__c, Next_Action__c, Next_Action_Date__c,
           StageName, LastModifiedDate
    FROM Opportunity WHERE Id IN ('{ids_str}')
""")['records']:
    opp_data[o['Id']] = o

# IronClad linkage per Opp
ic_by_opp = defaultdict(list)  # opp_id -> [(agreement_name, ic_id_or_None)]
for r in sf.query_all(f"""
    SELECT Opportunity__c, Name, IronClad_ID__c, Agreement_Type__c
    FROM Agreement__c WHERE Opportunity__c IN ('{ids_str}')
""")['records']:
    ic_by_opp[r['Opportunity__c']].append({
        'Name': r['Name'],
        'IC': r.get('IronClad_ID__c'),
        'Type': r.get('Agreement_Type__c'),
    })

def ic_summary(opp_id):
    agrs = ic_by_opp.get(opp_id, [])
    if not agrs:
        return ('No Agreements', '')
    linked = [a for a in agrs if a['IC']]
    if not linked:
        return (f'No (0 of {len(agrs)})', '')
    flag = 'Yes' if len(linked) == len(agrs) else f'Partial ({len(linked)} of {len(agrs)})'
    detail = '; '.join(f"{a['Name']} {a['Type']}={a['IC']}" for a in linked)
    return (flag, detail)

# ---- build workbook ----
wb = Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(left=Side(style='thin', color='CCCCCC'),
                     right=Side(style='thin', color='CCCCCC'),
                     top=Side(style='thin', color='CCCCCC'),
                     bottom=Side(style='thin', color='CCCCCC'))
ALT_FILL = PatternFill('solid', fgColor='F2F2F2')

COLUMNS = ['Stage', 'Owner', 'Opp Name', 'SF Id', 'IC Linked', 'IronClad IDs', 'Current State', 'Action Required', 'Decision', 'Notes']
COL_WIDTHS = [22, 22, 32, 22, 16, 38, 50, 60, 18, 30]

def write_header(ws):
    for i, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i-1]
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

def write_row(ws, row_idx, stage, owner, opp_name, sfid, current_state, action_req):
    if sfid:
        ic_flag, ic_detail = ic_summary(sfid)
    else:
        ic_flag, ic_detail = ('N/A', '')
    values = [stage, owner, opp_name, sfid or '(no SF Id — backfill)', ic_flag, ic_detail, current_state, action_req, '', '']
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=i, value=v)
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.border = THIN_BORDER
        if row_idx % 2 == 0:
            c.fill = ALT_FILL

# ---- Summary tab — Stage-led pivot ----
summary_ws = wb.create_sheet('Summary')
summary_ws['A1'] = 'MDU Stage Cleanup — Team Review'
summary_ws['A1'].font = Font(bold=True, size=14)
summary_ws['A2'] = 'Compiled during stage-by-stage cleanup of MDU pipeline (post 2026-04-29 restructure).'
summary_ws['A3'] = 'Each row in Review Items is an Opp where the current stage placement needs the owner to confirm or correct.'

# canonical stage order (matches MDU pipeline)
STAGE_ORDER = [
    'Engaged', 'Contract Negotiations', 'PAL/ROE Complete',
    'EMA/Bulk In Progress', 'EMA/Bulk Complete', 'On Hold',
]
def stage_sort_key(s):
    return (STAGE_ORDER.index(s) if s in STAGE_ORDER else 999, s)

stage_counts = defaultdict(int)
stage_owner_counts = defaultdict(lambda: defaultdict(int))
stage_ic_counts = defaultdict(lambda: Counter())  # stage -> Counter(ic_flag)
for stage, owner, sfid, cs, act in REVIEW_ITEMS:
    stage_counts[stage] += 1
    stage_owner_counts[stage][owner] += 1
    if sfid:
        flag, _ = ic_summary(sfid)
    else:
        flag = 'N/A'
    # collapse partial/no into a coarser bucket
    if flag == 'Yes':
        ic_bucket = 'IC Linked'
    elif flag.startswith('Partial'):
        ic_bucket = 'Partial'
    elif flag.startswith('No (') or flag == 'No Agreements':
        ic_bucket = 'No IC link'
    else:
        ic_bucket = 'N/A'
    stage_ic_counts[stage][ic_bucket] += 1

stages_sorted = sorted(stage_counts.keys(), key=stage_sort_key)

# ---- Pivot 1: Stage × Owner ----
summary_ws['A5'] = 'Pivot 1 — Items by Stage × Owner'
summary_ws['A5'].font = Font(bold=True, size=12)

all_owners = sorted({o for s in stage_owner_counts.values() for o in s})
header_row = 7
summary_ws.cell(row=header_row, column=1, value='Stage').font = Font(bold=True)
for i, owner in enumerate(all_owners, 2):
    c = summary_ws.cell(row=header_row, column=i, value=owner)
    c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True, horizontal='center')
total_col = len(all_owners) + 2
summary_ws.cell(row=header_row, column=total_col, value='Total').font = Font(bold=True)

r = header_row + 1
for stage in stages_sorted:
    summary_ws.cell(row=r, column=1, value=stage).font = Font(bold=True)
    total = 0
    for i, owner in enumerate(all_owners, 2):
        n = stage_owner_counts[stage].get(owner, 0)
        if n:
            summary_ws.cell(row=r, column=i, value=n)
        total += n
    summary_ws.cell(row=r, column=total_col, value=total).font = Font(bold=True)
    r += 1
summary_ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True)
for i, owner in enumerate(all_owners, 2):
    total = sum(stage_owner_counts[s].get(owner, 0) for s in stages_sorted)
    if total:
        summary_ws.cell(row=r, column=i, value=total).font = Font(bold=True)
summary_ws.cell(row=r, column=total_col, value=sum(stage_counts.values())).font = Font(bold=True)

# ---- Pivot 2: Stage × IronClad linkage ----
ic_pivot_start = r + 3
summary_ws.cell(row=ic_pivot_start, column=1, value='Pivot 2 — Items by Stage × IronClad linkage').font = Font(bold=True, size=12)
ic_buckets = ['IC Linked', 'Partial', 'No IC link', 'N/A']
ic_header = ic_pivot_start + 2
summary_ws.cell(row=ic_header, column=1, value='Stage').font = Font(bold=True)
for i, b in enumerate(ic_buckets, 2):
    summary_ws.cell(row=ic_header, column=i, value=b).font = Font(bold=True)
summary_ws.cell(row=ic_header, column=len(ic_buckets) + 2, value='Total').font = Font(bold=True)
for j, stage in enumerate(stages_sorted, ic_header + 1):
    summary_ws.cell(row=j, column=1, value=stage).font = Font(bold=True)
    total = 0
    for i, b in enumerate(ic_buckets, 2):
        n = stage_ic_counts[stage].get(b, 0)
        if n:
            summary_ws.cell(row=j, column=i, value=n)
        total += n
    summary_ws.cell(row=j, column=len(ic_buckets) + 2, value=total).font = Font(bold=True)

# Column widths
summary_ws.column_dimensions['A'].width = 28
for i in range(2, max(len(all_owners), len(ic_buckets)) + 3):
    summary_ws.column_dimensions[get_column_letter(i)].width = 16

by_owner = defaultdict(list)
for item in REVIEW_ITEMS:
    by_owner[item[1]].append(item)

# ---- All Items tab (single filterable list — owner column filters in place of per-owner tabs) ----
ws_all = wb.create_sheet('Review Items')
write_header(ws_all)
sorted_items = sorted(REVIEW_ITEMS, key=lambda x: (x[1], x[0], x[2] or ''))
for i, (stage, owner, sfid, cs, act) in enumerate(sorted_items, 2):
    opp = opp_data.get(sfid)
    opp_name = opp['Name'] if opp else cs.split('—')[0].strip().split('.')[0]
    write_row(ws_all, i, stage, owner, opp_name, sfid, cs, act)
for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
    ws_all.row_dimensions[row[0].row].height = 60
ws_all.auto_filter.ref = ws_all.dimensions

# ---- On Hold full backlog tab (all 282 rows) ----
if on_hold_rows:
    ws_oh = wb.create_sheet('On Hold (all 282)')
    OH_COLS = ['Owner','Opp Name','SF Id','Hold Reason','Last Activity','Bucket','Action Required','Decision']
    OH_WIDTHS = [22, 32, 22, 22, 14, 12, 60, 18]
    for i, col in enumerate(OH_COLS, 1):
        c = ws_oh.cell(row=1, column=i, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = THIN_BORDER
        ws_oh.column_dimensions[get_column_letter(i)].width = OH_WIDTHS[i-1]
    ws_oh.row_dimensions[1].height = 24
    ws_oh.freeze_panes = 'A2'
    sorted_holds = sorted(on_hold_rows, key=lambda r: (r['Owner'], r['Name']))
    for i, r in enumerate(sorted_holds, 2):
        last_act = r['Last_Activity'][:10] if r['Last_Activity'] else ''
        has_signal = bool((r['Next_Action'] and r['Next_Action'].strip()) or r['Projected'])
        bucket_label = 'has signal' if has_signal else 'dormant'
        action = '— low priority — owner triage when Hold_Reason picklist is enriched.'
        values = [r['Owner'], r['Name'], r['Id'], r['Reason'], last_act, bucket_label, action, '']
        for j, v in enumerate(values, 1):
            c = ws_oh.cell(row=i, column=j, value=v)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = THIN_BORDER
        ws_oh.row_dimensions[i].height = 28
    ws_oh.auto_filter.ref = ws_oh.dimensions

# Move Summary first
wb.move_sheet('Summary', offset=-len(wb.sheetnames)+1)

out = 'C:/Users/cass/Work_Projects/SalesForce/stage_cleanup_team_review.xlsx'
wb.save(out)
print(f"Saved: {out}")
print(f"Total review items: {len(REVIEW_ITEMS)}")
print(f"Owner tabs: {sorted(by_owner.keys())}")
