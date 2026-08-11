"""Taylor 2026-05-04 review pass — Substatus push Phase 1: snapshot + plan.

Reads xlsx TM updates, identifies every cell where Taylor populated a value
(Substatus, Sales_Status, etc) that current SF is blank or different.
Snapshots current SF state, builds the push plan, writes targets JSON for Phase 2.
"""
from __future__ import annotations

import csv
import json
import os
from collections import defaultdict, Counter
from datetime import datetime, date

import openpyxl
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


XLSX_PATH = r'C:\Users\cass\Downloads\Take 2 MDU Sales Review (Living) - TM updates 5.4.xlsx'
AUDIT_DIR = r'C:\Users\cass\Work_Projects\SalesForce\audit_logs\2026-05-07_taylor_substatus_push'
SOURCE_TAG = 'TM_review_2026-05-04'

# Map old (xlsx) value -> existing SF picklist value when there's a rename.
SUBSTATUS_VALUE_MAP = {
    'Bulk/EMA Rejected': 'Bulk/Marketing Rejected',
}

# Expected new picklist values — Phase 2 deploy adds these.
NEW_SUBSTATUS_VALUES = {
    'ISP or Funding Needed',
    'Incumbent EMA',
    'No Marketing/Bulk Needed',
}

def jsonable(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v

# ---- read xlsx ------------------------------------------------------------

print('Reading xlsx...')
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active
hdr = [c.value for c in ws[1]]
ix = {h: i for i, h in enumerate(hdr)}

xlsx_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[ix['Opportunity Name']]:
        continue
    xlsx_rows.append(row)
print(f'  {len(xlsx_rows)} rows in xlsx')

# ---- SF connect -----------------------------------------------------------

print('Connecting to Salesforce...')
sf = Salesforce(
    username=_SF["username"],
    password=_SF["password"],
    security_token=_SF["token"],
)

all_names = list({r[ix['Opportunity Name']] for r in xlsx_rows})
chunk = 200
opps = []
for i in range(0, len(all_names), chunk):
    batch = all_names[i:i+chunk]
    nstr = ','.join("'" + n.replace("'", "\\'") + "'" for n in batch)
    rs = sf.query_all(f"""
        SELECT Id, Name, Owner.Name, StageName, Substatus__c, Sales_Status__c,
               Hold_Reason__c, Next_Action__c
        FROM Opportunity WHERE Name IN ({nstr})
    """)['records']
    opps.extend(rs)

# Index by (Name, Owner.Name) — same convention as Phase 1
sf_by_key = {}
for o in opps:
    sf_by_key.setdefault((o['Name'], o['Owner']['Name']), []).append(o)

# ---- build push plan ------------------------------------------------------

substatus_pushes = []   # {Id, Name, Stage, OwnerName, raw_xlsx, mapped_value, current, status}
sales_status_pushes = []
unresolved = []

for r in xlsx_rows:
    name = r[ix['Opportunity Name']]
    owner = r[ix['Opportunity Owner']]
    key = (name, owner)
    cands = sf_by_key.get(key, [])
    if len(cands) != 1:
        # Skip dupes/missing — those aren't where Taylor changed Substatus anyway
        continue
    sf_o = cands[0]

    # Substatus
    x_substatus = r[ix['Substatus']]
    sf_substatus = sf_o.get('Substatus__c')
    if x_substatus and not sf_substatus:
        mapped = SUBSTATUS_VALUE_MAP.get(x_substatus, x_substatus)
        needs_new_pl = mapped in NEW_SUBSTATUS_VALUES
        substatus_pushes.append({
            'Id': sf_o['Id'],
            'Name': name,
            'Owner': owner,
            'StageName': sf_o['StageName'],
            'raw_xlsx_value': x_substatus,
            'mapped_value': mapped,
            'current_sf_value': sf_substatus,
            'needs_new_picklist_value': needs_new_pl,
        })

    # Sales Status
    x_ss = r[ix['Sales Status']]
    sf_ss = sf_o.get('Sales_Status__c')
    if x_ss and not sf_ss:
        sales_status_pushes.append({
            'Id': sf_o['Id'],
            'Name': name,
            'Owner': owner,
            'StageName': sf_o['StageName'],
            'value': x_ss,
            'current_sf_value': sf_ss,
        })

# ---- write outputs --------------------------------------------------------

os.makedirs(AUDIT_DIR, exist_ok=True)

resolved = {
    'generated_at': datetime.now().isoformat(),
    'source': SOURCE_TAG,
    'substatus_pushes': substatus_pushes,
    'sales_status_pushes': sales_status_pushes,
}
with open(os.path.join(AUDIT_DIR, 'phase1_resolved_targets.json'), 'w', encoding='utf-8') as f:
    json.dump(resolved, f, indent=2, default=jsonable)

# Snapshot of current state — so we can revert
snapshot_ids = list({p['Id'] for p in substatus_pushes + sales_status_pushes})
snapshot = {o['Id']: {
    'Id': o['Id'], 'Name': o['Name'], 'Owner': o['Owner']['Name'],
    'StageName': o['StageName'],
    'Substatus__c': o.get('Substatus__c'),
    'Sales_Status__c': o.get('Sales_Status__c'),
} for o in opps if o['Id'] in snapshot_ids}
with open(os.path.join(AUDIT_DIR, 'phase1_snapshot.json'), 'w', encoding='utf-8') as f:
    json.dump(snapshot, f, indent=2)

# Match report
val_counts = Counter(p['mapped_value'] for p in substatus_pushes)
needs_new_count = sum(1 for p in substatus_pushes if p['needs_new_picklist_value'])
already_in_pl = len(substatus_pushes) - needs_new_count

report = [
    f'Substatus push — Phase 1 dry-run  ({datetime.now().isoformat()})',
    '',
    f'Source xlsx: {XLSX_PATH}',
    '',
    f'Substatus updates planned:  {len(substatus_pushes)}',
    f'  Maps to existing picklist value: {already_in_pl}',
    f'  Needs new picklist value:        {needs_new_count} (will fail until deploy lands)',
    '',
    f'Sales_Status updates planned:  {len(sales_status_pushes)}',
    '',
    'Substatus value distribution:',
]
for v, c in val_counts.most_common():
    new_marker = ' [NEEDS PL DEPLOY]' if v in NEW_SUBSTATUS_VALUES else ''
    report.append(f'  {v}: {c}{new_marker}')

report.append('')
report.append('Substatus values per stage:')
stage_x_val = Counter((p['StageName'], p['mapped_value']) for p in substatus_pushes)
for (stage, val), c in sorted(stage_x_val.items(), key=lambda x: (x[0][0], -x[1])):
    report.append(f'  [{stage}] {val}: {c}')

report.append('')
report.append('Sales_Status updates:')
for p in sales_status_pushes:
    report.append(f"  [{p['StageName']}] [{p['Owner']}] {p['Name']}: -> {p['value']}")

with open(os.path.join(AUDIT_DIR, 'phase1_match_report.txt'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(report))

print('\n' + '\n'.join(report))
print(f'\nFiles written to {AUDIT_DIR}')
