"""Check whether Taylor edited any non-TM-Note cells in the xlsx.

Substatus, Sales Status, Hold Reason, Next Action are all candidates.
Compare xlsx values to current SF values to find divergences.
"""
import openpyxl
from simple_salesforce import Salesforce
from collections import Counter

XLSX = r'C:\Users\cass\Downloads\Take 2 MDU Sales Review (Living) - TM updates 5.4.xlsx'

sf = Salesforce(
    username='cass1@ubiquitygp.com',
    password='Hawaiian1984',
    security_token='IBSKT6CFUpSUJWxq1CMm0HkFC',
)

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

# Header (column index map)
hdr = [c.value for c in ws[1]]
ix = {h: i for i, h in enumerate(hdr)}
print('Columns:', hdr)
print()

# Pull every xlsx row that has a non-null Substatus, Sales Status, Hold Reason, or Next Action
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    rows.append(row)

# Group rows by (Name, Owner) and pick the first
by_key = {}
for r in rows:
    by_key[(r[0], r[3])] = r

# Pull current SF state for ALL Opps in xlsx
all_names = list({r[0] for r in rows})
chunk = 200
opps = []
for i in range(0, len(all_names), chunk):
    batch = all_names[i:i+chunk]
    nstr = ','.join("'" + n.replace("'", "\\'") + "'" for n in batch)
    rs = sf.query_all(f"""
        SELECT Id, Name, Owner.Name, Substatus__c, Sales_Status__c,
               Hold_Reason__c, Next_Action__c
        FROM Opportunity WHERE Name IN ({nstr})
    """)['records']
    opps.extend(rs)

# Index by (Name, Owner)
sf_by_key = {}
for o in opps:
    sf_by_key.setdefault((o['Name'], o['Owner']['Name']), []).append(o)

# Compare
diffs = {'Substatus__c': [], 'Sales_Status__c': [], 'Hold_Reason__c': [], 'Next_Action__c': []}
xlsx_to_sf = {
    'Substatus': 'Substatus__c',
    'Sales Status': 'Sales_Status__c',
    'Hold Reason': 'Hold_Reason__c',
    'Next Action': 'Next_Action__c',
}

for (name, owner), xlsx_row in by_key.items():
    cands = sf_by_key.get((name, owner), [])
    if len(cands) != 1:
        continue
    sf_o = cands[0]
    for xcol, sfcol in xlsx_to_sf.items():
        xval = xlsx_row[ix[xcol]]
        sval = sf_o.get(sfcol)
        # Normalize for compare: strip whitespace, treat None and '' equal
        xn = (str(xval).strip() if xval is not None else '')
        sn = (str(sval).strip() if sval is not None else '')
        if xn != sn:
            diffs[sfcol].append((name, owner, xn, sn))

print('Cells where xlsx value differs from current SF value:\n')
for field, lst in diffs.items():
    print(f'== {field} ({len(lst)} rows) ==')
    new_in_xlsx = [d for d in lst if d[2] and not d[3]]
    new_in_sf = [d for d in lst if d[3] and not d[2]]
    differ = [d for d in lst if d[2] and d[3] and d[2] != d[3]]
    print(f'  xlsx has, SF blank: {len(new_in_xlsx)}')
    print(f'  SF has, xlsx blank: {len(new_in_sf)}')
    print(f'  both have, differ:  {len(differ)}')

# For Substatus, show value-frequency + cross-tab against current SF stage
print('\n== Substatus value distribution in xlsx (where SF is currently blank) ==')
substatus_diffs = diffs['Substatus__c']
val_counts = Counter(d[2] for d in substatus_diffs if d[2])
for v, c in val_counts.most_common():
    print(f'  {v!r}: {c}')

# Cross-tab against current SF stage for each diff
print('\n== Substatus xlsx-has-but-SF-blank by current SF stage ==')
sf_state_by_key = {(o['Name'], o['Owner']['Name']): o for o in opps}
stage_x_substatus = Counter()
for n, o, xv, sv in substatus_diffs:
    sf_o = sf_state_by_key.get((n, o))
    if sf_o:
        stage_x_substatus[(sf_o['StageName'] if 'StageName' in sf_o else '?', xv)] += 1
# Re-query stages now (StageName wasn't in original SELECT)
oid_to_stage = {}
ids_str = "','".join({sf_state_by_key[(n,o)]['Id'] for n,o,_,_ in substatus_diffs if (n,o) in sf_state_by_key})
if ids_str:
    rs = sf.query_all(f"SELECT Id, StageName FROM Opportunity WHERE Id IN ('{ids_str}')")['records']
    oid_to_stage = {r['Id']: r['StageName'] for r in rs}
stage_x_substatus = Counter()
for n, o, xv, sv in substatus_diffs:
    sf_o = sf_state_by_key.get((n, o))
    stage = oid_to_stage.get(sf_o['Id']) if sf_o else '(no match)'
    stage_x_substatus[(stage, xv)] += 1
for (stage, xv), c in sorted(stage_x_substatus.items(), key=lambda x: (-x[1], x[0])):
    print(f'  [{stage}] {xv!r}: {c}')

# Did Taylor TM-Note these rows or not?
print('\n== Substatus diffs: TM-Noted vs not ==')
xlsx_tm_noted_keys = {(r[0], r[3]) for r in rows if r[ix['TM Note']]}
n_tm = sum(1 for n,o,_,_ in substatus_diffs if (n,o) in xlsx_tm_noted_keys)
n_not = sum(1 for n,o,_,_ in substatus_diffs if (n,o) not in xlsx_tm_noted_keys)
print(f'  TM-Noted: {n_tm}')
print(f'  Not TM-Noted (separate Substatus edits): {n_not}')
