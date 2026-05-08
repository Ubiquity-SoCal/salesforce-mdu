"""
Match SAQ and ROE Excel tracker records to Salesforce Opportunities.
- SAQ records: set Owner to Bill Holick
- ROE records: set RE_Assigned__c based on initials (TF/JB/RS)
"""
import openpyxl
from simple_salesforce import Salesforce

USERNAME = "cass1@ubiquitygp.com"
PASSWORD = "Karate88!"
SECURITY_TOKEN = "Ktc1n9mLmD9vwEcVcl45q0iAD"

BILL_HOLICK_ID = "005WR00000DEU6oYAH"
RE_MAP = {
    "TF": "005WR0000030R1hYAE",   # Tanya Friese
    "JB": "005WR0000030RCzYAM",   # Justin Barry
    "RS": "005WR0000030R9lYAE",   # Rosemarie Shortino
}

SAQ_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\PMO_Projects - MDU SAQ\Master List MDU Assignments.xlsm"
ROE_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\PMO_Projects - MDU 9-25 Units ROE Project\MDU 9 - 25 Units.xlsx"


def main():
    sf = Salesforce(username=USERNAME, password=PASSWORD, security_token=SECURITY_TOKEN)
    print("Connected to Salesforce")

    # Build AgreeName -> Opp ID lookup (case-insensitive)
    opps = sf.query_all("SELECT Id, Agreement_Name__c FROM Opportunity WHERE Agreement_Name__c != null")
    agree_map = {}
    for rec in opps["records"]:
        key = rec["Agreement_Name__c"].strip().lower()
        agree_map[key] = rec["Id"]
    print(f"Loaded {len(agree_map)} Opportunities with Agreement_Name__c")

    # ── SAQ: Match and set Owner to Bill Holick ──
    print("\n=== SAQ Tracker ===")
    wb = openpyxl.load_workbook(SAQ_PATH, read_only=True, data_only=True)
    ws = wb["Opportunities"]
    headers = [cell.value for cell in list(ws.iter_rows(min_row=2, max_row=2))[0]]
    agree_idx = next(i for i, h in enumerate(headers) if h and "agree" in str(h).lower())

    saq_matched = 0
    saq_unmatched = []
    saq_updates = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        agree_name = row[agree_idx] if agree_idx < len(row) else None
        if not agree_name:
            continue
        agree_name = str(agree_name).strip()
        opp_id = agree_map.get(agree_name.lower())
        if opp_id:
            saq_updates.append({"Id": opp_id, "OwnerId": BILL_HOLICK_ID})
            saq_matched += 1
        else:
            saq_unmatched.append(agree_name)

    wb.close()
    print(f"Matched: {saq_matched}, Unmatched: {len(saq_unmatched)}")

    # Dedup (same opp might appear in multiple SAQ rows)
    seen = set()
    saq_deduped = []
    for u in saq_updates:
        if u["Id"] not in seen:
            seen.add(u["Id"])
            saq_deduped.append(u)
    print(f"Unique Opps to update: {len(saq_deduped)}")

    if saq_deduped:
        results = sf.bulk.Opportunity.update(saq_deduped, batch_size=200)
        ok = sum(1 for r in results if r.get("success"))
        print(f"Updated {ok} Opp owners to Bill Holick")

    if saq_unmatched[:10]:
        print(f"Sample unmatched ({len(saq_unmatched)} total):")
        for name in saq_unmatched[:10]:
            print(f"  {name}")

    # ── ROE: Match and set RE_Assigned__c ──
    print("\n=== ROE Tracker (9-25 Units) ===")
    wb = openpyxl.load_workbook(ROE_PATH, read_only=True, data_only=True)

    roe_matched = 0
    roe_unmatched = []
    roe_updates = []

    for sheet_name in ["Site Data", "TX Site Data"]:
        ws = wb[sheet_name]
        headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        agree_idx = next(i for i, h in enumerate(headers) if h and "agree" in str(h).lower())
        re_idx = next(i for i, h in enumerate(headers) if h and "re assign" in str(h).lower())

        for row in ws.iter_rows(min_row=2, values_only=True):
            agree_name = row[agree_idx] if agree_idx < len(row) else None
            re_initials = str(row[re_idx]).strip().upper() if re_idx < len(row) and row[re_idx] else None

            if not agree_name:
                continue
            agree_name = str(agree_name).strip()
            opp_id = agree_map.get(agree_name.lower())

            if opp_id and re_initials and re_initials in RE_MAP:
                roe_updates.append({"Id": opp_id, "RE_Assigned__c": RE_MAP[re_initials]})
                roe_matched += 1
            elif opp_id and not re_initials:
                roe_matched += 1  # Matched but no RE assigned
            else:
                roe_unmatched.append(agree_name)

    wb.close()
    print(f"Matched with RE: {roe_matched}, Unmatched: {len(roe_unmatched)}")

    # Dedup
    seen = set()
    roe_deduped = []
    for u in roe_updates:
        if u["Id"] not in seen:
            seen.add(u["Id"])
            roe_deduped.append(u)
    print(f"Unique Opps to set RE_Assigned: {len(roe_deduped)}")

    if roe_deduped:
        results = sf.bulk.Opportunity.update(roe_deduped, batch_size=200)
        ok = sum(1 for r in results if r.get("success"))
        print(f"Updated {ok} Opps with RE_Assigned__c")

    if roe_unmatched[:10]:
        print(f"Sample unmatched ({len(roe_unmatched)} total):")
        for name in roe_unmatched[:10]:
            print(f"  {name}")

    # Summary
    print("\n=== Summary ===")
    r = sf.query(f"SELECT COUNT() FROM Opportunity WHERE OwnerId = '{BILL_HOLICK_ID}'")
    print(f"Opps owned by Bill Holick: {r['totalSize']}")
    r2 = sf.query("SELECT COUNT() FROM Opportunity WHERE RE_Assigned__c != null")
    print(f"Opps with RE_Assigned: {r2['totalSize']}")


if __name__ == "__main__":
    main()
