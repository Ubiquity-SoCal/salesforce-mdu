"""
Push xlsx Parcel data to Property_Location__c.Parcel__c for the 245 SMB ROE buildings.

Source: SMB_ROE_Project.xlsx, sheet ROE_Tracking.
Target: Property_Location__c.Parcel__c (string, len=255).

Decisions
=========
- xlsx wins. If SF currently has a value and xlsx has a (different) value, overwrite.
  Reasoning: the team built the xlsx with SF visibility, so any divergence is intentional.
- Within xlsx, prefer the column with MORE comma-separated parcel IDs.
  Tiebreak (same count): prefer `Parcel`. If `Parcel` is blank, fall back to `Parcel #`.
- Cleanups applied to xlsx value before writing:
    * Strip leading "common parcel" / "Common parcel" prefixes.
    * Strip trailing notes like " - wrong PIN" (kept in audit Note column).
    * Collapse whitespace.
- Match xlsx rows -> PL via the same matching logic as the backfill (Business Buildings
  -> PL.Name with normalize+zip-strip + UUID/Building_Id_c fallback).

Usage:
  python push_smb_roe_parcels_2026-04-27.py            # preview
  python push_smb_roe_parcels_2026-04-27.py --apply    # writes + audit log
"""
import sys, io, re, csv, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from simple_salesforce import Salesforce

ap = argparse.ArgumentParser()
ap.add_argument('--apply', action='store_true')
args = ap.parse_args()
APPLY = args.apply

XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
SHEET = 'ROE_Tracking'
SCRIPT_NAME = 'push_smb_roe_parcels_2026-04-27.py'
TS = datetime.now().isoformat(timespec='seconds')
AUDIT_DIR = Path(r'C:\Users\cass\Work_Projects\SalesForce\audit_logs')
AUDIT_DIR.mkdir(parents=True, exist_ok=True)

sf = Salesforce(username='cass1@ubiquitygp.com', password='Hawaiian1984', security_token='IBSKT6CFUpSUJWxq1CMm0HkFC')

UUID_RE = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
ZIP_SUFFIX_RE = re.compile(r'\s+\d{5}(?:-\d{4})?\s*$')
COMMON_PARCEL_PREFIX_RE = re.compile(r'^\s*common\s+parcel\s+', re.I)
ANNOTATION_RE = re.compile(r'\s*[-–—]\s*(wrong\s+PIN|bad|incorrect|verify)\b.*$', re.I)


def normalize_addr(s):
    if not s:
        return ''
    return re.sub(r'\s+', ' ', s.strip().upper())


def strip_zip(s):
    return ZIP_SUFFIX_RE.sub('', s).strip() if s else ''


def parcel_count(v):
    """Count comma-separated parcel IDs in a value."""
    if not v:
        return 0
    return len([x for x in re.split(r'[,;]', str(v)) if x.strip()])


def clean_parcel(raw):
    """Apply prefix/annotation cleanups; return (cleaned, removed_note)."""
    if raw is None:
        return None, None
    # Collapse newlines + whitespace FIRST so annotation regex anchors work
    s = re.sub(r'\s+', ' ', str(raw)).strip()
    if not s:
        return None, None
    note = None
    # Strip "common parcel" prefix
    new_s = COMMON_PARCEL_PREFIX_RE.sub('', s)
    if new_s != s:
        note = (note or '') + 'stripped "common parcel" prefix; '
        s = new_s
    # Strip annotation pattern (e.g. " - wrong PIN" possibly with trailing alt-PIN)
    m = ANNOTATION_RE.search(s)
    if m:
        annotation_text = m.group(0).strip(' -')
        note = (note or '') + f'stripped annotation: "{annotation_text}"; '
        s = ANNOTATION_RE.sub('', s).strip()
    # Final whitespace + trailing-comma cleanup
    s = re.sub(r'\s+', ' ', s).strip(' ,')
    return (s or None), (note.rstrip('; ') if note else None)


# ── Load xlsx ──
print(f"[Load] {XLSX} sheet={SHEET}")
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[SHEET]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]
print(f"  Rows: {len(data)}")

# ── Pull PLs ──
print("\n[SF] Pulling Property_Locations + Business_ROE Opps")
states = set(r.get('State') for r in data if r.get('State'))
state_clause = "','".join(s for s in states if s)
all_pls = sf.query_all(f"SELECT Id, Name, Parcel__c, Business_Base_Address__c, Business_Building_Id__c FROM Property_Location__c WHERE State__c IN ('{state_clause}')")['records']
print(f"  Pulled {len(all_pls)} Property_Locations in target states")

pl_by_name, pl_by_norm, pl_by_norm_no_zip, pl_by_bid = {}, {}, {}, {}
for pl in all_pls:
    if pl.get('Name'):
        pl_by_name[pl['Name']] = pl
        pl_by_norm.setdefault(normalize_addr(pl['Name']), pl)
        pl_by_norm_no_zip.setdefault(strip_zip(normalize_addr(pl['Name'])), pl)
    if pl.get('Business_Base_Address__c'):
        pl_by_norm.setdefault(normalize_addr(pl['Business_Base_Address__c']), pl)
        pl_by_norm_no_zip.setdefault(strip_zip(normalize_addr(pl['Business_Base_Address__c'])), pl)
    if pl.get('Business_Building_Id__c'):
        pl_by_bid[pl['Business_Building_Id__c']] = pl


def find_pl(r):
    bn = r.get('Business Buildings') or ''
    if bn in pl_by_name:
        return pl_by_name[bn]
    norm = normalize_addr(bn)
    pl = pl_by_norm.get(norm) or pl_by_norm_no_zip.get(strip_zip(norm))
    if not pl and UUID_RE.match(bn):
        pl = pl_by_bid.get(bn)
        if not pl:
            uba = r.get('Updated Business Address') or ''
            pl = pl_by_norm.get(normalize_addr(uba)) or pl_by_norm_no_zip.get(strip_zip(normalize_addr(uba)))
    return pl


# ── Plan updates ──
print("\n[Plan] Picking best parcel value per row + comparing to current SF state")

planned = []   # (pl_id, pl_name, current_value, new_value, source_col, internal_xlsx_disagreement, cleanup_note)
flagged_for_review = []  # rows where the team annotated the parcel as wrong/needs verification
no_match = 0
no_xlsx_data = 0
no_change = 0
overwrites = 0
fills = 0

for r in data:
    pl = find_pl(r)
    if not pl:
        no_match += 1
        continue

    p1 = r.get('Parcel #')
    p2 = r.get('Parcel')
    p1_clean, _ = clean_parcel(p1)
    p2_clean, p2_note = clean_parcel(p2)
    p1_count = parcel_count(p1_clean)
    p2_count = parcel_count(p2_clean)

    # Pick which xlsx column wins
    chosen_value = None
    chosen_col = None
    cleanup_note = None
    if p1_clean and p2_clean:
        if p1_count > p2_count:
            chosen_value, chosen_col = p1_clean, 'Parcel #'
        elif p2_count > p1_count:
            chosen_value, chosen_col, cleanup_note = p2_clean, 'Parcel', p2_note
        else:
            # Same count — prefer `Parcel`
            chosen_value, chosen_col, cleanup_note = p2_clean, 'Parcel', p2_note
    elif p2_clean:
        chosen_value, chosen_col, cleanup_note = p2_clean, 'Parcel', p2_note
    elif p1_clean:
        chosen_value, chosen_col = p1_clean, 'Parcel #'

    if not chosen_value:
        no_xlsx_data += 1
        continue

    # Skip rows where annotation indicates the team flagged the parcel as wrong/incorrect.
    # We can't safely guess the right value; leave SF untouched + flag for human review.
    if cleanup_note and re.search(r'wrong PIN|incorrect|verify', cleanup_note, re.I):
        flagged_for_review.append({
            'pl_id': pl['Id'], 'pl_name': pl['Name'],
            'current': pl.get('Parcel__c'),
            'p1_raw': p1, 'p2_raw': p2,
            'note': cleanup_note,
        })
        continue

    # Detect internal xlsx disagreement (informational, doesn't change action)
    xlsx_disagreement = bool(p1_clean and p2_clean and p1_clean != p2_clean)

    current = pl.get('Parcel__c')
    if current and current.strip() == chosen_value:
        no_change += 1
        continue

    if current:
        overwrites += 1
    else:
        fills += 1

    planned.append({
        'pl_id': pl['Id'],
        'pl_name': pl['Name'],
        'current': current,
        'new': chosen_value,
        'source_col': chosen_col,
        'xlsx_disagreement': xlsx_disagreement,
        'cleanup_note': cleanup_note,
        'p1_raw': p1,
        'p2_raw': p2,
    })

# ── Summary ──
print("\n" + "="*70)
print("PLAN SUMMARY")
print("="*70)
print(f"  xlsx rows:                       {len(data)}")
print(f"  No PL match:                     {no_match}")
print(f"  No xlsx parcel data:             {no_xlsx_data}")
print(f"  PL value already matches xlsx:   {no_change}")
print(f"  PL value will be FILLED (was empty):     {fills}")
print(f"  PL value will be OVERWRITTEN (was diff): {overwrites}")
print(f"  Total updates planned:           {len(planned)}")
print(f"  Skipped — flagged 'wrong PIN'/needs review: {len(flagged_for_review)}")
for f in flagged_for_review:
    print(f"    {f['pl_name'][:55]:55s}  raw={f['p2_raw']!s}")

# Source column distribution
src_dist = Counter(p['source_col'] for p in planned)
print(f"\n  Source column distribution of chosen values:")
for k, v in src_dist.most_common():
    print(f"    {v:4d}  {k}")

# How many planned updates carry an internal xlsx disagreement
xlsx_disagree = sum(1 for p in planned if p['xlsx_disagreement'])
print(f"\n  Planned updates where xlsx Parcel# != Parcel: {xlsx_disagree}")

# How many are overwrites worth eyeballing
print(f"\n  --- Sample overwrites (first 15) ---")
samples = [p for p in planned if p['current']][:15]
for p in samples:
    note = ''
    if p['xlsx_disagreement']:
        note += '  [xlsx disagree]'
    if p['cleanup_note']:
        note += f'  [{p["cleanup_note"]}]'
    print(f"    {p['pl_name'][:50]:50s} SF={p['current']!s:25s} -> xlsx={p['new']!s:30s}{note}")

# Sample fills
print(f"\n  --- Sample fills (first 10) ---")
fills_samples = [p for p in planned if not p['current']][:10]
for p in fills_samples:
    note = ''
    if p['xlsx_disagreement']:
        note += '  [xlsx disagree]'
    if p['cleanup_note']:
        note += f'  [{p["cleanup_note"]}]'
    print(f"    {p['pl_name'][:50]:50s} -> {p['new']!s:30s}{note}")

if not APPLY:
    print(f"\n[Preview only — re-run with --apply to push {len(planned)} updates]")
    sys.exit(0)

# ── Apply ──
print("\n" + "="*70)
print(f"APPLYING — updating {len(planned)} Property_Locations")
print("="*70)

audit_rows = []
batch_records = [{'Id': p['pl_id'], 'Parcel__c': p['new']} for p in planned]
for i in range(0, len(batch_records), 200):
    batch = batch_records[i:i+200]
    plan_slice = planned[i:i+200]
    print(f"  Batch {i//200 + 1}: {len(batch)} ...")
    results = sf.bulk.Property_Location__c.update(batch)
    for j, res in enumerate(results):
        p = plan_slice[j]
        if res.get('success'):
            note_parts = [f"source={p['source_col']}"]
            if p['xlsx_disagreement']:
                note_parts.append(f"xlsx-Parcel#={p['p1_raw']!s} vs xlsx-Parcel={p['p2_raw']!s}")
            if p['cleanup_note']:
                note_parts.append(p['cleanup_note'])
            audit_rows.append({
                'SF_Id': p['pl_id'],
                'Name': p['pl_name'],
                'Field': 'Parcel__c',
                'Before': p['current'] or '',
                'After': p['new'],
                'Source': SCRIPT_NAME,
                'Timestamp': TS,
                'Action': 'OVERWRITE' if p['current'] else 'FILL',
                'Note': '; '.join(note_parts),
            })
        else:
            print(f"    ⚠ FAILED: {p['pl_name']} — {res.get('errors', res)}")

audit_path = AUDIT_DIR / f'smb_roe_parcels_audit_{TS.replace(":","-")}.csv'
with audit_path.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['SF_Id','Name','Field','Before','After','Source','Timestamp','Action','Note'])
    w.writeheader()
    w.writerows(audit_rows)
print(f"\n✓ Done. Audit log: {audit_path} ({len(audit_rows)} rows)")
