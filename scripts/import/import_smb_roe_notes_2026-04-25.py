"""
Import notes from SMB_ROE_Project.xlsx ROE_Tracking sheet as ContentNotes
attached to the matching Business_ROE Opportunity records.

One consolidated note per Opp with structured body containing all non-empty
note columns. Title: "Imported from SMB ROE Project — 2026-04-25".

Idempotent: skips Opps that already have a ContentNote with this title.

Usage:
  python import_smb_roe_notes_2026-04-25.py --preview
  python import_smb_roe_notes_2026-04-25.py --apply
"""
import sys, io, base64, csv, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from simple_salesforce import Salesforce

ap = argparse.ArgumentParser()
ap.add_argument('--apply', action='store_true')
args = ap.parse_args()
APPLY = args.apply

XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
NOTE_TITLE = 'RE Notes 2 — Imported from SMB ROE Project 2026-04-25'

sf = Salesforce(username=_SF["username"], password=_SF["password"], security_token=_SF["token"])
AUDIT_DIR = Path(r'C:\Users\cass\Work_Projects\SalesForce\audit_logs')
AUDIT_DIR.mkdir(parents=True, exist_ok=True)
TS = datetime.now().isoformat(timespec='seconds')

# Per Koa 2026-04-25: only import RE Notes2 column (the actual team-added notes
# from xlsx work). RE Notes column was sourced FROM Salesforce so already present.
# Management Company + Owner Contact → separate Account/Contact creation task (Monday).
# Parcel data → Property_Location update task (Monday).
NOTE_COLS = [
    ('RE Notes2', 'RE Notes 2'),
]

# Load all Business_ROE Opps named ROE - * (created today)
print('[Setup] Loading Business_ROE Opps')
opps = sf.query_all(
    "SELECT Id, Name, Property_Location__r.Name "
    "FROM Opportunity "
    "WHERE RecordType.DeveloperName='Business_ROE'"
)['records']
print(f"  Found {len(opps)} Business_ROE Opps")

# Index by Property_Location.Name for matching xlsx rows
opps_by_pl_name = {}
for o in opps:
    pl_name = (o.get('Property_Location__r') or {}).get('Name')
    if pl_name:
        opps_by_pl_name.setdefault(pl_name, []).append(o)

# Find which Opps already have a note with this title (idempotency)
# ContentDocumentLink can't be filtered by Title — go via ContentNote first.
print('\n[Setup] Checking which Opps already have this note imported')
existing_notes = sf.query_all(
    f"SELECT Id, LatestPublishedVersion.ContentDocumentId FROM ContentNote WHERE Title = '{NOTE_TITLE}'"
)['records']
already_imported_opp_ids = set()
if existing_notes:
    cdids = [n['LatestPublishedVersion']['ContentDocumentId'] for n in existing_notes if n.get('LatestPublishedVersion')]
    if cdids:
        # Filter by Opp prefix (006) to limit scope
        for batch_start in range(0, len(cdids), 200):
            batch = cdids[batch_start:batch_start+200]
            quoted = "','".join(batch)
            links = sf.query_all(
                f"SELECT LinkedEntityId FROM ContentDocumentLink WHERE ContentDocumentId IN ('{quoted}')"
            )['records']
            for l in links:
                if l['LinkedEntityId'].startswith('006'):  # Opportunity
                    already_imported_opp_ids.add(l['LinkedEntityId'])
print(f"  {len(already_imported_opp_ids)} Opps already have this note (skip)")

# Load xlsx
print(f'\n[Load] {XLSX}')
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb['ROE_Tracking']
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = [dict(zip(header, r)) for r in rows[1:] if any(c is not None for c in r)]
print(f"  {len(data)} xlsx rows")

# Build planned notes
planned = []
import re

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()

def normalize(s):
    return re.sub(r'\s+', ' ', (s or '').strip().upper())
def strip_zip(s):
    return re.sub(r'\s+\d{5}(?:-\d{4})?\s*$', '', s).strip() if s else ''

# Pre-build normalized indices for fast lookup (zip-stripped too)
opps_by_normalized_pl = {}
opps_by_normalized_no_zip = {}
for pl_name, opps_list in opps_by_pl_name.items():
    n = normalize(pl_name)
    opps_by_normalized_pl[n] = opps_list[0]
    opps_by_normalized_no_zip[strip_zip(n)] = opps_list[0]

def find_opp_for_row(r):
    candidates = [
        r.get('Business Buildings') or '',
        r.get('Updated Business Address') or '',
        r.get('Address') or '',
    ]
    for c in candidates:
        if not c:
            continue
        n = normalize(c)
        if n in opps_by_normalized_pl:
            return opps_by_normalized_pl[n]
        nz = strip_zip(n)
        if nz and nz in opps_by_normalized_no_zip:
            return opps_by_normalized_no_zip[nz]
    return None

skipped = 0
for r in data:
    opp = find_opp_for_row(r)
    if not opp:
        skipped += 1
        continue
    if opp['Id'] in already_imported_opp_ids:
        continue

    # Collect non-empty note fields
    sections = []
    for col, label in NOTE_COLS:
        v = r.get(col)
        if v is not None and str(v).strip() and str(v).strip() not in ('----', '--', '-'):
            sections.append((label, str(v).strip()))
    if not sections:
        continue

    # Build HTML body
    body = f'<p><b>Source:</b> SMB ROE Project xlsx ROE_Tracking sheet</p>'
    body += f'<p><b>Imported:</b> {TS}</p><hr/>'
    for label, val in sections:
        # Escape basic HTML chars
        v_html = val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br/>')
        body += f'<p><b>{label}</b></p><p>{v_html}</p>'

    planned.append({
        'opp_id': opp['Id'],
        'opp_name': opp['Name'],
        'body_html': body,
        'section_count': len(sections),
    })

print(f'\nPlanned notes: {len(planned)}')
print(f'Skipped (no Opp match): {skipped}')
print(f'Already imported: {len(already_imported_opp_ids)}')

# Sample
print(f'\nSample (first 3):')
for p in planned[:3]:
    print(f"  {p['opp_id']}  {p['opp_name'][:55]}  [{p['section_count']} sections]")
    # Preview first 200 chars of body
    preview = p['body_html'][:300].replace('<', ' <')
    print(f"    body preview: {preview}...")

if not APPLY:
    print(f"\n[Preview only — re-run with --apply to create {len(planned)} ContentNotes]")
    sys.exit(0)

# Apply: create ContentNote + ContentDocumentLink per planned record
print(f'\nApplying {len(planned)} note creates...')
audit_rows = []
errors = []
for i, p in enumerate(planned):
    try:
        # Create ContentNote
        b64 = base64.b64encode(p['body_html'].encode('utf-8')).decode('utf-8')
        note = sf.ContentNote.create({'Title': NOTE_TITLE, 'Content': b64})
        note_id = note['id']
        # Get ContentDocumentId
        cdid_q = sf.query(f"SELECT LatestPublishedVersion.ContentDocumentId FROM ContentNote WHERE Id='{note_id}'")
        cdid = cdid_q['records'][0]['LatestPublishedVersion']['ContentDocumentId']
        # Link to Opp
        sf.ContentDocumentLink.create({
            'ContentDocumentId': cdid,
            'LinkedEntityId': p['opp_id'],
            'ShareType': 'V',
            'Visibility': 'AllUsers',
        })
        audit_rows.append({
            'SF_Id': p['opp_id'], 'Name': p['opp_name'],
            'Field': '(ContentNote)', 'Before': '', 'After': note_id,
            'Source': 'import_smb_roe_notes_2026-04-25.py',
            'Timestamp': TS, 'Action': 'CREATE',
            'Note': f'Imported note with {p["section_count"]} sections from SMB ROE Project xlsx',
        })
        if (i + 1) % 50 == 0:
            print(f"  {i+1}/{len(planned)} done...")
    except Exception as e:
        errors.append((p['opp_id'], p['opp_name'], str(e)))
        print(f"  ⚠ {p['opp_id']} ({p['opp_name'][:40]}): {e}")

print(f"\n  ✓ Notes created: {len(audit_rows)}")
print(f"  ⚠ Errors: {len(errors)}")

audit_path = AUDIT_DIR / f'smb_roe_notes_audit_{TS.replace(":","-")}.csv'
with audit_path.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['SF_Id','Name','Field','Before','After','Source','Timestamp','Action','Note'])
    w.writeheader()
    w.writerows(audit_rows)
print(f"  ✓ Audit log: {audit_path}")
