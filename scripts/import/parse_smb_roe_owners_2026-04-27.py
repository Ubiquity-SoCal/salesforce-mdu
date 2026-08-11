"""
Parse SMB ROE Project xlsx Management Company + Owner/Property Contact columns
into Account / Contact / Opportunity_Contact__c records and link to the 245
Business_ROE Opps backfilled 2026-04-25.

Strategy
========
- Account.Type fits existing picklist values: 'Management Company', 'Investor',
  'REIT', 'Law Firm', 'Other'.
- Capture ALL contacts found in each cell (multi-contact rows -> multiple
  Opportunity_Contact__c junction rows).
- Lossless backstop: raw original text from both columns is concatenated into
  Opportunity.Description so nothing is dropped if the parser misses something.
- Dedup Accounts by normalized name (exact only). Dedup Contacts by Email if
  present, else (lower(LastName) + Phone).

Roles assigned to Opportunity_Contact__c.Role__c:
  Management Company column            -> 'Property Manager'
  Owner column (default)                -> 'Property Owner'
  Cell starts with 'leasing'            -> 'Leasing Contact'
  Cell mentions 'POA' / 'HOA' / 'Owners Association' -> 'HOA Contact'
  Cell mentions 'broker'                -> 'Broker'

Usage:
  python parse_smb_roe_owners_2026-04-27.py            # preview, no writes
  python parse_smb_roe_owners_2026-04-27.py --apply    # executes + audit log
  python parse_smb_roe_owners_2026-04-27.py --sample 5 # show full payload for N rows
"""
import sys, io, re, csv, argparse, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime
from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


ap = argparse.ArgumentParser()
ap.add_argument('--apply', action='store_true')
ap.add_argument('--sample', type=int, default=10, help='Number of rows to display in detail')
ap.add_argument('--dump-plan', type=str, default=None,
                help='Write the planned state (new_contacts + junctions) to a JSON file and exit')
args = ap.parse_args()
APPLY = args.apply

XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
SHEET = 'ROE_Tracking'

sf = Salesforce(username=_SF["username"], password=_SF["password"], security_token=_SF["token"])

AUDIT_DIR = Path(r'C:\Users\cass\Work_Projects\SalesForce\audit_logs')
AUDIT_DIR.mkdir(parents=True, exist_ok=True)
TS = datetime.now().isoformat(timespec='seconds')
SCRIPT_NAME = 'parse_smb_roe_owners_2026-04-27.py'

# ── regex toolkit ────────────────────────────────────────────────────────────
PHONE_RE = re.compile(r'\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}')
EMAIL_RE = re.compile(r'[\w.+\-]+@[\w\-]+\.[\w\-.]+')
DATE_PARENS_RE = re.compile(r'\(\d{1,2}/\d{1,4}\)')
URL_RE = re.compile(r'\bwww\.[\w.\-]+|\b[\w\-]+\.(?:com|net|org|us|co|edu|info)\b', re.I)
COMPANY_SUFFIX_RE = re.compile(r'\b(LLC|L\.L\.C\.|Inc|Inc\.|Corp|Corporation|Ltd|Co\.|Company|Trust|LP|L\.P\.|Partners|Properties|Investments|Capital|Group|Holdings|Realty|Real Estate|Management|Mgmt)\b', re.I)
# Words that indicate a phrase is a company, not a person, even without LLC/Inc suffix
COMPANY_HINT_WORDS = re.compile(r'\b(LLC|L\.L\.C\.|Inc|Corp|Corporation|Ltd|Co\.|Company|Trust|LP|L\.P\.|Partners|Properties|Investments|Capital|Group|Holdings|Realty|Real Estate|Management|Mgmt|Realtors|Bank|Insurance|REIT|Foundation|Estate|Associates|Associate|Assoc|Industrial|Commercial|Hospitality|Hotels|Funds|Equity|Bancshares|Realty|Ventures|Plaza|Center|Centers|Mall|Square|Plaza|Tower|Towers|Building|Buildings|Office)\b', re.I)
# Allow leading digit (e.g. "1855 Baseline LLC", "990 Highland LLC") as well as letters.
# Prefer endings: LLC/Inc/Corp/Ltd/LP first; only fall back to Partners/Properties/etc if no stricter match
STRICT_LLC_RE = re.compile(r'((?:[A-Z]|\d)[\w&\.\-/ ]{2,100}?\s+(?:LLC|L\.L\.C\.|Inc\.?|Corp\.?|Corporation|Ltd\.?|LP|L\.P\.))\b(?!\w)')
LOOSE_LLC_RE = re.compile(r'([A-Z][\w&\.\-/ ]{2,100}?\s+(?:Trust|Partners|Properties|Investments|Capital|Group|Holdings|Realty|Real Estate|Foundation|Estate|Associates|Bank|Bancshares))\b(?!\w)')
# Generic noise that should never become an Account on its own
GARBAGE_ACCT_NAMES = {
    'owners association', 'owner association', 'property owners association',
    'common parcel', 'association', 'tenants', 'owners',
}
# A person: 2-3 capitalized words optionally with middle initial. Stop on lowercase, digits, or special.
PERSON_NEAR_CONTACT_RE = re.compile(
    r'((?:[A-Z][a-z]+|[A-Z]\.)\s+(?:[A-Z]\.\s+)?[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)'
    r'(?=[\s,]*(?:\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}|[\w.+\-]+@))'
)
# Person before a phone/email when separated by misc text
COMMA_PERSON_RE = re.compile(r',\s*((?:[A-Z][a-z]+\s+)+[A-Z][a-z]+)\b')

# Order alternation longest-first so "leasing & management" wins over "leasing"
LEASING_PREFIX_RE = re.compile(r'^\s*(?:\([\d/]+\)\s*)?(?:leasing\s*&\s*management|leasing\s*only|leasing\s*&|leasing)\s*[:.,]?\s*', re.I)
PM_PREFIX_RE = re.compile(r'^\s*(?:\([\d/]+\)\s*)?(?:property\s+management|property\s+manager|management|PM)\s*[:.,]?\s*', re.I)
# Word-bound 'owner' so we don't strip "Owners Association" or "Owners' Association"
OWNER_PREFIX_RE = re.compile(r"^\s*owner(?![a-z'])(?:\s*\([\d/]+\))?\s*[:.,]?\s*", re.I)
DATE_PREFIX_RE = re.compile(r'^\s*\(\d{1,2}/\d{1,4}\)\s*[:.,]?\s*')
CO_RE = re.compile(r'\bc\s*/\s*o\b', re.I)
HOA_RE = re.compile(r'\b(HOA|POA|Owners Association|Property Owners Association|homeowners association)\b', re.I)
BROKER_RE = re.compile(r'\bbroker\b', re.I)
TRUST_RE = re.compile(r'\bTrust\b')
REIT_RE = re.compile(r'\bREIT\b', re.I)
LAW_RE = re.compile(r'\b(law firm|attorney|esq\.?)\b', re.I)


def normalize_phone(s):
    digits = re.sub(r'\D', '', s)
    if len(digits) == 11 and digits.startswith('1'):
        digits = digits[1:]
    if len(digits) == 10:
        return f'({digits[0:3]}) {digits[3:6]}-{digits[6:10]}'
    return s.strip()


def normalize_account_name(name):
    n = (name or '').strip().lower()
    n = re.sub(r'[\.,]', '', n)
    n = re.sub(r'\s+', ' ', n)
    return n


def clean_company_name(s):
    """Trim trailing/leading noise from a candidate company name."""
    if not s:
        return s
    s = s.strip()
    # Strip trailing punctuation
    s = re.sub(r'[,.\s]+$', '', s)
    # Drop trailing "for common parcel" / "owns common parcel" etc
    s = re.sub(r'\b(owns|for)\s+common\s+parcel\b.*', '', s, flags=re.I).strip()
    return s


def extract_company_name(text, owner_column=False):
    """Pull the most likely company name out of a cell. Returns (name, account_type)."""
    if not text:
        return None, None
    s = str(text).strip()
    # strip leading prefixes
    s = LEASING_PREFIX_RE.sub('', s)
    s = OWNER_PREFIX_RE.sub('', s)
    s = DATE_PREFIX_RE.sub('', s)
    s = PM_PREFIX_RE.sub('', s)
    s = re.sub(r'^\s*(\d{1,2}/\d{1,4})[:\.,]\s*', '', s)  # leading date "5/2021:"

    # Multi-line: take first non-empty line that has substantive text
    first_chunk = next((line for line in s.splitlines() if line.strip()), s)

    # Try strict LLC/Inc/Corp first; loose Partners/Properties/Investments only as fallback.
    m = STRICT_LLC_RE.search(first_chunk) or LOOSE_LLC_RE.search(first_chunk)
    if m:
        name = clean_company_name(m.group(1))
        atype = 'Investor' if owner_column else 'Management Company'
        # Trust override
        if TRUST_RE.search(name):
            atype = 'Other'
        if REIT_RE.search(first_chunk):
            atype = 'REIT'
        if LAW_RE.search(first_chunk):
            atype = 'Law Firm'
        return name, atype

    # Trust / Family Trust / Foundation pattern
    trust_m = re.search(r'([A-Z][\w& ]{2,60}\s+(?:Family\s+)?(?:Trust|Foundation|Estate))\b', first_chunk)
    if trust_m:
        return clean_company_name(trust_m.group(1)), 'Other'

    # No suffix; take the words BEFORE the first phone/email/url as the name
    cut_at = len(first_chunk)
    for rgx in (PHONE_RE, EMAIL_RE, URL_RE):
        m2 = rgx.search(first_chunk)
        if m2 and m2.start() < cut_at:
            cut_at = m2.start()
    head = first_chunk[:cut_at].strip(' ,.')
    # Strip trailing connectors
    head = re.sub(r'\b(and|c/o|c\s*/\s*o)\s*$', '', head, flags=re.I).strip(' ,.')

    if not head or len(head) < 3:
        return None, None
    # Reject only if head looks like a pure person name AND has no business hint words
    is_person_shape = bool(re.fullmatch(r'(?:[A-Z][a-z]+\s+){1,2}[A-Z][a-z]+', head))
    has_business_hint = bool(COMPANY_HINT_WORDS.search(head))
    if is_person_shape and not has_business_hint:
        return None, None
    # Collapse extra noise
    head = clean_company_name(head)
    if not head:
        return None, None
    atype = 'Investor' if owner_column else 'Management Company'
    if HOA_RE.search(first_chunk):
        atype = 'Other'
    elif REIT_RE.search(first_chunk):
        atype = 'REIT'
    elif LAW_RE.search(first_chunk):
        atype = 'Law Firm'
    return head, atype


NAME_CHUNK_RE = re.compile(r'(?:(?<=^)|(?<=[\s,;\.\-:/]))((?:[A-Z][a-z]+\s+)?[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)(?=\W|$)')

# Reject if any of these tokens appears in a candidate person-name — they indicate
# a company-suffix word, role title, or honorific, not a person.
PERSON_REJECT_WORDS = {
    'llc', 'inc', 'corp', 'corporation', 'ltd', 'lp', 'company', 'co', 'companies',
    'partners', 'partnership', 'properties', 'property', 'investments', 'investment',
    'capital', 'group', 'holdings', 'realty', 'estate', 'management', 'mgmt',
    'manager', 'managers', 'leasing', 'broker', 'brokerage', 'agent', 'realtor',
    'realtors', 'associates', 'associate', 'assoc', 'foundation', 'trust',
    'industrial', 'commercial', 'hospitality', 'hotels', 'bank', 'bancshares',
    'ventures', 'plaza', 'center', 'centers', 'mall', 'square', 'tower', 'towers',
    'building', 'buildings', 'office', 'offices', 'church', 'ministry',
    'president', 'director', 'officer', 'manager', 'principal', 'partner',
    'owner', 'owners', 'association', 'associations', 'assn', 'lord',
    'business', 'enterprises', 'enterprise', 'services', 'service', 'solutions',
    'school', 'school', 'unified', 'district', 'medical', 'institute',
    'mesa', 'mesaaz', 'corp', 'group', 'realty',
}


def looks_like_person(text):
    """Check if a 2-3 word capitalized phrase is likely a person name (not a company)."""
    if not text:
        return False
    words = [w.strip(",.").lower() for w in text.split() if w.strip(",.")]
    if not (2 <= len(words) <= 4):
        return False
    if any(w in PERSON_REJECT_WORDS for w in words):
        return False
    # Reject if all words are short uppercase ALL-CAPS (likely acronym)
    return True


def extract_contacts(text, role_default):
    """
    Emit zero-or-more contacts. Anchored on emails (each email -> a Contact when we
    can find a plausible person name nearby), then phones with names before them.
    """
    if not text:
        return []
    s = str(text)
    contacts = []
    seen = set()

    # ── 1. Emails ──
    for em in EMAIL_RE.finditer(s):
        email = em.group(0).rstrip('.,;')
        # Look back 80 chars for a capitalized name BEFORE the email
        window = s[max(0, em.start() - 80):em.start()]
        candidates = list(NAME_CHUNK_RE.finditer(window))
        name = None
        # Walk right-to-left, take first that looks like a person
        for c in reversed(candidates):
            cand = c.group(1).strip()
            if looks_like_person(cand):
                name = cand
                break
        # Phone nearby (within 60 chars before/after)
        ph_m = PHONE_RE.search(s[max(0, em.start()-60): em.end()+60])
        phone = normalize_phone(ph_m.group(0)) if ph_m else None
        # Fallback name from email local-part if first.last shape
        if not name:
            local = email.split('@')[0]
            if '.' in local:
                parts = local.split('.')
                if len(parts) >= 2 and all(p.isalpha() and len(p) > 1 for p in parts[:2]):
                    name = parts[0].capitalize() + ' ' + parts[1].capitalize()
        if not name:
            # Skip rather than create a junk Contact like "info@..." with no name
            continue
        ws = name.split()
        first, last = (ws[0], ws[-1]) if len(ws) >= 2 else (None, ws[0])
        key = (last.lower(), email.lower())
        if key in seen:
            continue
        seen.add(key)
        contacts.append({
            'FirstName': first,
            'LastName': last,
            'Phone': phone,
            'Email': email,
        })

    # ── 2. Phones with names before them, not already captured ──
    used_phones = {c['Phone'] for c in contacts if c.get('Phone')}
    for ph in PHONE_RE.finditer(s):
        phone_norm = normalize_phone(ph.group(0))
        if phone_norm in used_phones:
            continue
        window = s[max(0, ph.start()-60):ph.start()]
        candidates = list(NAME_CHUNK_RE.finditer(window))
        name = None
        for c in reversed(candidates):
            cand = c.group(1).strip()
            if looks_like_person(cand):
                name = cand
                break
        if not name:
            continue  # skip lonely phones — usually the company main line
        ws = name.split()
        first, last = (ws[0], ws[-1]) if len(ws) >= 2 else (None, ws[0])
        key = (last.lower(), phone_norm)
        if key in seen:
            continue
        seen.add(key)
        used_phones.add(phone_norm)
        contacts.append({
            'FirstName': first,
            'LastName': last,
            'Phone': phone_norm,
            'Email': None,
        })

    return contacts


def detect_role(text, default):
    """Refine role based on prefix words."""
    if not text:
        return default
    s = text.strip()
    if HOA_RE.search(s):
        return 'HOA Contact'
    if BROKER_RE.search(s):
        return 'Broker'
    if LEASING_PREFIX_RE.match(s) or re.match(r'\s*leasing\b', s, flags=re.I):
        return 'Leasing Contact'
    return default


# ── Load ─────────────────────────────────────────────────────────────────────
print(f"[Load] {XLSX} sheet={SHEET}")
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[SHEET]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = []
for r in rows[1:]:
    if not any(c is not None for c in r):
        continue
    rec = dict(zip(header, r))
    data.append(rec)
print(f"  Loaded {len(data)} data rows")

# ── Pull existing Accounts + 245 Opps ───────────────────────────────────────
print("\n[SF] Pulling existing Accounts and target Opps")
existing_accts = sf.query_all("SELECT Id, Name, Type FROM Account")['records']
acct_by_norm = {}
for a in existing_accts:
    acct_by_norm[normalize_account_name(a['Name'])] = a
print(f"  Existing Accounts: {len(existing_accts)}")

opps = sf.query_all("""
  SELECT Id, Name, AccountId, Management_Company__c, Property_Location__c,
         Property_Location__r.Name, Description, Closed_Notes__c
  FROM Opportunity
  WHERE RecordType.DeveloperName='Business_ROE' AND Name LIKE 'ROE - %'
""")['records']
print(f"  Business_ROE Opps to enrich: {len(opps)}")

# Build an index: Property_Location.Name -> Opp record (one Opp per PL for these 245)
opp_by_pl_name = {}
for o in opps:
    pl = o.get('Property_Location__r') or {}
    if pl.get('Name'):
        opp_by_pl_name[pl['Name']] = o

# Map xlsx rows to Opps via matching logic from the backfill script
# Re-use Property_Location lookups built into the backfill: match on `Business Buildings`
pl_lookup = {}  # PL Name -> Opp
states_in_xlsx = set(r.get('State') for r in data if r.get('State'))
state_clause = "','".join(s for s in states_in_xlsx if s)
all_pls = sf.query_all(f"SELECT Id, Name, Business_Base_Address__c, Business_Building_Id__c FROM Property_Location__c WHERE State__c IN ('{state_clause}')")['records']
print(f"  Pulled {len(all_pls)} Property_Locations in target states")

UUID_RE = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
ZIP_SUFFIX_RE = re.compile(r'\s+\d{5}(?:-\d{4})?\s*$')


def normalize_addr(s):
    if not s:
        return ''
    return re.sub(r'\s+', ' ', s.strip().upper())


def strip_zip(s):
    return ZIP_SUFFIX_RE.sub('', s).strip() if s else ''


pl_by_name, pl_by_norm, pl_by_norm_no_zip, pl_by_bid = {}, {}, {}, {}
for pl in all_pls:
    if pl.get('Name'):
        pl_by_name[pl['Name']] = pl
        pl_by_norm.setdefault(normalize_addr(pl['Name']), pl)
        pl_by_norm_no_zip.setdefault(strip_zip(normalize_addr(pl['Name'])), pl)
    if pl.get('Business_Base_Address__c'):
        pl_by_norm.setdefault(normalize_addr(pl['Business_Base_Address__c']), pl)
        pl_by_norm_no_zip.setdefault(strip_zip(normalize_addr(pl['Business_Base_Address__c'])), pl)
    if pl.get('Business_Building_Id__c'):
        pl_by_bid[pl['Business_Building_Id__c']] = pl


def find_opp(r):
    bn = r.get('Business Buildings') or ''
    pl = None
    if bn in pl_by_name:
        pl = pl_by_name[bn]
    else:
        norm = normalize_addr(bn)
        pl = pl_by_norm.get(norm) or pl_by_norm_no_zip.get(strip_zip(norm))
        if not pl and UUID_RE.match(bn):
            pl = pl_by_bid.get(bn)
            if not pl:
                uba = r.get('Updated Business Address') or ''
                pl = pl_by_norm.get(normalize_addr(uba)) or pl_by_norm_no_zip.get(strip_zip(normalize_addr(uba)))
    if not pl:
        return None, None
    return pl, opp_by_pl_name.get(pl['Name'])


# ── Parse rows ──────────────────────────────────────────────────────────────
print("\n[Parse] Analyzing each row")

# Bookkeeping
new_accts = {}        # normalized name -> {Name, Type, Phone, Description, _source}
new_contacts = {}     # dedup_key -> {FirstName, LastName, Phone, Email, AccountId placeholder, _role}
opp_updates = []      # [(opp_id, {AccountId, Management_Company__c, Description})]
junctions = []        # [{Opportunity__c, Contact_dedup_key, Role__c}]
unmatched_rows = []
parse_log = []        # for sample display


def get_or_create_acct(name, atype, phone=None, desc_extra=None, website=None):
    if not name:
        return None
    norm = normalize_account_name(name)
    if not norm:
        return None
    # Reject generic garbage names
    if norm in GARBAGE_ACCT_NAMES:
        return None
    # Existing first
    if norm in acct_by_norm:
        return ('existing', acct_by_norm[norm]['Id'], acct_by_norm[norm]['Name'])
    # Already-proposed in this batch
    if norm in new_accts:
        return ('new', norm, new_accts[norm]['Name'])
    new_accts[norm] = {
        'Name': name,
        'Type': atype,
        'Phone': phone,
        'Description': desc_extra,
        'Website': website,
    }
    return ('new', norm, name)


def get_or_create_contact(c, acct_ref, role):
    if not c:
        return None
    fn, ln, ph, em = c.get('FirstName'), c.get('LastName'), c.get('Phone'), c.get('Email')
    if not (ln and (ph or em)):
        return None
    # Dedup: email primary, else (lower(ln) + phone)
    key = ('email:' + em.lower()) if em else ('np:' + (ln or '').lower() + '|' + (ph or ''))
    if key in new_contacts:
        # Already proposed; merge phone/email if missing
        existing = new_contacts[key]
        if ph and not existing.get('Phone'):
            existing['Phone'] = ph
        if em and not existing.get('Email'):
            existing['Email'] = em
        if fn and not existing.get('FirstName'):
            existing['FirstName'] = fn
        return key
    new_contacts[key] = {
        'FirstName': fn,
        'LastName': ln,
        'Phone': ph,
        'Email': em,
        '_acct_ref': acct_ref,
        '_role': role,
    }
    return key


for r in data:
    pl, opp = find_opp(r)
    if not opp:
        unmatched_rows.append(r.get('Business Buildings'))
        continue

    mc_text = r.get('Management Company') or ''
    own_text = r.get('Owner/Property Contact') or ''

    # Parse MC column
    mc_acct_ref = None
    if mc_text and str(mc_text).strip():
        mc_name, mc_type = extract_company_name(mc_text, owner_column=False)
        if mc_name:
            ph_match = PHONE_RE.search(str(mc_text))
            mc_phone = normalize_phone(ph_match.group(0)) if ph_match else None
            url_m = URL_RE.search(str(mc_text))
            mc_website = url_m.group(0) if url_m else None
            if mc_website and not mc_website.startswith('http'):
                mc_website = 'https://' + mc_website
            mc_acct_ref = get_or_create_acct(mc_name, mc_type or 'Management Company',
                                             phone=mc_phone,
                                             desc_extra=str(mc_text)[:255],
                                             website=mc_website)

    # Parse Owner column
    own_acct_ref = None
    if own_text and str(own_text).strip():
        own_name, own_type = extract_company_name(own_text, owner_column=True)
        if own_name:
            ph_match = PHONE_RE.search(str(own_text))
            own_phone = normalize_phone(ph_match.group(0)) if ph_match else None
            url_m = URL_RE.search(str(own_text))
            own_website = url_m.group(0) if url_m else None
            if own_website and not own_website.startswith('http'):
                own_website = 'https://' + own_website
            own_acct_ref = get_or_create_acct(own_name, own_type or 'Investor',
                                              phone=own_phone,
                                              desc_extra=str(own_text)[:255],
                                              website=own_website)

    # Contacts from MC column
    mc_contacts_added = []
    if mc_text:
        for c in extract_contacts(str(mc_text), role_default='Property Manager'):
            role = detect_role(str(mc_text), 'Property Manager')
            ck = get_or_create_contact(c, mc_acct_ref, role)
            if ck:
                mc_contacts_added.append((ck, role))

    # Contacts from Owner column
    own_contacts_added = []
    if own_text:
        for c in extract_contacts(str(own_text), role_default='Property Owner'):
            role = detect_role(str(own_text), 'Property Owner')
            ck = get_or_create_contact(c, own_acct_ref, role)
            if ck:
                own_contacts_added.append((ck, role))

    # Junction rows for this Opp
    for ck, role in mc_contacts_added + own_contacts_added:
        junctions.append({
            'Opportunity__c': opp['Id'],
            '_contact_key': ck,
            'Role__c': role,
        })

    # Build Opp Description (lossless backstop) and links
    desc_parts = []
    if mc_text:
        desc_parts.append(f'[Management Company]\n{mc_text}')
    if own_text:
        desc_parts.append(f'[Owner/Property Contact]\n{own_text}')
    new_desc = ('\n\n'.join(desc_parts))[:32000] if desc_parts else None

    upd = {}
    if own_acct_ref:
        upd['_AccountId_ref'] = own_acct_ref
    if mc_acct_ref:
        upd['_MC_ref'] = mc_acct_ref
    if new_desc:
        upd['Description'] = new_desc
    if upd:
        opp_updates.append((opp['Id'], opp['Name'], upd))

    parse_log.append({
        'Opp': opp['Name'],
        'mc_text': str(mc_text)[:200],
        'own_text': str(own_text)[:200],
        'mc_acct': mc_acct_ref,
        'own_acct': own_acct_ref,
        'mc_contacts': [c[0] for c in mc_contacts_added],
        'own_contacts': [c[0] for c in own_contacts_added],
    })


# ── Summary ─────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("PARSE PLAN SUMMARY")
print("=" * 70)
print(f"  xlsx rows:                 {len(data)}")
print(f"  unmatched (no Opp):        {len(unmatched_rows)}")
print(f"  Opps targeted for update:  {len(opp_updates)}")
print(f"  NEW Accounts (proposed):   {len(new_accts)}")
print(f"  NEW Contacts (proposed):   {len(new_contacts)}")
print(f"  Junction rows (proposed):  {len(junctions)}")

# Account.Type breakdown
type_dist = Counter(a['Type'] for a in new_accts.values())
print(f"\n  New Account Type distribution:")
for t, n in type_dist.most_common():
    print(f"    {n:4d}  {t}")

# Account-existing matches
mc_matches = sum(1 for r in opp_updates if r[2].get('_MC_ref') and r[2]['_MC_ref'][0] == 'existing')
own_matches = sum(1 for r in opp_updates if r[2].get('_AccountId_ref') and r[2]['_AccountId_ref'][0] == 'existing')
print(f"\n  Linked to EXISTING Account:")
print(f"    {mc_matches:4d}  MC links")
print(f"    {own_matches:4d}  Owner links")

# Rows with no extracted MC and no extracted Owner — needs human review
no_anything = sum(1 for r in opp_updates if not r[2].get('_MC_ref') and not r[2].get('_AccountId_ref'))
print(f"\n  Opps with NO extractable Account (Description still saved): {no_anything}")

# Show first N samples
print(f"\n  --- First {args.sample} parses ---")
for log in parse_log[:args.sample]:
    print(f"\n  Opp: {log['Opp']}")
    print(f"    MC raw:   {log['mc_text']}")
    print(f"    Own raw:  {log['own_text']}")
    if log['mc_acct']:
        print(f"    -> MC Acct: ({log['mc_acct'][0]}) {log['mc_acct'][2]}")
    if log['own_acct']:
        print(f"    -> Own Acct: ({log['own_acct'][0]}) {log['own_acct'][2]}")
    print(f"    -> MC Contacts: {len(log['mc_contacts'])}")
    print(f"    -> Own Contacts: {len(log['own_contacts'])}")

# Show contact dedup samples
print(f"\n  --- Sample new Contacts (first 12) ---")
for i, (k, c) in enumerate(list(new_contacts.items())[:12]):
    print(f"    [{i+1}] {c.get('FirstName') or ''} {c.get('LastName') or ''}  "
          f"phone={c.get('Phone') or ''}  email={c.get('Email') or ''}  role={c['_role']}")

# Show new Account samples
print(f"\n  --- Sample new Accounts (first 12) ---")
for i, (n, a) in enumerate(list(new_accts.items())[:12]):
    print(f"    [{i+1}] {a['Name'][:55]:55s}  Type={a['Type']:20s} Phone={a.get('Phone') or ''}")

if args.dump_plan:
    plan = {
        'new_contacts': new_contacts,  # key -> dict (with FirstName/LastName/Phone/Email/_role/_acct_ref)
        'junctions': junctions,         # list of dicts
    }
    # _acct_ref is a tuple ('new'|'existing', ref, name) — JSON-safe
    Path(args.dump_plan).write_text(json.dumps(plan, default=list, indent=2), encoding='utf-8')
    print(f"\n[Plan dumped to {args.dump_plan}]  contacts={len(new_contacts)} junctions={len(junctions)}")
    sys.exit(0)

if not APPLY:
    print(f"\n[Preview only — re-run with --apply to write {len(new_accts)} Accounts, "
          f"{len(new_contacts)} Contacts, {len(opp_updates)} Opp updates, "
          f"{len(junctions)} junctions]")
    sys.exit(0)


# ── Apply ──────────────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("APPLYING")
print("=" * 70)

audit_rows = []

# 1. Insert new Accounts in batches of 200
print(f"\n  [1/4] Creating {len(new_accts)} Accounts")
acct_norm_to_id = {}
for a in existing_accts:
    acct_norm_to_id[normalize_account_name(a['Name'])] = a['Id']

acct_keys = list(new_accts.keys())
for i in range(0, len(acct_keys), 200):
    batch_keys = acct_keys[i:i+200]
    batch = []
    for k in batch_keys:
        a = new_accts[k]
        rec = {'Name': a['Name'], 'Type': a['Type']}
        if a.get('Phone'):
            rec['Phone'] = a['Phone']
        if a.get('Website'):
            rec['Website'] = a['Website']
        if a.get('Description'):
            rec['Description'] = a['Description']
        batch.append(rec)
    print(f"    Batch {i//200 + 1}: {len(batch)} ...")
    results = sf.bulk.Account.insert(batch)
    for j, res in enumerate(results):
        rec = batch[j]
        if res.get('success'):
            acct_norm_to_id[acct_keys[i+j]] = res['id']
            audit_rows.append({
                'SF_Id': res['id'], 'Name': rec['Name'], 'Field': '(created)',
                'Before': '', 'After': f"Account Type={rec['Type']}",
                'Source': SCRIPT_NAME, 'Timestamp': TS, 'Action': 'CREATE',
                'Note': 'Account from SMB ROE owner/MC parse',
            })
        else:
            print(f"    ⚠ FAILED Account: {rec['Name']} — {res.get('errors', res)}")

# 2. Insert new Contacts
print(f"\n  [2/4] Creating {len(new_contacts)} Contacts")
ct_keys = list(new_contacts.keys())
ct_key_to_id = {}
for i in range(0, len(ct_keys), 200):
    batch_keys = ct_keys[i:i+200]
    batch = []
    for k in batch_keys:
        c = new_contacts[k]
        rec = {
            'FirstName': c.get('FirstName') or '',
            'LastName': c.get('LastName') or 'Unknown',
        }
        if c.get('Phone'):
            rec['Phone'] = c['Phone']
        if c.get('Email'):
            rec['Email'] = c['Email']
        # Resolve AccountId from ref tuple
        ar = c.get('_acct_ref')
        if ar:
            kind, ref, _ = ar
            aid = acct_norm_to_id.get(ref) if kind == 'new' else ref
            if aid:
                rec['AccountId'] = aid
        batch.append(rec)
    print(f"    Batch {i//200 + 1}: {len(batch)} ...")
    results = sf.bulk.Contact.insert(batch)
    for j, res in enumerate(results):
        rec = batch[j]
        if res.get('success'):
            ct_key_to_id[batch_keys[j]] = res['id']
            audit_rows.append({
                'SF_Id': res['id'],
                'Name': f"{rec.get('FirstName','')} {rec.get('LastName','')}".strip(),
                'Field': '(created)', 'Before': '', 'After': 'Contact',
                'Source': SCRIPT_NAME, 'Timestamp': TS, 'Action': 'CREATE',
                'Note': f"AccountId={rec.get('AccountId') or '(none)'}; Phone={rec.get('Phone') or ''}; Email={rec.get('Email') or ''}",
            })
        else:
            print(f"    ⚠ FAILED Contact: {rec.get('LastName')} — {res.get('errors', res)}")

# 3. Update Opps
print(f"\n  [3/4] Updating {len(opp_updates)} Opps with AccountId / MC / Description")
opp_batch = []
for opp_id, opp_name, upd in opp_updates:
    rec = {'Id': opp_id}
    if upd.get('Description'):
        rec['Description'] = upd['Description']
    ar = upd.get('_AccountId_ref')
    if ar:
        kind, ref, _ = ar
        aid = acct_norm_to_id.get(ref) if kind == 'new' else ref
        if aid:
            rec['AccountId'] = aid
    mr = upd.get('_MC_ref')
    if mr:
        kind, ref, _ = mr
        mid = acct_norm_to_id.get(ref) if kind == 'new' else ref
        if mid:
            rec['Management_Company__c'] = mid
    if len(rec) > 1:
        opp_batch.append((opp_id, opp_name, rec))

for i in range(0, len(opp_batch), 200):
    batch_records = [r for _, _, r in opp_batch[i:i+200]]
    print(f"    Batch {i//200 + 1}: {len(batch_records)} ...")
    results = sf.bulk.Opportunity.update(batch_records)
    for j, res in enumerate(results):
        opp_id, opp_name, rec = opp_batch[i+j]
        if res.get('success'):
            for f, v in rec.items():
                if f == 'Id':
                    continue
                audit_rows.append({
                    'SF_Id': opp_id, 'Name': opp_name, 'Field': f,
                    'Before': '(null)', 'After': str(v)[:200],
                    'Source': SCRIPT_NAME, 'Timestamp': TS, 'Action': 'UPDATE',
                    'Note': 'SMB ROE owner/MC parse',
                })
        else:
            print(f"    ⚠ FAILED Opp update: {opp_name} — {res.get('errors', res)}")

# 4. Insert Junctions
print(f"\n  [4/4] Creating {len(junctions)} Opportunity_Contact__c rows")
junc_batch = []
for j_rec in junctions:
    cid = ct_key_to_id.get(j_rec['_contact_key'])
    if not cid:
        continue
    junc_batch.append({
        'Opportunity__c': j_rec['Opportunity__c'],
        'Contact__c': cid,
        'Role__c': j_rec['Role__c'],
    })

for i in range(0, len(junc_batch), 200):
    batch = junc_batch[i:i+200]
    print(f"    Batch {i//200 + 1}: {len(batch)} ...")
    results = sf.bulk.Opportunity_Contact__c.insert(batch)
    for j, res in enumerate(results):
        rec = batch[j]
        if res.get('success'):
            audit_rows.append({
                'SF_Id': res['id'], 'Name': '(junction)', 'Field': '(created)',
                'Before': '', 'After': f"Opp={rec['Opportunity__c']} Contact={rec['Contact__c']} Role={rec['Role__c']}",
                'Source': SCRIPT_NAME, 'Timestamp': TS, 'Action': 'CREATE',
                'Note': 'Opportunity_Contact__c junction',
            })
        else:
            print(f"    ⚠ FAILED junction: {rec} — {res.get('errors', res)}")

# Audit log
audit_path = AUDIT_DIR / f'smb_roe_owners_audit_{TS.replace(":","-")}.csv'
with audit_path.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['SF_Id','Name','Field','Before','After','Source','Timestamp','Action','Note'])
    w.writeheader()
    w.writerows(audit_rows)

print(f"\n✓ Done. Audit log: {audit_path} ({len(audit_rows)} rows)")
