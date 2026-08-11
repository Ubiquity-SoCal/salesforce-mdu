"""
CA MDU Agreement Status + Opportunities_Prospects_RE -> Salesforce (DRY RUN PREVIEW)

Source of truth:
  CA MDU Agreement Status 04172026.xlsx  -- current Status, On Net?, ROE/PAL, ISP Tenant, Units
  Opportunities_Prospects_RE (6).xlsx    -- ONLY opportunities_market sheet: RE Assigned, Contacts,
                                            Comments (notes), PAL Signed Date, SiteTracker link,
                                            CX Estimate/Notes, Vetro status

Output: SalesForce/CA_MDU_Merge/preview.xlsx
  Tabs:
    Merged    - every property from either input, side-by-side w/ SF match + proposed action
    Create    - rows we'd net-new
    Update    - rows we'd update on existing SF Opportunity
    NoAction  - rows already aligned / nothing to do
    Issues    - rows with problems (multi-match, missing required, etc.)
    Schema    - fields we need to add + picklist values to extend
"""

import re
import sys
from collections import defaultdict, Counter
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from simple_salesforce import Salesforce
import os as _os

# ── Config ──────────────────────────────────────────────────────────────────
CA_MDU_XLSX = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\CA MDU Agreement Status 04172026.xlsx"
RE_XLSX     = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Opportunities_Prospects_RE (6).xlsx"
OUTPUT_XLSX = r"C:\Users\cass\Work_Projects\SalesForce\CA_MDU_Merge\preview.xlsx"

# Salesforce config -- read from the gitignored SalesForce/api/ creds file.
# Never hardcode the password here: this file is tracked in git.
def _sf_creds():
    _p = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                       "..", "..", "api", "Salesforce_Credentials.txt")
    _c = {}
    with open(_p) as _f:
        for _line in _f:
            if ":" in _line:
                _k, _v = _line.split(":", 1)
                _c[_k.strip()] = _v.strip()
    return _c


_SF = _sf_creds()
SF_USERNAME = _SF["Username"]
SF_PASSWORD = _SF["Password"]
SF_TOKEN = _SF["Security Token"]

JUSTIN_BARRY_ID    = "005WR0000030RCzYAM"
MDU_RECORD_TYPE_ID = "012WR00000Ra0mkYAB"
SFU_RECORD_TYPE_ID = "012WR00000S2ne1YAB"

# CA MDU bucket -> SF StageName
BUCKET_TO_STAGE = {
    "Prospects": "Prospecting",
    "Proposal Sent": "Engaged",
    "On Net - Access Agreement Complete": "ROE Secured",
    "Near Net - Access Agreement Complete": "ROE Secured",
    "ON Air Serviceable": "ROE Secured",  # MDU has no Closed Won
}

# CA Pipeline raw buckets + On Air bucket — go into proposed Pipeline_Bucket__c field
ALL_BUCKETS = [
    "Prospects",
    "Proposal Sent",
    "On Net - Access Agreement Complete",
    "Near Net - Access Agreement Complete",
    "ON Air Serviceable",
]

# SF Stage pipeline order (for regression detection flagging only)
STAGE_ORDER = [
    "Closed Lost", "On Hold", "Prospecting", "Engaged", "ROE Secured",
    "Contract Negotiations", "Under Contract", "Ready for Engineering",
    "Under Construction", "Activation", "Closed Won",
]


# ── Helpers ─────────────────────────────────────────────────────────────────
def norm(s):
    """Lowercase, strip, collapse whitespace, drop punctuation for matching."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).lower().strip()
    s = re.sub(r"[.,#'\"\\/()\\-]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def norm_name(s):
    """Name normalization — drops common suffixes/prefixes that vary between files."""
    n = norm(s)
    # Strip trailing _X markers (closed lost duplicates in SF)
    n = re.sub(r"\s*x\s*$", "", n)
    # Strip leading city prefix pattern from On Air names: "Encinitas_MDU_..."
    n = re.sub(r"^(carlsbad|encinitas|oceanside|solana beach)[\s_]+(mdu|sfu|hoa)[\s_]+", "", n)
    n = re.sub(r"^(mdu|sfu|hoa)[\s_]+", "", n)
    return n.strip()

def norm_addr(s):
    """Address normalization for matching — drops city/state/zip and unit labels."""
    n = norm(s)
    # Drop trailing state abbreviation + zip ("ca 92024")
    n = re.sub(r"\b(ca|california|az|arizona|tx|texas|ne|nebraska)\b\s*\d{0,5}(-\d{4})?\s*$", "", n)
    # Drop city suffixes
    n = re.sub(r"\b(carlsbad|encinitas|oceanside|solana beach|san diego)\b\s*$", "", n)
    # Common abbreviations
    n = n.replace(" street", " st").replace(" road", " rd").replace(" boulevard", " blvd")
    n = n.replace(" drive", " dr").replace(" avenue", " ave").replace(" place", " pl")
    n = n.replace(" north ", " n ").replace(" south ", " s ")
    n = n.replace(" east ", " e ").replace(" west ", " w ")
    n = re.sub(r"\s+", " ", n).strip()
    return n

def first_token(s):
    """Grab leading street number for address bucketing."""
    n = norm_addr(s)
    m = re.match(r"^(\d+)", n)
    return m.group(1) if m else ""

def excel_date(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (datetime, date)):
        try:
            return v.strftime("%Y-%m-%d")
        except (ValueError, AttributeError):
            return None
    try:
        d = pd.to_datetime(v, errors="coerce")
        if pd.isna(d):
            return None
        return d.strftime("%Y-%m-%d")
    except Exception:
        return None

def clean(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


# ── Load inputs ─────────────────────────────────────────────────────────────
def load_ca_mdu():
    """Load CA Pipeline + On Air, add Source column."""
    rows = []
    pipe = pd.read_excel(CA_MDU_XLSX, sheet_name="CA Pipeline")
    for _, r in pipe.iterrows():
        if not clean(r.get("Name")):
            continue
        rows.append({
            "Source": "CA Pipeline",
            "Name": clean(r["Name"]),
            "Bucket": clean(r["Status"]),
            "Units": r.get("Units"),
            "Address": clean(r["Address"]),
            "OnNet": clean(r.get("On Net?")),
            "Program": clean(r.get("Program")),
            "PropertyType": clean(r.get("Property Type")),
            "ISPTenant": clean(r.get("ISP Tenant")),
            "ROE_PAL": clean(r.get("ROE/PAL")),
        })
    on_air = pd.read_excel(CA_MDU_XLSX, sheet_name="On Air")
    for _, r in on_air.iterrows():
        if not clean(r.get("Name")):
            continue
        rows.append({
            "Source": "On Air",
            "Name": clean(r["Name"]),
            "Bucket": clean(r["Status"]) or "ON Air Serviceable",
            "Units": r.get("Units"),
            "Address": clean(r["Address"]),
            "OnNet": clean(r.get("On Net?")),
            "Program": clean(r.get("Program")),
            "PropertyType": clean(r.get("Property Type")),
            "ISPTenant": clean(r.get("ISP Tenant")),
            "ROE_PAL": clean(r.get("ROE/PAL")),
        })
    return rows


def load_market():
    """Load opportunities_market sheet only."""
    df = pd.read_excel(RE_XLSX, sheet_name="opportunities_market")
    rows = []
    for _, r in df.iterrows():
        name = clean(r.get("Name ") or r.get("Name"))
        if not name:
            continue
        rows.append({
            "UniqueID": clean(r.get("UniqueID")),
            "Name": name,
            "RE_Assigned": clean(r.get("RE Assigned")),
            "Status": clean(r.get("Status")),
            "Units": r.get("Units"),
            "Address": clean(r.get("Address")),
            "Team": clean(r.get("Team")),
            "Program": clean(r.get("Program")),
            "PropertyType": clean(r.get("Property Type")),
            "FDHPath": clean(r.get("FDH Path Name")),
            "ProjectID": clean(r.get("Project ID")),
            "Contacts": clean(r.get("Contacts")),
            "Comments": clean(r.get("Comments")),
            "PALSignedDate": excel_date(r.get("PAL Signed Date")),
            "SiteTrackerLink": clean(r.get("Signed File Uploaded link")),
            "CXEstimate": clean(r.get("CX Estimate")),
            "EstimateLink": clean(r.get("Estimate Link")),
            "SiteWalk": excel_date(r.get("Site Walk")),
            "CXNotes": clean(r.get("CX Notes")),
            "FDHActivated": clean(r.get("FDH Activated")),
            "VetroStatus": clean(r.get("Vetro Address Status")),
        })
    return rows


# ── Match logic ─────────────────────────────────────────────────────────────
def build_lookup(rows, keys):
    """Build dict: key -> list of row indexes for each key-producing function."""
    out = {}
    for idx, r in enumerate(rows):
        for k in keys(r):
            if k:
                out.setdefault(k, []).append(idx)
    return out


def match_ca_to_market(ca_rows, market_rows):
    """Match each CA MDU row to a market row. Returns list of (ca_idx -> market_idx or None)."""
    by_addr = build_lookup(market_rows, lambda r: [norm_addr(r["Address"])])
    by_name = build_lookup(market_rows, lambda r: [norm_name(r["Name"])])
    by_start = build_lookup(market_rows, lambda r: [first_token(r["Address"])] if first_token(r["Address"]) else [])

    matches = []
    for ca in ca_rows:
        na = norm_addr(ca["Address"])
        nn = norm_name(ca["Name"])
        hit = None
        # 1. exact address
        if na and na in by_addr and len(by_addr[na]) == 1:
            hit = by_addr[na][0]
        # 2. exact name
        elif nn and nn in by_name and len(by_name[nn]) == 1:
            hit = by_name[nn][0]
        # 3. address starts-with match with street number
        elif na:
            tok = first_token(ca["Address"])
            candidates = by_start.get(tok, [])
            for cidx in candidates:
                if norm_addr(market_rows[cidx]["Address"]).startswith(na[:20]) or na.startswith(norm_addr(market_rows[cidx]["Address"])[:20]):
                    hit = cidx
                    break
        matches.append(hit)
    return matches


def match_to_sf(ca_rows, market_rows, ca_to_market, sf_opps):
    """
    For each merged row, pick the best SF Opp match.
    sf_opps: list of SF records (dicts).
    Returns list of sf_record or None per ca_row.
    """
    sf_by_name = defaultdict(list)
    sf_by_addr = defaultdict(list)
    sf_by_agree = defaultdict(list)
    sf_by_street_num = defaultdict(list)
    for o in sf_opps:
        sf_by_name[norm_name(o.get("Name") or "")].append(o)
        sf_by_addr[norm_addr(o.get("Property_Address__c") or "")].append(o)
        if o.get("Agreement_Name__c"):
            sf_by_agree[norm_name(o["Agreement_Name__c"])].append(o)
        tok = first_token(o.get("Property_Address__c") or "")
        if tok:
            sf_by_street_num[tok].append(o)

    results = []
    for i, ca in enumerate(ca_rows):
        ca_na = norm_addr(ca["Address"])
        ca_nn = norm_name(ca["Name"])

        # Include market match name too if we have one
        mkt_nn = ""
        mkt_na = ""
        m_idx = ca_to_market[i]
        if m_idx is not None:
            mkt_nn = norm_name(market_rows[m_idx]["Name"])
            mkt_na = norm_addr(market_rows[m_idx]["Address"])

        candidates = []

        # 1. Exact address
        for key in (ca_na, mkt_na):
            if key and key in sf_by_addr:
                candidates.extend(sf_by_addr[key])

        # 2. Name (drop _X suffix, _ splits etc)
        for key in (ca_nn, mkt_nn):
            if key and key in sf_by_name:
                candidates.extend(sf_by_name[key])

        # 3. Agreement name
        for key in (ca_nn, mkt_nn):
            if key and key in sf_by_agree:
                candidates.extend(sf_by_agree[key])

        # 4. Street number + fuzzy
        if not candidates:
            tok = first_token(ca["Address"]) or first_token(market_rows[m_idx]["Address"] if m_idx is not None else "")
            if tok:
                for opp in sf_by_street_num.get(tok, []):
                    opp_na = norm_addr(opp.get("Property_Address__c") or "")
                    if ca_na and opp_na and (ca_na.startswith(opp_na[:15]) or opp_na.startswith(ca_na[:15])):
                        candidates.append(opp)

        # De-dup and prefer non-Closed-Lost
        seen = set()
        unique = []
        for o in candidates:
            if o["Id"] in seen:
                continue
            seen.add(o["Id"])
            unique.append(o)
        # Prefer Open stages
        unique.sort(key=lambda o: (1 if (o.get("StageName") or "").startswith("Closed") else 0, o["Name"]))

        if not unique:
            results.append(None)
        elif len(unique) == 1:
            results.append(unique[0])
        else:
            # Multiple candidates -- flag and take the first non-closed
            results.append({"_multi": True, "candidates": unique, **unique[0]})
    return results


# ── Build merged/proposed record ────────────────────────────────────────────
def propose_record(ca, mkt, sf_opp):
    """Build the proposed action + merged fields for the preview.

    Stage rule (per Koa 2026-04-21):
      - If ROE/PAL column shows PAL → keep SF stage (PAL = Under Contract tier)
      - Otherwise apply the CA MDU bucket → SF stage mapping (including regressions)
    """
    bucket = ca["Bucket"] or ""
    mapped_stage = BUCKET_TO_STAGE.get(bucket, "")
    if not mapped_stage and ca["Source"] == "On Air":
        mapped_stage = "ROE Secured"

    rp = (ca.get("ROE_PAL") or "").upper()
    is_pal = "PAL" in rp and "ROE" not in rp  # e.g. "PAL" vs "Ting / FF / ROE"

    if is_pal and sf_opp and sf_opp.get("StageName"):
        # PAL exception: keep SF's current stage
        proposed_stage = sf_opp["StageName"]
        pal_override = True
    else:
        proposed_stage = mapped_stage
        pal_override = False

    # Units — prefer CA MDU, fallback market
    units = ca.get("Units")
    if units is None or (isinstance(units, float) and pd.isna(units)):
        units = mkt.get("Units") if mkt else None

    # Agreement type — On Air has ROE_PAL filled; else inferred from bucket
    agreement_type = ""
    if ca.get("ROE_PAL"):
        # "ROE", "PAL", "Ting / FF" variants
        rp = ca["ROE_PAL"].upper()
        if "ROE" in rp:
            agreement_type = "ROE"
        elif "PAL" in rp:
            agreement_type = "PAL"
    elif "Access Agreement Complete" in bucket:
        agreement_type = "ROE"  # best guess; user may override

    # Signed date — from market's PALSignedDate if present
    signed_date = mkt.get("PALSignedDate") if mkt else None

    # RE Assigned — default Justin Barry
    re_assigned_initials = (mkt.get("RE_Assigned") if mkt else "") or "Justin Barry"

    # Combined Note
    note_parts = []
    note_parts.append(f"Source: CA MDU Agreement Status ({ca['Source']}) + Opportunities_Market")
    note_parts.append(f"CA Pipeline Bucket: {bucket or '(none)'}")
    if ca.get("OnNet"): note_parts.append(f"On Net: {ca['OnNet']}")
    if ca.get("ISPTenant"): note_parts.append(f"ISP Tenant: {ca['ISPTenant']}")
    if ca.get("ROE_PAL"): note_parts.append(f"Agreement (CA MDU col): {ca['ROE_PAL']}")
    if mkt:
        if mkt.get("UniqueID"): note_parts.append(f"Market UniqueID: {mkt['UniqueID']}")
        if mkt.get("Status"): note_parts.append(f"Market Status: {mkt['Status']}")
        if mkt.get("RE_Assigned"): note_parts.append(f"RE Assigned: {mkt['RE_Assigned']}")
        if mkt.get("Contacts"): note_parts.append(f"Contacts: {mkt['Contacts']}")
        if mkt.get("PALSignedDate"): note_parts.append(f"PAL Signed Date: {mkt['PALSignedDate']}")
        if mkt.get("SiteTrackerLink"): note_parts.append(f"SiteTracker: {mkt['SiteTrackerLink']}")
        if mkt.get("EstimateLink"): note_parts.append(f"Estimate Link: {mkt['EstimateLink']}")
        if mkt.get("CXNotes"): note_parts.append(f"CX Notes: {mkt['CXNotes']}")
        if mkt.get("FDHActivated"): note_parts.append(f"FDH Activated: {mkt['FDHActivated']}")
        if mkt.get("VetroStatus"): note_parts.append(f"Vetro Status: {mkt['VetroStatus']}")
        if mkt.get("Comments"):
            note_parts.append(f"\nComments:\n{mkt['Comments']}")
    combined_note = "\n".join(note_parts)

    # Action
    if sf_opp:
        action = "UPDATE"
        sf_id = sf_opp["Id"]
        sf_name = sf_opp["Name"]
        sf_stage = sf_opp.get("StageName") or ""
    else:
        action = "CREATE"
        sf_id = ""
        sf_name = ""
        sf_stage = ""

    issues = []
    if sf_opp and sf_opp.get("_multi"):
        issues.append(f"MULTI-MATCH: {len(sf_opp['candidates'])} candidates")
    if not bucket:
        issues.append("No CA bucket")
    if not proposed_stage:
        issues.append(f"No stage mapping for bucket {bucket!r}")
    if agreement_type and not signed_date:
        issues.append(f"Agreement type {agreement_type} proposed but no signed date (placeholder needed)")
    if not mkt:
        issues.append("No market-sheet match (note will be from CA MDU only)")
    if pal_override:
        issues.append(f"PAL exception — keeping SF stage {sf_opp.get('StageName')!r} (mapping would have been {mapped_stage})")
    # Flag reopens (Closed Lost -> open stage) so user can spot-check
    if sf_opp and (sf_opp.get("StageName") == "Closed Lost") and proposed_stage and proposed_stage != "Closed Lost":
        issues.append(f"REOPEN: Closed Lost -> {proposed_stage}")
    # Flag other regressions
    cur_rank = STAGE_ORDER.index(sf_opp["StageName"]) if (sf_opp and sf_opp.get("StageName") in STAGE_ORDER) else -1
    new_rank = STAGE_ORDER.index(proposed_stage) if proposed_stage in STAGE_ORDER else -1
    if sf_opp and cur_rank >= 0 and new_rank >= 0 and new_rank < cur_rank and sf_opp.get("StageName") != "Closed Lost":
        issues.append(f"REGRESSION: {sf_opp['StageName']} -> {proposed_stage}")

    return {
        "Action": action,
        "SF Id": sf_id,
        "SF Name (current)": sf_name,
        "SF Stage (current)": sf_stage,
        "Proposed Stage": proposed_stage,
        "Proposed Bucket (new field)": bucket,
        "Proposed Agreement Type": agreement_type,
        "Proposed Signed Date": signed_date or "",
        "Proposed Owner": "Justin Barry",
        "RE Assigned": re_assigned_initials,
        "Units": units,
        "Property Type": ca.get("PropertyType") or (mkt.get("PropertyType") if mkt else ""),
        "Program/City": ca.get("Program") or (mkt.get("Program") if mkt else ""),
        "Address": ca["Address"] or (mkt.get("Address") if mkt else ""),
        "CA MDU Name": ca["Name"],
        "Market Name": mkt.get("Name") if mkt else "",
        "Market Contacts": mkt.get("Contacts") if mkt else "",
        "Combined Note (preview)": combined_note,
        "Issues": " | ".join(issues),
    }


# ── Excel output ────────────────────────────────────────────────────────────
def write_workbook(merged, sf_opps, ca_rows, market_rows, ca_to_market, sf_matches):
    wb = Workbook()

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(name, rows, columns):
        ws = wb.create_sheet(title=name)
        for j, c in enumerate(columns, 1):
            cell = ws.cell(row=1, column=j, value=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for i, r in enumerate(rows, 2):
            for j, c in enumerate(columns, 1):
                v = r.get(c, "")
                if isinstance(v, float) and pd.isna(v):
                    v = ""
                cell = ws.cell(row=i, column=j, value=v)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        # Column widths
        widths = {
            "Action": 10, "SF Id": 20, "SF Name (current)": 35, "SF Stage (current)": 18,
            "Proposed Stage": 18, "Proposed Bucket (new field)": 28,
            "Proposed Agreement Type": 14, "Proposed Signed Date": 14,
            "Proposed Owner": 14, "RE Assigned": 14, "Units": 8,
            "Property Type": 12, "Program/City": 14, "Address": 35,
            "CA MDU Name": 30, "Market Name": 30, "Market Contacts": 30,
            "Combined Note (preview)": 70, "Issues": 35,
        }
        for j, c in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(j)].width = widths.get(c, 18)
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    cols_main = [
        "Action", "SF Id", "SF Name (current)", "SF Stage (current)",
        "Proposed Stage", "Proposed Bucket (new field)",
        "Proposed Agreement Type", "Proposed Signed Date",
        "Proposed Owner", "RE Assigned", "Units",
        "Property Type", "Program/City", "Address",
        "CA MDU Name", "Market Name", "Market Contacts",
        "Combined Note (preview)", "Issues",
    ]

    # Merged tab (all rows)
    wb.remove(wb.active)
    make_sheet("Merged", merged, cols_main)
    make_sheet("Create", [r for r in merged if r["Action"] == "CREATE"], cols_main)
    make_sheet("Update", [r for r in merged if r["Action"] == "UPDATE"], cols_main)
    make_sheet("Issues", [r for r in merged if r["Issues"]], cols_main)

    # Schema tab — changes we need in SF before live run
    schema_rows = [
        {
            "Item": "New field on Opportunity",
            "Details": "API: Pipeline_Bucket__c  |  Type: Picklist  |  Label: 'CA Pipeline Bucket'",
            "Values": ", ".join(ALL_BUCKETS),
            "Why": "CA MDU tracks raw bucket distinct from standard Stage; keep both.",
        },
        {
            "Item": "Existing field Sales_Status__c — LEAVE ALONE",
            "Details": "Current values: 'Contact Pending', 'Reached Out - Pending Response'",
            "Values": "(do not reuse — different semantic)",
            "Why": "Sales_Status__c is for outreach state; Pipeline_Bucket__c is the imported CA MDU value.",
        },
        {
            "Item": "Justin Barry as Owner",
            "Details": "OwnerId = 005WR0000030RCzYAM (Justin Barry)",
            "Values": "Applied to every CREATE + UPDATE touched",
            "Why": "Per Koa: Justin owns all CA MDU records.",
        },
        {
            "Item": "Stage mapping",
            "Details": "See BUCKET_TO_STAGE in script",
            "Values": "Prospects→Prospecting | Proposal Sent→Engaged | On Net/Near Net/ON Air → ROE Secured",
            "Why": "MDU pipeline has no Closed Won; ON Air = ROE Secured.",
        },
        {
            "Item": "Agreement__c",
            "Details": "Create when ROE/PAL populated or bucket 'Access Agreement Complete'",
            "Values": "Type from ROE/PAL col | Status=Completed when signed | Signed_Date__c from market sheet",
            "Why": "Signed agreements should exist as Agreement__c children on the Opp.",
        },
        {
            "Item": "Notes",
            "Details": "One combined ContentNote per Opp with all merged context",
            "Values": "See 'Combined Note (preview)' column",
            "Why": "Per Koa: single combined note is fine.",
        },
        {
            "Item": "Name protection",
            "Details": "Never overwrite existing Opp.Name on UPDATE",
            "Values": "Name cleanup Phase 1 done on 21 records",
            "Why": "Protect cleaned names from stomp.",
        },
        {
            "Item": "SFU RecordType (open question)",
            "Details": "CA Pipeline has SFU rows; SFU RT exists but has 0 records in SF",
            "Values": f"SFU RT Id: {SFU_RECORD_TYPE_ID}",
            "Why": "Decide: land SFU rows under MDU RT or SFU RT?",
        },
    ]
    cols_schema = ["Item", "Details", "Values", "Why"]
    make_sheet("Schema", schema_rows, cols_schema)

    # Summary tab
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = "CA MDU Merge — Dry Run Summary"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws.merge_cells("A1:D1")

    total = len(merged)
    creates = sum(1 for r in merged if r["Action"] == "CREATE")
    updates = sum(1 for r in merged if r["Action"] == "UPDATE")
    issues = sum(1 for r in merged if r["Issues"])
    with_market = sum(1 for r in merged if r["Market Name"])

    stats = [
        ("Total rows in CA MDU file", total),
        ("  - CA Pipeline", sum(1 for r in merged if "CA Pipeline" in (r.get("_source") or ""))),
        ("  - On Air",      sum(1 for r in merged if "On Air" in (r.get("_source") or ""))),
        ("Rows matched to opportunities_market", with_market),
        ("Rows WITHOUT market match", total - with_market),
        ("Rows matched to existing SF Opportunity (UPDATE)", updates),
        ("Rows with NO SF match (CREATE)", creates),
        ("Rows flagged with issues", issues),
        ("", ""),
        ("Total existing CA MDU Opps in SF (for reference)", len(sf_opps)),
    ]
    for i, (k, v) in enumerate(stats, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=k.startswith(("Total", "Rows matched", "Rows WITHOUT", "Rows matched to existing", "Rows with")))
        ws.cell(row=i, column=2, value=v)

    # Bucket breakdown
    bucket_counts = Counter(r["Proposed Bucket (new field)"] for r in merged)
    r0 = len(stats) + 5
    ws.cell(row=r0, column=1, value="Proposed Bucket breakdown").font = Font(bold=True, size=13, color="1F4E78")
    for i, (k, v) in enumerate(sorted(bucket_counts.items(), key=lambda x: -x[1]), r0 + 1):
        ws.cell(row=i, column=1, value=k or "(blank)")
        ws.cell(row=i, column=2, value=v)

    # Stage breakdown
    stage_counts = Counter(r["Proposed Stage"] for r in merged)
    r1 = r0 + len(bucket_counts) + 3
    ws.cell(row=r1, column=1, value="Proposed Stage breakdown").font = Font(bold=True, size=13, color="1F4E78")
    for i, (k, v) in enumerate(sorted(stage_counts.items(), key=lambda x: -x[1]), r1 + 1):
        ws.cell(row=i, column=1, value=k or "(blank)")
        ws.cell(row=i, column=2, value=v)

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 15

    wb.save(OUTPUT_XLSX)
    print(f"\n[OK] Preview workbook written to:\n  {OUTPUT_XLSX}")


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("CA MDU + Market Merge — DRY RUN PREVIEW")
    print("=" * 70)

    print("\n1. Loading CA MDU Agreement Status...")
    ca_rows = load_ca_mdu()
    print(f"   {len(ca_rows)} rows (CA Pipeline + On Air)")

    print("\n2. Loading opportunities_market...")
    market_rows = load_market()
    print(f"   {len(market_rows)} rows")

    print("\n3. Matching CA MDU -> opportunities_market (by address/name)...")
    ca_to_market = match_ca_to_market(ca_rows, market_rows)
    matched = sum(1 for m in ca_to_market if m is not None)
    print(f"   {matched}/{len(ca_rows)} matched to market sheet")

    print("\n4. Connecting to Salesforce and pulling existing CA MDU Opps...")
    sf = Salesforce(username=SF_USERNAME, password=SF_PASSWORD, security_token=SF_TOKEN)

    # Pull CA MDU Opps (state=CA) plus any Opps in the 4 cities regardless of state (belt+suspenders)
    q = sf.query_all("""
        SELECT Id, Name, StageName, Property_Address__c, Property_City__c,
               Property_State__c, Agreement_Name__c, OwnerId, RecordType.DeveloperName
        FROM Opportunity
        WHERE RecordType.DeveloperName IN ('MDU','SFU')
          AND (Property_State__c = 'CA'
               OR Property_City__c IN ('Carlsbad','Encinitas','Oceanside','Solana Beach'))
    """)
    sf_opps = q["records"]
    print(f"   {len(sf_opps)} existing CA-area Opps loaded")

    print("\n5. Matching CA MDU rows -> existing SF Opps...")
    sf_matches = match_to_sf(ca_rows, market_rows, ca_to_market, sf_opps)
    updates = sum(1 for m in sf_matches if m is not None)
    print(f"   {updates} would UPDATE existing SF Opps")
    print(f"   {len(ca_rows) - updates} would CREATE new SF Opps")

    print("\n6. Building proposed records...")
    merged = []
    for i, ca in enumerate(ca_rows):
        mkt = market_rows[ca_to_market[i]] if ca_to_market[i] is not None else None
        sf_opp = sf_matches[i]
        rec = propose_record(ca, mkt, sf_opp)
        rec["_source"] = ca["Source"]
        merged.append(rec)

    print("\n7. Writing preview workbook...")
    write_workbook(merged, sf_opps, ca_rows, market_rows, ca_to_market, sf_matches)

    # Print inline summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    creates = sum(1 for r in merged if r["Action"] == "CREATE")
    updates = sum(1 for r in merged if r["Action"] == "UPDATE")
    issues = sum(1 for r in merged if r["Issues"])
    print(f"  Total rows:       {len(merged)}")
    print(f"  Would CREATE:     {creates}")
    print(f"  Would UPDATE:     {updates}")
    print(f"  Flagged issues:   {issues}")
    print()
    bc = Counter(r["Proposed Bucket (new field)"] for r in merged)
    print("  Bucket breakdown:")
    for k, v in bc.most_common():
        print(f"    {k or '(blank)'}: {v}")


if __name__ == "__main__":
    main()
