"""
Backfill SMB ROE Project building-level pursuits as Business_ROE Opportunities (2026-04-25).

Source: SMB_ROE_Project.xlsx, sheet "ROE_Tracking" (246 rows)
Target: Opportunity, RecordType=Business_ROE, linked to Property_Location

Mappings:
  RE Status (xlsx)                  → SF Stage             | SF Sales_Status
  ─────────────────────────────────────────────────────────────────────────────
  Research Completed                → Prospecting          | Research Completed
  Engaged                           → Engaged              | (blank)
  Assigned FiberFirst               → Engaged              | FF Sales - Tenant Interest Required
  Proposal Sent                     → Contract Negotiations| (blank)
  Pending Signature                 → Contract Negotiations| (blank)
  Completed                         → ROE Secured          | (blank)
  Closed - Lost                     → Closed Lost          | (blank, Loss_Reason=Other)
  Closed - Contact Info             → Closed Lost          | (blank, Loss_Reason=No Contact Info)
  Data Issue                        → Closed Lost          | (blank, Loss_Reason=Other, note added)

Naming: "ROE - {Business Buildings}"
Owner: RE_Assigned user (TF=Tanya Friese, JB=Justin Barry, RS=Rosemarie Shortino)
Property_Location__c: matched by Name = xlsx "Business Buildings" value (verified Friday)

Usage:
  python backfill_smb_roe_opps_2026-04-25.py --preview     (default, no writes)
  python backfill_smb_roe_opps_2026-04-25.py --apply       (executes + audit log)
"""
import sys, io, re, csv, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from simple_salesforce import Salesforce

ap = argparse.ArgumentParser()
ap.add_argument('--apply', action='store_true')
args = ap.parse_args()
APPLY = args.apply

XLSX = Path(r'C:\Users\cass\Work_Projects\SMB_ROE_Project.xlsx')
SHEET = 'ROE_Tracking'

sf = Salesforce(username=_SF["username"], password=_SF["password"], security_token=_SF["token"])
AUDIT_DIR = Path(r'C:\Users\cass\Work_Projects\SalesForce\audit_logs')
AUDIT_DIR.mkdir(parents=True, exist_ok=True)
TS = datetime.now().isoformat(timespec='seconds')
SCRIPT_NAME = 'backfill_smb_roe_opps_2026-04-25.py'

# ── Lookup tables ──
print("[Setup] Loading SF lookup tables")
RT = {r['DeveloperName']: r['Id'] for r in sf.query("SELECT Id, DeveloperName FROM RecordType WHERE SObjectType='Opportunity'")['records']}
print(f"  Record Types: {RT}")

# RE Assigned initials → User Id
USERS_BY_NAME = {u['Name']: u['Id'] for u in sf.query("SELECT Id, Name FROM User WHERE IsActive=true")['records']}
INITIALS_TO_USER = {
    'TF': USERS_BY_NAME.get('Tanya Friese'),
    'JB': USERS_BY_NAME.get('Justin Barry'),
    'RS': USERS_BY_NAME.get('Rosemarie Shortino'),
}
for k, v in INITIALS_TO_USER.items():
    print(f"  {k} → {v}")

# Property_Type cleanup map
PT_MAP = {
    'Strip Mall': 'Strip Mall',
    'Strip Center': 'Strip Mall',
    'Shopping Plaza': 'Strip Mall',
    'Westroads Mall': 'Strip Mall',
    'Strip Mall. 7 tenants, 2115 w guadalupe building attached': 'Strip Mall',
    'Business Park': 'Business Park',
    'Office': 'Office',
    'Office ': 'Office',
    'office': 'Office',
    'Office/Marketing': 'Office',
    'Industrial': 'Commercial / Business',
    'Medical': 'Commercial / Business',
    'Government / Municipal': 'Commercial / Business',
    'Old Nursing Home': 'Commercial / Business',
    'Other': 'Commercial / Business',
    'Mixed Use': 'Mixed Use',
    '': None,  # leave blank
}

# RE Status (xlsx) → (Stage, Sales_Status, Loss_Reason)
STATUS_MAP = {
    'Research Completed': ('Prospecting', 'Research Completed', None),
    'Engaged': ('Engaged', None, None),
    'Assigned FiberFirst': ('Engaged', 'FF Sales - Tenant Interest Required', None),
    'Proposal Sent': ('Contract Negotiations', None, None),
    'Pending Signature': ('Contract Negotiations', None, None),
    'Completed': ('ROE Secured', None, None),
    'Closed - Lost': ('Closed Lost', None, 'Other'),
    'Closed - Contact Info': ('Closed Lost', None, 'No Contact Info'),
    'Data Issue': ('Closed Lost', None, 'Other'),
}

# ── Load xlsx ──
print(f"\n[Load] {XLSX} sheet={SHEET}")
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[SHEET]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
data = []
for r in rows[1:]:
    if not any(c is not None for c in r):
        continue
    rec = dict(zip(header, r))
    data.append(rec)
print(f"  Loaded {len(data)} data rows")

# ── Pre-fetch all Property_Locations matching the xlsx Building names ──
print("\n[Match] Fetching Property_Locations")
building_names = sorted(set(r['Business Buildings'] for r in data if r.get('Business Buildings')))
print(f"  Unique building names in xlsx: {len(building_names)}")

UUID_RE = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
ZIP_SUFFIX_RE = re.compile(r'\s+\d{5}(?:-\d{4})?\s*$')

def normalize_addr(s):
    """Collapse whitespace, uppercase, strip."""
    if not s:
        return ''
    return re.sub(r'\s+', ' ', s.strip().upper())

def strip_zip(s):
    """Remove trailing 5-digit zip from an address string."""
    return ZIP_SUFFIX_RE.sub('', s).strip() if s else ''

# Pull ALL Property_Locations with relevant states (one query, then index multiple ways)
states_in_xlsx = set(r.get('State') for r in data if r.get('State'))
print(f"  States in xlsx: {sorted(states_in_xlsx)}")
state_clause = "','".join(s for s in states_in_xlsx if s)
all_pls = sf.query_all(f"SELECT Id, Name, City__c, State__c, Business_Base_Address__c, Business_Building_Id__c FROM Property_Location__c WHERE State__c IN ('{state_clause}')")['records']
print(f"  Pulled {len(all_pls)} Property_Locations in those states")

# Build several lookup tables
pl_by_name = {}            # exact Name
pl_by_normalized = {}      # normalized Name OR Business_Base_Address__c
pl_by_normalized_no_zip = {}  # normalized name with zip stripped
pl_by_business_id = {}     # Business_Building_Id__c

for pl in all_pls:
    if pl.get('Name'):
        pl_by_name[pl['Name']] = pl
        n = normalize_addr(pl['Name'])
        pl_by_normalized.setdefault(n, pl)
        pl_by_normalized_no_zip.setdefault(strip_zip(n), pl)
    if pl.get('Business_Base_Address__c'):
        n = normalize_addr(pl['Business_Base_Address__c'])
        pl_by_normalized.setdefault(n, pl)
        pl_by_normalized_no_zip.setdefault(strip_zip(n), pl)
    if pl.get('Business_Building_Id__c'):
        pl_by_business_id[pl['Business_Building_Id__c']] = pl

# Resolve via multiple strategies. Build a per-source-row lookup since UUIDs need
# to fall back to the Updated Business Address column.
data_pl_lookup = {}  # row index → matched PL

def find_pl_for_row(r):
    bn = r.get('Business Buildings') or ''
    # Strategy 1: exact name match
    if bn in pl_by_name:
        return pl_by_name[bn]
    # Strategy 2: normalized
    norm = normalize_addr(bn)
    if norm in pl_by_normalized:
        return pl_by_normalized[norm]
    # Strategy 3: normalized + zip stripped
    no_zip = strip_zip(norm)
    if no_zip in pl_by_normalized_no_zip:
        return pl_by_normalized_no_zip[no_zip]
    # Strategy 4: UUID — match against Business_Building_Id__c, OR fall back to Updated Business Address
    if UUID_RE.match(bn):
        if bn in pl_by_business_id:
            return pl_by_business_id[bn]
        # Fallback: use Updated Business Address
        uba = r.get('Updated Business Address') or ''
        if uba:
            uba_norm = normalize_addr(uba)
            if uba_norm in pl_by_normalized:
                return pl_by_normalized[uba_norm]
            uba_no_zip = strip_zip(uba_norm)
            if uba_no_zip in pl_by_normalized_no_zip:
                return pl_by_normalized_no_zip[uba_no_zip]
    # Strategy 5: try the Address field too (with zip stripped)
    addr = r.get('Address') or ''
    if addr:
        addr_norm = normalize_addr(addr)
        if addr_norm in pl_by_normalized:
            return pl_by_normalized[addr_norm]
        addr_no_zip = strip_zip(addr_norm)
        if addr_no_zip in pl_by_normalized_no_zip:
            return pl_by_normalized_no_zip[addr_no_zip]
    return None

matched_count = 0
unmatched_buildings = []
for r in data:
    pl = find_pl_for_row(r)
    if pl:
        matched_count += 1
        # Cache by Business Buildings for the proposed-record builder
        bn = r.get('Business Buildings') or ''
        data_pl_lookup[bn] = pl
    else:
        bn = r.get('Business Buildings') or '(blank)'
        unmatched_buildings.append(bn)

# Replace the old pl_map for downstream use (mapping building name → PL)
pl_map = data_pl_lookup
print(f"  Rows matched to a Property_Location: {matched_count} of {len(data)}")
unmatched = unmatched_buildings

unmatched = [b for b in building_names if b not in pl_map]
if unmatched:
    print(f"\n  ⚠ Unmatched buildings ({len(unmatched)}):")
    for b in unmatched[:20]:
        print(f"    {b}")
    if len(unmatched) > 20:
        print(f"    ...and {len(unmatched)-20} more")

# ── Pre-fetch existing Opps with names like ROE - {addr} (avoid duplicates) ──
existing_roe_names = set()
for o in sf.query_all("SELECT Name FROM Opportunity WHERE Name LIKE 'ROE - %'")['records']:
    existing_roe_names.add(o['Name'])
print(f"\n  Existing 'ROE - %' Opps in SF: {len(existing_roe_names)}")

# ── Build the proposed records ──
ZIP_RE = re.compile(r'\b(\d{5})(?:-\d{4})?\b')

proposed = []
skipped = []
for r in data:
    bn = r.get('Business Buildings')
    if not bn:
        skipped.append((r, 'no Business Buildings value'))
        continue
    pl = pl_map.get(bn)
    if not pl:
        skipped.append((r, f'no Property_Location match for {bn!r}'))
        continue
    # Choose a sensible name source — if Business Buildings is a UUID, use the matched PL's Name instead
    name_source = bn if not UUID_RE.match(bn) else (pl.get('Name') or r.get('Updated Business Address') or bn)
    # Strip trailing zip from name to keep it consistent with Property_Location naming
    name_source = strip_zip(name_source) if name_source else name_source
    new_name = f"ROE - {name_source}"
    if new_name in existing_roe_names:
        skipped.append((r, f'duplicate — Opp named {new_name!r} already exists'))
        continue

    # Map RE Status → Stage / Sales_Status / Loss_Reason
    re_status = r.get('RE Status')
    if not re_status or re_status not in STATUS_MAP:
        skipped.append((r, f'unmapped RE Status: {re_status!r}'))
        continue
    stage, sales_status, loss_reason = STATUS_MAP[re_status]

    # Property_Type cleanup
    raw_pt = (r.get('Property Type') or '').strip()
    pt = PT_MAP.get(raw_pt, PT_MAP.get(raw_pt.title())) if raw_pt else None
    if raw_pt and pt is None:
        # Unknown — fall back to Commercial / Business
        pt = 'Commercial / Business'

    # Owner from RE Assigned initials
    initials = (r.get('RE Assigned') or '').strip().upper()
    owner_id = INITIALS_TO_USER.get(initials)
    re_user_id = owner_id
    if not owner_id:
        # Fall back to current user (Cass) for unassigned
        owner_id = sf.User_id__c if hasattr(sf, 'User_id__c') else None
    if not owner_id:
        # Use cass1's user id
        owner_id = '005WR000002ieYTYAY'  # cass1 from connection.md

    # City / State / Zip
    city = (r.get('City') or pl.get('City__c') or '').strip().upper() if (r.get('City') or pl.get('City__c')) else None
    state = (r.get('State') or pl.get('State__c') or '').strip()
    addr = (r.get('Address') or '').strip()
    zip_match = ZIP_RE.search(addr) if addr else None
    zip_code = zip_match.group(1) if zip_match else None

    # CloseDate logic:
    #   Closed Lost / ROE Secured: use xlsx Closed Date if present, else today
    #   Open stages: today + 90 days (placeholder forecast)
    from datetime import timedelta
    today = datetime.now().date()
    if stage in ('Closed Lost', 'ROE Secured'):
        cd_xlsx = r.get('Closed Date')
        if cd_xlsx and hasattr(cd_xlsx, 'strftime'):
            close_date = cd_xlsx.strftime('%Y-%m-%d')
        elif cd_xlsx:
            close_date = str(cd_xlsx)[:10]
        else:
            close_date = str(today)
    else:
        close_date = str(today + timedelta(days=90))

    rec = {
        'Name': new_name,
        'RecordTypeId': RT['Business_ROE'],
        'Property_Location__c': pl['Id'],
        'StageName': stage,
        'OwnerId': owner_id,
        'CloseDate': close_date,
    }
    if re_user_id:
        rec['RE_Assigned__c'] = re_user_id
    if pt:
        rec['Property_Type__c'] = pt
    if city:
        rec['Property_City__c'] = city
    if state:
        rec['Property_State__c'] = state
    if zip_code:
        rec['Property_Zip__c'] = zip_code
    if r.get('Units'):
        rec['Units__c'] = r['Units']
    if sales_status:
        rec['Sales_Status__c'] = sales_status
    if loss_reason:
        rec['Loss_Reason__c'] = loss_reason
    if r.get('Closed Notes'):
        rec['Closed_Notes__c'] = str(r['Closed Notes'])[:32768]
    if r.get('FF Notes'):
        rec['FF_Notes__c'] = str(r['FF Notes'])[:32768]
    if r.get('Off Hold/Exp Date'):
        # Format as YYYY-MM-DD
        v = r['Off Hold/Exp Date']
        rec['Off_Hold_Date__c'] = v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else str(v)[:10]
    if r.get('Assign to FF Sales Date'):
        v = r['Assign to FF Sales Date']
        rec['Sales_Handoff_Date__c'] = v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else str(v)[:10]
    if r.get('Existing Fiber Provider'):
        rec['Incumbent_Provider__c'] = str(r['Existing Fiber Provider'])[:255]

    # Build Closed_Notes addendum for Data Issue / various contexts
    notes_parts = []
    if re_status == 'Data Issue':
        notes_parts.append('SOURCE NOTE: imported from xlsx with RE Status=Data Issue')
    if r.get('RE Notes'):
        notes_parts.append(f'RE Notes: {r["RE Notes"]}')
    if r.get('RE Notes2'):
        notes_parts.append(f'RE Notes (alt): {r["RE Notes2"]}')
    if r.get('Build  Notes'):
        notes_parts.append(f'Build Notes: {r["Build  Notes"]}')
    if notes_parts:
        existing = rec.get('Closed_Notes__c', '')
        rec['Closed_Notes__c'] = (existing + '\n\n' + '\n\n'.join(notes_parts))[:32768]

    proposed.append((rec, r))

# ── Summary ──
print("\n" + "=" * 70)
print("BACKFILL PLAN SUMMARY")
print("=" * 70)
print(f"  Source rows:            {len(data)}")
print(f"  Proposed for create:    {len(proposed)}")
print(f"  Skipped:                {len(skipped)}")

if skipped:
    print(f"\n  Skip reasons:")
    reasons = Counter(s[1].split(':')[0].strip() for s in skipped)
    for r, c in reasons.most_common():
        print(f"    {c:3d}  {r}")

print(f"\n  Stage distribution of planned creates:")
for s, c in Counter(p[0]['StageName'] for p in proposed).most_common():
    print(f"    {c:4d}  {s}")

print(f"\n  Owner distribution:")
owner_names = {v: k for k, v in INITIALS_TO_USER.items() if v}
owner_names['005WR000002ieYTYAY'] = 'Cass (fallback)'
for o, c in Counter(p[0].get('OwnerId') for p in proposed).most_common():
    print(f"    {c:4d}  {owner_names.get(o, o)}")

print(f"\n  First 3 proposed records (full payload):")
import json

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()

for p, _src in proposed[:3]:
    print(f"\n  --- {p['Name']} ---")
    for k, v in p.items():
        sval = str(v)[:80] + ('...' if len(str(v)) > 80 else '')
        print(f"    {k:25s} = {sval}")

if not APPLY:
    print(f"\n[Preview only — re-run with --apply to insert {len(proposed)} Opp records]")
    sys.exit(0)

# ── Apply ──
print("\n" + "=" * 70)
print("APPLYING — creating Opps in batches of 200")
print("=" * 70)

audit_rows = []
created_ids = []
errors = []

records_to_create = [p[0] for p in proposed]
for i in range(0, len(records_to_create), 200):
    batch = records_to_create[i:i+200]
    print(f"  Batch {i//200 + 1}: {len(batch)} records...")
    results = sf.bulk.Opportunity.insert(batch)
    for j, res in enumerate(results):
        rec = batch[j]
        if res.get('success'):
            created_ids.append(res['id'])
            audit_rows.append({
                'SF_Id': res['id'], 'Name': rec['Name'],
                'Field': '(created)', 'Before': '', 'After': 'Business_ROE Opp',
                'Source': SCRIPT_NAME, 'Timestamp': TS, 'Action': 'CREATE',
                'Note': f'Stage={rec["StageName"]}, RT=Business_ROE, PL={rec["Property_Location__c"]}',
            })
        else:
            errors.append((rec['Name'], res))
            print(f"    ⚠ FAILED: {rec['Name']} — {res.get('errors', res)}")

# Write audit log
audit_path = AUDIT_DIR / f'smb_roe_backfill_audit_{TS.replace(":","-")}.csv'
with audit_path.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['SF_Id','Name','Field','Before','After','Source','Timestamp','Action','Note'])
    w.writeheader()
    w.writerows(audit_rows)

print(f"\n✓ Created: {len(created_ids)}")
print(f"⚠ Errors:  {len(errors)}")
print(f"✓ Audit log: {audit_path} ({len(audit_rows)} rows)")
