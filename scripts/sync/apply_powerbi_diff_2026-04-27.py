"""
Apply only the actual diff between the 2026-04-27 PowerBI export and current SF state.

Avoids running the full production sync script's 45K idempotent upserts when only
~50 records actually changed. Same end state, surgical writes, full audit log.

Steps:
  1. Insert 3 new PLs
  2. Update 1 PL whose Circuit_ID changed
  3. Insert 2 new Units (linked to their parent PLs)
  4. Update Units whose activation fields changed (Activated, Activation/Deactivation Dates, etc.)
  5. Flag newly-stale PLs/Units with Import_Delete=true + dated note (precedent: 4/1/2026 batch)
  6. Audit log
  7. Archive the source xlsx

Usage:
  python apply_powerbi_diff_2026-04-27.py            # preview
  python apply_powerbi_diff_2026-04-27.py --apply
"""
import sys, io, re, csv, shutil, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from simple_salesforce import Salesforce

import sys as _sys
_sys.path.insert(0, r"C:\Users\cass\Work_Projects")
from _shared.sf_auth import creds as _sf_creds  # single source of truth for SF creds
_SF = _sf_creds()


ap = argparse.ArgumentParser()
ap.add_argument('--apply', action='store_true')
args = ap.parse_args()
APPLY = args.apply

XLSX = Path(r'C:\Users\cass\Work_Projects\SalesForce\PowerBI_Report\data - 2026-04-27T103900.191.xlsx')
ARCHIVE_DIR = Path(r'C:\Users\cass\Work_Projects\SalesForce\PowerBI_Report\Previously_Imported')
AUDIT_DIR = Path(r'C:\Users\cass\Work_Projects\SalesForce\audit_logs')
AUDIT_DIR.mkdir(parents=True, exist_ok=True)
SCRIPT_NAME = 'apply_powerbi_diff_2026-04-27.py'
TS = datetime.now().isoformat(timespec='seconds')
TODAY = datetime.now().date()
STALE_NOTE = f'Not in PowerBI export as of {TODAY.month}/{TODAY.day}/{TODAY.year}. Flagged for review.'
NOW_UTC = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')

sf = Salesforce(username=_SF["username"], password=_SF["password"], security_token=_SF["token"])


def to_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None


def to_date(v):
    if v is None: return None
    if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
    s = str(v).strip()[:10]
    return s if s else None


def normalize_addr(v):
    s = to_str(v)
    return re.sub(r'\s+', ' ', s) if s else None


# ── Load xlsx ──
print(f'[Load] {XLSX.name}')
wb = load_workbook(XLSX, data_only=True)
ws = wb['Sheet1']
headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
hi = {h: i for i, h in enumerate(headers) if h}

export_pls = {}   # bba -> dict of fields
export_units = {} # cid -> dict of fields with parent_bba

for row_num in range(4, ws.max_row + 1):
    cell = lambda h: ws.cell(row=row_num, column=hi[h] + 1).value if h in hi else None
    bba = normalize_addr(cell('Business Base Address'))
    cid = to_str(cell('Circuit ID'))
    if not bba or not cid:
        continue

    if bba not in export_pls:
        export_pls[bba] = {
            'Business_Base_Address__c': bba,
            'Name': bba[:80],
            'Market__c': to_str(cell('Market')),
            'State__c': to_str(cell('State')),
            'FDH_Activated_Date__c': to_date(cell('FDH Activated Date')),
            'FDH_Name__c': to_str(cell('FDH Name')),
            'Serving_Area__c': to_str(cell('Serving Area')),
            'City__c': to_str(cell('City')),
            'Business_Building_Id__c': to_str(cell('Business Building Id')),
            'Circuit_ID__c': cid,
            'Import_DateTime__c': NOW_UTC,
        }

    export_units[cid] = {
        'Circuit_ID__c': cid,
        'Name': (to_str(cell('Address')) or '')[:80],
        'Unit__c': to_str(cell('Unit #')),
        'Activated__c': to_str(cell('Activated')),
        'Address_Activation_Date__c': to_date(cell('Address Activation Date')),
        'ValidForFF__c': to_str(cell('ValidForFF')),
        'Address_De_activation_Date__c': to_date(cell('Address De-activation Date')),
        'Address_Deactivated__c': to_str(cell('Address Deactivated')),
        'Ordered_Product__c': to_str(cell('Ordered Product')),
        'AreaId__c': to_str(cell('AreaId')),
        '_parent_bba': bba,
        'Import_DateTime__c': NOW_UTC,
    }
wb.close()
print(f'  {len(export_pls)} unique PLs, {len(export_units)} units in export')

# ── Pull SF state ──
print('\n[SF] Loading current state...')
sf_pls = sf.query_all("""
  SELECT Id, Business_Base_Address__c, Market__c, State__c, City__c, Circuit_ID__c,
         FDH_Name__c, Serving_Area__c, FDH_Activated_Date__c, Business_Building_Id__c,
         Import_Delete_Property__c
  FROM Property_Location__c
""")['records']
sf_pl_by_bba = {p['Business_Base_Address__c']: p for p in sf_pls if p.get('Business_Base_Address__c')}
print(f'  PLs: {len(sf_pls)}')

sf_units = sf.query_all("""
  SELECT Id, Name, Circuit_ID__c, Unit__c, Activated__c, Address_Activation_Date__c,
         Address_De_activation_Date__c, Address_Deactivated__c, Ordered_Product__c,
         ValidForFF__c, Property_Location__c, Property_Location__r.Business_Base_Address__c,
         Import_Delete_Unit__c
  FROM Property_Unit__c
""")['records']
sf_unit_by_cid = {u['Circuit_ID__c']: u for u in sf_units if u.get('Circuit_ID__c')}
print(f'  Units: {len(sf_units)}')

# ── Plan ──
PL_DIFF_FIELDS = ['Market__c','State__c','City__c','Circuit_ID__c','FDH_Name__c',
                  'Serving_Area__c','FDH_Activated_Date__c','Business_Building_Id__c']
UNIT_DIFF_FIELDS = ['Activated__c','Address_Activation_Date__c','Address_De_activation_Date__c',
                    'Address_Deactivated__c','Ordered_Product__c','ValidForFF__c','Unit__c','Name']

new_pls, pl_updates = [], []
for bba, ep in export_pls.items():
    sp = sf_pl_by_bba.get(bba)
    if not sp:
        new_pls.append(ep)
        continue
    diffs = {k: ep.get(k) for k in PL_DIFF_FIELDS if (sp.get(k) or ep.get(k)) and sp.get(k) != ep.get(k)}
    if diffs:
        diffs['Id'] = sp['Id']
        diffs['_bba'] = bba
        diffs['_old'] = {k: sp.get(k) for k in PL_DIFF_FIELDS}
        pl_updates.append(diffs)

new_units, unit_updates = [], []
for cid, eu in export_units.items():
    su = sf_unit_by_cid.get(cid)
    if not su:
        new_units.append(eu)
        continue
    diffs = {k: eu.get(k) for k in UNIT_DIFF_FIELDS if (su.get(k) or eu.get(k)) and su.get(k) != eu.get(k)}
    if diffs:
        diffs['Id'] = su['Id']
        diffs['_cid'] = cid
        diffs['_old'] = {k: su.get(k) for k in UNIT_DIFF_FIELDS}
        unit_updates.append(diffs)

# Newly-stale (in SF, not in export, not yet flagged)
sf_only_pls = [p for p in sf_pls
               if p.get('Business_Base_Address__c') not in export_pls
               and not p.get('Import_Delete_Property__c')]
sf_only_units = [u for u in sf_units
                 if u.get('Circuit_ID__c') not in export_units
                 and not u.get('Import_Delete_Unit__c')]

# ── Summary ──
print('\n' + '='*70)
print('SURGICAL APPLY PLAN')
print('='*70)
print(f'  Insert new PLs:                {len(new_pls)}')
print(f'  Update PLs (field diffs):      {len(pl_updates)}')
print(f'  Insert new Units:              {len(new_units)}')
print(f'  Update Units (field diffs):    {len(unit_updates)}')
print(f'  Flag newly-stale PLs:          {len(sf_only_pls)}')
print(f'  Flag newly-stale Units:        {len(sf_only_units)}')

print(f'\n  --- New PLs ---')
for p in new_pls:
    print(f'    {p["State__c"]} {p["Business_Base_Address__c"][:60]}')
print(f'\n  --- PL updates ---')
for u in pl_updates:
    changes = {k: v for k, v in u.items() if not k.startswith('_') and k != 'Id'}
    print(f'    {u["_bba"][:55]}: {changes}')
print(f'\n  --- New Units ---')
for u in new_units:
    print(f'    cid={u["Circuit_ID__c"]} parent={u["_parent_bba"][:50]} name={u.get("Name")[:40]}')
print(f'\n  --- Newly-stale PLs ---')
for p in sf_only_pls:
    print(f'    {p["Business_Base_Address__c"][:60]}')
print(f'\n  --- Newly-stale Units ---')
for u in sf_only_units:
    parent = u.get('Property_Location__r') or {}
    print(f'    cid={u["Circuit_ID__c"]} parent={parent.get("Business_Base_Address__c")}')

print(f'\n  Activation-field unit changes breakdown:')
field_count = Counter()
for u in unit_updates:
    for k in UNIT_DIFF_FIELDS:
        if k in u: field_count[k] += 1
for k, c in field_count.most_common():
    print(f'    {c:5d}  {k}')

if not APPLY:
    print(f'\n[Preview only — re-run with --apply to write]')
    sys.exit(0)

# ── Apply ──
print('\n' + '='*70)
print('APPLYING')
print('='*70)
audit_rows = []

def log(sf_id, name, field, before, after, action, note=''):
    audit_rows.append({
        'SF_Id': sf_id, 'Name': name, 'Field': field,
        'Before': before, 'After': after,
        'Source': SCRIPT_NAME, 'Timestamp': TS, 'Action': action, 'Note': note,
    })

# 1. New PLs
print(f'\n[1/6] Creating {len(new_pls)} new PLs')
if new_pls:
    rec_for_insert = [{k: v for k, v in p.items() if not k.startswith('_')} for p in new_pls]
    res = sf.bulk.Property_Location__c.insert(rec_for_insert)
    new_pl_id_by_bba = {}
    for r, p in zip(res, new_pls):
        if r.get('success'):
            new_pl_id_by_bba[p['Business_Base_Address__c']] = r['id']
            log(r['id'], p['Business_Base_Address__c'], '(created)', '', 'PL created', 'CREATE')
        else:
            print(f"   ⚠ FAIL: {p['Business_Base_Address__c']} -- {r.get('errors', r)}")
else:
    new_pl_id_by_bba = {}

# 2. Update PLs
print(f'\n[2/6] Updating {len(pl_updates)} PLs')
if pl_updates:
    rec = [{k: v for k, v in u.items() if not k.startswith('_')} for u in pl_updates]
    res = sf.bulk.Property_Location__c.update(rec)
    for r, u in zip(res, pl_updates):
        if r.get('success'):
            for k in PL_DIFF_FIELDS:
                if k in u:
                    log(u['Id'], u['_bba'], k, u['_old'].get(k), u.get(k), 'UPDATE')
        else:
            print(f"   ⚠ FAIL: {u['_bba']} -- {r.get('errors', r)}")

# 3. New Units (need parent PL Id)
print(f'\n[3/6] Creating {len(new_units)} new Units')
if new_units:
    rec_for_insert = []
    for u in new_units:
        parent_id = new_pl_id_by_bba.get(u['_parent_bba'])
        if not parent_id:
            sp = sf_pl_by_bba.get(u['_parent_bba'])
            if sp: parent_id = sp['Id']
        if not parent_id:
            print(f"   ⚠ Skipping Unit {u['Circuit_ID__c']} — no parent PL Id for {u['_parent_bba']!r}")
            continue
        new_rec = {k: v for k, v in u.items() if not k.startswith('_')}
        new_rec['Property_Location__c'] = parent_id
        rec_for_insert.append((new_rec, u))
    if rec_for_insert:
        recs = [r[0] for r in rec_for_insert]
        res = sf.bulk.Property_Unit__c.insert(recs)
        for r, (rec, u) in zip(res, rec_for_insert):
            if r.get('success'):
                log(r['id'], u['Circuit_ID__c'], '(created)', '', 'Unit created', 'CREATE',
                    note=f'parent_bba={u["_parent_bba"]}')
            else:
                print(f"   ⚠ FAIL: {u['Circuit_ID__c']} -- {r.get('errors', r)}")

# 4. Update Units
print(f'\n[4/6] Updating {len(unit_updates)} Units')
if unit_updates:
    rec = [{k: v for k, v in u.items() if not k.startswith('_')} for u in unit_updates]
    for i in range(0, len(rec), 200):
        batch = rec[i:i+200]
        ucut = unit_updates[i:i+200]
        res = sf.bulk.Property_Unit__c.update(batch)
        for r, u in zip(res, ucut):
            if r.get('success'):
                for k in UNIT_DIFF_FIELDS:
                    if k in u:
                        log(u['Id'], u['_cid'], k, u['_old'].get(k), u.get(k), 'UPDATE')
            else:
                print(f"   ⚠ FAIL: {u['_cid']} -- {r.get('errors', r)}")

# 5. Flag newly-stale PLs
print(f'\n[5/6] Flagging {len(sf_only_pls)} newly-stale PLs')
if sf_only_pls:
    rec = [{'Id': p['Id'], 'Import_Delete_Property__c': True, 'Import_Delete_Note__c': STALE_NOTE} for p in sf_only_pls]
    res = sf.bulk.Property_Location__c.update(rec)
    for r, p in zip(res, sf_only_pls):
        if r.get('success'):
            log(p['Id'], p['Business_Base_Address__c'], 'Import_Delete_Property__c', False, True, 'FLAG_STALE', note=STALE_NOTE)
        else:
            print(f"   ⚠ FAIL: {p['Business_Base_Address__c']} -- {r.get('errors', r)}")

# 6. Flag newly-stale Units
print(f'\n[6/6] Flagging {len(sf_only_units)} newly-stale Units')
if sf_only_units:
    rec = [{'Id': u['Id'], 'Import_Delete_Unit__c': True, 'Import_Delete_Note__c': STALE_NOTE} for u in sf_only_units]
    res = sf.bulk.Property_Unit__c.update(rec)
    for r, u in zip(res, sf_only_units):
        if r.get('success'):
            log(u['Id'], u['Circuit_ID__c'], 'Import_Delete_Unit__c', False, True, 'FLAG_STALE', note=STALE_NOTE)
        else:
            print(f"   ⚠ FAIL: {u['Circuit_ID__c']} -- {r.get('errors', r)}")

# Audit log
audit_path = AUDIT_DIR / f'powerbi_diff_apply_{TS.replace(":","-")}.csv'
with audit_path.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['SF_Id','Name','Field','Before','After','Source','Timestamp','Action','Note'])
    w.writeheader()
    w.writerows(audit_rows)
print(f'\n✓ Audit log: {audit_path} ({len(audit_rows)} rows)')

# Archive the xlsx
ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
ts_str = datetime.now().strftime('%Y%m%d_%H%M%S')
archive_path = ARCHIVE_DIR / f'{XLSX.stem}_imported_{ts_str}{XLSX.suffix}'
shutil.move(str(XLSX), str(archive_path))
print(f'✓ Archived xlsx: {archive_path.name}')
