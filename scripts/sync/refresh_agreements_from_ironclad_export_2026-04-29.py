"""
Read fresh IronClad export, sync IronClad__c records, then update Agreement__c
Status / Signed_Date__c for any drift.

Match key: IronClad__c.Name == IronClad export's "Ironclad Id" (col 2 / "IC-XXXX")

Status mapping: IronClad stage -> SF Status__c
  create     -> Create
  review     -> Review
  sign       -> Sign
  completed  -> Completed
  archive    -> Archive
  paused     -> Paused
  cancelled  -> Cancelled

Run: python refresh_agreements_from_ironclad_export_2026-04-29.py [--apply]
"""

import sys
import csv
from collections import Counter
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from simple_salesforce import Salesforce

USERNAME = "cass1@ubiquitygp.com"
PASSWORD = "Hawaiian1984"
SECURITY_TOKEN = "IBSKT6CFUpSUJWxq1CMm0HkFC"

EXPORT = Path("C:/Users/cass/Downloads/ironclad_export_2026-05-05_165706_all.xlsx")
LOG_DIR = Path("C:/Users/cass/Work_Projects/SalesForce/audit_logs")
LOG_DIR.mkdir(parents=True, exist_ok=True)

APPLY = "--apply" in sys.argv

STAGE_TO_SF_STATUS = {
    "create": "Create",
    "review": "Review",
    "sign": "Sign",
    "completed": "Completed",
    "archive": "Archive",
    "paused": "Paused",
    "cancelled": "Cancelled",
}


def load_export():
    wb = load_workbook(EXPORT, read_only=True, data_only=True)
    ws = wb["export"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    col_idx = {name: i for i, name in enumerate(headers) if name}

    def g(row, name):
        i = col_idx.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    out = {}
    for row in rows:
        ic_id = g(row, "Ironclad Id")
        if not ic_id:
            continue
        out[ic_id] = {
            "stage": g(row, "Stage"),
            "contract_status": g(row, "Contract Status"),
            "executed_date": g(row, "Executed Date"),
            "effective_date": g(row, "Effective Date"),
            "workflow_completed_date": g(row, "Workflow Completed Date"),
            "agreement_date": g(row, "Agreement Date"),  # set even on evergreen/repository docs
            "agreename": g(row, "AgreeName"),
            "agreement_type": g(row, "Agreement Type"),
            "record_type": g(row, "Record Type"),
            "counterparty_signer": g(row, "Counterparty Signer Name"),
            "record_name": g(row, "Record Name"),
        }
    print(f"Export rows parsed: {len(out)}")
    return out


def fmt_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    # Already YYYY-MM-DD?
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return s


def main():
    sf = Salesforce(username=USERNAME, password=PASSWORD, security_token=SECURITY_TOKEN)
    export = load_export()

    # Pull MDU Agreements that have an IronClad_Record__c lookup.
    # Match the export by IronClad_Record__r.IronClad_Id__c (the real IronClad workflow ID),
    # NOT by Name (which is a SF auto-number).
    soql = """
        SELECT Id, Name, Agreement_Type__c, Status__c, Signed_Date__c,
               IronClad_ID__c, IronClad_Record__c, IronClad_Record__r.IronClad_Id__c,
               IronClad_Record__r.Record_Type_IC__c,
               Opportunity__c, Opportunity__r.Name, Opportunity__r.StageName,
               Opportunity__r.RecordType.DeveloperName
        FROM Agreement__c
        WHERE IronClad_Record__c != null
    """
    res = sf.query_all(soql)["records"]
    print(f"Agreements with IronClad lookup: {len(res)}")

    SUPPORTED_RTS = {"MDU", "Business_ROE"}

    diffs = []
    ic_type_diffs = {}  # IronClad__c.Id -> {ic_name, from, to} (dedup: many Agr can share one IC)
    no_match = 0
    no_change = 0
    skipped_other_rt = 0
    matched_ic_record_ids = set()  # IronClad__c IDs to stamp Last_Synced
    rt_counts = Counter()

    for a in res:
        rt = (a.get("Opportunity__r") or {}).get("RecordType", {}).get("DeveloperName")
        if rt not in SUPPORTED_RTS:
            skipped_other_rt += 1
            continue
        rt_counts[rt] += 1
        ic_name = (a.get("IronClad_Record__r") or {}).get("IronClad_Id__c")
        if not ic_name or ic_name not in export:
            no_match += 1
            continue
        # Track the IronClad__c parent for Last_Synced stamp regardless of whether status changed
        if a.get("IronClad_Record__c"):
            matched_ic_record_ids.add(a["IronClad_Record__c"])
        e = export[ic_name]

        # Detect drift on IronClad__c.Record_Type_IC__c (IC's authoritative agreement type).
        # Track once per IronClad__c parent so we don't double-queue when multiple Agreements share one IC record.
        cur_ic_type = (a.get("IronClad_Record__r") or {}).get("Record_Type_IC__c")
        new_ic_type = e.get("record_type")
        ic_id = a.get("IronClad_Record__c")
        if new_ic_type and new_ic_type != cur_ic_type and ic_id and ic_id not in ic_type_diffs:
            ic_type_diffs[ic_id] = {
                "ic_name": ic_name,
                "from": cur_ic_type,
                "to": new_ic_type,
            }

        stage = (e["stage"] or "").lower()
        target_status = STAGE_TO_SF_STATUS.get(stage)
        if not target_status:
            no_change += 1
            continue

        # IronClad is authoritative for Signed_Date — overwrite SF's manual entry whenever
        # IC has a date, regardless of current Status. Priority: Executed > Effective >
        # Workflow Completed > Agreement Date. Agreement Date catches evergreen/repository
        # docs (PDFs uploaded post-signing, which never have Executed/Effective populated).
        target_signed = (
            fmt_date(e.get("executed_date"))
            or fmt_date(e.get("effective_date"))
            or fmt_date(e.get("workflow_completed_date"))
            or fmt_date(e.get("agreement_date"))
        )

        cur_status = a.get("Status__c")
        cur_signed = a.get("Signed_Date__c")

        same_status = (cur_status == target_status)
        # signed_date: IC authoritative. Update if IC has a date and SF differs.
        # If IC has no date, leave SF's value alone.
        same_signed = (target_signed is None) or (cur_signed == target_signed)

        if same_status and same_signed:
            no_change += 1
            continue

        diffs.append({
            "agr_id": a["Id"],
            "agr_name": a["Name"],
            "agr_type": a["Agreement_Type__c"],
            "opp_name": (a.get("Opportunity__r") or {}).get("Name", ""),
            "opp_stage": (a.get("Opportunity__r") or {}).get("StageName", ""),
            "ic_name": ic_name,
            "from_status": cur_status,
            "to_status": target_status,
            "from_signed": cur_signed,
            "to_signed": target_signed,
        })

    print(f"\nNo IC match in export:        {no_match}")
    print(f"Skipped (other RT):           {skipped_other_rt}")
    print(f"By RT processed:              {dict(rt_counts)}")
    print(f"No change needed:             {no_change}")
    print(f"DIFFS to apply:               {len(diffs)}")
    print(f"IronClad type diffs:          {len(ic_type_diffs)}")
    print(f"IronClad__c Last_Synced stamps: {len(matched_ic_record_ids)}")

    if ic_type_diffs:
        print("\nIronClad Record Type drift (sample 15):")
        for ic_id, d in list(ic_type_diffs.items())[:15]:
            print(f"  {d['ic_name']:<10} {str(d['from'])[:35]:<35} -> {d['to']}")

    # Transition summary
    transitions = Counter((d["from_status"], d["to_status"]) for d in diffs)
    print("\nStatus transitions:")
    for (frm, to), c in transitions.most_common():
        print(f"  {str(frm):<14} -> {str(to):<14} {c}")

    # Sample diffs
    print("\nSample diffs (first 15):")
    for d in diffs[:15]:
        print(f"  {d['agr_name']}  {d['opp_name'][:35]:<35}  {d['agr_type']:<14} "
              f"{str(d['from_status']):<10} -> {str(d['to_status']):<10}  "
              f"signed: {d['from_signed']} -> {d['to_signed']}  IC={d['ic_name']}")

    # Audit
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    audit = LOG_DIR / f"agr_refresh_from_ironclad_{ts}.csv"
    with open(audit, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["SF_Id", "Agreement_Name", "Field", "Before", "After", "Source", "Action", "Timestamp"])
        action = "UPDATE" if APPLY else "PREVIEW"
        for d in diffs:
            w.writerow([d["agr_id"], d["agr_name"], "Status__c", d["from_status"], d["to_status"],
                        f"ironclad_export_2026-04-29 ({d['ic_name']})", action, datetime.now().isoformat()])
            if d["to_signed"] and d["to_signed"] != d["from_signed"]:
                w.writerow([d["agr_id"], d["agr_name"], "Signed_Date__c", d["from_signed"], d["to_signed"],
                            f"ironclad_export_2026-04-29 ({d['ic_name']})", action, datetime.now().isoformat()])
        for ic_id, d in ic_type_diffs.items():
            w.writerow([ic_id, d["ic_name"], "Record_Type_IC__c", d["from"], d["to"],
                        f"ironclad_export_2026-04-29 ({d['ic_name']})", action, datetime.now().isoformat()])
    print(f"\nAudit: {audit}")

    if not APPLY:
        print("\nPREVIEW only. Re-run with --apply to push.")
        return

    # Apply Agreement diffs
    ok = fail = 0
    for d in diffs:
        body = {"Status__c": d["to_status"]}
        if d["to_signed"] and d["to_signed"] != d["from_signed"]:
            body["Signed_Date__c"] = d["to_signed"]
        try:
            sf.Agreement__c.update(d["agr_id"], body)
            ok += 1
        except Exception as e:
            fail += 1
            print(f"  ! {d['agr_name']}: {e}")
    print(f"\nAgreement updates: ok={ok} fail={fail}")

    # Push IronClad type fixes one-by-one (small N, want clear errors)
    if ic_type_diffs:
        ict_ok = ict_fail = 0
        for ic_id, d in ic_type_diffs.items():
            try:
                sf.IronClad__c.update(ic_id, {"Record_Type_IC__c": d["to"]})
                ict_ok += 1
            except Exception as e:
                ict_fail += 1
                print(f"  ! IC type update {d['ic_name']}: {e}")
        print(f"IronClad Record Type updates: ok={ict_ok} fail={ict_fail}")

    # Stamp IronClad__c.Last_Synced__c via composite API (200/batch)
    import requests
    now_iso = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000Z')
    headers = {'Authorization': f'Bearer {sf.session_id}', 'Content-Type': 'application/json'}
    url = f'{sf.base_url}composite/sobjects'
    ic_ids = list(matched_ic_record_ids)
    ic_ok = ic_fail = 0
    for i in range(0, len(ic_ids), 200):
        chunk = ic_ids[i:i+200]
        records = [{'attributes': {'type': 'IronClad__c'}, 'Id': rid, 'Last_Synced__c': now_iso} for rid in chunk]
        r = requests.patch(url, headers=headers, json={'allOrNone': False, 'records': records}, timeout=120)
        if r.status_code == 200:
            for res in r.json():
                if res.get('success'):
                    ic_ok += 1
                else:
                    ic_fail += 1
                    print(f"  ! IC stamp failed: {res.get('errors')}")
        else:
            ic_fail += len(chunk)
            print(f"  ! batch HTTP {r.status_code}: {r.text[:200]}")
    print(f"IronClad__c Last_Synced stamps: ok={ic_ok} fail={ic_fail}")


if __name__ == "__main__":
    main()
