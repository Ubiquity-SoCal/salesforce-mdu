"""
Inspect the Encinitas PBI enrichment template file in detail.
Examines all sheets, headers, formulas, formatting, structure, etc.
"""
import openpyxl
from openpyxl.utils import get_column_letter
import json

FILE_PATH = r"C:\Users\cass\OneDrive - Ubiquity Management\Desktop\Ting\PowerBI_vs_Build\CompleteEncinitasView_PBI.xlsx"

# Load workbook twice: once for values, once for formulas
wb_values = openpyxl.load_workbook(FILE_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False)

print("=" * 80)
print("WORKBOOK OVERVIEW")
print("=" * 80)
print(f"File: {FILE_PATH}")
print(f"Sheet names: {wb_values.sheetnames}")
print()

# ============================================================
# EXAMINE EACH SHEET
# ============================================================
for sheet_name in wb_values.sheetnames:
    ws_val = wb_values[sheet_name]
    ws_form = wb_formulas[sheet_name]

    print("=" * 80)
    print(f"SHEET: '{sheet_name}'")
    print("=" * 80)
    print(f"  Dimensions: {ws_val.dimensions}")
    print(f"  Max row: {ws_val.max_row}, Max col: {ws_val.max_column}")
    print(f"  Min row: {ws_val.min_row}, Min col: {ws_val.min_column}")
    print()

    # --- Column widths ---
    print("  COLUMN WIDTHS:")
    for col_letter, col_dim in ws_val.column_dimensions.items():
        if col_dim.width is not None and col_dim.width != 8.43:  # 8.43 is default
            print(f"    Column {col_letter}: width={col_dim.width}, hidden={col_dim.hidden}, "
                  f"bestFit={col_dim.bestFit}, customWidth={col_dim.customWidth}")
    print()

    # --- Row heights ---
    print("  ROW HEIGHTS (non-default):")
    for row_num, row_dim in ws_val.row_dimensions.items():
        if row_dim.height is not None and row_dim.height != 15:
            print(f"    Row {row_num}: height={row_dim.height}, hidden={row_dim.hidden}")
    print()

    # --- Merged cells ---
    if ws_val.merged_cells.ranges:
        print("  MERGED CELLS:")
        for mc in ws_val.merged_cells.ranges:
            print(f"    {mc}")
        print()

    # --- Frozen panes ---
    if ws_val.freeze_panes:
        print(f"  FREEZE PANES: {ws_val.freeze_panes}")
        print()

    # --- Auto filters ---
    if ws_val.auto_filter and ws_val.auto_filter.ref:
        print(f"  AUTO FILTER: {ws_val.auto_filter.ref}")
        print()

    # --- Conditional formatting ---
    if ws_val.conditional_formatting:
        print("  CONDITIONAL FORMATTING:")
        for cf in ws_val.conditional_formatting:
            print(f"    Range: {cf}, Rules: {len(cf.rules)}")
            for rule in cf.rules:
                print(f"      Type: {rule.type}, Formula: {rule.formula}, "
                      f"Priority: {rule.priority}")
        print()

    # --- Data validations ---
    if ws_val.data_validations and ws_val.data_validations.dataValidation:
        print("  DATA VALIDATIONS:")
        for dv in ws_val.data_validations.dataValidation:
            print(f"    Range: {dv.sqref}, Type: {dv.type}, Formula1: {dv.formula1}")
        print()


# ============================================================
# POWERBI_DATA SHEET - DETAILED
# ============================================================
print("\n" + "=" * 80)
print("DETAILED: PowerBI_Data SHEET")
print("=" * 80)

for candidate in wb_values.sheetnames:
    if "power" in candidate.lower() or "pbi" in candidate.lower() or "data" in candidate.lower():
        data_sheet_name = candidate
        break
else:
    # Try the first sheet
    data_sheet_name = wb_values.sheetnames[0]

ws_dv = wb_values[data_sheet_name]
ws_df = wb_formulas[data_sheet_name]

print(f"Using sheet: '{data_sheet_name}'")
print(f"Rows: {ws_dv.max_row}, Cols: {ws_dv.max_column}")
print()

# Headers
print("COLUMN HEADERS (Row 1):")
for col in range(1, ws_dv.max_column + 1):
    val_cell = ws_dv.cell(row=1, column=col)
    form_cell = ws_df.cell(row=1, column=col)
    col_letter = get_column_letter(col)
    header = val_cell.value
    formula = form_cell.value
    is_formula = isinstance(formula, str) and formula.startswith("=")
    print(f"  Col {col_letter} ({col}): '{header}'"
          + (f"  [FORMULA: {formula}]" if is_formula else ""))
print()

# Sample data rows (rows 2-6) - show values AND formulas
print("SAMPLE DATA (Rows 2-6):")
for row in range(2, min(7, ws_dv.max_row + 1)):
    print(f"  --- Row {row} ---")
    for col in range(1, ws_dv.max_column + 1):
        val_cell = ws_dv.cell(row=row, column=col)
        form_cell = ws_df.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        value = val_cell.value
        formula = form_cell.value
        is_formula = isinstance(formula, str) and str(formula).startswith("=")
        header = ws_dv.cell(row=1, column=col).value
        if value is not None or is_formula:
            line = f"    {col_letter} ({header}): value='{value}'"
            if is_formula:
                line += f"  FORMULA='{formula}'"
            # formatting info
            font = val_cell.font
            if font.bold:
                line += " [BOLD]"
            print(line)
    print()

# Check for formulas in enriched columns across ALL rows
print("FORMULA CHECK IN ENRICHED COLUMNS (checking all rows):")
# Find the enriched columns by header name
enriched_headers = ["address type", "phase", "serviceable from date", "serviceable_from_date",
                    "serviceable date", "fdh", "fdh name", "fdh_name"]
enriched_cols = {}
for col in range(1, ws_dv.max_column + 1):
    header = str(ws_dv.cell(row=1, column=col).value or "").lower().strip()
    for eh in enriched_headers:
        if eh in header:
            enriched_cols[ws_dv.cell(row=1, column=col).value] = col
            break

print(f"  Enriched columns found: {enriched_cols}")
for header_name, col_idx in enriched_cols.items():
    col_letter = get_column_letter(col_idx)
    formula_count = 0
    static_count = 0
    sample_formulas = []
    sample_values = []
    for row in range(2, ws_dv.max_row + 1):
        form_cell = ws_df.cell(row=row, column=col_idx)
        val_cell = ws_dv.cell(row=row, column=col_idx)
        formula = form_cell.value
        value = val_cell.value
        if isinstance(formula, str) and formula.startswith("="):
            formula_count += 1
            if len(sample_formulas) < 3:
                sample_formulas.append((row, formula))
        elif value is not None:
            static_count += 1
            if len(sample_values) < 3:
                sample_values.append((row, value))
    print(f"\n  Column {col_letter} '{header_name}':")
    print(f"    Formula cells: {formula_count}, Static value cells: {static_count}")
    if sample_formulas:
        print(f"    Sample formulas:")
        for r, f in sample_formulas:
            print(f"      Row {r}: {f}")
    if sample_values:
        print(f"    Sample values:")
        for r, v in sample_values:
            print(f"      Row {r}: {v}")


# ============================================================
# SUMMARY SHEET - DETAILED
# ============================================================
print("\n\n" + "=" * 80)
print("DETAILED: Summary SHEET")
print("=" * 80)

for candidate in wb_values.sheetnames:
    if "summary" in candidate.lower():
        summary_sheet_name = candidate
        break
else:
    summary_sheet_name = None

if summary_sheet_name:
    ws_sv = wb_values[summary_sheet_name]
    ws_sf = wb_formulas[summary_sheet_name]
    print(f"Using sheet: '{summary_sheet_name}'")
    print(f"Rows: {ws_sv.max_row}, Cols: {ws_sv.max_column}")
    print()

    # Column widths
    print("COLUMN WIDTHS:")
    for col_letter, col_dim in ws_sv.column_dimensions.items():
        w = col_dim.width
        print(f"  Column {col_letter}: width={w}, hidden={col_dim.hidden}, "
              f"customWidth={col_dim.customWidth}")
    print()

    # Row-by-row analysis for first 40 rows
    print("ROW-BY-ROW ANALYSIS (all rows up to max):")
    max_check = min(ws_sv.max_row + 1, 100)
    for row in range(1, max_check):
        cells_info = []
        has_content = False
        for col in range(1, ws_sv.max_column + 1):
            val_cell = ws_sv.cell(row=row, column=col)
            form_cell = ws_sf.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            value = val_cell.value
            formula = form_cell.value
            is_formula = isinstance(formula, str) and str(formula).startswith("=")

            if value is not None or is_formula:
                has_content = True
                info = {
                    "col": col_letter,
                    "value": value,
                }
                if is_formula:
                    info["formula"] = formula

                # Font
                font = val_cell.font
                font_info = {}
                if font.bold:
                    font_info["bold"] = True
                if font.italic:
                    font_info["italic"] = True
                if font.size and font.size != 11:
                    font_info["size"] = font.size
                if font.name and font.name != "Calibri":
                    font_info["name"] = font.name
                if font.color and font.color.rgb and font.color.rgb != "00000000":
                    font_info["color"] = str(font.color.rgb)
                if font_info:
                    info["font"] = font_info

                # Alignment
                align = val_cell.alignment
                align_info = {}
                if align.horizontal:
                    align_info["horizontal"] = align.horizontal
                if align.vertical:
                    align_info["vertical"] = align.vertical
                if align.indent and align.indent > 0:
                    align_info["indent"] = align.indent
                if align.wrap_text:
                    align_info["wrap_text"] = True
                if align_info:
                    info["alignment"] = align_info

                # Fill
                fill = val_cell.fill
                if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb not in ("00000000", "0"):
                    info["fill_fg"] = str(fill.fgColor.rgb)
                if fill.bgColor and fill.bgColor.rgb and fill.bgColor.rgb not in ("00000000", "0"):
                    info["fill_bg"] = str(fill.bgColor.rgb)
                if fill.patternType and fill.patternType != "none":
                    info["fill_pattern"] = fill.patternType

                # Border
                border = val_cell.border
                border_info = {}
                for side_name in ["left", "right", "top", "bottom"]:
                    side = getattr(border, side_name)
                    if side and side.style:
                        border_info[side_name] = side.style
                if border_info:
                    info["border"] = border_info

                # Number format
                if val_cell.number_format and val_cell.number_format != "General":
                    info["number_format"] = val_cell.number_format

                cells_info.append(info)

        if has_content:
            print(f"\n  --- Row {row} ---")
            # Check row height
            if row in ws_sv.row_dimensions:
                rd = ws_sv.row_dimensions[row]
                if rd.height and rd.height != 15:
                    print(f"    [Row height: {rd.height}]")
            for ci in cells_info:
                parts = [f"    {ci['col']}: value={repr(ci['value'])}"]
                if "formula" in ci:
                    parts.append(f"formula={repr(ci['formula'])}")
                if "font" in ci:
                    parts.append(f"font={ci['font']}")
                if "alignment" in ci:
                    parts.append(f"align={ci['alignment']}")
                if "fill_fg" in ci:
                    parts.append(f"fill_fg={ci['fill_fg']}")
                if "fill_bg" in ci:
                    parts.append(f"fill_bg={ci['fill_bg']}")
                if "fill_pattern" in ci:
                    parts.append(f"fill_pattern={ci['fill_pattern']}")
                if "border" in ci:
                    parts.append(f"border={ci['border']}")
                if "number_format" in ci:
                    parts.append(f"numfmt={ci['number_format']}")
                print("  ".join(parts))

    # Check for any named ranges
    print("\n\nNAMED RANGES:")
    for name in wb_values.defined_names.definedName:
        print(f"  {name.name}: {name.attr_text}")

    # Tables
    print("\nTABLES IN SHEETS:")
    for sn in wb_values.sheetnames:
        ws = wb_values[sn]
        if ws.tables:
            for table_name, table in ws.tables.items():
                print(f"  Sheet '{sn}': Table '{table_name}' ref={table.ref}")

else:
    print("No Summary sheet found!")

# ============================================================
# CHECK ALL SHEETS FOR EXTRA INFO
# ============================================================
print("\n\n" + "=" * 80)
print("ALL SHEETS - QUICK SCAN")
print("=" * 80)
for sn in wb_values.sheetnames:
    ws = wb_values[sn]
    print(f"\n  Sheet '{sn}': {ws.max_row} rows x {ws.max_column} cols")
    # Show first row (headers)
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            headers.append(f"{get_column_letter(col)}:{v}")
    print(f"    Headers: {headers}")

wb_values.close()
wb_formulas.close()
print("\nDone.")
